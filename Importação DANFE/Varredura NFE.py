import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import sys
from difflib import SequenceMatcher
import unicodedata
import xml.etree.ElementTree as ET
import re
from glob import glob
from collections import defaultdict, deque
from pathlib import Path
import json

# ─────────────────────────────────────────────────────────────
# Força stdout/stderr em UTF-8 (Windows não dá mole…)
# ─────────────────────────────────────────────────────────────
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# =====================================================================
# TENTAR OPENPYXL
# =====================================================================
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ Aviso: openpyxl não instalado. Formatação de tabelas não estará disponível.")
    print("Instale com: pip install openpyxl")

# =====================================================================
# CONFIG / PERFIL
# =====================================================================
def _load_json_config():
    """
    Lê json/config.json (mesma pasta do script) e devolve:
    - base_dados_path
    - testes_path
    - active_owner (CLEUBER, GILSON, ADRIANA, LUCAS)
    """
    try:
        cfg_path = Path(__file__).parent / "json" / "config.json"
        if cfg_path.exists():
            with open(cfg_path, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            base_dados_path = cfg.get("base_dados_path")
            testes_path = cfg.get("testes_path")
            active_owner = (cfg.get("active_owner") or "CLEUBER").strip().upper()
            return base_dados_path, testes_path, active_owner
    except Exception as e:
        print(f"⚠️ Aviso: falha ao ler config.json: {e}")
    # fallback
    return None, None, "CLEUBER"

def _resolve_paths_and_owner():
    """
    Ordem de prioridade:
      1) python script.py BASE TESTES DONO [NOTAS]
      2) python script.py BASE TESTES [NOTAS]  -> DONO vem do config
      3) nada na linha de comando -> tudo do config ou fallback
    """
    cfg_base, cfg_testes, cfg_owner = _load_json_config()

    base = cfg_base or r"\\rilkler\\LIVRO CAIXA\\TESTE\\BASE DE DADOS.xlsx"
    testes = cfg_testes or r"\\rilkler\\LIVRO CAIXA\\TESTE\\TESTES.xlsx"
    owner = (cfg_owner or "CLEUBER").strip().upper()

    argc = len(sys.argv)

    if argc >= 2:
        base = sys.argv[1]
    if argc >= 3:
        testes = sys.argv[2]
    if argc >= 4:
        cand = sys.argv[3].strip()
        # 👇 se o 3º argumento for um caminho de arquivo, NÃO é dono
        if cand.lower().endswith(".xlsx") or "\\" in cand or "/" in cand:
            # ignora como dono, mantém o do config
            pass
        else:
            owner = cand.upper()

    return base, testes, owner

def _resolve_notas_recebidas_path(testes_path: str) -> str | None:
    """
    Tenta obter o caminho do arquivo 'NOTAS RECEBIDAS.xlsx'.
    Ordem:
      1) 4º argumento do script
      2) mesmo diretório do testes_path
    """
    try:
        if len(sys.argv) >= 5 and os.path.exists(sys.argv[4]):
            return sys.argv[4]
    except Exception:
        pass
    try:
        cand = Path(testes_path).parent / "NOTAS RECEBIDAS.xlsx"
        if cand.exists():
            return str(cand)
    except Exception:
        pass
    return None

# =====================================================================
# MAPEAMENTOS DE CONTAS SEPARADOS POR PRODUTOR
# =====================================================================
# CLEUBER
MAP_CONTAS_CLEUBER = {
    "Caixa Geral": "001",
    "Cheques a Compensar": "001",
    "Fundo Fixo - Gildevan": "001",
    "Fundo Fixo - Cleidson Alves": "001",
    "Fundo Fixo - Rodrigo": "001",
    "Fundo Fixo - Wandres": "001",
    "Fundo Fixo - Cezar Dias": "001",
    "Fundo Fixo - Geraldo": "001",
    "Fundo Fixo - Daniel": "001",
    "Fundo Fixo - Hadlaim": "001",
    "Fundo Fixo - Lourival": "001",
    "Fundo Fixo - Rogeris": "001",
    "Fundo Fixo - Joaquim": "001",
    "Caixa Dobrado": "001",
    "Fundo Fxo - Douglas": "001",
    "Fundo Fixo - Samuel": "001",
    "Fundo Fixo - Adarildo": "001",
    "Fundo Fixo - Fabricio": "001",
    "Fundo Fixo - Fernando": "001",
    "Fundo Fixo - Orivan": "001",
    "Fundo Fixo - Saimon": "001",
    "Fundo Fxo - Eduardo": "001",
    "Fundo Fixo - Melquiades": "001",
    "Fundo Fixo - Anivaldo": "001",
    "Fundo Fixo - Cida": "001",
    "Caixa Dobrado - Cobrança": "001",
    "Fundo Fixo - Neto": "001",
    "Fundo Fixo - Osvaldo": "001",
    "Fundo Fixo - Cleto Zanatta": "001",
    "Fundo Fixo - Edison": "001",
    "Fundo Fixo - Phelipe": "001",
    "Caixa Deposito": "001",
    "Fundo Fixo - Valdivino": "001",
    "Fundo Fixo - Jose Domingos": "001",
    "Fudo Fixo - Stenyo": "001",
    "Fundo Fixo - Marcos": "001",
    "Fundo Fixo - ONR": "001",
    "Fundo Fixo - Marcelo Dutra": "001",
    "Fundo Fixo - Gustavo": "001",
    "Fundo Fixo - Delimar": "001",
    "Caixa Contábil": "001",
    "Banco Sicoob_Frutacc_597": "001",
    "Banco Bradesco_Frutacc_28.751": "001",
    # contas de CLEUBER:
    "Banco do Brasil_Cleuber_24585": "004",
    "Banco da Amazonia_Cleuber_34472": "001",
    "Caixa Economica_Cleuber_20573": "001",
    "Banco Bradesco_Cleuber_22102": "003",
    "Banco Sicoob_Cleuber_052": "002",
    "Caixa Economica_Cleuber_25766": "001",
    "Banco Santander_Cleuber_1008472": "001",
    "Banco Sicredi_Cleuber_36120": "001",
    "Banco Itau_Cleuber_63206": "001",
    "Banco Sicoob_Cleuber_81934": "002",
    "Caixa Economica_Cleuber_20177": "001",
    # contas “gerais”
    "Banco Itau_Frutacc_16900": "001",
    "Banco Sicredi_Anne_27012": "001",

    # (as que estavam só no GILSON)
    "Conta Rotative Gilson": "001",
    "Banco Itau_Gilson_26059": "001",
    "Banco do Brasil_Gilson_21252": "001",
    "Banco Bradesco_Gilson_27014": "001",
    "Banco Sicoob_Gilson_781": "001",
    "Banco Sicredi_Gilson_39644": "001",

    # (as que estavam só na ADRIANA)
    "Caixa Economica_Adriana_20590": "001",
    "Banco Bradesco_Adriana_29260": "001",

    # (a que estava só no LUCAS)
    "Banco Bradesco_Lucas 29620": "001",

    # fallback
    "Não Mapeado": "0000",
}

# GILSON
MAP_CONTAS_GILSON = {
    "Caixa Geral": "001",
    "Cheques a Compensar": "001",
    "Fundo Fixo - Gildevan": "001",
    "Fundo Fixo - Cleidson Alves": "001",
    "Fundo Fixo - Rodrigo": "001",
    "Fundo Fixo - Wandres": "001",
    "Fundo Fixo - Cezar Dias": "001",
    "Fundo Fixo - Geraldo": "001",
    "Fundo Fixo - Daniel": "001",
    "Fundo Fixo - Hadlaim": "001",
    "Fundo Fixo - Lourival": "001",
    "Fundo Fixo - Rogeris": "001",
    "Fundo Fixo - Joaquim": "001",
    "Caixa Dobrado": "001",
    "Fundo Fxo - Douglas": "001",
    "Fundo Fixo - Samuel": "001",
    "Fundo Fixo - Adarildo": "001",
    "Fundo Fixo - Fabricio": "001",
    "Fundo Fixo - Fernando": "001",
    "Fundo Fixo - Orivan": "001",
    "Fundo Fixo - Saimon": "001",
    "Fundo Fxo - Eduardo": "001",
    "Fundo Fixo - Melquiades": "001",
    "Fundo Fixo - Anivaldo": "001",
    "Fundo Fixo - Cida": "001",
    "Caixa Dobrado - Cobrança": "001",
    "Fundo Fixo - Neto": "001",
    "Fundo Fixo - Osvaldo": "001",
    "Fundo Fixo - Cleto Zanatta": "001",
    "Fundo Fixo - Edison": "001",
    "Fundo Fixo - Phelipe": "001",
    "Caixa Deposito": "001",
    "Fundo Fixo - Valdivino": "001",
    "Fundo Fixo - Jose Domingos": "001",
    "Fudo Fixo - Stenyo": "001",
    "Fundo Fixo - Marcos": "001",
    "Fundo Fixo - ONR": "001",
    "Fundo Fixo - Marcelo Dutra": "001",
    "Fundo Fixo - Gustavo": "001",
    "Fundo Fixo - Delimar": "001",
    "Caixa Contábil": "001",
    "Banco Sicoob_Frutacc_597": "001",
    "Banco Bradesco_Frutacc_28.751": "001",
    # contas de CLEUBER:
    "Banco do Brasil_Cleuber_24585": "001",
    "Banco da Amazonia_Cleuber_34472": "001",
    "Caixa Economica_Cleuber_20573": "001",
    "Banco Bradesco_Cleuber_22102": "001",
    "Banco Sicoob_Cleuber_052": "001",
    "Caixa Economica_Cleuber_25766": "001",
    "Banco Santander_Cleuber_1008472": "001",
    "Banco Sicredi_Cleuber_36120": "001",
    "Banco Itau_Cleuber_63206": "001",
    "Banco Sicoob_Cleuber_81934": "001",
    "Caixa Economica_Cleuber_20177": "001",
    # contas “gerais”
    "Banco Itau_Frutacc_16900": "001",
    "Banco Sicredi_Anne_27012": "001",

    # (as que estavam só no GILSON)
    "Conta Rotative Gilson": "001",
    "Banco Itau_Gilson_26059": "002",
    "Banco do Brasil_Gilson_21252": "001",
    "Banco Bradesco_Gilson_27014": "005",
    "Banco Sicoob_Gilson_781": "003",
    "Banco Sicredi_Gilson_39644": "004",

    # (as que estavam só na ADRIANA)
    "Caixa Economica_Adriana_20590": "001",
    "Banco Bradesco_Adriana_29260": "001",

    # (a que estava só no LUCAS)
    "Banco Bradesco_Lucas 29620": "001",

    # fallback
    "Não Mapeado": "0000",
}

# ADRIANA
MAP_CONTAS_ADRIANA = {
    "Caixa Geral": "001",
    "Cheques a Compensar": "001",
    "Fundo Fixo - Gildevan": "001",
    "Fundo Fixo - Cleidson Alves": "001",
    "Fundo Fixo - Rodrigo": "001",
    "Fundo Fixo - Wandres": "001",
    "Fundo Fixo - Cezar Dias": "001",
    "Fundo Fixo - Geraldo": "001",
    "Fundo Fixo - Daniel": "001",
    "Fundo Fixo - Hadlaim": "001",
    "Fundo Fixo - Lourival": "001",
    "Fundo Fixo - Rogeris": "001",
    "Fundo Fixo - Joaquim": "001",
    "Caixa Dobrado": "001",
    "Fundo Fxo - Douglas": "001",
    "Fundo Fixo - Samuel": "001",
    "Fundo Fixo - Adarildo": "001",
    "Fundo Fixo - Fabricio": "001",
    "Fundo Fixo - Fernando": "001",
    "Fundo Fixo - Orivan": "001",
    "Fundo Fixo - Saimon": "001",
    "Fundo Fxo - Eduardo": "001",
    "Fundo Fixo - Melquiades": "001",
    "Fundo Fixo - Anivaldo": "001",
    "Fundo Fixo - Cida": "001",
    "Caixa Dobrado - Cobrança": "001",
    "Fundo Fixo - Neto": "001",
    "Fundo Fixo - Osvaldo": "001",
    "Fundo Fixo - Cleto Zanatta": "001",
    "Fundo Fixo - Edison": "001",
    "Fundo Fixo - Phelipe": "001",
    "Caixa Deposito": "001",
    "Fundo Fixo - Valdivino": "001",
    "Fundo Fixo - Jose Domingos": "001",
    "Fundo Fixo - Stenyo": "001",
    "Fundo Fixo - Marcos": "001",
    "Fundo Fixo - ONR": "001",
    "Fundo Fixo - Marcelo Dutra": "001",
    "Fundo Fixo - Gustavo": "001",
    "Fundo Fixo - Delimar": "001",
    "Caixa Contábil": "001",
    "Banco Sicoob_Frutacc_597": "001",
    "Banco Bradesco_Frutacc_28.751": "001",
    # contas de CLEUBER:
    "Banco do Brasil_Cleuber_24585": "001",
    "Banco da Amazonia_Cleuber_34472": "001",
    "Caixa Economica_Cleuber_20573": "001",
    "Banco Bradesco_Cleuber_22102": "001",
    "Banco Sicoob_Cleuber_052": "001",
    "Caixa Economica_Cleuber_25766": "001",
    "Banco Santander_Cleuber_1008472": "001",
    "Banco Sicredi_Cleuber_36120": "001",
    "Banco Itau_Cleuber_63206": "001",
    "Banco Sicoob_Cleuber_81934": "001",
    "Caixa Economica_Cleuber_20177": "001",
    # contas “gerais”
    "Banco Itau_Frutacc_16900": "001",
    "Banco Sicredi_Anne_27012": "001",

    # (as que estavam só no GILSON)
    "Conta Rotative Gilson": "001",
    "Banco Itau_Gilson_26059": "001",
    "Banco do Brasil_Gilson_21252": "001",
    "Banco Bradesco_Gilson_27014": "001",
    "Banco Sicoob_Gilson_781": "001",
    "Banco Sicredi_Gilson_39644": "001",

    # (as que estavam só na ADRIANA)
    "Caixa Economica_Adriana_20590": "001",
    "Banco Bradesco_Adriana_29260": "002",

    # (a que estava só no LUCAS)
    "Banco Bradesco_Lucas 29620": "001",

    # fallback
    "Não Mapeado": "0000",
}

# LUCAS
MAP_CONTAS_LUCAS = {
    "Caixa Geral": "001",
    "Cheques a Compensar": "001",
    "Fundo Fixo - Gildevan": "001",
    "Fundo Fixo - Cleidson Alves": "001",
    "Fundo Fixo - Rodrigo": "001",
    "Fundo Fixo - Wandres": "001",
    "Fundo Fixo - Cezar Dias": "001",
    "Fundo Fixo - Geraldo": "001",
    "Fundo Fixo - Daniel": "001",
    "Fundo Fixo - Hadlaim": "001",
    "Fundo Fixo - Lourival": "001",
    "Fundo Fixo - Rogeris": "001",
    "Fundo Fixo - Joaquim": "001",
    "Caixa Dobrado": "001",
    "Fundo Fxo - Douglas": "001",
    "Fundo Fixo - Samuel": "001",
    "Fundo Fixo - Adarildo": "001",
    "Fundo Fixo - Fabricio": "001",
    "Fundo Fixo - Fernando": "001",
    "Fundo Fixo - Orivan": "001",
    "Fundo Fixo - Saimon": "001",
    "Fundo Fxo - Eduardo": "001",
    "Fundo Fixo - Melquiades": "001",
    "Fundo Fixo - Anivaldo": "001",
    "Fundo Fixo - Cida": "001",
    "Caixa Dobrado - Cobrança": "001",
    "Fundo Fixo - Neto": "001",
    "Fundo Fixo - Osvaldo": "001",
    "Fundo Fixo - Cleto Zanatta": "001",
    "Fundo Fixo - Edison": "001",
    "Fundo Fixo - Phelipe": "001",
    "Caixa Deposito": "001",
    "Fundo Fixo - Valdivino": "001",
    "Fundo Fixo - Jose Domingos": "001",
    "Fudo Fixo - Stenyo": "001",
    "Fundo Fixo - Marcos": "001",
    "Fundo Fixo - ONR": "001",
    "Fundo Fixo - Marcelo Dutra": "001",
    "Fundo Fixo - Gustavo": "001",
    "Fundo Fixo - Delimar": "001",
    "Caixa Contábil": "001",
    "Banco Sicoob_Frutacc_597": "001",
    "Banco Bradesco_Frutacc_28.751": "001",
    # contas de CLEUBER:
    "Banco do Brasil_Cleuber_24585": "001",
    "Banco da Amazonia_Cleuber_34472": "001",
    "Caixa Economica_Cleuber_20573": "001",
    "Banco Bradesco_Cleuber_22102": "001",
    "Banco Sicoob_Cleuber_052": "001",
    "Caixa Economica_Cleuber_25766": "001",
    "Banco Santander_Cleuber_1008472": "001",
    "Banco Sicredi_Cleuber_36120": "001",
    "Banco Itau_Cleuber_63206": "001",
    "Banco Sicoob_Cleuber_81934": "001",
    "Caixa Economica_Cleuber_20177": "001",
    # contas “gerais”
    "Banco Itau_Frutacc_16900": "001",
    "Banco Sicredi_Anne_27012": "001",

    # (as que estavam só no GILSON)
    "Conta Rotative Gilson": "001",
    "Banco Itau_Gilson_26059": "001",
    "Banco do Brasil_Gilson_21252": "001",
    "Banco Bradesco_Gilson_27014": "001",
    "Banco Sicoob_Gilson_781": "001",
    "Banco Sicredi_Gilson_39644": "001",

    # (as que estavam só na ADRIANA)
    "Caixa Economica_Adriana_20590": "001",
    "Banco Bradesco_Adriana_29260": "001",

    # (a que estava só no LUCAS)
    "Banco Bradesco_Lucas 29620": "002",

    # fallback
    "Não Mapeado": "0000",
}

def _select_map_contas_by_owner(owner: str) -> dict:
    owner = (owner or "CLEUBER").upper()
    if owner == "GILSON":
        return MAP_CONTAS_GILSON
    elif owner == "ADRIANA":
        return MAP_CONTAS_ADRIANA
    elif owner == "LUCAS":
        return MAP_CONTAS_LUCAS
    # default / desconhecido → CLEUBER
    return MAP_CONTAS_CLEUBER

# =====================================================================
# CONSTANTES E FUNÇÕES DE APOIO
# =====================================================================
XML_DIRS = [
    r"\\rilkler\LIVRO CAIXA\ISENTOS",
    r"\\rilkler\LIVRO CAIXA\OUTROS_XMLS"
]

SIMILARIDADE_MIN_NOME = 0.80
SIMILARIDADE_MIN_NOME_STRICT = 0.90
VAL_TOL = 0.10  # tolerância para valor
DATE_WINDOW_DAYS_BEFORE = 10
DATE_WINDOW_DAYS_AFTER = 45

_nf_re = re.compile(r'NF\D*(\d+)', re.IGNORECASE)
_chave44_re = re.compile(r'\b(\d{44})\b')

NS_NFE = {"nfe": "http://www.portalfiscal.inf.br/nfe"}

def _extract_nf_num(texto: str) -> str | None:
    if not texto:
        return None
    m = _nf_re.search(texto)
    if m:
        return m.group(1).lstrip("0") or "0"
    m2 = _chave44_re.search(texto)
    if m2:
        chave = m2.group(1)
        return chave[35:44].lstrip("0") or "0"
    return None

def _norm_txt(s: str) -> str:
    s = str(s or "").upper().strip()
    s = unicodedata.normalize("NFKD", s).encode("ASCII", "ignore").decode("ASCII")
    s = re.sub(r"\s+", " ", s)
    return s

def _sim(a: str, b: str) -> float:
    return SequenceMatcher(None, _norm_txt(a), _norm_txt(b)).ratio()

def _parse_xml_info(xml_path: str):
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
        inf = root.find(".//nfe:infNFe", NS_NFE)

        # tenta NFSe nacional
        if inf is None:
            NS_NFSE = {"nfse": "http://www.sped.fazenda.gov.br/nfse"}
            infs = root.find(".//nfse:infNFSe", NS_NFSE)
            if infs is not None:
                cnpj_emit = (infs.findtext(".//nfse:emit/nfse:CNPJ", default="", namespaces=NS_NFSE) or "").strip()
                xnome_emit = infs.findtext(".//nfse:emit/nfse:xNome", default="", namespaces=NS_NFSE) or ""
                nnf = (infs.findtext(".//nfse:nNFSe", default="", namespaces=NS_NFSE) or "").strip()
                if not nnf:
                    nnf = (infs.findtext(".//nfse:nDFSe", default="", namespaces=NS_NFSE) or "").strip()
                vliq_txt = infs.findtext(".//nfse:valores/nfse:vLiq", default="", namespaces=NS_NFSE) or ""
                if vliq_txt:
                    vnf = float(str(vliq_txt).replace(",", "."))
                else:
                    vserv_txt = root.findtext(".//nfse:DPS//nfse:valores//nfse:vServ", default="0", namespaces=NS_NFSE) or "0"
                    vnf = float(str(vserv_txt).replace(",", "."))
                infcpl = root.findtext(".//nfse:DPS//nfse:xDescServ", default="", namespaces=NS_NFSE) or ""
                return {
                    "cnpj_emit": re.sub(r"\D", "", cnpj_emit).zfill(14),
                    "xnome_emit": xnome_emit,
                    "nnf": nnf,
                    "vnf": vnf,
                    "infcpl": infcpl,
                    "ref_list": [],
                    "path": xml_path,
                }

        if inf is None:
            return None

        cnpj_emit = (inf.findtext(".//nfe:emit/nfe:CNPJ", default="", namespaces=NS_NFE) or "").strip()
        xnome_emit = inf.findtext(".//nfe:emit/nfe:xNome", default="", namespaces=NS_NFE) or ""
        nnf = (inf.findtext(".//nfe:ide/nfe:nNF", default="", namespaces=NS_NFE) or "").strip()
        vnf_txt = inf.findtext(".//nfe:total/nfe:ICMSTot/nfe:vNF", default="0", namespaces=NS_NFE) or "0"
        vnf = float(str(vnf_txt).replace(",", "."))
        infcpl = inf.findtext(".//nfe:infAdic/nfe:infCpl", default="", namespaces=NS_NFE) or ""
        ref_list = []
        for r in inf.findall(".//nfe:NFref/nfe:refNFe", NS_NFE):
            if r is not None and r.text:
                ref_list.append(r.text.strip())

        return {
            "cnpj_emit": re.sub(r"\D", "", cnpj_emit).zfill(14),
            "xnome_emit": xnome_emit,
            "nnf": nnf,
            "vnf": vnf,
            "infcpl": infcpl,
            "ref_list": ref_list,
            "path": xml_path,
        }
    except Exception:
        return None

def _iter_xmls(dirs):
    for d in dirs:
        p = Path(d)
        if not p.exists():
            continue
        for xmlf in p.rglob("*.xml"):
            info = _parse_xml_info(str(xmlf))
            if info:
                yield info

def _xml_menciona_nf_do_mesmo_fornecedor(cnpj_emit: str, nnf_procurada: str, xml_info: dict) -> bool:
    if xml_info["cnpj_emit"] != cnpj_emit:
        return False
    alvo = re.sub(r"\D", "", str(nnf_procurada))
    infcpl = xml_info.get("infcpl", "") or ""
    if alvo and alvo in re.sub(r"\D", "", infcpl):
        return True
    for ch in xml_info.get("ref_list", []):
        if alvo and alvo in (re.sub(r"\D", "", ch) or ""):
            return True
    return False

def _buscar_valor_devolucao_relacionada(nota: dict) -> float | None:
    try:
        cnpj = str(nota.get('cnpj_busca', '')).strip()
        nnf = str(nota.get('num_nf_busca', '')).strip()
        if not cnpj or not nnf:
            return None
        for xml_info in _iter_xmls(XML_DIRS):
            try:
                if _xml_menciona_nf_do_mesmo_fornecedor(cnpj, nnf, xml_info):
                    v = xml_info.get('vnf')
                    if v is None:
                        continue
                    return float(v)
            except Exception:
                continue
    except Exception:
        return None
    return None

def _should_use_data_nota(nota_row, data_pagamento):
    try:
        alvo = nota_row.get('ANO', None)
        if pd.notna(alvo):
            alvo = int(str(alvo).strip())
        else:
            alvo = None
    except Exception:
        alvo = None
    if alvo is None and pd.notna(nota_row.get('data_nota', pd.NaT)):
        alvo = int(nota_row['data_nota'].year)
    if alvo is not None and pd.notna(data_pagamento):
        return int(data_pagamento.year) < int(alvo)
    return False

def _tolerancia_valor_para_cnpj_diferente(v):
    return max(10.0, 0.10 * max(v, 0.0))

# =====================================================================
# CARREGAR PATHS E PERFIL
# =====================================================================
base_dados_path, testes_path, ACTIVE_OWNER = _resolve_paths_and_owner()
ACTIVE_MAP_CONTAS = _select_map_contas_by_owner(ACTIVE_OWNER)
notas_recebidas_path = _resolve_notas_recebidas_path(testes_path)

print("")
print(f"🗂️  base_dados_path = {base_dados_path}")
print(f"🗂️  testes_path     = {testes_path}")
print(f"👤  perfil ativo     = {ACTIVE_OWNER}")
print(f"📄  notas recebidas  = {notas_recebidas_path or '— não encontrado —'}")

if not os.path.exists(base_dados_path):
    print(f"❌ Base de dados não encontrada: {base_dados_path}")
    sys.exit(1)
if not os.path.exists(testes_path):
    print(f"❌ Planilha de testes/relatório não encontrada: {testes_path}")
    sys.exit(1)

# =====================================================================
# FUNÇÕES DE CONTA (AGORA USANDO O PERFIL ATIVO)
# =====================================================================
def _norm_simple(s: str) -> str:
    return unicodedata.normalize("NFKD", str(s or "")).encode("ASCII", "ignore").decode("ASCII").upper().strip()

def _build_norm_map_from(active_map: dict) -> dict:
    return {
        unicodedata.normalize("NFKD", k).encode("ASCII","ignore").decode("ASCII").upper().strip(): v
        for k, v in active_map.items()
    }

_ACTIVE_MAP_CONTAS_NORM = _build_norm_map_from(ACTIVE_MAP_CONTAS)

def _conta_codigo(nome: str) -> str:
    """
    Faz o lookup de conta dentro do MAPEAMENTO DO PERFIL ATIVO.
    1) match exato normalizado
    2) "contém"
    3) similaridade
    4) devolve "Não Mapeado" do perfil
    """
    if not nome:
        return ACTIVE_MAP_CONTAS.get("Não Mapeado", "0000")
    n = _norm_simple(nome)
    # 1) exato
    if n in _ACTIVE_MAP_CONTAS_NORM:
        return _ACTIVE_MAP_CONTAS_NORM[n]
    # 2) contém
    for k_norm, cod in _ACTIVE_MAP_CONTAS_NORM.items():
        if n in k_norm or k_norm in n:
            return cod
    # 3) similaridade
    best_cod, best_sc = None, 0.0
    for k_norm, cod in _ACTIVE_MAP_CONTAS_NORM.items():
        sc = SequenceMatcher(None, n, k_norm).ratio()
        if sc > best_sc:
            best_sc, best_cod = sc, cod
    if best_sc >= 0.85 and best_cod:
        return best_cod
    # 4) fallback
    return ACTIVE_MAP_CONTAS.get("Não Mapeado", "0000")

# 🔹 Helper p/ banco padrão por perfil (Caixa Geral -> código do perfil)
DEFAULT_CONTA_NOME = "Caixa Geral"
def _banco_padrao_cod() -> str:
    try:
        cod = _conta_codigo(DEFAULT_CONTA_NOME)
        if cod and cod.strip():
            return cod
    except Exception:
        pass
    return ACTIVE_MAP_CONTAS.get("Não Mapeado", "0000")

# =====================================================================
# MAPEAMENTO DE FAZENDAS
# =====================================================================
MAP_FAZENDAS = {
    "Frutacc": "1",
    "União": "2",
    "L3": "3",
    "Primavera": "4",
    "Aliança": "5",
    "Arm. Primavera": "6",
    "Estrela": "7",
    "Barragem": "8",
    "Guara": "9",
    "B. Grande": "8",
    "Frutacc III": "1",
    "Primavera Retiro": "4",
    "Siganna": "1",
    "Formiga": "1",
    "Gabriela": "2",
    "Pouso da Anta": "1",
    "Aliança 2": "1",
    "Primavera Retiro lucas": "1",
}

PRODUTOS_ESPECIAIS = [
    "GASOLINA COMUM",
    "GASOLINA C-COMUM",
    "GASOLINA ADITIVADA",
    "GASOLINA C COMUM",
    "BC:03- GASOLINA C ADITIVADA",
    "OLEO DIESEL",
    "DIESEL",
    "ETANOL",
    "MOBILGREASE",
]

# =====================================================================
# LEITURA DA PLANILHA (DESPESAS)
# =====================================================================
try:
    print("Processando planilha de notas...")
    df_notas = pd.read_excel(testes_path, sheet_name='RELATORIO', header=5)
    df_despesas = df_notas[df_notas['DESPESAS'].notna()].copy()

    # colunas auxiliares
    df_despesas['num_nf_busca'] = (
        df_despesas['Nº NF'].astype(str)
        .str.strip()
        .str.replace(' ', '')
        .str.replace('.', '')
        .str.upper()
    )

    def _cnpj_from_row(row):
        cnpj = re.sub(r'\D', '', str(row.get('CNPJ', '')))
        if len(cnpj) == 14 and cnpj != '0'*14:
            return cnpj
        chave = re.sub(r'\D', '', str(row.get('XML', '')))
        if len(chave) == 44:
            cnpj_xml = chave[6:20]
            if len(cnpj_xml) == 14:
                return cnpj_xml
        for v in row.values:
            s = re.sub(r'\D', '', str(v))
            m = re.search(r'(?<!\d)(\d{14})(?!\d)', s)
            if m:
                return m.group(1)
        return cnpj.zfill(14)

    df_despesas['cnpj_busca'] = df_despesas.apply(_cnpj_from_row, axis=1)
    df_despesas['valor_busca'] = pd.to_numeric(df_despesas['DESPESAS'], errors='coerce')
    df_despesas['data_nota'] = pd.to_datetime(df_despesas['DATA'], dayfirst=True, errors='coerce')
    df_despesas['fornecedor'] = df_despesas['EMITENTE'].astype(str).str.strip()
    df_despesas['fornecedor_norm'] = df_despesas['fornecedor'].apply(_norm_txt)
    df_despesas['cod_fazenda'] = df_despesas['FAZENDA'].map(MAP_FAZENDAS).fillna('0')
    df_despesas['produto_upper'] = df_despesas['PRODUTO'].astype(str).str.upper()
    df_despesas['produto_especial'] = df_despesas['produto_upper'].apply(
        lambda x: any(produto in x for produto in PRODUTOS_ESPECIAIS)
    )

    # pular linhas já marcadas de verde
    df_to_process = df_despesas
    try:
        if OPENPYXL_AVAILABLE:
            wb_chk = load_workbook(testes_path, data_only=True)
            ws_chk = wb_chk['RELATORIO']
            deslocamento = 7
            col_inicio, col_fim = 3, 17
            HEX_VERDE = "C6EFCE"
            verdes_idx = set()
            for idx in df_despesas.index:
                row_excel = deslocamento + idx
                is_green = False
                for col in range(col_inicio, col_fim + 1):
                    cell = ws_chk.cell(row=row_excel, column=col)
                    rgb = (
                        getattr(cell.fill.start_color, "rgb", None)
                        or getattr(cell.fill.fgColor, "rgb", None)
                        or ""
                    )
                    if isinstance(rgb, str) and rgb.endswith(HEX_VERDE):
                        is_green = True
                        break
                if is_green:
                    verdes_idx.add(idx)
            if verdes_idx:
                print(f"ℹ️ Pulando {len(verdes_idx)} notas já marcadas em verde (pagas).")
            df_to_process = df_despesas.loc[~df_despesas.index.isin(verdes_idx)]
            if df_to_process.empty:
                print("✅ Todas as notas já possuem pagamento associado (marcadas em verde). Nada a processar agora.")
        else:
            print("⚠️ openpyxl indisponível — processando todas.")
    except Exception as e:
        print(f"⚠️ Não foi possível verificar marcações verdes: {e}")
        df_to_process = df_despesas

    print(f"✅ {len(df_despesas)} despesas encontradas | ➡️ a processar: {len(df_to_process)}")

except Exception as e:
    raise ValueError(f"❌ Erro em notas: {str(e)}")

# =====================================================================
# LEITURA DA BASE DE PAGAMENTOS
# =====================================================================
try:
    print("\nProcessando base de pagamentos...")
    df_base = pd.read_excel(base_dados_path, sheet_name='Planilha1', header=None)
    header_row = None
    for idx, row in df_base.iterrows():
        if 'Nº NF' in row.values:
            header_row = idx
            break
    if header_row is None:
        raise ValueError("Cabeçalho não encontrado na planilha")
    df_base = pd.read_excel(base_dados_path, sheet_name='Planilha1', header=header_row)
    df_base = df_base.dropna(subset=['Nº NF']).reset_index(drop=True)

    col_map = {
        'Nº NF': 'num_nf',
        'CPF/CNPJ': 'cnpj',
        'ValorParcela': 'valor',
        'NF Cancelada': 'nota_cancelada',
        'Pagamento Cancelado': 'pagamento_cancelado',
        'Conta pag.': 'banco',
        'Data do Pagamento': 'data_pagamento',
        'Data venc.': 'data_vencimento'
    }
    presentes = {k: v for k, v in col_map.items() if k in df_base.columns}
    df_base.rename(columns=presentes, inplace=True)

    if 'N° Primário' in df_base.columns:
        df_base.rename(columns={'N° Primário': 'num_primario'}, inplace=True)
    else:
        df_base['num_primario'] = np.nan

    col_for = None
    for cand in ['Fornecedor', 'Favorecido', 'Razão Social', 'Emitente', 'Nome Fornecedor']:
        if cand in df_base.columns:
            col_for = cand
            break
    if col_for is None:
        df_base['fornecedor_base'] = ''
    else:
        df_base.rename(columns={col_for: 'fornecedor_base'}, inplace=True)
    df_base['fornecedor_base_norm'] = df_base['fornecedor_base'].apply(_norm_txt)

    df_base['num_nf'] = (
        df_base['num_nf']
        .astype(str)
        .str.strip()
        .str.replace(' ', '')
        .str.replace('.', '')
        .str.upper()
    )
    df_base['cnpj'] = df_base['cnpj'].astype(str).apply(
        lambda x: ''.join(filter(str.isdigit, x)).zfill(14)
    )
    df_base['valor'] = pd.to_numeric(df_base['valor'], errors='coerce')
    df_base['data_pagamento'] = pd.to_datetime(df_base['data_pagamento'], dayfirst=True, errors='coerce')
    df_base['data_vencimento'] = pd.to_datetime(df_base['data_vencimento'], dayfirst=True, errors='coerce')
    df_base['pagamento_cancelado'] = df_base['pagamento_cancelado'].astype(str).str.strip().str.upper()
    df_base['nota_cancelada'] = df_base['nota_cancelada'].astype(str).str.strip().str.upper()
    df_base['associada'] = False
    df_base['total_parcelas'] = df_base.groupby(['num_nf', 'cnpj'])['num_nf'].transform('size')
    df_base['_ord_data'] = df_base['data_pagamento'].fillna(df_base['data_vencimento'])
    df_base.sort_values(['num_nf', 'cnpj', '_ord_data', 'valor', 'num_primario'], inplace=True, na_position='last')
    df_base['parcela_idx'] = df_base.groupby(['num_nf', 'cnpj']).cumcount() + 1

    print(f"✅ {len(df_base)} pagamentos encontrados")

except Exception as e:
    print("Erro detalhado:", str(e))
    raise ValueError(f"❌ Erro em pagamentos: {str(e)}")

# =====================================================================
# PROCESSAMENTO
# =====================================================================
results = []
txt_lines = []
pagamentos_associados = 0
parcelas_nao_pagas = 0
produtos_especiais = 0
produtos_especiais_cancelados = 0
linhas_pagas_idx = []
linhas_receitas_pagas_idx = []

print("\nAssociando pagamentos usando datas de vencimento...")

for i, nota in df_to_process.iterrows():
    result_row = nota.to_dict()
    data_nota = nota['data_nota']
    nome_nota_norm = _norm_txt(nota.get('fornecedor', ''))
    valor_nota = float(nota.get('valor_busca') or 0.0)
    num_nf_nota = str(nota.get('num_nf_busca', '')).strip()
    cnpj_nota = str(nota.get('cnpj_busca', '')).strip()

    parcela_encontrada = None

    # CAMADA inicial: NF + CNPJ + não associado + não cancelado
    mask1 = (
        (df_base['num_nf'] == nota['num_nf_busca']) &
        (df_base['cnpj'] == nota['cnpj_busca']) &
        (~df_base['associada']) &
        (df_base['pagamento_cancelado'] != 'SIM')
    )
    cands = df_base[mask1].copy()

    # tirar primários cancelados
    if 'num_primario' in df_base.columns and not cands.empty:
        grupo_nf = df_base.loc[(df_base['num_nf'] == nota['num_nf_busca'])].copy()
        primarios_cancelados = set(
            grupo_nf.loc[grupo_nf['pagamento_cancelado'] == 'SIM', 'num_primario']
                    .dropna().astype(str).unique().tolist()
        )
        if primarios_cancelados:
            cands = cands.loc[~cands['num_primario'].astype(str).isin(primarios_cancelados)].copy()

    # ================== BLOCO DE MULTIPLAS PARCELAS ==================
    if not cands.empty:
        cands = cands.copy()
        cands['valor'] = cands['valor'].astype(float)
        cands['data_pagamento'] = pd.to_datetime(cands['data_pagamento'], errors='coerce')
        cands['data_vencimento'] = pd.to_datetime(cands.get('data_vencimento', pd.NaT), errors='coerce')

        if not pd.isna(data_nota):
            dmin = (data_nota - timedelta(days=DATE_WINDOW_DAYS_BEFORE)).normalize()
            dmax = (data_nota + timedelta(days=DATE_WINDOW_DAYS_AFTER)).normalize() + pd.Timedelta(days=1)
            cands = cands[(cands['data_pagamento'] >= dmin) & (cands['data_pagamento'] < dmax)]

        cands['score'] = 0.0
        cands['score'] += (cands['cnpj'].astype(str) == cnpj_nota).astype(float) * 2.0
        if not pd.isna(data_nota):
            cands['score'] += (cands['data_vencimento'].dt.date == data_nota.date()).astype(float) * 1.5
        cands['diff_val'] = (cands['valor'] - valor_nota).abs()
        cands['score'] += np.isclose(cands['valor'], valor_nota, atol=VAL_TOL).astype(float) * 1.0
        cands['score'] += (~cands['data_pagamento'].isna()).astype(float) * 1.0
        cands['score'] += (cands['banco'].astype(str).str.strip() != '').astype(float) * 1.0
        cands = cands.sort_values(['score', 'diff_val', 'data_pagamento'], ascending=[False, True, False])

        parcelas_escolhidas = []
        saldo = valor_nota

        for idx_sel, row in cands.iterrows():
            v = float(row.get('valor', 0.0) or 0.0)
            if str(row.get('cnpj', '')) != cnpj_nota:
                sim = row.get('sim_nome', None)
                if sim is not None:
                    try:
                        if float(sim) < SIMILARIDADE_MIN_NOME:
                            continue
                    except Exception:
                        pass
            if v <= saldo + VAL_TOL:
                parcelas_escolhidas.append((idx_sel, row))
                saldo -= v
                df_base.at[idx_sel, 'associada'] = True
            if abs(saldo) <= VAL_TOL:
                break

        if parcelas_escolhidas:
            num_nf = num_nf_nota
            cod_faz = str(nota.get('cod_fazenda', '') or '').zfill(3)
            for k, (idx_sel, parcela) in enumerate(parcelas_escolhidas, start=1):
                data_pgto = parcela.get('data_pagamento')
                usa_data_nota = _should_use_data_nota(nota, data_pgto)
                data_base = nota.get('data_nota') if (usa_data_nota or pd.isna(data_pgto)) else data_pgto
                data_fmt = (pd.to_datetime(data_base).strftime('%d-%m-%Y') if not pd.isna(data_base) else '')
                cod_banco = _conta_codigo(str(parcela.get('banco', '')).strip()) or _banco_padrao_cod()
                valor = float(parcela.get('valor', 0.0) or 0.0)
                valor_cent = str(int(round(valor * 100)))
                descricao = f"PAGAMENTO NF {num_nf} (PARCELA {k} de {len(parcelas_escolhidas)})"
                txt_lines.append("|".join([
                    data_fmt,
                    cod_faz,
                    cod_banco,
                    f"{num_nf}-{k}",
                    str(k),
                    descricao,
                    cnpj_nota,
                    "2", "000",
                    valor_cent, valor_cent,
                    "N"
                ]))
                pagamentos_associados += 1
                linhas_pagas_idx.append(nota.name)

            ult = parcelas_escolhidas[-1][1]
            data_pgto = ult.get('data_pagamento')
            usa_data_nota = _should_use_data_nota(nota, data_pgto)
            data_base = nota.get('data_nota') if (usa_data_nota or pd.isna(data_pgto)) else data_pgto
            valor_aceito = sum(float(p[1].get('valor', 0.0) or 0.0) for p in parcelas_escolhidas)
            saldo_final = valor_nota - valor_aceito

            if abs(saldo_final) <= VAL_TOL:
                result_row.update({
                    'Status Nota': 'Ativa',
                    'Status Pagamento': 'Pago',
                    'Banco': _conta_codigo(str(ult.get('banco', '')).strip()) or _banco_padrao_cod(),
                    'Data Pagamento': (pd.to_datetime(data_base).strftime('%d%m%Y') if not pd.isna(data_base) else ''),
                    'Observações': f'Pagto fracionado: {len(parcelas_escolhidas)} parcela(s)'
                })
                results.append(result_row)
                continue
            else:
                result_row.update({
                    'Status Nota': 'Ativa',
                    'Status Pagamento': 'Parcial',
                    'Banco': _conta_codigo(str(ult.get('banco', '')).strip()) or _banco_padrao_cod(),
                    'Data Pagamento': (pd.to_datetime(data_base).strftime('%d%m%Y') if not pd.isna(data_base) else ''),
                    'Observações': f'Parcial: {len(parcelas_escolhidas)} parcela(s); falta R$ {saldo_final:.2f}'
                })

    # ======== ABATE POR DEVOLUÇÃO ========
    try:
        valor_aceito = 0.0
        if 'parcelas_escolhidas' in locals() and parcelas_escolhidas:
            valor_aceito = float(sum(float(p[1].get('valor', 0.0) or 0.0) for p in parcelas_escolhidas))
        saldo_remanescente = max(0.0, valor_nota - valor_aceito)
    except Exception:
        saldo_remanescente = valor_nota

    if saldo_remanescente > VAL_TOL:
        vnf_dev = _buscar_valor_devolucao_relacionada(nota)
        if vnf_dev and vnf_dev > 0:
            diff = abs(vnf_dev - saldo_remanescente)
            if diff <= VAL_TOL:
                cands_diff = df_base[
                    (df_base['associada'] != True) &
                    (df_base['pagamento_cancelado'].astype(str).str.upper() != 'SIM') &
                    (df_base['cnpj'].astype(str) == cnpj_nota) &
                    np.isclose(df_base['valor'].astype(float), saldo_remanescente, atol=VAL_TOL)
                ].copy()
                if not pd.isna(data_nota) and not cands_diff.empty:
                    dmin = (data_nota - timedelta(days=DATE_WINDOW_DAYS_BEFORE)).normalize()
                    dmax = (data_nota + timedelta(days=DATE_WINDOW_DAYS_AFTER)).normalize() + pd.Timedelta(days=1)
                    cands_diff['data_pagamento'] = pd.to_datetime(cands_diff['data_pagamento'], errors='coerce')
                    cands_diff = cands_diff[(cands_diff['data_pagamento'] >= dmin) & (cands_diff['data_pagamento'] < dmax)]
                if not cands_diff.empty:
                    parcela_k = 1
                    if 'parcelas_escolhidas' in locals() and parcelas_escolhidas:
                        parcela_k = len(parcelas_escolhidas) + 1
                    row = cands_diff.sort_values('data_pagamento').iloc[0]
                    data_pgto = pd.to_datetime(row.get('data_pagamento'), errors='coerce')
                    usa_data_nota = _should_use_data_nota(nota, data_pgto)
                    data_base = nota.get('data_nota') if (usa_data_nota or pd.isna(data_pgto)) else data_pgto
                    data_fmt = (pd.to_datetime(data_base).strftime('%d-%m-%Y') if not pd.isna(data_base) else '')
                    cod_banco = _conta_codigo(str(row.get('banco', '')).strip()) or _banco_padrao_cod()
                    valor_cent = str(int(round(float(saldo_remanescente) * 100)))
                    descricao = f"PAGAMENTO NF {num_nf_nota} (ABATE DEVOLUÇÃO)"
                    txt_lines.append("|".join([
                        data_fmt,
                        str(nota.get('cod_fazenda')).zfill(3),
                        cod_banco,
                        f"{num_nf_nota}-{parcela_k}",
                        str(parcela_k),
                        descricao,
                        cnpj_nota,
                        "2", "000",
                        valor_cent, valor_cent,
                        "N"
                    ]))
                    try:
                        df_base.at[row.name, 'associada'] = True
                    except Exception:
                        pass
                    result_row.update({
                        'Status Nota': 'Ativa',
                        'Status Pagamento': 'Pago',
                        'Banco': cod_banco,
                        'Data Pagamento': (pd.to_datetime(data_base).strftime('%d%m%Y') if not pd.isna(data_base) else ''),
                        'Observações': (result_row.get('Observações') or '') + f" | Abate por devolução: {saldo_remanescente:.2f}"
                    })
                    results.append(result_row)
                    continue

    # ======== CAMADA 1 / FALLBACKS ========
    if 'num_nf' not in cands.columns:
        cands['num_nf'] = None
    if 'historico' in cands.columns:
        cands.loc[cands['num_nf'].isna(), 'num_nf'] = (
            cands.loc[cands['num_nf'].isna(), 'historico']
                 .astype(str).apply(_extract_nf_num)
        )

    cands_layer1 = cands[
        (cands['cnpj'].astype(str) == cnpj_nota) &
        (cands['num_nf'].astype(str) == num_nf_nota)
    ]
    if not cands_layer1.empty:
        cands = cands_layer1.copy()
    else:
        cands = cands[(cands['cnpj'].astype(str) == cnpj_nota)].copy()
        if cands.empty and 'sim_nome' in df_base.columns:
            cands = df_base[
                (df_base['associada'] != True) &
                (df_base['pagamento_cancelado'].astype(str).str.upper() != 'SIM') &
                (df_base['sim_nome'].astype(float) >= SIMILARIDADE_MIN_NOME_STRICT)
            ].copy()
        if 'num_nf' in cands.columns:
            cands = cands[
                (cands['num_nf'].isna()) |
                (cands['num_nf'].astype(str) == '') |
                (num_nf_nota == '') |
                (cands['num_nf'].astype(str) == num_nf_nota)
            ]

    if not pd.isna(data_nota) and not cands.empty:
        dmin = (data_nota - timedelta(days=DATE_WINDOW_DAYS_BEFORE)).normalize()
        dmax = (data_nota + timedelta(days=DATE_WINDOW_DAYS_AFTER)).normalize() + pd.Timedelta(days=1)
        cands['data_pagamento'] = pd.to_datetime(cands['data_pagamento'], errors='coerce')
        cands = cands[(cands['data_pagamento'] >= dmin) & (cands['data_pagamento'] < dmax)]

    cands = cands.copy()
    if not cands.empty:
        cands['valor'] = cands['valor'].astype(float)
        cands['score'] = 0.0
        cands['score'] += (cands['cnpj'].astype(str) == cnpj_nota).astype(float) * 2.0
        if not pd.isna(data_nota) and 'data_vencimento' in cands.columns:
            cands['data_vencimento'] = pd.to_datetime(cands['data_vencimento'], errors='coerce')
            cands['score'] += (cands['data_vencimento'].dt.date == data_nota.date()).astype(float) * 1.5
        cands['diff_val'] = (cands['valor'] - valor_nota).abs()
        cands['isclose_val'] = np.isclose(cands['valor'], valor_nota, atol=VAL_TOL)
        cands['score'] += cands['isclose_val'].astype(float) * 1.0
        cands['score'] += (~cands['data_pagamento'].isna()).astype(float) * 1.0
        if 'banco' in cands.columns:
            cands['score'] += (cands['banco'].astype(str).str.strip() != '').astype(float) * 1.0
        if 'sim_nome' in cands.columns and num_nf_nota:
            cands = cands[
                (cands['cnpj'].astype(str) == cnpj_nota) |
                (
                    (cands['sim_nome'].astype(float) >= SIMILARIDADE_MIN_NOME_STRICT) &
                    ((cands['num_nf'].isna()) | (cands['num_nf'].astype(str) == '') | (cands['num_nf'].astype(str) == num_nf_nota))
                )
            ]
        cands = cands.sort_values(['score', 'diff_val', 'data_pagamento'],
                                  ascending=[False, True, False])
        if not cands.empty:
            idx_sel = cands.index[0]
            cand = cands.loc[idx_sel]

            if str(cand.get('cnpj','')) != cnpj_nota:
                sim_ok = False
                if 'sim_nome' in cand.index and not pd.isna(cand['sim_nome']):
                    sim_ok = float(cand['sim_nome']) >= SIMILARIDADE_MIN_NOME
                if not sim_ok:
                    parcela_encontrada = None
                else:
                    parcela_encontrada = cand
                    df_base.at[idx_sel, 'associada'] = True
            else:
                parcela_encontrada = cand
                df_base.at[idx_sel, 'associada'] = True

    # ===== PROCESSAR RESULTADO =====
    if parcela_encontrada is not None:
        if pd.isna(parcela_encontrada['data_pagamento']) or str(parcela_encontrada['banco']).strip() == '':
            result_row.update({
                'Status Nota': "Ativa",
                'Status Pagamento': 'Não pago',
                'Banco': '',
                'Data Pagamento': '',
                'Observações': 'Parcela encontrada sem dados de pagamento'
            })
            parcelas_nao_pagas += 1
        else:
            status_nota = "Cancelada" if "CANCELADA" in str(parcela_encontrada['nota_cancelada']).upper() else "Ativa"
            status_pag = "Pago"
            data_pgto = parcela_encontrada['data_pagamento']
            usa_data_nota = _should_use_data_nota(nota, data_pgto)
            data_base = nota['data_nota'] if (usa_data_nota or pd.isna(data_pgto)) else data_pgto
            data_str = data_base.strftime('%d%m%Y') if not pd.isna(data_base) else ""

            banco_nome = str(parcela_encontrada['banco']).strip()
            cod_banco = _conta_codigo(banco_nome) or _banco_padrao_cod()

            origem = (
                "Associada por CNPJ"
                if str(parcela_encontrada.get('cnpj', '')) == cnpj_nota
                else "Associada por NF + Nome≈" + f"{float(parcela_encontrada.get('sim_nome',0)): .0%}".replace(" ","")
            )

            if (
                hasattr(parcela_encontrada, "index")
                and 'num_primario' in parcela_encontrada.index
                and pd.notna(parcela_encontrada['num_primario'])
                and str(parcela_encontrada.get('cnpj', '')) != cnpj_nota
            ):
                origem = "Associada por NF (N° Primário distinto)"

            if (
                hasattr(parcela_encontrada, "index")
                and 'sim_nome' in parcela_encontrada.index
                and not pd.isna(parcela_encontrada['sim_nome'])
            ):
                try:
                    origem += " + Nome≈" + f"{float(parcela_encontrada['sim_nome']):.0%}"
                except Exception:
                    pass

            result_row.update({
                'Status Nota': status_nota,
                'Status Pagamento': status_pag,
                'Banco': cod_banco,
                'Data Pagamento': data_str,
                'Observações': origem
            })

            if status_nota == "Ativa" and status_pag == "Pago" and data_str:
                data_fmt = data_base.strftime('%d-%m-%Y') if not pd.isna(data_base) else nota['data_nota'].strftime('%d-%m-%Y')
                cod_fazenda3 = str(nota['cod_fazenda']).zfill(3)
                num_nf = num_nf_nota
                cnpj = cnpj_nota
                valor_cent = str(int(round(float(parcela_encontrada['valor']) * 100)))
                parcela_num = int(parcela_encontrada.get('parcela_idx', 1))
                num_nf_txt = f"{num_nf}-{parcela_num}"
                descricao = f"PAGAMENTO NF {num_nf}".upper()
                parcela_txt = str(parcela_num)
                conta_cod_pg = _conta_codigo(parcela_encontrada.get('banco', '')) or _banco_padrao_cod()
                txt_line = [
                    data_fmt,
                    cod_fazenda3,
                    conta_cod_pg,
                    num_nf_txt,
                    parcela_txt,
                    descricao,
                    cnpj,
                    "2",
                    "000",
                    valor_cent,
                    valor_cent,
                    "N"
                ]
                txt_lines.append("|".join(txt_line))
                pagamentos_associados += 1
                linhas_pagas_idx.append(nota.name)
    else:
        mask_canceladas = (
            (df_base['num_nf'] == num_nf_nota) &
            (df_base['cnpj'] == cnpj_nota) &
            (~df_base['associada']) &
            (df_base['pagamento_cancelado'] == 'SIM')
        )
        parcelas_canceladas = df_base[mask_canceladas]
        if not parcelas_canceladas.empty:
            result_row.update({
                'Status Nota': "Ativa",
                'Status Pagamento': 'Cancelado',
                'Banco': '',
                'Data Pagamento': '',
                'Observações': 'Parcela encontrada mas cancelada'
            })
        else:
            result_row.update({
                'Status Nota': "Ativa",
                'Status Pagamento': 'Não pago',
                'Banco': '',
                'Data Pagamento': '',
                'Observações': 'Pagamento não realizado para esta nota'
            })
        parcelas_nao_pagas += 1

    # PRODUTOS ESPECIAIS → pagar automático (usando mapeamento do perfil para banco padrão)
    if nota['produto_especial'] and result_row['Status Pagamento'] == 'Não pago':
        produtos_especiais += 1
        data_nota2 = nota['data_nota']
        data_str = data_nota2.strftime('%d%m%Y') if not pd.isna(data_nota2) else ""
        _cod_banco_padrao = _banco_padrao_cod()
        result_row.update({
            'Status Nota': "Ativa",
            'Status Pagamento': 'Pago',
            'Banco': _cod_banco_padrao,
            'Data Pagamento': data_str,
            'Observações': 'Produto especial (combustível/lubrificante) - Pagamento automático'
        })
        data_fmt = data_nota2.strftime('%d-%m-%Y')
        cod_fazenda3 = str(nota['cod_fazenda']).zfill(3)
        num_nf = num_nf_nota
        cnpj = cnpj_nota
        valor_cent = str(int(round(float(valor_nota) * 100)))
        txt_lines.append("|".join([
            data_fmt,
            cod_fazenda3,
            _cod_banco_padrao,
            num_nf,
            "1",
            f"PAGAMENTO NF {num_nf}".upper(),
            cnpj,
            "2",
            "000",
            valor_cent,
            valor_cent,
            "N"
        ]))
        pagamentos_associados += 1
        linhas_pagas_idx.append(nota.name)

    results.append(result_row)

print(f"\n🔍 Resultados da associação:")
print(f"- Total de despesas processadas: {len(df_despesas)}")
print(f"- Produtos especiais (combustível/lubrificante): {produtos_especiais}")
print(f"  > Produtos especiais cancelados: {produtos_especiais_cancelados}")
print(f"- Parcelas associadas: {pagamentos_associados}")
print(f"- Parcelas não pagas: {parcelas_nao_pagas}")

# =====================================================================
# PASSO EXTRA: AJUSTES POR XML
# =====================================================================
print("\nAnalisando XMLs para diferenças (devolução/reefat)…")
idx_to_results_pos = {}
for pos, r in enumerate(results):
    idx_to_results_pos[df_to_process.index[pos]] = pos

xml_infos = list(_iter_xmls(XML_DIRS))
ajustes_por_xml = 0

for i, nota in df_to_process.iterrows():
    pos_res = idx_to_results_pos.get(i)
    if pos_res is None:
        continue
    if results[pos_res].get('Status Pagamento') == 'Pago':
        continue

    nf_alvo = str(nota['num_nf_busca'])
    cnpj_alvo = nota['cnpj_busca']
    valor_nf = float(nota['valor_busca'] or 0.0)

    achado = None
    for x in xml_infos:
        if _xml_menciona_nf_do_mesmo_fornecedor(cnpj_alvo, nf_alvo, x):
            achado = x
            break
    if not achado:
        continue

    diferenca = round(abs(valor_nf - float(achado['vnf'])), 2)
    if diferenca <= VAL_TOL:
        continue

    cand = df_base.loc[
        (~df_base['associada']) &
        (df_base['pagamento_cancelado'] != 'SIM') &
        np.isclose(df_base['valor'], diferenca, atol=VAL_TOL)
    ].copy()
    if cand.empty:
        continue

    cand['score'] = 0.0
    cand['score'] += (cand['num_nf'] == nf_alvo).astype(float) * 2.0
    cand['score'] += (cand['cnpj'] == cnpj_alvo).astype(float) * 1.5
    if 'fornecedor_base_norm' in cand.columns:
        cand['sim_nome'] = cand['fornecedor_base_norm'].apply(lambda x: _sim(x, nota['fornecedor_norm']))
        cand.loc[cand['sim_nome'] >= SIMILARIDADE_MIN_NOME, 'score'] += cand['sim_nome']
    cand = cand.sort_values('score', ascending=False)
    j = cand.index[0]
    pgto = cand.loc[j]
    df_base.at[j, 'associada'] = True

    try:
        ref_ano = int(nota['ANO']) if ('ANO' in nota.index and pd.notna(nota['ANO'])) else None
    except Exception:
        ref_ano = None
    if ref_ano is None:
        ref_ano = int(nota['data_nota'].year) if ('data_nota' in nota.index and pd.notna(nota['data_nota'])) else None

    data_pg = pgto['data_pagamento']
    usar_data_nota = (pd.notna(data_pg) and ref_ano is not None and int(data_pg.year) < ref_ano)
    data_base = nota['data_nota'] if (usar_data_nota or pd.isna(data_pg)) else data_pg
    data_str = data_base.strftime('%d%m%Y') if pd.notna(data_base) else ""
    banco_nome = str(pgto['banco']).strip()
    cod_banco = _conta_codigo(banco_nome) or _banco_padrao_cod()

    results[pos_res].update({
        'Status Nota': "Ativa",
        'Status Pagamento': "Pago",
        'Banco': cod_banco,
        'Data Pagamento': data_str,
        'Observações': f"Diferença via XML (NF ref.: {achado.get('nnf','?')} | {Path(achado['path']).name})"
    })

    data_fmt = (data_base.strftime('%d-%m-%Y') if pd.notna(data_base)
                else (nota['data_nota'].strftime('%d-%m-%Y') if pd.notna(nota.get('data_nota', pd.NaT)) else ""))
    cod_fazenda3 = str(nota['cod_fazenda']).zfill(3)
    num_nf = nf_alvo
    cnpj = cnpj_alvo
    valor_cent = str(int(round(float(diferenca) * 100)))
    descricao = f"PAGAMENTO NF {num_nf}".upper()
    parcela_txt = "1"

    # 🔸 Linha TXT usa o cod_banco do pagamento (mapeado pelo perfil); fallback para padrão do perfil
    txt_lines.append("|".join([
        data_fmt,
        cod_fazenda3,
        (cod_banco or _banco_padrao_cod()),
        num_nf,
        parcela_txt,
        descricao,
        cnpj,
        "2",
        "000",
        valor_cent,
        valor_cent,
        "N"
    ]))
    pagamentos_associados += 1
    linhas_pagas_idx.append(i)
    ajustes_por_xml += 1

print(f"✅ Ajustes por XML (diferença): {ajustes_por_xml}")

# =====================================================================
# SALVAR RESULTADOS
# =====================================================================
try:
    print("\nSalvando resultados formatados...")
    df_result = pd.DataFrame(results)
    colunas_remover = [
        'num_nf_busca', 'cnpj_busca', 'valor_busca', 'data_nota',
        'produto_upper', 'produto_especial'
    ]
    df_result = df_result.drop(columns=[c for c in colunas_remover if c in df_result.columns], errors='ignore')
    df_result = df_result.dropna(axis=1, how='all')

    colunas_relevantes = [
        'DATA', 'MÊS', 'ANO', 'Nº NF', 'EMITENTE', 'CNPJ', 'PRODUTO',
        'CFOP', 'DESPESAS', 'NATUREZA', 'XML', 'FAZENDA', 'Status Nota',
        'Status Pagamento', 'Banco', 'Data Pagamento', 'Observações'
    ]
    colunas_finais = [c for c in colunas_relevantes if c in df_result.columns]
    df_result = df_result[colunas_finais]

    if 'Data Pagamento' in df_result.columns:
        df_result['Data Pagamento'] = df_result['Data Pagamento'].apply(
            lambda x: f"{x[:2]}/{x[2:4]}/{x[4:]}" if x and isinstance(x, str) and len(x) == 8 else x
        )

    output_excel = "RESULTADO_PAGAMENTOS.xlsx"
    df_result.to_excel(output_excel, index=False)

    if OPENPYXL_AVAILABLE:
        try:
            wb = load_workbook(output_excel)
            ws = wb.active
            tab = Table(displayName="TabelaResultados", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
            for cell in ws[1]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)
            wb.save(output_excel)
            print(f"✅ Planilha formatada salva como tabela: {output_excel}")
        except Exception as e:
            print(f"⚠️ Erro ao formatar tabela: {str(e)}")
    else:
        print(f"✅ Planilha formatada salva: {output_excel}")

    # renumeração das linhas TXT por NF + CNPJ
    if txt_lines:
        from datetime import datetime as _dt
        recs = []
        for i_line, line in enumerate(txt_lines):
            parts = line.split("|")
            if len(parts) < 12:
                continue
            data_txt = parts[0]
            num_nf_field = parts[3]
            cnpj = parts[6]
            base_nf = num_nf_field.split("-", 1)[0]
            try:
                dt = _dt.strptime(data_txt, "%d-%m-%Y")
            except Exception:
                dt = None
            recs.append({"i": i_line, "parts": parts, "base_nf": base_nf, "cnpj": cnpj, "dt": dt})
        grupos = defaultdict(list)
        for r in recs:
            grupos[(r["base_nf"], r["cnpj"])].append(r)
        new_lines = [""] * len(txt_lines)
        for (nf, cnpj), lst in grupos.items():
            lst.sort(key=lambda r: (r["dt"] is None, r["dt"]))
            if len(lst) == 1:
                r = lst[0]
                r["parts"][3] = r["base_nf"]
                r["parts"][5] = re.sub(r"\s*\(PARCELA\s+\d+\)\s*", "", r["parts"][5], flags=re.I)
                new_lines[r["i"]] = "|".join(r["parts"])
            else:
                for idx, r in enumerate(lst, start=1):
                    r["parts"][3] = f"{r['base_nf']}-{idx}"
                    r["parts"][4] = str(idx)
                    r["parts"][5] = re.sub(r"\s*\(PARCELA\s+\d+\)\s*", "", r["parts"][5], flags=re.I)
                    r["parts"][5] = f"{r['parts'][5]} (PARCELA {idx})"
                    new_lines[r["i"]] = "|".join(r["parts"])
        txt_lines = new_lines

    if txt_lines:
        with open("PAGAMENTOS.txt", "w", encoding="utf-8") as f:
            f.write("\n".join(txt_lines))
        print(f"✅ Arquivo TXT gerado com {len(txt_lines)} pagamentos válidos")
    else:
        print("⚠️ Nenhum pagamento válido para gerar TXT")

    # =====================================================================
    # RECEBIMENTOS (NOTAS RECEBIDAS.xlsx)
    # =====================================================================
    try:
        notas_recebidas_path = _resolve_notas_recebidas_path(testes_path)
        if not notas_recebidas_path or not os.path.exists(notas_recebidas_path):
            print("ℹ️ NOTAS RECEBIDAS.xlsx não encontrado — pulando RECEBIMENTOS.")
        else:
            print(f"\nProcessando RECEBIMENTOS a partir de: {notas_recebidas_path}")
            if 'RECEITAS' not in df_notas.columns:
                print("ℹ️ A aba RELATORIO não possui coluna 'RECEITAS' — pulando RECEBIMENTOS.")
            else:
                df_receitas = df_notas.copy()
                df_receitas['valor_receita'] = pd.to_numeric(df_receitas['RECEITAS'], errors='coerce').fillna(0.0)
                df_receitas = df_receitas[df_receitas['valor_receita'] > 0].copy()
                if df_receitas.empty:
                    print("ℹ️ Não há RECEITAS (>0) na RELATORIO. Nada a pagar via recebimentos.")
                else:
                    col_nf = 'Nº NF' if 'Nº NF' in df_receitas.columns else ('NF' if 'NF' in df_receitas.columns else None)
                    if not col_nf:
                        print("⚠️ Não encontrei coluna de 'Nº NF' para RECEITAS — pulando RECEBIMENTOS.")
                    else:
                        cand_part_cols = ['CLIENTE','Destinatário','DESTINATÁRIO','DESTINATARIO','PN','Participante','Favorecido','EMITENTE']
                        col_pn = next((c for c in cand_part_cols if c in df_receitas.columns), None) or df_receitas.columns[0]
                        col_data = 'DATA' if 'DATA' in df_receitas.columns else None
                        col_faz = 'FAZENDA' if 'FAZENDA' in df_receitas.columns else None
                        col_cnpj = 'CNPJ' if 'CNPJ' in df_receitas.columns else None

                        df_receitas['__pn'] = df_receitas[col_pn].astype(str)
                        df_receitas['__pn_norm'] = df_receitas['__pn'].apply(_norm_txt)
                        df_receitas['__nf_ord'] = pd.to_numeric(
                            df_receitas[col_nf].astype(str).str.replace(r'\D','', regex=True),
                            errors='coerce'
                        )
                        df_receitas = df_receitas.sort_values(['__pn_norm','__nf_ord'], kind='stable')

                        df_r = pd.read_excel(notas_recebidas_path, sheet_name=0, header=1)
                        col_pn_r = next((c for c in ['PN','Participante','Cliente'] if c in df_r.columns), None)
                        if col_pn_r is None:
                            col_pn_r = 'Unnamed: 4' if 'Unnamed: 4' in df_r.columns else None
                        col_valor_r = next((c for c in ['Valor','VALOR'] if c in df_r.columns), None)
                        if col_valor_r is None:
                            col_valor_r = 'Unnamed: 6' if 'Unnamed: 6' in df_r.columns else None

                        if not col_pn_r or not col_valor_r:
                            print("⚠️ NOTAS RECEBIDAS.xlsx sem colunas de PN/Valor — pulando RECEBIMENTOS.")
                        else:
                            df_r['__pn_norm'] = df_r[col_pn_r].astype(str).apply(_norm_txt)
                            df_r['__valor'] = pd.to_numeric(df_r[col_valor_r], errors='coerce').fillna(0.0)
                            df_r = df_r[df_r['__valor'] > 0].copy()
                            col_conta_r = next((c for c in ['CONTA','Conta','conta'] if c in df_r.columns), None)

                            # NOVO: CNPJ no NOTAS RECEBIDAS
                            col_cnpj_r = next((c for c in ['CPF/CNPJ','CNPJ','CPF','Documento'] if c in df_r.columns), None)
                            cnpj_por_pn = {}
                            if col_cnpj_r:
                                df_r['__cnpj_r'] = df_r[col_cnpj_r].astype(str).str.replace(r'\D','', regex=True).str.zfill(14)
                                mask_valid = df_r['__cnpj_r'].ne('0'*14)
                                if mask_valid.any():
                                    cnpj_por_pn = (
                                        df_r.loc[mask_valid]
                                            .groupby('__pn_norm')['__cnpj_r']
                                            .agg(lambda s: s.value_counts().idxmax())
                                            .to_dict()
                                    )

                            OFFSET_NR = 3
                            VALOR_COL_XL = df_r.columns.get_loc(col_valor_r) + 1

                            def _map_conta_receb(c):
                                nome = str(c or '').strip()
                                if not nome:
                                    return _banco_padrao_cod(), ""
                                cod = _conta_codigo(nome) or _banco_padrao_cod()
                                return cod, nome

                            if '__data_r' not in df_r.columns and 'Data' in df_r.columns:
                                df_r['__data_r'] = pd.to_datetime(df_r['Data'], dayfirst=True, errors='coerce')
                            elif '__data_r' not in df_r.columns:
                                df_r['__data_r'] = pd.NaT

                            wb_nr = load_workbook(notas_recebidas_path) if OPENPYXL_AVAILABLE else None
                            ws_nr = wb_nr.worksheets[0] if wb_nr else None

                            filas_por_pn = defaultdict(deque)
                            rows_receb_atualizados = set()
                            rows_receb_zerados = set()

                            for i_r, r in df_r.reset_index(drop=True).iterrows():
                                if float(r["__valor"]) > 0:
                                    excel_row = OFFSET_NR + i_r
                                    conta_cod, conta_nome = _map_conta_receb(r[col_conta_r] if col_conta_r else "")
                                    filas_por_pn[r["__pn_norm"]].append({
                                        "row": excel_row,
                                        "valor": float(r["__valor"]),
                                        "data": r.get("__data_r", pd.NaT),
                                        "conta_cod": conta_cod,
                                        "conta_nome": conta_nome
                                    })

                            if OPENPYXL_AVAILABLE:
                                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            else:
                                green_fill = None

                            col_data_r = next((c for c in ['Data','DATA','Data Pagamento','Dt'] if c in df_r.columns), None)
                            if col_data_r:
                                df_r['__data_r'] = pd.to_datetime(df_r[col_data_r], dayfirst=True, errors='coerce')
                            else:
                                df_r['__data_r'] = pd.NaT

                            def saldo_total_pn(pn_norm: str) -> float:
                                return sum(item["valor"] for item in filas_por_pn.get(pn_norm, []))

                            totais_por_pn = df_r.groupby('__pn_norm')['__valor'].sum().to_dict()
                            txt_recebimentos = []
                            resumo_receb = []

                            for pn_norm, grupo in df_receitas.groupby('__pn_norm', sort=False):
                                disponivel = float(totais_por_pn.get(pn_norm, 0.0))
                                pn_nome = str(grupo.iloc[0][col_pn])
                                if disponivel <= 0:
                                    resumo_receb.append(f"• {pn_nome}: sem recebimentos. Nenhuma nota paga.")
                                    continue
                                faltam = 0
                                pagos = 0

                                for idx, row in grupo.iterrows():
                                    valor_nf = float(row['valor_receita'])
                                    if saldo_total_pn(pn_norm) + 1e-9 >= valor_nf:
                                        restante = valor_nf
                                        conta_cod_usada = ""
                                        conta_nome_usado = ""
                                        dt_receb_usada = pd.NaT

                                        while restante > 1e-9 and filas_por_pn[pn_norm]:
                                            topo = filas_por_pn[pn_norm][0]
                                            usar = min(topo["valor"], restante)
                                            topo["valor"] -= usar
                                            restante -= usar
                                            if not conta_cod_usada:
                                                conta_cod_usada = topo.get("conta_cod","")
                                                conta_nome_usado = topo.get("conta_nome","")
                                            if pd.isna(dt_receb_usada):
                                                dt_receb_usada = topo.get("data", pd.NaT)
                                            if OPENPYXL_AVAILABLE and ws_nr:
                                                ws_nr.cell(row=topo["row"], column=VALOR_COL_XL).value = round(topo["valor"], 2)
                                            rows_receb_atualizados.add(topo["row"])
                                            if topo["valor"] <= 1e-9:
                                                filas_por_pn[pn_norm].popleft()
                                                rows_receb_zerados.add(topo["row"])
                                                if OPENPYXL_AVAILABLE and ws_nr and green_fill:
                                                    for col in range(2, 10):
                                                        ws_nr.cell(row=topo["row"], column=col).fill = green_fill

                                        pagos += 1
                                        linhas_receitas_pagas_idx.append(idx)

                                        if pd.isna(dt_receb_usada):
                                            dt_receb_usada = pd.Timestamp.today()
                                        data_fmt = dt_receb_usada.strftime('%d-%m-%Y')
                                        cod_faz = "000"
                                        if col_faz and pd.notna(row.get(col_faz, None)):
                                            cod = str(MAP_FAZENDAS.get(str(row[col_faz]).strip(), "0"))
                                            cod_faz = cod.zfill(3)
                                        cnpj_rec = cnpj_por_pn.get(pn_norm, "")
                                        if (not cnpj_rec or cnpj_rec == "0"*14) and col_cnpj:
                                            cnpj_rec = "".join(ch for ch in str(row[col_cnpj]) if ch.isdigit()).zfill(14)
                                        num_nf = str(row[col_nf]).strip()
                                        conta_info = f" [CONTA: {conta_nome_usado} — {conta_cod_usada}]" if conta_nome_usado else ""
                                        descricao = f"RECEBIMENTO NF {num_nf} {pn_nome}{conta_info}".upper()
                                        from decimal import Decimal, ROUND_HALF_UP
                                        valor_cent = str(
                                            int(
                                                (Decimal(str(valor_nf)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) * 100)
                                                .to_integral_value(rounding=ROUND_HALF_UP)
                                            )
                                        )
                                        linha = [
                                            data_fmt,
                                            cod_faz,
                                            (conta_cod_usada or _conta_codigo(conta_nome_usado) or _banco_padrao_cod()),
                                            num_nf,
                                            "1",
                                            descricao,
                                            cnpj_rec,
                                            "1",
                                            valor_cent,
                                            "000",
                                            valor_cent,
                                            "N"
                                        ]
                                        txt_recebimentos.append("|".join(linha))
                                    else:
                                        faltam += 1

                                restante_pn = saldo_total_pn(pn_norm)
                                if restante_pn > 1e-6:
                                    resumo_receb.append(f"• {pn_nome}: pagas {pagos} nota(s), sobrou R$ {restante_pn:,.2f}.")
                                else:
                                    if faltam > 0:
                                        resumo_receb.append(f"• {pn_nome}: pagas {pagos} nota(s), faltam {faltam} nota(s).")
                                    else:
                                        resumo_receb.append(f"• {pn_nome}: todas as notas pagas.")

                            if txt_recebimentos and not linhas_receitas_pagas_idx:
                                txt_recebimentos = []

                            if txt_recebimentos:
                                with open("RECEBIMENTOS.txt", "w", encoding="utf-8") as f:
                                    f.write("\n".join(txt_recebimentos))
                                print(f"✅ Arquivo TXT gerado com {len(txt_recebimentos)} recebimento(s)")
                            else:
                                print("ℹ️ Nenhum recebimento gerado (valores insuficientes).")

                            if OPENPYXL_AVAILABLE and wb_nr and (rows_receb_atualizados or rows_receb_zerados):
                                try:
                                    wb_nr.save(notas_recebidas_path)
                                    print(f"✅ NOTAS RECEBIDAS atualizada: {len(rows_receb_atualizados)} linha(s) alterada(s); "
                                          f"{len(rows_receb_zerados)} zerada(s) e destacada(s).")
                                except Exception as e:
                                    print(f"⚠️ Falha ao salvar NOTAS RECEBIDAS atualizada: {e}")

                            if OPENPYXL_AVAILABLE and linhas_receitas_pagas_idx:
                                try:
                                    wb = load_workbook(testes_path)
                                    ws = wb['RELATORIO']
                                    green_fill2 = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                    deslocamento = 7
                                    col_inicio, col_fim = 3, 17
                                    for idx in linhas_receitas_pagas_idx:
                                        row_excel = deslocamento + idx
                                        for col in range(col_inicio, col_fim + 1):
                                            ws.cell(row=row_excel, column=col).fill = green_fill2
                                    wb.save(testes_path)
                                    print(f"✅ {len(linhas_receitas_pagas_idx)} linha(s) de RECEITA marcadas em verde.")
                                except Exception as e:
                                    print(f"⚠️ Falha ao marcar RECEITAS em verde: {e}")

                            if resumo_receb:
                                print("\nResumo RECEBIMENTOS por participante:")
                                for msg in resumo_receb:
                                    print(msg)

    except Exception as e:
        print(f"❌ Erro no bloco de RECEBIMENTOS: {str(e)}")

    # =====================================================================
    # MARCAR DESPESAS PAGAS EM VERDE
    # =====================================================================
    if OPENPYXL_AVAILABLE and linhas_pagas_idx:
        try:
            print("\nMarcando notas pagas na planilha original (colunas C-Q)...")
            wb_original = load_workbook(testes_path)
            ws_original = wb_original['RELATORIO']
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            deslocamento = 7
            col_inicio = 3
            col_fim = 17
            for idx in linhas_pagas_idx:
                row_idx = deslocamento + idx
                for col in range(col_inicio, col_fim + 1):
                    ws_original.cell(row=row_idx, column=col).fill = green_fill
            wb_original.save(testes_path)
            print(f"✅ Planilha original atualizada com notas pagas destacadas: {testes_path}")
            print(f"   - {len(linhas_pagas_idx)} notas marcadas como pagas (colunas C-Q)")
        except Exception as e:
            print(f"⚠️ Erro ao marcar notas pagas: {str(e)}")
    elif not OPENPYXL_AVAILABLE:
        print("⚠️ Openpyxl não disponível - não foi possível destacar notas pagas")
    elif not linhas_pagas_idx:
        print("⚠️ Nenhuma nota paga para destacar")

    # =====================================================================
    # MARCAR PRODUTOS ESPECIAIS
    # =====================================================================
    if OPENPYXL_AVAILABLE:
        wb_original = load_workbook(testes_path)
        ws_original = wb_original['RELATORIO']
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        deslocamento = 7
        col_inicio, col_fim = 3, 17
        idx_pagas = set(linhas_pagas_idx)
        idx_especiais = set(df_despesas.loc[df_despesas['produto_especial']].index)
        idx_para_destacar = idx_pagas.union(idx_especiais)
        for idx in idx_para_destacar:
            row_excel = deslocamento + idx
            for col in range(col_inicio, col_fim + 1):
                ws_original.cell(row=row_excel, column=col).fill = green_fill
        wb_original.save(testes_path)
        print(f"✅ Destacadas {len(idx_para_destacar)} linhas (pagas + especiais)")

    print("\n✅ Processo concluído com sucesso!")

except Exception as e:
    print(f"❌ Erro ao salvar resultados: {str(e)}")
