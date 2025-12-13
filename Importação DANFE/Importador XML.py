import sys
import os
import json
import traceback
from datetime import datetime
from xml.etree import ElementTree as ET
from pathlib import Path

from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton, QTextEdit,
    QFileDialog, QMessageBox, QCheckBox, QLabel, QFrame, QHBoxLayout,
    QProgressDialog, QDialog, QLineEdit, QDialogButtonBox,
    QFormLayout, QGroupBox, QSplitter, QSizePolicy, QToolButton,
    QStyle, QSpacerItem, QGraphicsDropShadowEffect, QTabWidget
)
from PySide6.QtGui import (
    QFont, QColor, QPalette, QIcon, QTextCursor, QPixmap, QAction
)
from PySide6.QtCore import (
    Qt, QTimer, QCoreApplication, QProcess, QProcessEnvironment, QSize
)
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from copy import copy

ICON_PATH = Path(__file__).parent / "image" / "logo.png"

STYLE_SHEET = """
QMainWindow, QWidget { background-color: #1B1D1E; font-family: 'Segoe UI', Arial, sans-serif; color: #E0E0E0; }
QLineEdit, QDateEdit, QComboBox, QTextEdit { color: #E0E0E0; background-color: #2B2F31; border: 1px solid #1e5a9c; border-radius: 6px; padding: 6px; }
QLineEdit::placeholder { color: #5A5A5A; }
QPushButton { background-color: #1e5a9c; color: #FFFFFF; border: none; border-radius: 6px; padding: 8px 16px; font-weight: bold; }
QPushButton:hover, QPushButton:pressed { background-color: #002a54; }
QPushButton#danger { background-color: #C0392B; }
QPushButton#danger:hover { background-color: #E74C3C; }
QPushButton#success { background-color: #27AE60; }
QPushButton#success:hover { background-color: #2ECC71; }
QGroupBox { border: 1px solid #11398a; border-radius: 6px; margin-top: 10px; font-weight: bold; background-color: #0d1b3d; }
QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px; color: #ffffff; }
QTableWidget { background-color: #222426; color: #E0E0E0; border: 1px solid #1e5a9c; border-radius: 4px; gridline-color: #3A3C3D; alternate-background-color: #2A2C2D; }
QHeaderView::section { background-color: #1e5a9c; color: #FFFFFF; padding: 6px; border: none; }
QTabWidget::pane { border: 1px solid #1e5a9c; border-radius: 4px; background: #212425; margin-top: 5px; }
QTabBar::tab { background: #2A2C2D; color: #E0E0E0; padding: 8px 16px; border: 1px solid #1e5a9c; border-top-left-radius: 4px; border-top-right-radius: 4px; margin-right: 2px; }
QTabBar::tab:selected { background: #1e5a9c; color: #FFFFFF; border-bottom: 2px solid #002a54; }
QStatusBar { background-color: #212425; color: #7F7F7F; border-top: 1px solid #1e5a9c; }
"""

# =========================
#  MAPEAMENTOS POR PRODUTOR
# =========================
# CPFs por produtor
def _digits(s: str) -> str:
    return ''.join(ch for ch in (s or '') if ch.isdigit())

CLEUBER_CPF = "42276950153"
GILSON_CPF  = "54860253191"
ADRIANA_CPF = "47943246187"
LUCAS_CPF   = "03886681130"

# Tabela de owners (rótulos aceitos para alternância de perfil)
OWNERS = {
    "CLEUBER": {"CPF": CLEUBER_CPF, "PROFILE_LABELS": ("Cleuber", "CLEUBER")},
    "GILSON":  {"CPF": GILSON_CPF,  "PROFILE_LABELS": ("Gilson", "GILSON")},
    "ADRIANA": {"CPF": ADRIANA_CPF, "PROFILE_LABELS": ("Adriana", "ADRIANA")},
    "LUCAS":   {"CPF": LUCAS_CPF,   "PROFILE_LABELS": ("Lucas", "LUCAS")},
}
for k in list(OWNERS.keys()):
    OWNERS[k]["CPF_D"] = _digits(OWNERS[k]["CPF"])

# ---- Mapeamentos por produtor ----
FARM_MAPPING_CLEUBER = {
    "115149210": "Arm. Primavera",
    "111739837": "Aliança",
    "114436720": "B. Grande",
    "115449965": "Estrela",
    "294186832": "Frutacc",
    "294907068": "Frutacc III",
    "295057386": "L3",
    "112877672": "Primavera",
    "113135521": "Primavera Retiro",
    "294904093": "União",
    "295359790": "Frutacc V",
    "295325704": "Siganna"
}
FARM_MAPPING_GILSON  = {
    "112725503": "Formiga",
}
FARM_MAPPING_ADRIANA = {
    "113348037": "Pouso da Anta",
}
FARM_MAPPING_LUCAS   = {
    "115563083": "Aliança 2",
    "115008810": "Primavera Retiro lucas",
}

CODIGOS_CIDADES_CLEUBER = {
    "Lagoa da Confusao": "Frutacc",
    "Rialma - GO": "Aliança",
    "Rialma": "Aliança", "Lizarda - TO": "Frutacc",
    "TROMBAS": "Primavera", "Trombas - GO": "Primavera",
    "DUERE": "L3", "DUERÉ": "L3", "DUERE TO": "L3", "Duere": "L3",
    "Ceres": "Aliança", "Ceres - GO": "Aliança", "Rianapolis": "Aliança", "NOVA GLORIA": "Aliança",
    "MONTIVIDIU": "Barragem", "MONTIVIDIU DO NORTE - GO": "Barragem",
    "Nova Glória": "Aliança", "Nova Gloria": "Aliança",
    "Lagoa da Confusão": "Frutacc", "MONTIVIDIU DO NORTE": "Barragem",
    "LAGOA DA CONFUSAO": "Frutacc", "LAGOA DA CONFUSÃO": "Frutacc",
    "LAGOA CONFUSAO": "Frutacc", "LAGOA DA CONFUSAO - TO": "Frutacc",
    "RIALMA": "Aliança", "Trombas": "Primavera", "CERES": "Aliança",
    "Formoso do Araguaia": "União", "FORMOSO DO ARAGUAIA": "União",
    "APARECIDA DO RIO NEGRO": "Primavera", "Gurupi - TO": "Frutacc",
    "Goianésia - GO": "Aliança", "Palmas - TO": "Guara",
    "Tasso Fragoso": "Guara", "BALSAS": "Guara", "Balsas": "Guara",
    "Montividiu": "Barragem", "Uruaçu - GO": "Aliança", 'Goianira - GO': "Aliança",

    # --- Porangatu (GO) ---
    "Porangatu - GO": "Primavera",
    "PORANGATU - GO": "Primavera",
    "Porangatu": "Primavera",
    "PORANGATU": "Primavera",

    # --- Montividiu do Norte (inclui variante com erro de digitação) ---
    "Montivldiu do Norte - GO": "Barragem",
    "Montividiu do Norte - GO": "Barragem",
    "Montividiu do Norte": "Barragem",

    # --- Cristalândia (TO) ---
    "Cristalândia - TO": "Frutacc",
    "Cristalandia - TO": "Frutacc",
    "Cristalândia": "Frutacc",
    "Cristalandia": "Frutacc",

    # --- Paraíso do Tocantins (TO) ---
    "Paraíso do Tocantins - TO": "Frutacc",
    "Paraiso do Tocantins - TO": "Frutacc",
    "Paraíso do Tocantins": "Frutacc",
    "Paraiso do Tocantins": "Frutacc",

    # --- Nova Crixás (GO) ---
    "Nova Crixás - GO": "Aliança",
    "Nova Crixas - GO": "Aliança",
    "Nova Crixás": "Aliança",
    "Nova Crixas": "Aliança",
}
CODIGOS_CIDADES_GILSON  = {
    "Nova Glória - GO": "Formiga",
    "RIALMA - GO": "Formiga",
    "Rialma - GO": "Formiga",
    "Rialma": "Formiga",
    "RIALMA": "Formiga",
    "Ceres": "NAO LANÇAR",
    "CERES": "NAO LANÇAR",
    "Goiania": "NAO LANÇAR",
    "GOIANIA": "NAO LANÇAR",
    "GOIÂNIA": "NAO LANÇAR",
    "aparecida de goiania": "NAO LANÇAR",
    "Aparecida de Goiania": "NAO LANÇAR",
    "APARECIDA DE GOIANIA": "NAO LANÇAR",
    "Ceres - GO": "NAO LANÇAR",
    "MONTIVIDIU DO NORTE - GO": "Gabriela",
    "Montividiu do Norte - GO": "Gabriela",
    "Montividiu do Norte": "Gabriela",
    "MONTIVIDIU DO NORTE": "Gabriela",
    "Uruaçu - GO": "Formiga",
}
CODIGOS_CIDADES_ADRIANA = {
    "GOIANIA": "NAO LANÇAR",
    "Goiania": "NAO LANÇAR",
    "GOIÂNIA": "NAO LANÇAR",
    "Goiania - GO": "NAO LANÇAR",
    "MONTIVIDIU DO NORTE - GO": "Pouso da Anta",
    "Montividiu do Norte - GO": "Pouso da Anta",
    "Montividiu do Norte": "Pouso da Anta",
    "MONTIVIDIU DO NORTE": "Pouso da Anta",
}
CODIGOS_CIDADES_LUCAS   = {
    "SAO MIGUEL DO ARAGUAIA": "Aliança 2",
    "São Miguel do Araguaia": "Aliança 2",
    "SÃO MIGUEL DO ARAGUAIA": "Aliança 2",
    "Sao Miguel do Araguaia": "Aliança 2",
    "TROMBAS": "Primavera Retiro lucas",
    "Trombas": "Primavera Retiro lucas",
    "Trombas - GO": "Primavera Retiro lucas",
    "TROMBAS - GO": "Primavera Retiro lucas",
    "NOVA GLORIA": "Aliança 2",
}

FARM_MAPPING_BY_OWNER = {
    "CLEUBER": FARM_MAPPING_CLEUBER,
    "GILSON":  FARM_MAPPING_GILSON,
    "ADRIANA": FARM_MAPPING_ADRIANA,
    "LUCAS":   FARM_MAPPING_LUCAS,
}
CODIGOS_CIDADES_BY_OWNER = {
    "CLEUBER": CODIGOS_CIDADES_CLEUBER,
    "GILSON":  CODIGOS_CIDADES_GILSON,
    "ADRIANA": CODIGOS_CIDADES_ADRIANA,
    "LUCAS":   CODIGOS_CIDADES_LUCAS,
}

# ===== Normalização de cidades =====
import unicodedata
import re

def _norm_city_key(s: str) -> str:
    if not s:
        return ""
    s = re.sub(r"[–—−]", "-", s)
    s_noacc = unicodedata.normalize("NFKD", s)
    s_noacc = "".join(ch for ch in s_noacc if not unicodedata.combining(ch))
    s_noacc = re.sub(r"\s+", " ", s_noacc)
    s_noacc = re.sub(r"\s*-\s*", " - ", s_noacc).strip()
    return s_noacc.upper()

# versão normalizada por owner
CODIGOS_CIDADES_NORM_BY_OWNER = {}
for owner_key, mapa in CODIGOS_CIDADES_BY_OWNER.items():
    CODIGOS_CIDADES_NORM_BY_OWNER[owner_key] = {_norm_city_key(k): v for k, v in mapa.items()}

# usado quando não achar produtor
PRODUTOR_PADRAO = "Produtor"

# -----------------------------
# Janelas auxiliares (Loading / Config / Associação)
# -----------------------------
class GlobalProgress:
    _dlg = None

    @classmethod
    def _ensure(cls, parent=None):
        if cls._dlg is None:
            cls._dlg = QProgressDialog("", "Cancelar", 0, 0, parent)
            cls._dlg.setWindowTitle("Processando…")
            cls._dlg.setAutoClose(False)
            cls._dlg.setAutoReset(False)
            cls._dlg.setWindowModality(Qt.ApplicationModal)

    @classmethod
    def begin(cls, texto: str, maximo: int = 0, parent=None):
        cls._ensure(parent or QApplication.activeWindow())
        dlg = cls._dlg
        dlg.setLabelText(texto)
        dlg.setRange(0, maximo if maximo and maximo > 0 else 0)
        dlg.setValue(0)
        dlg.show()
        QCoreApplication.processEvents()

    @classmethod
    def set_max(cls, maximo: int):
        cls._ensure()
        cls._dlg.setRange(0, maximo if maximo and maximo > 0 else 0)
        QCoreApplication.processEvents()

    @classmethod
    def set_value(cls, valor: int):
        if cls._dlg:
            cls._dlg.setValue(valor)
            QCoreApplication.processEvents()

    @classmethod
    def step(cls, inc: int = 1):
        if not cls._dlg:
            return
        if cls._dlg.maximum() == 0:
            return
        cls._dlg.setValue(cls._dlg.value() + (inc or 1))
        QCoreApplication.processEvents()

    @classmethod
    def end(cls):
        if cls._dlg:
            cls._dlg.reset()
            cls._dlg.hide()
            QCoreApplication.processEvents()

class ConfigDialog(QDialog):
    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.setWindowTitle("⚙️ Configurações")
        self.setFixedSize(560, 320)
        self.config = config

        lay = QVBoxLayout(self)
        self.setObjectName("cfgDlg")
        self.setStyleSheet("#cfgDlg QLineEdit, #cfgDlg QComboBox, #cfgDlg QDateEdit, #cfgDlg QTextEdit { border:none; }")

        grp = QGroupBox("Caminhos de Trabalho")
        grp.setObjectName("cfgGrp")
        grp.setStyleSheet("#cfgGrp{border:1px solid #1e5a9c; border-radius:12px; background:transparent;} #cfgGrp::title{left:10px; padding:0 6px; color:#E0E0E0;}")
        form = QFormLayout(grp)

        self.excel_path_edit = QLineEdit(self.config.get('excel_path', ''))
        self.excel_path_edit.setPlaceholderText("Caminho para a planilha Excel (RELATORIO)")
        btn_excel = QPushButton("Procurar"); btn_excel.clicked.connect(self.browse_excel)
        row1 = QHBoxLayout(); row1.addWidget(self.excel_path_edit); row1.addWidget(btn_excel)
        form.addRow("Planilha Excel:", row1)

        self.isento_path_edit = QLineEdit(self.config.get('isento_path', ''))
        self.isento_path_edit.setPlaceholderText("Pasta com XMLs para identificar ISENTO")
        btn_isento = QPushButton("Procurar"); btn_isento.clicked.connect(self.browse_isento)
        row2 = QHBoxLayout(); row2.addWidget(self.isento_path_edit); row2.addWidget(btn_isento)
        form.addRow("Pasta XMLs ISENTO:", row2)

        self.notas_receb_path_edit = QLineEdit(self.config.get('notas_recebidas_path', ''))
        self.notas_receb_path_edit.setPlaceholderText("Caminho para a planilha NOTAS RECEBIDAS.xlsx")
        btn_notas = QPushButton("Procurar"); btn_notas.clicked.connect(self.browse_notas_recebidas)
        row3 = QHBoxLayout(); row3.addWidget(self.notas_receb_path_edit); row3.addWidget(btn_notas)
        form.addRow("Planilha NOTAS RECEBIDAS:", row3)

        lay.addWidget(grp)
        btns = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept); btns.rejected.connect(self.reject)
        lay.addWidget(btns)

    def browse_excel(self):
        file, _ = QFileDialog.getOpenFileName(self, "Selecione a planilha Excel", "", "Excel (*.xlsx)")
        if file: self.excel_path_edit.setText(file)

    def browse_isento(self):
        folder = QFileDialog.getExistingDirectory(self, "Selecione a pasta de XMLs ISENTOS")
        if folder: self.isento_path_edit.setText(folder)

    def browse_notas_recebidas(self):
        file, _ = QFileDialog.getOpenFileName(self, "Selecione a planilha NOTAS RECEBIDAS", "", "Excel (*.xlsx)")
        if file:
            self.notas_receb_path_edit.setText(file)

    def get_config(self):
        return {
            'excel_path': self.excel_path_edit.text(),
            'isento_path': self.isento_path_edit.text(),
            'notas_recebidas_path': self.notas_receb_path_edit.text(),
        }

class AssocPagDialog(QDialog):
    def __init__(self, base_default="", testes_default="", parent=None):
        super().__init__(parent)
        self.setWindowTitle("🔗 Associar Pagamentos")
        self.setModal(True)
        self.setFixedSize(640, 190)

        layout = QVBoxLayout(self)
        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignRight)

        self.base_edit = QLineEdit(base_default)
        btn_base = QPushButton("..."); btn_base.setFixedWidth(40); btn_base.clicked.connect(self.browse_base)
        row_base = QHBoxLayout(); row_base.addWidget(self.base_edit, 1); row_base.addWidget(btn_base, 0)
        form.addRow("Planilha BASE DE DADOS:", row_base)

        self.testes_edit = QLineEdit(testes_default)
        btn_testes = QPushButton("..."); btn_testes.setFixedWidth(40); btn_testes.clicked.connect(self.browse_testes)
        row_testes = QHBoxLayout(); row_testes.addWidget(self.testes_edit, 1); row_testes.addWidget(btn_testes, 0)
        form.addRow("Planilha TESTES (RELATÓRIO):", row_testes)

        layout.addLayout(form)
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.validate_and_accept); btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _suggest_dir(self, path_text: str) -> str:
        try:
            p = Path(path_text)
            if p.exists(): return str(p.parent if p.is_file() else p)
            if p.parent.exists(): return str(p.parent)
        except Exception:
            pass
        return ""

    def browse_base(self):
        start_dir = self._suggest_dir(self.base_edit.text())
        file, _ = QFileDialog.getOpenFileName(self, "Selecione a planilha BASE DE DADOS", start_dir, "Excel (*.xlsx)")
        if file: self.base_edit.setText(file)

    def browse_testes(self):
        start_dir = self._suggest_dir(self.testes_edit.text())
        file, _ = QFileDialog.getOpenFileName(self, "Selecione a planilha TESTES (RELATÓRIO)", start_dir, "Excel (*.xlsx)")
        if file: self.testes_edit.setText(file)

    def validate_and_accept(self):
        base_file = self.base_edit.text().strip()
        testes_file = self.testes_edit.text().strip()
        if not base_file or not testes_file:
            QMessageBox.warning(self, "Campos obrigatórios", "Informe os dois arquivos (.xlsx)."); return
        if Path(base_file).suffix.lower() != ".xlsx" or Path(testes_file).suffix.lower() != ".xlsx":
            QMessageBox.warning(self, "Formato inválido", "Os arquivos devem ser .xlsx."); return
        if not Path(base_file).exists():
            QMessageBox.warning(self, "Arquivo não encontrado", f"Base de dados não existe:\n{base_file}"); return
        if not Path(testes_file).exists():
            QMessageBox.warning(self, "Arquivo não encontrado", f"Planilha de testes/relatório não existe:\n{testes_file}"); return
        self._paths = (base_file, testes_file); self.accept()

    def get_paths(self):
        return getattr(self, "_paths", ("", ""))

# -----------------------------
# App principal
# -----------------------------
class RuralXmlImporter(QWidget):
    def __init__(self):
        super().__init__()

        # Estado
        self.proc = None
        self._cancel_import = False
        self.loading_window = None
        self.isento_keys = {}
        self.key_xml = {}
        self.stat_total = 0
        self.stat_ok = 0
        self.stat_err = 0

        self.setWindowTitle("Importador Rural de XML")
        self.resize(940, 700)
        self.setWindowIcon(QIcon(str(ICON_PATH)))

        self._apply_global_styles()

        root = QVBoxLayout(self)
        root.setContentsMargins(14, 14, 14, 14)
        root.setSpacing(12)

        header = self._build_header()
        root.addWidget(header)

        controls_card = self._build_controls_card()
        stats_card = self._build_stats_card()

        top_row = QHBoxLayout()
        top_row.setSpacing(12)
        top_row.addWidget(controls_card, 3)
        top_row.addWidget(stats_card, 2)
        root.addLayout(top_row)

        self.splitter = QSplitter(Qt.Vertical)
        self.splitter.setChildrenCollapsible(False)

        log_card = self._build_log_card()
        self.splitter.addWidget(log_card)
        self.splitter.setStretchFactor(0, 1)
        self.splitter.setStretchFactor(1, 3)
        root.addWidget(self.splitter)

        footer = QLabel("🌱 Desenvolvido para produtores rurais — v.1.0")
        footer.setAlignment(Qt.AlignCenter)
        footer.setStyleSheet("font-size:11px; color:#7F7F7F; padding-top:4px;")
        root.addWidget(footer)

        self.config = self.load_config()
        self.base_dados_path = self.config.get('base_dados_path', '')
        self.testes_path = self.config.get('testes_path', '')
        self.excel_path = self.config.get('excel_path', r"\\rilkler\LIVRO CAIXA\TESTE\TESTES.xlsx")
        self.isento_path = self.config.get('isento_path', '')
        self.notas_recebidas_path = self.config.get('notas_recebidas_path', '')
        # PERFIL ATIVO (não muda interface; só vem do config.json)
        self.active_owner = self.config.get('active_owner', 'CLEUBER').upper()
        if self.active_owner not in OWNERS:
            self.active_owner = 'CLEUBER'

        # >>> ADICIONE ESTA LINHA:
        self._apply_owner_paths(self.active_owner)

    # ---------- UI helpers ----------
    def _apply_global_styles(self):
        self.setStyleSheet(STYLE_SHEET)

    def _add_shadow(self, widget: QWidget, radius=18, blur=24, color=QColor(0,0,0,60), y_offset=6):
        eff = QGraphicsDropShadowEffect(self)
        eff.setBlurRadius(blur)
        eff.setColor(color)
        eff.setOffset(0, y_offset)
        widget.setGraphicsEffect(eff)
        widget.setStyleSheet(widget.styleSheet() + f"; border-radius:{radius}px;")

    def _build_header(self) -> QFrame:
        header = QFrame()
        header.setStyleSheet("QFrame{border:1px solid #1e5a9c; border-radius:16px;}")

        lay = QHBoxLayout(header)
        lay.setContentsMargins(18, 16, 18, 16)
        lay.setSpacing(14)

        icon = QLabel()
        if ICON_PATH.exists():
            pix = QPixmap(str(ICON_PATH)).scaled(44, 44, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            icon.setPixmap(pix)
            icon.setStyleSheet("border:none;")
        else:
            icon.setText("🚜")
            icon.setStyleSheet("font-size:34px; border:none;")
        lay.addWidget(icon, 0, Qt.AlignVCenter)

        title = QLabel("IMPORTADOR RURAL DE XML")
        f = QFont(); f.setPointSize(20); f.setBold(True)
        title.setFont(f)
        subtitle = QLabel("Importe notas, associe pagamentos e acompanhe tudo em tempo real.")

        title.setStyleSheet("border:none;")
        subtitle.setStyleSheet("border:none;")

        title_box = QVBoxLayout()
        title_box.addWidget(title)
        title_box.addWidget(subtitle)

        lay.addLayout(title_box, 1)

        btn_cfg = QToolButton()
        btn_cfg.setText("⚙️ Configurar")
        btn_cfg.clicked.connect(self.open_config)

        btn_help = QToolButton()
        btn_help.setText("❓ Ajuda")
        btn_help.clicked.connect(lambda: QMessageBox.information(
            self, "Ajuda",
            "1) Clique em ⚙️ Configurar para definir planilha e parâmetros.\n"
            "2) Use 🔗 Associar Pagamentos para gerar relatórios e TXT.\n"
            "3) Use 📤 Importar XMLs para lançar notas na planilha.\n"
            "4) Acompanhe os logs e use 'Salvar Log' para guardar o histórico."
        ))

        btn_close = QToolButton()
        btn_close.setText("✖ Fechar")
        btn_close.clicked.connect(self._close_self_tab)

        row = QHBoxLayout()
        row.setSpacing(8)
        row.addWidget(btn_cfg)
        row.addWidget(btn_help)
        row.addWidget(btn_close)
        lay.addLayout(row, 0)

        self._add_shadow(header, radius=16, blur=24, color=QColor(0,0,0,50), y_offset=5)
        return header

    def _apply_owner_paths(self, owner: str):
        """
        Ajusta excel_path/testes_path/isento_path conforme o dono ativo
        e salva no json/config.json para refletir na interface.
        """
        owner = (owner or "CLEUBER").upper()

        excel_file = rf"\\rilkler\LIVRO CAIXA\TESTE\LIVRO CAIXA {owner}.xlsx"
        self.config["excel_path"]  = excel_file
        self.config["testes_path"] = excel_file

        self.config["isento_path"] = (
            rf"C:\Users\conta\OneDrive\Área de Trabalho\Documentos Automacao\NOTAS LIVRO CAIXA\{owner}"
        )

        # mantém se já houver em config; não forçamos mudança
        # self.config.setdefault("notas_recebidas_path", r"\\rilkler\LIVRO CAIXA\TESTE\NOTAS RECEBIDAS.xlsx")

        # Atualiza atributos usados pela UI/lógica
        self.excel_path  = self.config.get('excel_path', '')
        self.testes_path = self.config.get('testes_path', '')
        self.isento_path = self.config.get('isento_path', '')

        try:
            self.save_config()  # persiste para o diálogo de Configurações ler já atualizado
        except Exception:
            pass

    def _close_self_tab(self):
        parent = self.parent()
        while parent and not isinstance(parent, QTabWidget):
            parent = parent.parent()
        if parent:
            idx = parent.indexOf(self)
            if idx != -1:
                parent.removeTab(idx)
        else:
            self.close()

    def _build_controls_card(self) -> QFrame:
        card = QFrame()
        card.setStyleSheet("QFrame{border:1px solid #1e5a9c; border-radius:12px;}")

        lay = QVBoxLayout(card)
        lay.setContentsMargins(14, 12, 14, 12)
        lay.setSpacing(10)

        actions = QHBoxLayout()
        actions.setSpacing(10)

        self.btn_import = QPushButton("📤 Importar XMLs")
        self.btn_import.clicked.connect(self.import_xmls)
        actions.addWidget(self.btn_import)

        self.btn_assoc = QPushButton("🔗 Associar Pagamentos")
        self.btn_assoc.clicked.connect(self.associar_pagamentos)
        actions.addWidget(self.btn_assoc)

        self.btn_import_lanc = QPushButton("📥 Importar Lançamentos")
        self.btn_import_lanc.clicked.connect(self.importar_lancamentos_simples)
        actions.addWidget(self.btn_import_lanc)

        self.btn_cancel = QPushButton("⛔ Cancelar")
        self.btn_cancel.setEnabled(False)
        self.btn_cancel.setObjectName("danger")
        self.btn_import.setObjectName("success")
        self.btn_cancel.clicked.connect(self.cancelar_processos)
        actions.addWidget(self.btn_cancel)

        actions.addStretch()

        self.btn_log_clear = QToolButton()
        self.btn_log_clear.setText("🧹 Limpar Log")
        self.btn_log_clear.setStyleSheet("QToolButton{background:#0d1b3d; border:1px solid #1e5a9c; border-radius:8px; padding:6px 10px;} QToolButton:hover{background:#123764;}")
        self.btn_log_clear.clicked.connect(self._log_clear)
        actions.addWidget(self.btn_log_clear)

        self.btn_log_save = QToolButton()
        self.btn_log_save.setText("💾 Salvar Log")
        self.btn_log_save.setStyleSheet("QToolButton{background:#0d1b3d; border:1px solid #1e5a9c; border-radius:8px; padding:6px 10px;} QToolButton:hover{background:#123764;}")
        self.btn_log_save.clicked.connect(self._log_save)
        actions.addWidget(self.btn_log_save)

        lay.addLayout(actions)

        opt = QHBoxLayout()
        opt.setSpacing(12)
        self.chk_delete = QCheckBox("Excluir notas existentes antes de importar")
        opt.addWidget(self.chk_delete)
        opt.addStretch()
        lay.addLayout(opt)

        self._add_shadow(card, radius=14, blur=20, color=QColor(0,0,0,45), y_offset=4)
        return card

    def _build_stats_card(self) -> QFrame:
        card = QFrame()
        card.setObjectName("statsCard")
        card.setStyleSheet("#statsCard{border:1px solid #1e5a9c; border-radius:14px;} #statsCard *{border:none; background:transparent;}")
        lay = QVBoxLayout(card)
        lay.setContentsMargins(14, 12, 14, 12)
        lay.setSpacing(6)

        title = QLabel("📊 Último Status da Sessão")
        f = QFont(); f.setPointSize(12); f.setBold(True)
        title.setStyleSheet("")
        lay.addWidget(title)

        self.lbl_last_status = QLabel("—")
        self.lbl_last_status.setStyleSheet("font-weight:600; border:none; background:transparent;")
        self.lbl_last_status_time = QLabel(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        self.lbl_last_status_time.setStyleSheet("border:none; background:transparent;")
        self.lbl_last_status_time.setAlignment(Qt.AlignRight)
        status_row = QHBoxLayout()
        status_row.setSpacing(10)
        status_row.addWidget(self.lbl_last_status)
        status_row.addStretch()
        status_row.addWidget(self.lbl_last_status_time)
        lay.addLayout(status_row)

        chips = QHBoxLayout()
        chips.setSpacing(10)

        self.lbl_stat_total = self._make_chip("Total", "#2B2F31", "#E0E0E0")
        self.lbl_stat_ok    = self._make_chip("Sucesso", "#183d2a", "#A7F3D0")
        self.lbl_stat_err   = self._make_chip("Erros", "#3b1f1f", "#FF6B6B")

        chips.addWidget(self.lbl_stat_total)
        chips.addWidget(self.lbl_stat_ok)
        chips.addWidget(self.lbl_stat_err)
        chips.addStretch()

        lay.addLayout(chips)

        self._add_shadow(card, radius=14, blur=20, color=QColor(0,0,0,45), y_offset=4)
        return card

    def _make_chip(self, label: str, bg: str, fg: str) -> QLabel:
        w = QLabel(f"{label}: 0")
        w.setAlignment(Qt.AlignCenter)
        w.setStyleSheet(f"QLabel {{ background:{bg}; color:{fg}; border-radius:10px; padding:8px 12px; font-weight:600; }}")
        return w

    def _build_log_card(self) -> QFrame:
        card = QFrame()
        card.setObjectName("logCard")
        card.setStyleSheet("#logCard{background:#212425; border:1px solid #1e5a9c; border-radius:10px;} #logCard QLabel{border:none; background:transparent; color:#E0E0E0;}")
        lay = QVBoxLayout(card)
        lay.setContentsMargins(12, 10, 12, 12)
        lay.setSpacing(8)

        title = QLabel("📝 Histórico")
        f = QFont(); f.setBold(True); f.setPointSize(12)
        title.setFont(f)
        title.setStyleSheet("padding:2px 6px;")
        lay.addWidget(title, alignment=Qt.AlignLeft)

        body = QFrame()
        body.setObjectName("logBody")
        body.setStyleSheet("#logBody{background:#2B2F31; border:none; border-radius:8px;}")
        body_lay = QVBoxLayout(body)
        body_lay.setContentsMargins(12, 12, 12, 12)
        body_lay.setSpacing(0)

        self.log = QTextEdit(readOnly=True)
        self.log.setMinimumHeight(260)
        self.log.setFrameStyle(QFrame.NoFrame)
        self.log.setStyleSheet("QTextEdit{background:transparent; border:none;} QTextEdit::viewport{background:transparent; border:none;}")

        body_lay.addWidget(self.log)
        lay.addWidget(body)

        return card

    def _update_stats(self):
        self.lbl_stat_total.setText(f"Total: {self.stat_total}")
        self.lbl_stat_ok.setText(f"Sucesso: {self.stat_ok}")
        self.lbl_stat_err.setText(f"Erros: {self.stat_err}")

    # ---------- Utilidades de Log ----------
    def _log_clear(self):
        self.log.clear()
        self.log_msg("Log limpo.", "info")

    def _log_save(self):
        try:
            out_dir = Path(__file__).parent / "logs"
            out_dir.mkdir(exist_ok=True, parents=True)
            fname = out_dir / f"importador_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            with open(fname, "w", encoding="utf-8") as f:
                f.write(self.log.toPlainText())
            self.log_msg(f"Log salvo em: {fname}", "success")
        except Exception as e:
            self.log_msg(f"Falha ao salvar log: {e}", "error")

    def _log_find_next(self):
        term = self.search_edit.text().strip()
        if not term:
            return
        doc = self.log.document()
        cursor = self.log.textCursor()
        pos = cursor.position()
        found = doc.find(term, pos)
        if not found.isNull():
            self.log.setTextCursor(found)
            self.log.ensureCursorVisible()
        else:
            start = QTextCursor(doc)
            start.movePosition(QTextCursor.Start)
            found = doc.find(term, start)
            if not found.isNull():
                self.log.setTextCursor(found)
                self.log.ensureCursorVisible()
            else:
                self.log_msg(f"'{term}' não encontrado no log.", "warning")

    # ---------- Varredura NFE (QProcess) ----------
    def run_varredura(self, base_file: str, testes_file: str):
        try:
            script_path = Path(__file__).parent / "Varredura NFE.py"

            if self.proc and self.proc.state() != QProcess.NotRunning:
                self.log_msg("Já existe uma varredura em execução.", "warning")
                return

            if getattr(self, "proc", None):
                try:
                    if self.proc.state() != QProcess.NotRunning:
                        self.proc.kill()
                        self.proc.waitForFinished(800)
                except Exception:
                    pass
                self.proc.deleteLater()

            self.proc = QProcess(self)
            self.proc.setWorkingDirectory(str(script_path.parent))
            self.proc.setProcessChannelMode(QProcess.MergedChannels)

            self.proc.started.connect(self.on_proc_started)
            self.proc.readyReadStandardOutput.connect(self.on_proc_output)
            self.proc.readyReadStandardError.connect(self.on_proc_output)
            self.proc.finished.connect(self.on_proc_finished)
            self.proc.errorOccurred.connect(self.on_proc_error)

            env = QProcessEnvironment.systemEnvironment()
            env.insert("PYTHONUTF8", "1")
            env.insert("PYTHONIOENCODING", "utf-8")
            self.proc.setProcessEnvironment(env)

            program = sys.executable
            args = ["-u", str(script_path), base_file, testes_file]

            nr_cfg = self.config.get("notas_recebidas_path", "") or getattr(self, "notas_recebidas_path", "")
            if nr_cfg:
                p = Path(nr_cfg)
                if p.exists() and p.suffix.lower() == ".xlsx":
                    args.append(str(p))
                else:
                    self.log_msg(f"AVISO: caminho de NOTAS RECEBIDAS inválido ou inexistente:\n{nr_cfg}", "warning")

            self.btn_assoc.setEnabled(False)
            self.btn_cancel.setEnabled(True)
            self.log_msg("Abrindo 'Varredura NFE.py' para associar pagamentos/recebimentos.", "info")
            self.proc.start(program, args)

        except Exception as e:
            self.log_msg(f"Erro ao iniciar varredura: {e}", "error")
            self.btn_assoc.setEnabled(True)
            self.btn_cancel.setEnabled(False)

    def on_proc_started(self):
        self.log_msg("✅ Varredura iniciada.", "success")

    def on_proc_output(self):
        try:
            data = bytes(self.proc.readAllStandardOutput())
            if not data:
                data = bytes(self.proc.readAllStandardError())
            if data:
                text = data.decode("utf-8", errors="ignore")
                self._append_plain(text)
        except Exception as e:
            self.log_msg(f"Falha lendo saída do processo: {e}", "error")

    def on_proc_finished(self, exit_code: int, exit_status):
        if exit_code == 0:
            self.log_msg("Varredura finalizada com sucesso.", "success")
            try:
                pag_file = Path(__file__).parent / "PAGAMENTOS.txt"
                if pag_file.exists():
                    with open(pag_file, "r", encoding="utf-8") as f:
                        lines = [ln for ln in (l.strip() for l in f) if ln]
                    count = len(lines)
                    if count > 0:
                        plural = "s" if count != 1 else ""
                        self.lbl_last_status.setText(f"{count} nota{plural} associada{plural} ao pagamento ✅")
                    else:
                        self.lbl_last_status.setText("Nenhuma nota associada ao pagamento")
                else:
                    self.lbl_last_status.setText("Nenhum arquivo PAGAMENTOS.txt encontrado")
            except Exception as e:
                self.lbl_last_status.setText("Erro ao contar notas associadas")
                self.log_msg(f"Erro ao ler PAGAMENTOS.txt para contar notas: {e}", "error")

            try:
                from pathlib import Path as _Path
                main_win = self.window()
                if main_win and hasattr(main_win, "_import_lancamentos_txt"):
                    resp = QMessageBox.question(
                        self,
                        "Gerar Lançamentos",
                        "Gerar Lançamentos para o sistema?",
                        QMessageBox.Yes | QMessageBox.No
                    )
                    if resp == QMessageBox.Yes:
                        pag_file = _Path(__file__).parent / "PAGAMENTOS.txt"
                        if pag_file.exists():
                            try:
                                main_win._import_lancamentos_txt(str(pag_file))
                                main_win.carregar_lancamentos()
                                main_win.dashboard.load_data()
                                QMessageBox.information(self, "Concluído", "Lançamentos importados com sucesso.")
                            except Exception as _e:
                                QMessageBox.critical(self, "Erro ao importar", f"Falha ao importar PAGAMENTOS.txt:\n{_e}")
                        else:
                            QMessageBox.warning(self, "Arquivo não encontrado", f"Não encontrei o arquivo:\n{pag_file}")

                        receb_file = _Path(__file__).parent / "RECEBIMENTOS.txt"
                        if receb_file.exists():
                            try:
                                main_win._import_lancamentos_txt(str(receb_file))
                                main_win.carregar_lancamentos()
                                main_win.dashboard.load_data()
                                QMessageBox.information(self, "Concluído", "Recebimentos importados com sucesso.")
                            except Exception as _e:
                                QMessageBox.critical(self, "Erro ao importar", f"Falha ao importar RECEBIMENTOS.txt:\n{_e}")
                        else:
                            self.log_msg(f"RECEBIMENTOS.txt não encontrado em {receb_file}", "warning")
            except Exception:
                pass
        else:
            self.log_msg(f"Varredura finalizada com código {exit_code}.", "error")
            self.lbl_last_status.setText("ERRO POR CONTA DISSO E DAQUILO")
        self.lbl_last_status_time.setText(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        self.btn_assoc.setEnabled(True)
        self.btn_cancel.setEnabled(False)

    def on_proc_error(self, err):
        self.log_msg(f"Erro ao executar varredura: {err}", "error")
        self.btn_assoc.setEnabled(True)
        self.btn_cancel.setEnabled(False)

    def _append_plain(self, text: str):
        if not text:
            return
        text = text.replace('\r\n', '\n').replace('\r', '\n')
        self.log.moveCursor(QTextCursor.End)
        self.log.insertPlainText(text)
        self.log.moveCursor(QTextCursor.End)
        self.log.ensureCursorVisible()

    def cancelar_processos(self):
        cancelou = False
        if getattr(self, "proc", None) and self.proc.state() != QProcess.NotRunning:
            self.log_msg("Cancelando varredura de pagamentos...", "warning")
            self.proc.terminate()
            if not self.proc.waitForFinished(1500):
                self.proc.kill()
                self.proc.waitForFinished(1500)
            cancelou = True

        if GlobalProgress._dlg:
            self.log_msg("Cancelando importação de XMLs...", "warning")
            self._cancel_import = True
            GlobalProgress.end()
            cancelou = True

        if cancelou:
            self.btn_cancel.setEnabled(False)
            self.log_msg("Processo(s) cancelado(s).", "success")
            self.lbl_last_status.setText("PROCESSO CANCELADO PELO USUARIO")
            self.lbl_last_status_time.setText(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        else:
            self.log_msg("Nenhum processo em execução para cancelar.", "info")

    # ---------- Ações principais ----------
    def associar_pagamentos(self):
        try:
            script_path = Path(__file__).parent / "Varredura NFE.py"
            if not script_path.exists():
                msg = f"Arquivo não encontrado:\n{script_path}"
                self.log_msg("Script 'Varredura NFE.py' não encontrado.", "error")
                QMessageBox.warning(self, "Arquivo não encontrado", msg)
                return

            base_default = self.config.get("base_dados_path", "")
            testes_default = self.config.get("testes_path", "")

            dlg = AssocPagDialog(base_default=base_default, testes_default=testes_default, parent=self)
            if dlg.exec() != QDialog.Accepted:
                self.log_msg("Associação cancelada pelo usuário.", "warning")
                self.lbl_last_status.setText("PROCESSO CANCELADO PELO USUARIO")
                self.lbl_last_status_time.setText(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
                return
            base_file, testes_file = dlg.get_paths()

            self.config["base_dados_path"] = base_file
            self.config["testes_path"] = testes_file
            self.save_config()
            self.base_dados_path = base_file
            self.testes_path = testes_file
            self.log_msg("Caminhos salvos em json/config.json", "success")

            self.run_varredura(base_file, testes_file)

        except Exception as e:
            self.log_msg(f"Falha ao preparar 'Varredura NFE.py': {e}", "error")
            QMessageBox.critical(self, "Erro ao executar", f"Ocorreu um erro:\n{e}")

    def import_xmls(self):
        try:
            start_dir = self.config.get("last_xml_dir", os.getcwd())
            files, _ = QFileDialog.getOpenFileNames(self, "Selecione os arquivos XML", start_dir, "XML (*.xml)")
            if not files:
                self.log_msg("Nenhum arquivo selecionado. Operação cancelada.", "warning")
                return

            try:
                self.config["last_xml_dir"] = str(Path(files[0]).parent)
                self.save_config()
            except Exception:
                pass

            self.stat_total = len(files)
            self.stat_ok = 0
            self.stat_err = 0
            self._update_stats()

            self._cancel_import = False
            self.btn_cancel.setEnabled(True)

            GlobalProgress.begin("Importando XMLs…", maximo=len(files), parent=self)
            QTimer.singleShot(0, lambda: self.process_files(files))

            self.log_msg(f"{len(files)} arquivo(s) selecionado(s) para importação.", "info")
        except Exception as e:
            self.log_msg(f"Erro ao iniciar importação: {e}", "error")
            QMessageBox.critical(self, "Erro", f"Ocorreu um erro ao iniciar a importação:\n{e}")

    def importar_lancamentos_simples(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Importar Lançamentos",
            "",
            "Textos (*.txt *.TXT);;Planilhas Excel (*.xlsx *.xls);;Todos os arquivos (*.*)"
        )
        if not path:
            return

        try:
            main_win = self.window()
            if not main_win or not hasattr(main_win, "_import_lancamentos_txt"):
                QMessageBox.warning(self, "Aviso", "Janela principal não disponível para importar.")
                return

            if path.lower().endswith(".txt"):
                main_win._import_lancamentos_txt(path)
            else:
                main_win._import_lancamentos_excel(path)

            if hasattr(main_win, "carregar_lancamentos"):
                main_win.carregar_lancamentos()
            if hasattr(main_win, "dashboard"):
                try:
                    main_win.dashboard.load_data()
                except Exception:
                    pass

            self.log_msg(f"Lançamentos importados de {os.path.basename(path)}", "success")
        except Exception as e:
            QMessageBox.warning(self, "Importação Falhou", f"{e}")

    # ---------- Config ----------
    def load_config(self):
        config_dir = Path(__file__).parent / "json"
        config_file = config_dir / "config.json"
        config_dir.mkdir(parents=True, exist_ok=True)
        if config_file.exists():
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                self.log_msg(f"Erro ao carregar configurações: {e}", "error")
                return {}
        return {}

    def save_config(self):
        config_dir = Path(__file__).parent / "json"
        config_file = config_dir / "config.json"
        try:
            with open(config_file, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=4)
            self.log_msg("Configurações salvas com sucesso", "success")
        except Exception as e:
            self.log_msg(f"Erro ao salvar configurações: {e}", "error")

    def open_config(self):
        dialog = ConfigDialog(self.config, self)
        if dialog.exec() == QDialog.Accepted:
            new_cfg = dialog.get_config() or {}
            self.config.update(new_cfg)
            self.notas_recebidas_path = self.config.get('notas_recebidas_path', '')
            self.excel_path  = self.config.get('excel_path', '')
            self.isento_path = self.config.get('isento_path', '')
            # mantém active_owner se já existe
            if 'active_owner' not in self.config:
                self.config['active_owner'] = self.active_owner
            self.save_config()
            self.log_msg("Configurações atualizadas", "success")

    # ---------- Log / mensagens ----------
    def log_msg(self, message: str, msg_type: str = "info"):
        now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        palette = {
            "info":   {"emoji": "ℹ️", "text": "#FFFFFF", "accent": "#3A3C3D", "weight": "500"},
            "success":{"emoji": "✅", "text": "#A7F3D0", "accent": "#2F7D5D", "weight": "700"},
            "warning":{"emoji": "⚠️", "text": "#FFFFFF", "accent": "#8A6D3B", "weight": "600"},
            "error":  {"emoji": "❌", "text": "#FF6B6B", "accent": "#7A2E2E", "weight": "800"},
            "title":  {"emoji": "📌", "text": "#FFFFFF", "accent": "#1e5a9c", "weight": "800"},
            "divider":{"emoji": "",   "text": "",       "accent": "#3A3C3D", "weight": "400"},
        }

        if msg_type == "divider":
            self.log.append('<div style="border-top:1px solid #3A3C3D; margin:10px 0;"></div>')
            return

        p = palette.get(msg_type, palette["info"])

        html = (
            f'<div style="border-left:3px solid {p["accent"]};'
            f' padding:6px 10px; margin:2px 0;">'
            f'<span style="opacity:.7; font-family:monospace;">[{now}]</span>'
            f' <span style="margin:0 6px 0 8px;">{p["emoji"]}</span>'
            f'<span style="color:{p["text"]}; font-weight:{p["weight"]};">{message}</span>'
            f'</div>'
        )
        self.log.append(html)

    # ---------- OpenPyXL helpers ----------
    def copy_row_style(self, ws, src_row, dest_row, cols):
        for col in cols:
            src = ws.cell(row=src_row, column=col)
            dest = ws.cell(row=dest_row, column=col)
            dest.font = copy(src.font)
            dest.border = copy(src.border)
            dest.fill = copy(src.fill)
            dest.protection = copy(src.protection)
            dest.alignment = copy(src.alignment)
            dest.number_format = '"R$"#,##0.00;[Red]"R$"-#,##0.00' if col in [13,14] else copy(src.number_format)
        ws.row_dimensions[dest_row].height = ws.row_dimensions[src_row].height

    def extend_table(self, ws, header_row, new_row):
        for table in ws.tables.values():
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            if min_row == header_row:
                table.ref = f"{get_column_letter(min_col)}{header_row}:{get_column_letter(max_col)}{new_row}"
                self.log_msg(f"Tabela '{table.name}' estendida até {new_row}.", "info")

    # ---------- DETECÇÃO DE DONO / PERFIL ----------
    def _resolve_owner_from_parties(self,
                                    emit_id: str,
                                    dest_id: str,
                                    emit_ie: str,
                                    dest_ie: str,
                                    emit_city: str,
                                    emit_uf: str,
                                    dest_city: str,
                                    dest_uf: str) -> tuple:
        """
        Retorna (owner_key, owner_side, farm_ie, farm_name)

        owner_key: "CLEUBER" | "GILSON" | "ADRIANA" | "LUCAS" | None
        owner_side: "emit" | "dest" | None
        farm_ie: ie usado
        farm_name: nome da fazenda (ou 'ISENTO')
        """
        active = (self.config.get('active_owner') or self.active_owner or 'CLEUBER').upper()
        if active not in OWNERS:
            active = 'CLEUBER'

        emit_id_d = _digits(emit_id)
        dest_id_d = _digits(dest_id)

        emit_matches = [k for k, v in OWNERS.items() if emit_id_d and emit_id_d == v["CPF_D"]]
        dest_matches = [k for k, v in OWNERS.items() if dest_id_d and dest_id_d == v["CPF_D"]]

        # 1) se os dois lados são produtores → prioriza perfil ativo
        if emit_matches and dest_matches:
            if active in emit_matches:
                owner_key = active
                owner_side = "emit"
            elif active in dest_matches:
                owner_key = active
                owner_side = "dest"
            else:
                # se o perfil ativo não está em nenhum, fica com emit
                owner_key = emit_matches[0]
                owner_side = "emit"
        # 2) só o emitente é produtor
        elif emit_matches:
            owner_key = emit_matches[0]
            owner_side = "emit"
        # 3) só o destinatário é produtor
        elif dest_matches:
            owner_key = dest_matches[0]
            owner_side = "dest"
        else:
            owner_key = None
            owner_side = None

        # agora tenta IE / cidade
        farm_ie = ''
        farm_name = 'ISENTO'

        # se já tenho o dono, uso o mapa dele
        if owner_key:
            farm_map = FARM_MAPPING_BY_OWNER.get(owner_key, {})
            city_map = CODIGOS_CIDADES_BY_OWNER.get(owner_key, {})
            city_norm_map = CODIGOS_CIDADES_NORM_BY_OWNER.get(owner_key, {})

            # IE do lado do dono
            if owner_side == "emit" and emit_ie and emit_ie in farm_map:
                farm_ie = emit_ie
                farm_name = farm_map[emit_ie]
            elif owner_side == "dest" and dest_ie and dest_ie in farm_map:
                farm_ie = dest_ie
                farm_name = farm_map[dest_ie]
            else:
                # tenta cidade (emit/dest conforme lado do dono; se não achar, tenta os dois)
                candidates = []
                if owner_side == "emit":
                    candidates = [f"{emit_city} - {emit_uf}".strip(), emit_city]
                elif owner_side == "dest":
                    candidates = [f"{dest_city} - {dest_uf}".strip(), dest_city]
                else:
                    candidates = [
                        f"{emit_city} - {emit_uf}".strip(), emit_city,
                        f"{dest_city} - {dest_uf}".strip(), dest_city
                    ]
                found = False
                for raw in candidates:
                    if not raw:
                        continue
                    for key_try in self._variants_city(raw, emit_uf or dest_uf):
                        if key_try in city_map:
                            farm_name = city_map[key_try]
                            found = True
                            break
                        k_norm = _norm_city_key(key_try)
                        if k_norm in city_norm_map:
                            farm_name = city_norm_map[k_norm]
                            found = True
                            break
                    if found:
                        break
        else:
            # não sabemos quem é → usa o perfil ativo como fallback de mapeamento de cidade
            fallback_map = CODIGOS_CIDADES_BY_OWNER.get(active, {})
            fallback_norm = CODIGOS_CIDADES_NORM_BY_OWNER.get(active, {})
            candidates = [
                f"{emit_city} - {emit_uf}".strip(), emit_city,
                f"{dest_city} - {dest_uf}".strip(), dest_city
            ]
            for raw in candidates:
                if not raw:
                    continue
                for key_try in self._variants_city(raw, emit_uf or dest_uf):
                    if key_try in fallback_map:
                        farm_name = fallback_map[key_try]
                        break
                    k_norm = _norm_city_key(key_try)
                    if k_norm in fallback_norm:
                        farm_name = fallback_norm[k_norm]
                        break

        return owner_key, owner_side, farm_ie, farm_name

    def _variants_city(self, city: str, uf: str):
        city = (city or '').strip()
        if not city:
            return []
        s_noacc = unicodedata.normalize("NFKD", city)
        s_noacc = ''.join(ch for ch in s_noacc if not unicodedata.combining(ch))
        base = [city, city.upper(), s_noacc, s_noacc.upper()]
        if uf:
            clean = re.sub(r'\s*-\s*[A-Z]{2}$', '', s_noacc, flags=re.I).strip()
            if clean:
                base += [f"{clean} - {uf}", f"{clean.upper()} - {uf}"]
        return base

    # ---------- Importação de XML ----------
    def process_files(self, files):
        self.isento_keys = {}
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            for col in ['M','N']:
                for row in range(7, ws.max_row + 1):
                    cell = ws[f"{col}{row}"]
                    if cell.value is not None:
                        cell.number_format = '"R$"#,##0.00;[Red]"R$"-#,##0.00'
            self.log_msg(f"Planilha aberta com sucesso: {self.excel_path}", "success")
        except Exception as e:
            self.log_msg(f"Falha ao abrir a planilha: {e}", "error")
            QMessageBox.critical(self, "Erro", f"Falha ao abrir a planilha:\n{e}")
            GlobalProgress.end()
            self.btn_cancel.setEnabled(False)
            return

        header = 6
        if ws.cell(header, 3).value is None or str(ws.cell(header, 3).value).strip().upper() != "DATA":
            self.log_msg("Cabeçalho 'DATA' não encontrado na linha 6 da planilha", "error")
            QMessageBox.critical(self, "Erro", "Cabeçalho 'DATA' não encontrado na linha 6.")
            GlobalProgress.end()
            self.btn_cancel.setEnabled(False)
            return

        if self.chk_delete.isChecked():
            self.log_msg("Opção 'Excluir notas existentes' está marcada", "info")
            start_row = header + 1
            if ws.max_row >= start_row:
                rows_to_delete = ws.max_row - start_row + 1
                ws.delete_rows(start_row, rows_to_delete)
                self.log_msg(f"Notas existentes excluídas: {rows_to_delete} linha(s) removida(s)", "success")
            else:
                self.log_msg("Nenhuma nota existente para excluir", "info")
        else:
            self.log_msg("Opção 'Excluir notas existentes' NÃO está marcada", "info")

        start = header + 1
        self.log_msg(f"Importação iniciará na linha: {start}", "info")
        self.log_msg("--------------------------------", "divider")

        cols_style = list(range(3, 18))
        current = start
        last = start - 1
        imported_files = 0
        total_files = len(files)

        for i, xml_file in enumerate(files):
            if self._cancel_import or (GlobalProgress._dlg and GlobalProgress._dlg.wasCanceled()):
                self.log_msg("Importação cancelada pelo usuário", "warning")
                break

            filename = os.path.basename(xml_file)
            GlobalProgress.set_value(i + 1)
            QCoreApplication.processEvents()

            try:
                self.log_msg(f"Processando arquivo: {filename}", "title")
                tree = ET.parse(xml_file)
                root = tree.getroot()
                ns = {'n': 'http://www.portalfiscal.inf.br/nfe'}
                ns_nfse = {'s': 'http://www.sped.fazenda.gov.br/nfse'}

                ide = root.find('.//n:ide', ns)

                is_nfse = False
                if ide is not None:
                    # ========== NFE ==========
                    dh = (ide.findtext('n:dhEmi', default='', namespaces=ns) or '').strip()
                    nNF = (ide.findtext('n:nNF', default='', namespaces=ns) or '').strip()
                    tp  = (ide.findtext('n:tpNF', default='0', namespaces=ns) or '0').strip()
                    nat = (ide.findtext('n:natOp', default='', namespaces=ns) or '').strip()

                    try:
                        dt = datetime.fromisoformat(dh)
                    except Exception:
                        dt = datetime.strptime(dh, "%Y-%m-%dT%H:%M:%S%z")
                    date  = dt.strftime("%d/%m/%Y")
                    month = dt.month
                    year  = dt.year

                    emit = root.find('.//n:emit', ns)
                    dest = root.find('.//n:dest', ns)
                    emit_name = emit.findtext('n:xNome', default='', namespaces=ns) if emit is not None else ''
                    dest_name = dest.findtext('n:xNome', default='', namespaces=ns) if dest is not None else ''
                    emit_id_node = emit.find('n:CNPJ', ns) if emit is not None else None
                    if emit_id_node is None and emit is not None:
                        emit_id_node = emit.find('n:CPF', ns)
                    dest_id_node = dest.find('n:CNPJ', ns) if dest is not None else None
                    if dest_id_node is None and dest is not None:
                        dest_id_node = dest.find('n:CPF', ns)
                    emit_id = emit_id_node.text.strip() if (emit_id_node is not None and emit_id_node.text) else ''
                    dest_id = dest_id_node.text.strip() if (dest_id_node is not None and dest_id_node.text) else ''

                    # fallback geral
                    if not emit_id or not dest_id:
                        all_ids = [n.text.strip() for n in root.findall('.//n:CNPJ', ns) + root.findall('.//n:CPF', ns) if n.text]
                        if not emit_id and all_ids: emit_id = all_ids[0]
                        if not dest_id and len(all_ids) > 1: dest_id = all_ids[-1]

                    emit_ie_node = emit.find('n:IE', ns) if emit is not None else None
                    dest_ie_node = dest.find('n:IE', ns) if dest is not None else None
                    emit_ie = emit_ie_node.text.strip() if (emit_ie_node is not None and emit_ie_node.text) else ''
                    dest_ie = dest_ie_node.text.strip() if (dest_ie_node is not None and dest_ie_node.text) else ''

                    emit_city = emit.findtext('n:enderEmit/n:xMun', default='', namespaces=ns) if emit is not None else ''
                    emit_uf   = emit.findtext('n:enderEmit/n:UF',  default='', namespaces=ns) if emit is not None else ''
                    dest_city = dest.findtext('n:enderDest/n:xMun', default='', namespaces=ns) if dest is not None else ''
                    dest_uf   = dest.findtext('n:enderDest/n:UF',  default='', namespaces=ns) if dest is not None else ''

                    owner_key, owner_side, farm_ie, farm_name = self._resolve_owner_from_parties(
                        emit_id, dest_id, emit_ie, dest_ie, emit_city, emit_uf, dest_city, dest_uf
                    )

                    # dados de produto/total
                    prod  = root.findtext('.//n:det/n:prod/n:xProd', default='', namespaces=ns) or ''
                    cfop  = root.findtext('.//n:det/n:prod/n:CFOP', default='', namespaces=ns) or ''
                    v_total_txt = root.findtext('.//n:ICMSTot/n:vNF', default='0', namespaces=ns) or '0'
                    try:
                        total = float(str(v_total_txt).replace(',', '.'))
                    except Exception:
                        total = 0.0

                    key = (root.findtext('.//n:protNFe//n:chNFe', default='', namespaces=ns)
                           or root.findtext('.//n:infProt/n:chNFe', default='', namespaces=ns) or '')
                    if key:
                        self.key_xml[key] = xml_file

                    # define contraparte e coluna
                    if owner_key and owner_side == "emit":
                        final_name, final_id = dest_name, dest_id
                        valor_col = 13 if tp == '1' else 14
                        operation_type = f"RECEITA ({owner_key} emitente, tpNF={tp})" if tp == '1' else f"DESPESA ({owner_key} emitente, tpNF={tp})"
                    elif owner_key and owner_side == "dest":
                        final_name, final_id = emit_name, emit_id
                        valor_col = 14 if tp == '1' else 13
                        operation_type = f"DESPESA ({owner_key} destinatário, tpNF={tp})" if tp == '1' else f"RECEITA ({owner_key} destinatário, tpNF={tp})"
                    else:
                        final_name, final_id = emit_name, emit_id
                        valor_col = 13 if tp == '1' else 14
                        operation_type = f"RECEITA ({PRODUTOR_PADRAO} não identificado, tpNF={tp})" if tp == '1' else f"DESPESA ({PRODUTOR_PADRAO} não identificado, tpNF={tp})"
                        self.log_msg(f"{PRODUTOR_PADRAO} não identificado como emitente/destinatário. Usando fallback genérico.", "warning")

                    dups = root.findall('.//n:dup', ns)
                    if dups:
                        self.log_msg(f"Nota fiscal {nNF} possui {len(dups)} parcela(s)", "info")
                        for dup in dups:
                            while ws.cell(current, 3).value:
                                current += 1

                            dVenc = dup.find('n:dVenc', ns).text
                            vDup = float(dup.find('n:vDup', ns).text)

                            dt_venc = datetime.strptime(dVenc, "%Y-%m-%d")
                            date_parc = dt_venc.strftime("%d/%m/%Y")
                            month_parc = dt_venc.month
                            year_parc = dt_venc.year

                            src_row = start if current == start else current - 1
                            self.copy_row_style(ws, src_row, current, cols_style)

                            ws.cell(current, 3, date_parc)
                            ws.cell(current, 4, month_parc)
                            ws.cell(current, 5, year_parc)
                            ws.cell(current, 6, nNF)
                            ws.cell(current, 7, final_name)
                            ws.cell(current, 8, farm_ie)
                            ws.cell(current, 9, farm_name)
                            ws.cell(current, 10, final_id)
                            ws.cell(current, 11, prod)
                            ws.cell(current, 12, cfop)
                            ws.cell(current, valor_col, vDup)
                            ws.cell(current, 15, nat)
                            ws.cell(current, 16, key)
                            ws.cell(current, 17, None)

                            if farm_name == "ISENTO":
                                if key not in self.isento_keys: self.isento_keys[key] = []
                                self.isento_keys[key].append(current)
                                self.log_msg(f"Nota ISENTO registrada (linha {current})", "info")

                            last = max(last, current)
                            self.log_msg(f"Linha {current}: Parcela R$ {vDup:.2f} vencendo em {date_parc} ({operation_type})", "success")
                            current += 1
                    else:
                        while ws.cell(current, 3).value:
                            current += 1

                        src_row = start if current == start else current - 1
                        self.copy_row_style(ws, src_row, current, cols_style)

                        ws.cell(current, 3, date)
                        ws.cell(current, 4, month)
                        ws.cell(current, 5, year)
                        ws.cell(current, 6, nNF)
                        ws.cell(current, 7, final_name)
                        ws.cell(current, 8, farm_ie)
                        ws.cell(current, 9, farm_name)
                        ws.cell(current, 10, final_id)
                        ws.cell(current, 11, prod)
                        ws.cell(current, 12, cfop)
                        ws.cell(current, valor_col, total)
                        ws.cell(current, 15, nat)
                        ws.cell(current, 16, key)
                        ws.cell(current, 17, None)

                        if farm_name == "ISENTO":
                            if key not in self.isento_keys: self.isento_keys[key] = []
                            self.isento_keys[key].append(current)
                            self.log_msg(f"Nota ISENTO registrada (linha {current})", "info")

                        last = max(last, current)
                        self.log_msg(f"Linha {current}: Nota completa R$ {total:.2f} emitida em {date} ({operation_type})", "success")

                else:
                    # ========== NFSE ==========
                    is_nfse = True
                    infs = root.find('.//s:infNFSe', ns_nfse)
                    if infs is None:
                        raise ValueError("XML não parece ser NF-e nem NFSe suportado")

                    nfse_id = (infs.get('Id') or '').strip()

                    nNF = (infs.findtext('.//s:nNFSe', default='', namespaces=ns_nfse) or '').strip()
                    if not nNF:
                        nNF = (infs.findtext('.//s:nDFSe', default='', namespaces=ns_nfse) or '').strip()

                    dh = (infs.findtext('.//s:dhEmiNFSe', default='', namespaces=ns_nfse)
                          or infs.findtext('.//s:dhEmi', default='', namespaces=ns_nfse)
                          or root.findtext('.//s:DPS//s:dhEmi', default='', namespaces=ns_nfse) or '')
                    dt = None
                    for fmt in ("%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
                        if not dh: break
                        try:
                            dt = datetime.strptime(dh, fmt)
                            break
                        except Exception:
                            continue
                    if dt is None:
                        dt = datetime.now()
                    date  = dt.strftime("%d/%m/%Y")
                    month = dt.month
                    year  = dt.year

                    emit = infs.find('.//s:emit', ns_nfse)
                    toma = infs.find('.//s:toma', ns_nfse)
                    emit_name = (emit.findtext('s:xNome', default='', namespaces=ns_nfse) or '') if emit is not None else ''
                    toma_name = (toma.findtext('s:xNome', default='', namespaces=ns_nfse) or '') if toma is not None else ''

                    emit_id = ((emit.findtext('s:CNPJ', default='', namespaces=ns_nfse) or
                                emit.findtext('s:CPF',  default='', namespaces=ns_nfse) or '').strip()) if emit is not None else ''

                    toma_id = ((toma.findtext('s:CNPJ', default='', namespaces=ns_nfse) or
                                toma.findtext('s:CPF',  default='', namespaces=ns_nfse) or '').strip()) if toma is not None else ''

                    interm = (root.find('.//s:interm', ns_nfse) or root.find('.//s:DPS//s:interm', ns_nfse))
                    if (not toma_id or not toma_name) and interm is not None:
                        toma = interm
                        toma_id = ((interm.findtext('s:CNPJ', default='', namespaces=ns_nfse) or
                                    interm.findtext('s:CPF',  default='', namespaces=ns_nfse) or '').strip())
                        toma_name = (interm.findtext('s:xNome', default='', namespaces=ns_nfse) or '')

                    vliq = (infs.findtext('.//s:valores/s:vLiq', default='', namespaces=ns_nfse) or '').strip()
                    if vliq:
                        total = float(str(vliq).replace(',', '.'))
                    else:
                        vserv = (root.findtext('.//s:DPS//s:valores//s:vServ', default='0', namespaces=ns_nfse) or '0')
                        total = float(str(vserv).replace(',', '.'))

                    prod = (root.findtext('.//s:DPS//s:xDescServ', default='', namespaces=ns_nfse) or '')
                    cfop = ""

                    emit_ie_nfse = emit.findtext('s:IE', default='', namespaces=ns_nfse) if emit is not None else ''
                    toma_ie_nfse = toma.findtext('s:IE', default='', namespaces=ns_nfse) if toma is not None else ''

                    # cidades para fallback
                    def _nfse_get_uf(root, infs, emit, toma, ns_nfse) -> str:
                        xpaths = [
                            's:UFIncid',
                            './/s:ender/s:UF', './/s:enderNac/s:UF',
                            './/s:prest//s:end//s:endNac/s:UF',
                            './/s:prest//s:enderNac/s:UF',
                            './/s:toma//s:end//s:endNac/s:UF',
                            './/s:toma//s:enderNac/s:UF',
                            './/s:emit//s:enderNac/s:UF',
                            './/s:DPS//s:prest//s:end//s:endNac/s:UF',
                            './/s:DPS//s:toma//s:end//s:endNac/s:UF',
                        ]
                        for xp in xpaths:
                            v = ''
                            if infs is not None:
                                try:
                                    v = infs.findtext(xp, default='', namespaces=ns_nfse)
                                except Exception:
                                    v = ''
                            if not v:
                                v = root.findtext(xp, default='', namespaces=ns_nfse)
                            if v:
                                return v.strip().upper()
                        return ''

                    uf = _nfse_get_uf(root, infs, emit, toma, ns_nfse)

                    emit_city = ''
                    dest_city = ''
                    if emit is not None:
                        emit_city = (emit.findtext('s:ender/s:xMun', default='', namespaces=ns_nfse) or
                                     emit.findtext('s:enderNac/s:xMun', default='', namespaces=ns_nfse) or '')
                    if toma is not None:
                        dest_city = (toma.findtext('s:ender/s:xMun', default='', namespaces=ns_nfse) or
                                     toma.findtext('s:enderNac/s:xMun', default='', namespaces=ns_nfse) or '')

                    owner_key, owner_side, farm_ie, farm_name = self._resolve_owner_from_parties(
                        emit_id, toma_id, emit_ie_nfse, toma_ie_nfse, emit_city, uf, dest_city, uf
                    )

                    nat = "SERVIÇO"
                    if nfse_id:
                        import re as _re
                        key = _re.sub(r'^\D+', '', nfse_id)
                    else:
                        key = nNF or ""
                    if key:
                        self.key_xml[key] = xml_file

                    if owner_key and owner_side == "emit":
                        final_name, final_id = toma_name, toma_id
                        valor_col = 13
                        operation_type = f"RECEITA (NFSe — {owner_key} Prestador)"
                    elif owner_key and owner_side == "dest":
                        final_name, final_id = emit_name, emit_id
                        valor_col = 14
                        operation_type = f"DESPESA (NFSe — {owner_key} Tomador)"
                    else:
                        final_name, final_id = emit_name, emit_id
                        valor_col = 14
                        operation_type = f"DESPESA (NFSe — {PRODUTOR_PADRAO} não identificado)"
                        self.log_msg(f"{PRODUTOR_PADRAO} não identificado na NFSe. Usando fallback genérico.", "warning")

                    while ws.cell(current, 3).value:
                        current += 1

                    src_row = start if current == start else current - 1
                    self.copy_row_style(ws, src_row, current, cols_style)

                    ws.cell(current, 3, date)
                    ws.cell(current, 4, month)
                    ws.cell(current, 5, year)
                    ws.cell(current, 6, nNF)
                    ws.cell(current, 7, final_name)
                    ws.cell(current, 8, farm_ie)
                    ws.cell(current, 9, farm_name)
                    ws.cell(current, 10, final_id)
                    ws.cell(current, 11, prod)
                    ws.cell(current, 12, cfop)
                    ws.cell(current, valor_col, total)
                    ws.cell(current, 15, nat)
                    ws.cell(current, 16, key)
                    ws.cell(current, 17, None)

                    if farm_name == "ISENTO":
                        if key not in self.isento_keys: self.isento_keys[key] = []
                        self.isento_keys[key].append(current)
                        self.log_msg(f"Nota ISENTO registrada (linha {current})", "info")

                    last = max(last, current)
                    self.log_msg(f"Linha {current}: Nota completa R$ {total:.2f} emitida em {date} ({operation_type})", "success")

                imported_files += 1
                self.stat_ok += 1
                self._update_stats()
                self.log_msg("Processamento concluído para este arquivo ✅", "info")
                self.log_msg("--------------------------------", "divider")

            except Exception:
                self.stat_err += 1
                self._update_stats()
                self.log_msg(f"Erro durante o processamento do arquivo: {traceback.format_exc()}", "error")
                self.log_msg("--------------------------------", "divider")

            # ====== CORREÇÃO PÓS-LOOP — REVER APENAS NOTAS ISENTAS PELA CIDADE DO PRESTADOR ======
            try:
                if self.isento_keys:
                    self.log_msg(f"Revisando {len(self.isento_keys)} nota(s) ISENTA(s) pela cidade do prestador…", "info")

                for key, linhas in self.isento_keys.items():
                    xml_path = self.key_xml.get(key, "")
                    if not xml_path or not os.path.exists(xml_path):
                        continue

                    try:
                        tree_fix = ET.parse(xml_path)
                        root_fix = tree_fix.getroot()
                        ns_nfe  = {'n': 'http://www.portalfiscal.inf.br/nfe'}
                        ns_nfse = {'s': 'http://www.sped.fazenda.gov.br/nfse'}

                        is_nfse_fix = root_fix.find('.//s:infNFSe', ns_nfse) is not None

                        active = (self.config.get('active_owner') or self.active_owner or 'CLEUBER').upper()
                        if active not in CODIGOS_CIDADES_BY_OWNER:
                            active = 'CLEUBER'
                        city_map = CODIGOS_CIDADES_BY_OWNER[active]
                        city_norm_map = CODIGOS_CIDADES_NORM_BY_OWNER[active]

                        def _normalize_no_accents(s: str) -> str:
                            s = (s or '').strip()
                            if not s:
                                return ''
                            s_noacc = unicodedata.normalize('NFKD', s)
                            return ''.join(ch for ch in s_noacc if not unicodedata.combining(ch))

                        def _variants(city: str, uf: str):
                            city = (city or '').strip()
                            if not city:
                                return []
                            noacc = _normalize_no_accents(city)
                            base  = re.sub(r'\s*-\s*[A-Z]{2}$', '', noacc, flags=re.I).strip()
                            out = [city, city.upper(), noacc, noacc.upper()]
                            if uf and base:
                                out += [f"{base} - {uf}", f"{base.upper()} - {uf}"]
                            return out

                        def _nfse_get_uf(root, infs, emit, toma) -> str:
                            xps = [
                                's:UFIncid',
                                './/s:ender/s:UF', './/s:enderNac/s:UF',
                                './/s:prest//s:end//s:endNac/s:UF',
                                './/s:prest//s:enderNac/s:UF',
                                './/s:toma//s:end//s:endNac/s:UF',
                                './/s:toma//s:enderNac/s:UF',
                                './/s:emit//s:enderNac/s:UF',
                                './/s:DPS//s:prest//s:end//s:endNac/s:UF',
                            ]
                            for xp in xps:
                                v = infs.findtext(xp, default='', namespaces=ns_nfse) if infs is not None else ''
                                if not v:
                                    v = root.findtext(xp, default='', namespaces=ns_nfse)
                                if v:
                                    return v.strip().upper()
                            return ''

                        city_candidates = []
                        uf = ""

                        if not is_nfse_fix:
                            emit = root_fix.find('.//n:emit', ns_nfe)
                            if emit is not None:
                                city_candidates += [emit.findtext('n:enderEmit/n:xMun', default='', namespaces=ns_nfe) or '']
                                uf_txt = emit.findtext('n:enderEmit/n:UF', default='', namespaces=ns_nfe)
                                if uf_txt:
                                    uf = uf_txt.strip().upper()
                        else:
                            infs2 = root_fix.find('.//s:infNFSe', ns_nfse)
                            emit2 = infs2.find('.//s:emit', ns_nfse) if infs2 is not None else None
                            toma2 = infs2.find('.//s:toma', ns_nfse) if infs2 is not None else None

                            city_candidates = [
                                (infs2.findtext('s:xLocIncid', default='', namespaces=ns_nfse) if infs2 is not None else ''),
                                (toma2.findtext('s:ender/s:xMun', default='', namespaces=ns_nfse) if toma2 is not None else ''),
                                (toma2.findtext('s:enderNac/s:xMun', default='', namespaces=ns_nfse) if toma2 is not None else ''),
                                (infs2.findtext('s:xLocPrestacao', default='', namespaces=ns_nfse) if infs2 is not None else ''),
                                (infs2.findtext('s:xLocEmi', default='', namespaces=ns_nfse) if infs2 is not None else ''),
                                (emit2.findtext('s:ender/s:xMun', default='', namespaces=ns_nfse) if emit2 is not None else ''),
                                (emit2.findtext('s:enderNac/s:xMun', default='', namespaces=ns_nfse) if emit2 is not None else ''),
                            ]

                            uf = _nfse_get_uf(root_fix, infs2, emit2, toma2)

                        farm_name_new = None
                        for raw in city_candidates:
                            for key_try in _variants(raw, uf):
                                if not key_try:
                                    continue
                                if key_try in city_map:
                                    farm_name_new = city_map[key_try]
                                    break
                                k_norm = _norm_city_key(key_try)
                                if k_norm in city_norm_map:
                                    farm_name_new = city_norm_map[k_norm]
                                    break
                            if farm_name_new:
                                break

                        if farm_name_new:
                            for lin in linhas:
                                ws.cell(lin, 9, farm_name_new)
                            self.log_msg(f"ISENTO corrigido via cidade do prestador (key {key} → {farm_name_new})", "success")

                    except Exception as inner_e:
                        self.log_msg(f"Falha ao revisar key {key}: {inner_e}", "error")

            except Exception as _e:
                self.log_msg(f"Falha na correção pós-loop de ISENTOS: {_e}", "error")
            # ====== FIM DA CORREÇÃO PÓS-LOOP ======

        if last >= start:
            self.extend_table(ws, header, last)

        try:
            for col in ['M','N']:
                for row in range(start, ws.max_row + 1):
                    cell = ws[f"{col}{row}"]
                    if cell.value is not None:
                        cell.number_format = '"R$"#,##0.00;[Red]"R$"-#,##0.00'

            wb.save(self.excel_path)
            if imported_files > 0:
                self.log_msg("✅ IMPORTAÇÃO FINALIZADA COM SUCESSO", "title")
                self.log_msg(f"Total de arquivos processados: {imported_files}/{len(files)}", "success")
                if imported_files < len(files):
                    self.log_msg(f"{len(files) - imported_files} arquivo(s) com problemas não foram importados", "warning")
            else:
                self.log_msg("Importação cancelada ou sem arquivos processados.", "warning")
        except Exception as e:
            self.log_msg(f"❌ ERRO AO SALVAR PLANILHA: {e}", "error")
            QMessageBox.critical(self, "Erro", f"Falha ao salvar:\n{e}")

        if self.isento_keys and self.isento_path:
            self.processar_isento()

        GlobalProgress.end()
        self.btn_cancel.setEnabled(False)

    # ---------- ISENTO ----------
    def processar_isento(self):
        try:
            self.log_msg("Iniciando processamento de notas ISENTO", "title")
            self.log_msg(f"Procurando XMLs em: {self.isento_path} (recursivamente)", "info")

            resultados = {}
            erros = []
            encontrados = 0

            # usa perfil ativo para o mapeamento
            active = (self.config.get('active_owner') or self.active_owner or 'CLEUBER').upper()
            if active not in CODIGOS_CIDADES_BY_OWNER:
                active = 'CLEUBER'
            city_map = CODIGOS_CIDADES_BY_OWNER[active]

            for dirpath, _, filenames in os.walk(self.isento_path):
                for filename in filenames:
                    if not filename.lower().endswith(".xml"):
                        continue

                    caminho_arquivo = os.path.join(dirpath, filename)
                    try:
                        tree = ET.parse(caminho_arquivo)
                    except Exception:
                        continue

                    root = tree.getroot()
                    ns = {"nfe": "http://www.portalfiscal.inf.br/nfe"}

                    elem_ch = root.find(".//nfe:chNFe", ns)
                    if elem_ch is None or elem_ch.text not in self.isento_keys:
                        continue
                    chave = elem_ch.text

                    elem_mun = root.find(".//nfe:enderDest/nfe:xMun", ns)
                    if elem_mun is None or not elem_mun.text:
                        resultados[chave] = "Cidade não informada"
                        erros.append((chave, None))
                    else:
                        cidade = elem_mun.text
                        codigo = city_map.get(cidade)
                        if codigo:
                            resultados[chave] = codigo
                            encontrados += 1
                        else:
                            resultados[chave] = "Código não encontrado"
                            erros.append((chave, cidade))

            self.log_msg(f"Encontrados {encontrados} XMLs para notas ISENTO", "success")

            if resultados:
                try:
                    wb = load_workbook(self.excel_path)
                    ws = wb.active
                    atualizados = 0
                    for chave, codigo in resultados.items():
                        if chave in self.isento_keys:
                            for linha in self.isento_keys[chave]:
                                ws.cell(linha, 9, codigo)
                                atualizados += 1
                                self.log_msg(f"Linha {linha}: Atualizada fazenda para '{codigo}'", "success")
                    wb.save(self.excel_path)
                    self.log_msg(f"✅ {atualizados} NOTAS ISENTO ATUALIZADAS NA PLANILHA", "title")
                except Exception as e:
                    self.log_msg(f"❌ ERRO AO ATUALIZAR PLANILHA: {e}", "error")

            pasta_script = Path(__file__).parent
            arquivo_saida = pasta_script / "resultados_isento.txt"
            with open(arquivo_saida, "w", encoding="utf-8") as f:
                f.write("=== Resultados por Chave ===\n")
                for chave in self.isento_keys:
                    status = resultados.get(chave, "Não encontrado")
                    f.write(f"{chave}: {status}\n")
                f.write("\n=== XMLs sem cidade encontrada ou sem código ===\n")
                for chave, cidade in erros:
                    if cidade:
                        f.write(f"{chave}: {cidade}\n")
                    else:
                        f.write(f"{chave}: Cidade não informada\n")
            self.log_msg(f"Relatório detalhado salvo em: {arquivo_saida}", "info")

        except Exception as e:
            self.log_msg(f"❌ ERRO NO PROCESSAMENTO ISENTO: {e}", "error")

# -----------------------------
# Main
# -----------------------------
if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon(str(ICON_PATH)))
    app.setStyleSheet(STYLE_SHEET)

    w = RuralXmlImporter()
    w.show()
    sys.exit(app.exec())
