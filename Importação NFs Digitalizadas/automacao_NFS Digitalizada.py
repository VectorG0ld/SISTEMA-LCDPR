#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Leitor de NFS-e por RECORTES (layouts) — versão com **FILTROS POR LAYOUT e POR TÓPICO** + **Resumo por IA (OpenAI)**

O que mudou nesta versão:
- Cada **Layout** possui seu **próprio filtro padrão** (ConfigLimpezaScan) e, dentro de cada layout, 
  **cada TÓPICO** (DATA, MUNICIPIO_TOMADOR, etc.) tem **seu próprio filtro** independente.
- Mantida compatibilidade com `treat_strong()` como fallback.
- **Novo**: o **RESUMO FINAL DA NOTA** passa a ser gerado por **IA (OpenAI)** a partir do
  texto bruto extraído (o mesmo que vai para o TXT). O restante da leitura/extração 
  permanece idêntico.

Como ajustar filtros:
1) Vá na seção `# ================== LAYOUTS ==================`.
2) Para cada Layout, defina `filter_default=` e, se quiser, sobrescreva por tópico em
   `filters_by_topic={ 'DATA': ..., 'NUMERO_NFS': ..., ... }`.
3) Se não definir um tópico em `filters_by_topic`, ele herdará automaticamente
   uma cópia do `filter_default` daquele Layout.

Requisitos (exemplos):
    pip install pdfplumber pytesseract pillow openai

Para usar a IA:
    - Defina a variável de ambiente OPENAI_API_KEY.
    - (Opcional) Defina OPENAI_MODEL (padrão: gpt-4o-mini).
"""

from __future__ import annotations

import os
import re
import sys
import io
import glob
from dataclasses import dataclass, field, replace
from typing import Dict, List, Tuple, Optional, Union
import time  # ✅ usar perf_counter para cronometrar

import pdfplumber

from PIL import Image, ImageDraw, ImageFont, ImageOps, ImageFilter, ImageEnhance
import pytesseract

# ===== NOVO: planilha Excel e Tabela + geração de TXT a partir da planilha
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from pathlib import Path

def _find_in_this_or_parent(filename: str) -> Path:
    here = Path(__file__).resolve().parent
    cand1 = here / filename
    cand2 = here.parent / filename  # raiz do projeto (onde fica o main)
    if cand1.exists():
        return cand1
    if cand2.exists():
        return cand2
    raise FileNotFoundError(f"Não encontrei '{filename}' nem em {cand1} nem em {cand2}.")

# ================== CONFIG GERAL ==================
def _strip_accents(s: str) -> str:
    try:
        import unicodedata
        return "".join(ch for ch in unicodedata.normalize("NFD", s or "") if unicodedata.category(ch) != "Mn")
    except Exception:
        return s or ""

def _norm_simple(s: str) -> str:
    s = _strip_accents(s).lower()
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def _unique_path(base_path: str) -> str:
    if not os.path.exists(base_path):
        return base_path
    root, ext = os.path.splitext(base_path)
    i = 2
    while True:
        cand = f"{root} ({i}){ext}"
        if not os.path.exists(cand):
            return cand
        i += 1

# ——— CANCELAMENTO (worker injeta is_cancelled; o core consulta aqui) ———
def _cancelled() -> bool:
    try:
        cb = globals().get("is_cancelled", None)
        return bool(cb and callable(cb) and cb())
    except Exception:
        return False

BASE_DIR = r"C:\Users\conta\Downloads"
# Se precisar no Windows:
# pytesseract.pytesseract.tesseract_cmd = r"C:\\Program Files\\Tesseract-OCR\\tesseract.exe"

DPI = 800               # ajuste conforme sua velocidade/qualidade
CROP_PAD = 28           # margem extra em volta de cada recorte (px)
OCR_LANG = "por+eng"    # ajuda em cabeçalhos mistos

WRITE_TXT_OUTPUT = False  # se quiser TXT detalhado independente, ligue aqui
# Exibir no terminal apenas o resumo gerado pela IA
PRINT_ONLY_IA = True

# ================== MAPEAMENTOS (fornecidos) ==================

FARM_MAPPING = {
    "115149210": "Arm. Primavera",
    "111739837": "Aliança",
    "114436720": "B. Grande",
    "115449965": "Estrela",
    "294186832": "Frutacc",
    "294907068": "Frutacc III",
    "295057386": "L3",
    "112877672": "Primavera",
    "113135521": "Primavera Retiro",
    "294904093": "União",
    "295359790": "Frutacc V",
    "295325704": "Siganna"
}

CODIGOS_CIDADES = {
    "Lagoa da Confusao": "Frutacc",
    "Montividiu do Norte": "Barragem",
    "Rialma": "Aliança",
    "TROMBAS": "Primavera",
    "DUERE": "L3", "DUERÉ": "L3", "DUERE TO": "L3", "Duere": "L3",
    "Ceres": "Aliança", "Rianapolis": "Aliança", "NOVA GLORIA": "Aliança",
    "MONTIVIDIU": "Barragem", "MONTIVIDIU DO NORTE - GO": "Barragem",
    "Nova Glória": "Aliança", "Nova Gloria": "Aliança", "Nova Giéria": "Aliança",
    "Lagoa da Confusão": "Frutacc", "MONTIVIDIU DO NORTE": "Barragem",
    "LAGOA DA CONFUSAO": "Frutacc", "LAGOA DA CONFUSÃO": "Frutacc",
    "LAGOA CONFUSAO": "Frutacc", "LAGOA DA CONFUSAO - TO": "Frutacc",
    "RIALMA": "Aliança", "Trombas": "Primavera", "CERES": "Aliança",
    "Formoso do Araguaia": "União", "FORMOSO DO ARAGUAIA": "União",
    "APARECIDA DO RIO NEGRO": "Primavera",
    "Tasso Fragoso": "Guara", "BALSAS": "Guara", "Balsas": "Guara",
    "Montividiu": "Barragem",
}

# ===== NOVO: Mapeamento FAZENDA -> CÓDIGO (para planilha e TXT)
FAZENDA2COD_RAW = {
    "Fazenda Frutacc":"001","Fazenda Frutacc II":"001","Fazenda Frutacc III":"001",
    "Fazenda L3":"003","Armazem L3":"003","Fazenda Rio Formoso":"002",
    "Fazenda Siganna":"001","Armazem Frutacc":"006","Lagoa da Confusão":"001",
    "Fazenda Primavera":"004","Fazenda Primaveira":"004","Fazenda Estrela":"008",
    "Fazenda Ilhéus":"004","Sitio Boa Esperança":"007","Fazenda Retiro":"004",
    "Fazenda Barragem Grande":"007","Fazenda Ilha do Formoso":"001","Fazenda Pouso da Anta":"009",
    "Montividiu do Norte":"006","Nova Gloria":"005","Nova Glória":"005",
    "Formoso do Araguaia":"002","Duere":"003","Duerê":"003","Dueré":"003",
    "Trombas":"004","Lagoa da Confusao":"001",
    "FAZENDA FRUTACC":"001","FAZENDA UNIÃO":"002","FAZENDA L3":"003",
    "FAZENDA PRIMAVERA":"004","FAZENDA ALIANÇA":"005","ARMAZEM PRIMAVERA":"006",
    "FAZENDA BARRAGEM GRANDE":"007","FAZENDA ESTRELA":"008","FAZENDA GUARA":"009",
}

# Aliases (IA costuma devolver nomes curtos — ajudamos o mapeamento)
FAZENDA2COD_ALIASES = {
    "Aliança":"005","Alianca":"005","União":"002","Uniao":"002","L3":"003",
    "Primavera":"004","Estrela":"008","Barragem":"007","B. Grande":"007","Barragem Grande":"007",
    "Frutacc":"001","Siganna":"001","Guara":"009","Guará":"009","Retiro":"004",
    "Ilheus":"004","Ilhéus":"004","Arm. Primavera":"006","Armazem Primavera":"006","Armazém Primavera":"006",
}

def _norm_fazenda_key(s: str) -> str:
    # usa _norm_simple p/ tirar acentos, símbolos e padronizar
    s = _norm_simple((s or ""))  # lower, sem acentos, sem símbolos
    # mantém "ARMAZEM/ARMAZÉM" (NÃO remove)
    s = re.sub(r"\b(fazenda|sitio|sítio)\b", "", s).strip()
    return re.sub(r"\s+", " ", s).upper()

FAZENDA2COD: Dict[str, str] = {}
_DUP_CONFLICT: Dict[str, set] = {}
_DUP_SAME: Dict[str, set] = {}

for k, v in {**FAZENDA2COD_RAW, **FAZENDA2COD_ALIASES}.items():
    nk = _norm_fazenda_key(k)
    if nk in FAZENDA2COD:
        if FAZENDA2COD[nk] != v:
            _DUP_CONFLICT.setdefault(nk, set()).update({FAZENDA2COD[nk], v})
        else:
            _DUP_SAME.setdefault(nk, set()).add(v)
    FAZENDA2COD[nk] = v

if _DUP_CONFLICT:
    print("[ATENÇÃO] Mapeamento FAZENDA->CÓDIGO contém chaves CONFLITANTES (normalizadas):")
    for nk, vals in _DUP_CONFLICT.items():
        print(f"  - '{nk}' -> códigos: {sorted(list(vals))}")

def codigo_fazenda_from_nome(nome: str) -> str:
    nk = _norm_fazenda_key(nome)
    return FAZENDA2COD.get(nk, "-")

# ================== REGEX / HELPERS ==================
# --- Datas "flex": aceita dd/mm/2 025 (com espaços no ano) ---
RE_DATEPT_FLEX = re.compile(
    r"\b(\d{1,2})\s*[\/\.\-]\s*(\d{1,2})\s*[\/\.\-]\s*((?:\d\s*){4})\b"
)

# Rótulo de VENCIMENTO seguido de data (com tolerância de até ~30 chars entre eles)
RE_VENC_LABEL_DATE = re.compile(
    r"(?is)\b(?:dt\.?\s*(?:de\s*)?)?(?:vencimento|vencto|vcto|venc)\b.{0,30}?"
    r"(\d{1,2}\s*[\/\.\-]\s*\d{1,2}\s*[\/\.\-]\s*(?:\d\s*){4})"
)

def _norm_date_parts(d: str, mth: str, y_raw: str) -> str:
    y = re.sub(r"\s+", "", y_raw)  # remove espaços dentro do ano: "2 025" -> "2025"
    d = d.zfill(2); mth = mth.zfill(2)
    return f"{d}/{mth}/{y}"

def normalize_date_ddmmyyyy_flex(text: str) -> str:
    """Retorna primeira data dd/mm/aaaa tolerante a espaços no ano."""
    if not text:
        return ""
    m = RE_DATEPT_FLEX.search(text)
    if not m:
        return normalize_date_ddmmyyyy(text)
    d, mth, y = m.groups()
    return _norm_date_parts(d, mth, y)

def parse_vencimento_from_recebimento(text: str) -> str:
    """
    Extrai a DATA DE VENCIMENTO do texto do Recebimento:
    1) Prioriza data próxima ao rótulo 'Vencimento/Vencto/Vcto/Venc.'
    2) Se não houver rótulo, pega a ÚLTIMA data do texto (geralmente emissão vem antes).
    Aceita ano com espaços (ex.: 2 025).
    """
    if not text:
        return ""
    m = RE_VENC_LABEL_DATE.search(text)
    if m:
        m2 = RE_DATEPT_FLEX.search(m.group(1))
        if m2:
            d, mm, y = m2.groups()
            return _norm_date_parts(d, mm, y)
    all_matches = list(RE_DATEPT_FLEX.finditer(text))
    if not all_matches:
        all_matches = list(RE_DATEPT.finditer(text))
        if not all_matches:
            return ""
    d, mm, y = all_matches[-1].groups()
    return _norm_date_parts(d, mm, y)

def add_macro_button_and_save_xlsm(xlsx_path: str) -> Optional[str]:
    """
    Abre o .xlsx no Excel (COM), injeta módulo VBA com macro 'GerarTXT',
    adiciona um botão na planilha 'lancamentos' e salva como .xlsm.
    Retorna o caminho do .xlsm ou None se não conseguir (sem quebrar o fluxo).
    """
    try:
        import win32com.client as win32  # requer 'pip install pywin32' e Excel instalado
        from win32com.client import constants as xl

        excel = win32.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(xlsx_path)
        ws = wb.Worksheets("lancamentos")

        # 1) Injeta o módulo VBA com a macro GerarTXT
        vbcomp = wb.VBProject.VBComponents.Add(1)  # 1 = vbext_ct_StdModule
        code = r'''
Public Sub GerarTXT()
    Dim sh As Worksheet: Set sh = ThisWorkbook.Worksheets("lancamentos")
    Dim lastRow As Long: lastRow = sh.Cells(sh.Rows.Count, 1).End(xlUp).Row

    Dim fso As Object, ts As Object
    Dim outPath As String
    outPath = ThisWorkbook.Path & "\saida_lancamentos.txt"
    Set fso = CreateObject("Scripting.FileSystemObject")
    Set ts = fso.CreateTextFile(outPath, True, False) ' ASCII

    Dim r As Long
    For r = 2 To lastRow
        Dim gerar As String: gerar = UCase(Trim(CStr(sh.Cells(r, 8).Value))) ' Coluna H = "Gerar"
        If gerar = "" Then gerar = "SIM"
        If gerar = "S" Or gerar = "SIM" Or gerar = "X" Or gerar = "TRUE" Or gerar = "1" Then
            Dim data As String, codfaz As String, participante As String
            Dim numero As String, desc As String, cnpj As String, valorcent As String

            data = CStr(sh.Cells(r, 1).Value)
            codfaz = CStr(sh.Cells(r, 2).Value)
            participante = CStr(sh.Cells(r, 3).Value)
            numero = CStr(sh.Cells(r, 4).Value)
            desc = CStr(sh.Cells(r, 5).Value)
            cnpj = CStr(sh.Cells(r, 6).Value)
            If UCase(cnpj) <> "INVALIDO" Then
                cnpj = Replace(Replace(Replace(Replace(cnpj, ".", ""), "-", ""), "/", ""), " ", "")
            End If
            valorcent = CStr(sh.Cells(r, 7).Value)

            Dim linha As String
            linha = data & "|" & codfaz & "|" & participante & "|" & numero & "|1|" & desc & "|" & cnpj & "|2|000|" & valorcent & "|" & valorcent & "|N"
            ts.WriteLine linha
        End If
    Next r
    ts.Close
    MsgBox "TXT gerado em: " & outPath, vbInformation
End Sub
'''
        vbcomp.CodeModule.AddFromString(code)

        # 2) Adiciona um botão (form control) e atribui a macro
        #    Posição/tamanho: Left, Top, Width, Height (em pontos)
        btn = ws.Buttons().Add(10, 10, 140, 24)
        btn.OnAction = "GerarTXT"
        btn.Characters().Text = "Gerar TXT"

        # 3) Salva como .xlsm
        xlsm_path = xlsx_path[:-5] + ".xlsm" if xlsx_path.lower().endswith(".xlsx") else xlsx_path + ".xlsm"
        wb.SaveAs(xlsm_path, FileFormat=52)  # 52 = xlOpenXMLWorkbookMacroEnabled (.xlsm)
        wb.Close(SaveChanges=True)
        excel.Quit()
        return xlsm_path

    except Exception as e:
        print(f"[AVISO] Não foi possível criar macro/botão automaticamente ({e}). "
              f"Planilha .xlsx criada sem macro; instale 'pywin32' e habilite 'Confiar no acesso ao modelo de projeto VBA'.")
        return None

def create_xlsm_from_template(template_path: str, dest_xlsm_path: str, rows_lanc: List[Dict[str, str]]) -> Optional[str]:
    """
    Carrega um .xlsm de template (com macro/botão prontos), preenche a aba 'lancamentos'
    e salva em dest_xlsm_path (preservando o VBA).
    Requisitos do template:
      - Aba 'lancamentos' existente, com cabeçalhos em A1:H1 compatíveis:
        ["Data","CodFazenda","Participante","NumeroNF","Descricao","CNPJ","ValorCentavos","Gerar"]
      - Uma tabela (opcional) chamada 'lancamentos_tbl' para autoestilo.
      - Um botão já apontando para a macro 'GerarTXT'.
    """
    try:
        wb = load_workbook(template_path, keep_vba=True)
        # tenta 'lancamentos' e variações
        target_ws = None
        for name in wb.sheetnames:
            if name.strip().lower() == "lancamentos":
                target_ws = wb[name]
                break
        if target_ws is None:
            for name in wb.sheetnames:
                if "lancamentos" in name.strip().lower():
                    target_ws = wb[name]
                    break
        if target_ws is None:
            # cai na primeira planilha
            target_ws = wb[wb.sheetnames[0]]
        ws = target_ws

        # limpa linhas de dados (mantendo cabeçalho)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        # escreve linhas
        for r in rows_lanc:
            ws.append([
                r["data"], r["codfaz"], r["participante"], r["numero"],
                r["descricao"], r["cnpj"], r["valor_centavos"], "SIM"
            ])

        # ajusta tabela (se existir)
        try:
            from openpyxl.utils import get_column_letter
            end_row = ws.max_row
            end_col = 8
            ref = f"A1:{get_column_letter(end_col)}{end_row}"
            if "lancamentos_tbl" in ws.tables:
                ws.tables["lancamentos_tbl"].ref = ref
        except Exception:
            pass

        wb.save(dest_xlsm_path)
        return dest_xlsm_path
    except Exception as e:
        print(f"[AVISO] Falha ao usar template XLSM: {e}")
        return None

def resolve_codfaz_from_municipio_text(topic_text: str) -> Tuple[str, str]:
    """
    Usa APENAS o texto do tópico [MUNICIPIO_TOMADOR] para determinar a fazenda/código.
    Retorna (nome_fazenda_mapeada, codfaz). Se não achar, retorna ('-', '-').
    Regras:
      1) Se houver um código numérico que exista no FARM_MAPPING -> mapeia por FARM_MAPPING.
      2) Caso contrário, detecta a cidade via parse_cidade_from() e mapeia usando CODIGOS_CIDADES.
      3) Converte o nome mapeado em código usando FAZENDA2COD; se não achar, retorna "-".
    """
    if not topic_text:
        return "-", "-"

    # 1) tenta códigos numéricos “crus” que possam existir no FARM_MAPPING
    for tok in re.findall(r"\b\d{6,12}\b", topic_text):
        if tok in FARM_MAPPING:
            nome = FARM_MAPPING[tok]
            cod = codigo_fazenda_from_nome(nome)
            return (nome, cod if cod != "NAO ENCONTRADO" else "-")

    # 2) tenta por cidade (normaliza e equaliza)
    cidade = parse_cidade_from(topic_text)
    if cidade:
            nk_city = _city_key(cidade)
            for k, nome in CODIGOS_CIDADES.items():
                if _city_key(k) == nk_city:
                    cod = codigo_fazenda_from_nome(nome)
                    return (nome, cod if cod != "NAO ENCONTRADO" else "-")

    return "-", "-"

def _city_key(s: str) -> str:
    s = _norm_simple(s).upper()
    s = re.sub(r"\b[A-Z]{2}\b$", "", s).strip()  # remove UF ao final (ex.: " - GO")
    return re.sub(r"\s+", " ", s)

# ================== IMAGEM: CONFIGURÁVEL & ROBUSTO ==================

@dataclass
class ConfigLimpezaScan:
    """
    Parâmetros legíveis para tratamento da imagem escaneada:
      - preto: viés pró-preto no limiar de Otsu (↑ expande áreas pretas). 1.00–1.40
      - ruido: força do anti-ruído (0 desliga | 1 leve | 2 médio | 3 forte)
      - contraste: ganho de contraste global (1.0 = neutro)
      - gama: >1 escurece tons médios (deixa traço mais denso)
      - borda: nitidez (0 desliga | 1 padrão | 2 forte)
      - corte_baixo/corte_alto: autocontraste (percentis cortados do histograma)
      - fechar_furos: fechamento morfológico (0 desliga | 1 leve | 2 médio)
      - saida_1bit: True para B/N puro (modo '1'), False mantém 'L' binarizado
      - zoom: fator de ampliação do recorte antes do tratamento (1.0 = sem zoom)
      - matar_pretos_ate: remove componentes pretos com área ≤ X px (0 desliga)
      - manter_maiores_pretos: mantém apenas os N maiores componentes pretos (0 desliga)
      - matar_barras_horiz_altura_max: remove barras pretas horizontais com altura ≤ este valor (px). 0 = desliga
      - matar_barras_horiz_min_largura_frac: só remove a barra se ela cobrir ao menos esta fração da largura do recorte
      - matar_isolados_vizinhos: filtro “majority” 8-conexo (remove pontilhado); 0 desliga; 4–6 bom começo
      - repetir_isolados_passes: quantas passadas do filtro acima
    """
    preto: float = 1.20
    ruido: float = 1.0
    contraste: float = 1.10
    gama: float = 1.20
    borda: float = 1.0
    corte_baixo: int = 2
    corte_alto: int = 8
    fechar_furos: float = 1.0
    saida_1bit: bool = False
    zoom: float = 1.0
    matar_pretos_ate: int = 0
    manter_maiores_pretos: int = 0
    matar_barras_horiz_altura_max: int = 0
    matar_barras_horiz_min_largura_frac: float = 0.5
    matar_isolados_vizinhos: int = 0
    repetir_isolados_passes: int = 1

# Fallback global (usado por treat_strong e como base de cópia)
CFG_FALLBACK = ConfigLimpezaScan(
    preto=1.25,
    ruido=1.0,
    contraste=1.20,
    gama=1.15,
    borda=1.2,
    corte_baixo=2,
    corte_alto=8,
    fechar_furos=1.0,
    saida_1bit=False,
    # reforço anti-risco horizontal
    matar_barras_horiz_altura_max=100,
    matar_barras_horiz_min_largura_frac=0.25
)
# ---------- Heurísticas de "Recebimento de Mercadoria" ----------
from types import SimpleNamespace

def _make_recebimento_layout():
    """Mini layout só com o crop DATA para a página de Recebimento."""
    return SimpleNamespace(
        name="RECEBIMENTO",
        keywords=[],
        crops={"DATA": RECEBIMENTO_DATA_CROP},
        filter_default=CFG_RECEBIMENTO_DATA,
        filters_by_topic={"DATA": CFG_RECEBIMENTO_DATA},
    )

def is_recebimento(text_norm: str, raw_text: str = "") -> bool:
    if "recebimento de mercadoria" in text_norm:
        return True
    if re.search(r"\bn[ºo0]\s*recebimento\b", raw_text or "", flags=re.IGNORECASE):
        return True
    if "solicitacao de pagamento" in text_norm:
        return True
    if re.search(r"\brelatorio\s+impresso\s+por\s+sap\s+business(\s+one)?\b", text_norm):
        return True
    if re.search(r"\bimpresso\s+por\s+sap\s+business(\s+one)?\b", text_norm):
        return True
    return False

RECEBIMENTO_DATA_CROP = {"ESQUERDA": 55, "CIMA": 7, "DIREITA": 95, "BAIXO": 19}
CFG_RECEBIMENTO_DATA = replace(
    CFG_FALLBACK,
    preto=2.8,
    ruido=3.5,
    contraste=6.0,
    gama=1.10,
    borda=1.0,
    fechar_furos=1.5,
    matar_pretos_ate=250,
    matar_barras_horiz_altura_max=100,
    matar_barras_horiz_min_largura_frac=0.35
)

def _ocr_from_relbox(png_path: str, relbox: Dict[str, float], cfg: ConfigLimpezaScan) -> str:
    with Image.open(png_path) as im:
        w, h = im.size
        x0, y0, x1, y1 = _rel2abs(w, h, _as_box(relbox), pad=CROP_PAD)
        raw = im.crop((x0, y0, x1, y1))
        treated = limpar_imagem_escaner(raw, cfg)
    return _ocr_text(treated)

RE_CNPJ    = re.compile(r"\b\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}\b")
RE_CPF    = re.compile(r"\b\d{3}\.?\d{3}\.?\d{3}[-\/]?\d{2}\b")
RE_DATEPT  = re.compile(r"\b(\d{2})[\/\.\-](\d{2})[\/\.\-](\d{4})\b")
RE_MONEY_BR = re.compile(r"(?<!\d)(?:R\$\s*)?(\d{1,3}(?:\.\d{3})*,\d{2})(?!\d)")

UF_SIGLAS = {
    "AC","AL","AM","AP","BA","CE","DF","ES","GO","MA","MG","MS","MT",
    "PA","PB","PE","PI","PR","RJ","RN","RO","RR","RS","SC","SE","SP","TO"
}

def only_digits(s: str) -> str:
    return re.sub(r"\D+", "", s or "")

def mask_cnpj(cnpj_digits: str) -> str:
    d = only_digits(cnpj_digits)
    if len(d) != 14: return cnpj_digits or ""
    return f"{d[0:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:14]}"

# ===== NOVO: parsing do RESUMO FINAL (global, antes do main)
RESUMO_RE = {
    "numero": re.compile(r"(?im)^\s*Número\s+NFS-e.*?:\s*(.+)$"),
    "data": re.compile(r"(?im)^\s*Data\s+de\s+Emiss[aã]o.*?:\s*(\d{2}/\d{2}/\d{4})"),
    "fazenda": re.compile(r"(?im)^\s*Fazenda.*?:\s*(.+)$"),
    "prestador": re.compile(r"(?im)^\s*Prestador.*?:\s*(.+)$"),
    "cnpj": re.compile(r"(?im)^\s*CNPJ\s+Prestador.*?:\s*(.+)$"),
    "valor": re.compile(r"(?im)^\s*Valor\s+Total/L[ií]quido.*?:\s*([\d\.]+,\d{2})"),
}

def parse_campos_from_resumo(resumo: str) -> Dict[str, str]:
    out = {"numero":"", "data":"", "fazenda":"", "prestador":"", "cnpj":"", "valor":""}
    if not resumo:
        return out
    for k, rx in RESUMO_RE.items():
        m = rx.search(resumo or "")
        if m:
            out[k] = (m.group(1) or "").strip()
    if out["cnpj"] and out["cnpj"].upper() != "INVALIDO":
        d = only_digits(out["cnpj"])
        out["cnpj"] = mask_cnpj(d) if cnpj_is_valid(d) else "INVALIDO"
    return out

def _format_resumo_campos(campos: Dict[str, str]) -> str:
    def _nz(v, dash="-"): 
        v = (v or "").strip()
        return v if v else dash
    return (
        "📘 RESUMO FINAL DA NOTA\n"
        f"  Número.............: {_nz(campos.get('numero'))}\n"
        f"  Data de Emissão....: {_nz(campos.get('data'))}\n"
        f"  Fazenda............: {_nz(campos.get('fazenda'))}\n"
        f"  Prestador..........: {_nz(campos.get('prestador'))}\n"
        f"  CNPJ Prestador.....: {_nz(campos.get('cnpj'))}\n"
        f"  Valor Total/Líquido: {_nz(campos.get('valor'))}\n"
    )

def money_br_to_centavos(v: str) -> str:
    v = (v or "").strip()
    if not v:
        return ""
    d = v.replace(".", "").replace(",", ".")
    try:
        centavos = int(round(float(d) * 100))
        return str(centavos)
    except Exception:
        return ""
    
def date_slash_to_dash(d: str) -> str:
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", d or "")
    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}" if m else (d or "")

def normalize_date_ddmmyyyy(text: str) -> str:
    if not text: return ""
    m = RE_DATEPT.search(text)
    if not m: return ""
    d, mth, y = m.groups()
    return f"{d}/{mth}/{y}"

def to_image(page, dpi=DPI) -> Image.Image:
    return page.to_image(resolution=dpi).original  # PIL.Image

# ===== NOVO: Validação de CNPJ (com dígitos verificadores)
def cnpj_is_valid(cnpj: str) -> bool:
    d = only_digits(cnpj)
    if len(d) != 14 or d == d[0] * 14:
        return False
    pesos1 = [5,4,3,2,9,8,7,6,5,4,3,2]
    soma1 = sum(int(d[i]) * pesos1[i] for i in range(12))
    dv1 = 11 - (soma1 % 11)
    dv1 = 0 if dv1 >= 10 else dv1
    pesos2 = [6,5,4,3,2,9,8,7,6,5,4,3,2]
    soma2 = sum(int(d[i]) * pesos2[i] for i in range(13))
    dv2 = 11 - (soma2 % 11)
    dv2 = 0 if dv2 >= 10 else dv2
    return d[12] == str(dv1) and d[13] == str(dv2)

# -------- helpers de imagem --------

def _clamp(v, lo, hi):
    return max(lo, min(hi, v))


def _odd_from_level(level: float, base: int = 3, step: int = 2, max_val: int = 9) -> int:
    """Converte um level (0,1,2,3,...) para kernel ímpar (1,3,5,7,9...)."""
    if level <= 0:
        return 1
    size = base + int(round(level - 1)) * step
    size = _clamp(size, 1, max_val)
    if size % 2 == 0:
        size += 1
    return size


def _lookup_gamma(gama: float):
    inv = 1.0 / 255.0
    return [int(_clamp(((i * inv) ** gama) * 255.0, 0, 255)) for i in range(256)]


def _pct(v: float) -> float:
    if v is None: return 0.0
    v = float(v)
    return v/100.0 if v>1.0 else v


def _as_box(rel: Union[Tuple[float, float, float, float], Dict[str, float]]) -> Tuple[float, float, float, float]:
    if isinstance(rel, (tuple, list)) and len(rel) == 4:
        return float(rel[0]), float(rel[1]), float(rel[2]), float(rel[3])
    left   = _pct(rel.get("ESQUERDA", 0.0))
    top    = _pct(rel.get("CIMA", 0.0))
    right  = _pct(rel.get("DIREITA", 1.0))
    bottom = _pct(rel.get("BAIXO", 1.0))
    return left, top, right, bottom


def _rel2abs(w: int, h: int, box: Tuple[float, float, float, float], pad: int = CROP_PAD) -> Tuple[int, int, int, int]:
    x0 = max(0, int(box[0] * w) - pad)
    y0 = max(0, int(box[1] * h) - pad)
    x1 = min(w, int(box[2] * w) + pad)
    y1 = min(h, int(box[3] * h) + pad)
    return x0, y0, x1, y1


def _otsu_threshold(img_l: Image.Image) -> int:
    """Calcula limiar de Otsu (grayscale L)."""
    hist = img_l.histogram()[:256]
    total = sum(hist)
    sumB = 0.0
    wB = 0.0
    maximum = 0.0
    sum1 = sum(i * hist[i] for i in range(256))
    threshold = 127
    for i in range(256):
        wB += hist[i]
        if wB == 0:
            continue
        wF = total - wB
        if wF == 0:
            break
        sumB += i * hist[i]
        mB = sumB / wB
        mF = (sum1 - sumB) / wF
        between = wB * wF * (mB - mF) ** 2
        if between >= maximum:
            threshold = i
            maximum = between
    return threshold


def _noise_filter(img_l: Image.Image, k: int) -> Image.Image:
    """Aplica filtro anti-ruído preservando bordas (ModeFilter ou MedianFilter)."""
    if k <= 1:
        return img_l
    try:
        ModeFilter = ImageFilter.ModeFilter  # type: ignore[attr-defined]
        return img_l.filter(ModeFilter(size=k))
    except Exception:
        return img_l.filter(ImageFilter.MedianFilter(size=k))


def _remove_small_black_components(bw: Image.Image, area_min: int) -> Image.Image:
    # bw: imagem "L" binária 0/255 (preto/branco)
    w, h = bw.size
    px = bw.load()
    visited = [[False]*w for _ in range(h)]
    dirs = ((1,0),(-1,0),(0,1),(0,-1))
    for y in range(h):
        for x in range(w):
            if visited[y][x] or px[x,y] != 0:  # só preto
                continue
            # BFS do componente
            stack = [(x,y)]
            comp = []
            visited[y][x] = True
            while stack:
                cx, cy = stack.pop()
                comp.append((cx,cy))
                for dx,dy in dirs:
                    nx, ny = cx+dx, cy+dy
                    if 0 <= nx < w and 0 <= ny < h and not visited[ny][nx] and px[nx,ny] == 0:
                        visited[ny][nx] = True
                        stack.append((nx,ny))
            if len(comp) <= area_min:
                for (ux,uy) in comp:
                    px[ux,uy] = 255  # apaga (fica branco)
    return bw


def _keep_topk_black_components(bw: Image.Image, k: int, area_min: int = 0) -> Image.Image:
    # mantém só os k maiores componentes pretos (>= area_min); apaga o resto
    w, h = bw.size
    px = bw.load()
    visited = [[False]*w for _ in range(h)]
    dirs = ((1,0),(-1,0),(0,1),(0,-1))
    comps = []
    for y in range(h):
        for x in range(w):
            if visited[y][x] or px[x,y] != 0:
                continue
            stack = [(x,y)]
            comp = []
            visited[y][x] = True
            while stack:
                cx, cy = stack.pop()
                comp.append((cx,cy))
                for dx,dy in dirs:
                    nx, ny = cx+dx, cy+dy
                    if 0 <= nx < w and 0 <= ny < h and not visited[ny][nx] and px[nx,ny] == 0:
                        visited[ny][nx] = True
                        stack.append((nx,ny))
            if len(comp) >= max(1, area_min):
                comps.append(comp)
    comps.sort(key=len, reverse=True)
    to_keep = set()
    for comp in comps[:max(0,k)]:
        to_keep.update(comp)
    for y in range(h):
        for x in range(w):
            if px[x,y] == 0 and (x,y) not in to_keep:
                px[x,y] = 255
    return bw

def _remove_isolated_black_pixels(bw: Image.Image, min_neighbors: int, passes: int = 1) -> Image.Image:
    """
    Remove pixels pretos que não têm vizinhos pretos suficientes (majority filter 8-conexo).
    min_neighbors: 4–6 costuma funcionar muito bem em papel pontilhado/ruidoso.
    """
    w, h = bw.size
    for _ in range(max(1, passes)):
        px = bw.load()
        to_white = []
        for y in range(h):
            for x in range(w):
                if px[x, y] != 0:
                    continue
                cnt = 0
                for dy in (-1, 0, 1):
                    for dx in (-1, 0, 1):
                        if dx == 0 and dy == 0:
                            continue
                        nx, ny = x + dx, y + dy
                        if 0 <= nx < w and 0 <= ny < h and px[nx, ny] == 0:
                            cnt += 1
                if cnt < min_neighbors:
                    to_white.append((x, y))
        for x, y in to_white:
            px[x, y] = 255
    return bw

def _remove_horizontal_bars(bw: Image.Image, hmax: int, min_w_frac: float = 0.5, min_aspect: float = 6.0) -> Image.Image:
    """
    Remove componentes pretos horizontais (riscos/barras):
      - altura do bbox ≤ hmax
      - largura ≥ min_w_frac * largura_da_imagem
      - razão largura/altura ≥ min_aspect
    """
    if hmax <= 0:
        return bw
    w, h = bw.size
    px = bw.load()
    visited = [[False]*w for _ in range(h)]
    dirs = ((1,0),(-1,0),(0,1),(0,-1))
    for y in range(h):
        for x in range(w):
            if visited[y][x] or px[x,y] != 0:
                continue
            stack = [(x,y)]
            comp = []
            visited[y][x] = True
            minx = maxx = x
            miny = maxy = y
            while stack:
                cx, cy = stack.pop()
                comp.append((cx,cy))
                if cx < minx: minx = cx
                if cx > maxx: maxx = cx
                if cy < miny: miny = cy
                if cy > maxy: maxy = cy
                for dx,dy in dirs:
                    nx, ny = cx+dx, cy+dy
                    if 0 <= nx < w and 0 <= ny < h and not visited[ny][nx] and px[nx,ny] == 0:
                        visited[ny][nx] = True
                        stack.append((nx,ny))
            comp_w = (maxx - minx + 1)
            comp_h = (maxy - miny + 1)
            if comp_h <= hmax and comp_w >= int(min_w_frac * w) and (comp_w / max(1, comp_h)) >= min_aspect:
                for (ux,uy) in comp:
                    px[ux,uy] = 255
    return bw

def limpar_imagem_escaner(img: Image.Image, cfg: ConfigLimpezaScan) -> Image.Image:
    """
    Melhora leitura de NFs digitalizadas:
      - Reduz ruído (sal e pimenta) sem borrar bordas
      - Autocontraste + correção de gama + contraste
      - UnsharpMask com threshold para não realçar grão
      - Binarização por Otsu com viés pró-preto
      - Morfologia "abrir" e "fechar" sob medida
      - Remoção de barras horizontais (riscos de caneta)
      - Filtro de pontilhado por vizinhança (majority)
    """
    # 1) Cinza
    g = img.convert("L")

    # 2) Anti-ruído
    k_ruido = _odd_from_level(cfg.ruido)  # 1,3,5,7...
    g = _noise_filter(g, k_ruido)

    # 3) Autocontraste + contraste + gama
    g = ImageOps.autocontrast(g, cutoff=(cfg.corte_baixo, cfg.corte_alto))
    if abs(cfg.contraste - 1.0) > 1e-3:
        g = ImageEnhance.Contrast(g).enhance(_clamp(cfg.contraste, 0.1, 5.0))
    if abs(cfg.gama - 1.0) > 1e-3:
        g = g.point(_lookup_gamma(cfg.gama))

    # 4) Nitidez (unsharp com threshold para não amplificar ruído fino)
    if cfg.borda > 0:
        radius = 1.0 * _clamp(cfg.borda, 0.5, 3.0)
        percent = int(220 * _clamp(cfg.borda, 0.2, 3.0))
        g = g.filter(ImageFilter.UnsharpMask(radius=radius, percent=percent, threshold=2))

    # 5) Otsu com viés pró-preto
    t = _otsu_threshold(g)
    t = _clamp(int(t * _clamp(cfg.preto, 0.8, 1.6)), 0, 255)
    bw = g.point(lambda p, T=t: 0 if p < T else 255).convert("L")

    # 6) Morfologia: abrir (remove pontinhos) e fechar (consolida traço)
    if k_ruido > 1:
        bw = bw.filter(ImageFilter.MaxFilter(k_ruido))
        bw = bw.filter(ImageFilter.MinFilter(k_ruido))

    k_close = _odd_from_level(cfg.fechar_furos)
    if k_close > 1:
        bw = bw.filter(ImageFilter.MinFilter(k_close))
        bw = bw.filter(ImageFilter.MaxFilter(k_close))

    # 6.1) Remoção de pontinhos por componente conectado
    if getattr(cfg, "matar_pretos_ate", 0) > 0:
        bw = _remove_small_black_components(bw, cfg.matar_pretos_ate)

    # 6.2) (Opcional) manter apenas os N maiores blocos pretos (dígitos)
    if getattr(cfg, "manter_maiores_pretos", 0) > 0:
        bw = _keep_topk_black_components(bw, cfg.manter_maiores_pretos, area_min=getattr(cfg, "matar_pretos_ate", 0))

    # 6.3) Remoção específica de barras pretas horizontais (riscos de caneta)
    if getattr(cfg, "matar_barras_horiz_altura_max", 0) > 0:
        bw = _remove_horizontal_bars(
            bw,
            getattr(cfg, "matar_barras_horiz_altura_max", 0),
            getattr(cfg, "matar_barras_horiz_min_largura_frac", 0.5),
            6.0
        )
    # 6.4) Filtro "mata-pontilhado" por vizinhança (majority)
    if getattr(cfg, "matar_isolados_vizinhos", 0) > 0:
        bw = _remove_isolated_black_pixels(
            bw,
            int(getattr(cfg, "matar_isolados_vizinhos", 0)),
            int(getattr(cfg, "repetir_isolados_passes", 1))
        )

    if cfg.saida_1bit:
        bw = bw.convert("1")
    return bw

# Mantido por compatibilidade: a pipeline antiga agora chama a nova com fallback

def treat_strong(img: Image.Image) -> Image.Image:
    """Compat: mantém a assinatura, mas usa o pipeline configurável com CFG_FALLBACK."""
    return limpar_imagem_escaner(img, CFG_FALLBACK)

# ================== LAYOUTS ==================

@dataclass
class LayoutSpec:
    name: str
    keywords: List[str]
    crops: Dict[str, Union[Tuple[float, float, float, float], Dict[str, float]]]
    anti_keywords: List[str] = field(default_factory=list)
    filter_default: ConfigLimpezaScan = field(default_factory=lambda: replace(CFG_FALLBACK))
    filters_by_topic: Dict[str, ConfigLimpezaScan] = field(default_factory=dict)

# Layouts com crops + filtros por layout/tópico
LAYOUTS: List[LayoutSpec] = [
    LayoutSpec(
        name="PADRAO",
        keywords=["danfs", "DANFS", "Documento Auxiliar da NFS-e", "INTERMEDIÁRIO DO SERVIÇO", "Número Processo Suspensão",
                  "Cálculo do BM", "dps", "documento auxiliar da nfs-e"],
        crops={
            "DATA": {"ESQUERDA": 40, "CIMA": 7, "DIREITA": 70, "BAIXO": 15},
            "MUNICIPIO_TOMADOR": {"ESQUERDA": 2, "CIMA": 24, "DIREITA": 98, "BAIXO": 33},
            "NUMERO_NFS": {"ESQUERDA": 0, "CIMA": 5, "DIREITA": 24, "BAIXO": 15},
            "PRESTADOR_CNPJ": {"ESQUERDA": 3, "CIMA": 12, "DIREITA": 97, "BAIXO": 24},
            "VALOR_NFS": {"ESQUERDA": 55, "CIMA": 55, "DIREITA": 97, "BAIXO": 80},
        },
        filter_default=replace(CFG_FALLBACK, preto=1.25, ruido=1.0, contraste=1.20, gama=1.15, borda=1.2),
        filters_by_topic={
            # Replicado por tópico (você pode personalizar cada um):
            "DATA":            replace(CFG_FALLBACK, preto=2.8, ruido=1.8, contraste=5.0, gama=1.10, borda=1.0, fechar_furos=1.2),
            "MUNICIPIO_TOMADOR":replace(CFG_FALLBACK, preto=2.8, ruido=1.8, contraste=5.0, gama=1.12, borda=1.1, fechar_furos=1.2),
            "NUMERO_NFS":      replace(CFG_FALLBACK, preto=2.8, ruido=3.0, contraste=6.0, gama=1.2, borda=1.4, saida_1bit=True, fechar_furos=1.0),
            "PRESTADOR_CNPJ":  replace(CFG_FALLBACK, preto=2.8, ruido=2.3, contraste=6.0, gama=1.20, borda=1.5, fechar_furos=1.0),
            "VALOR_NFS":       replace(CFG_FALLBACK, preto=1.45, ruido=3.0, contraste=2.5, gama=1.22, borda=1.5, saida_1bit=True, matar_pretos_ate=15),
        },
    ),
    LayoutSpec(
        name="PORANGATU",
        keywords=["porangatu", "código de verifica", "codigo de verifica"],
        anti_keywords=[
            "danfs", "DANFS", "DANFSE", "DANFS", "Número da DPS", "Cálculo do BM", "INTERMEDIÁRIO DO SERVIÇO",
            "documento auxiliar da nfs-e"
        ],
        crops={
            "DATA": {"ESQUERDA": 65, "CIMA": 2, "DIREITA": 97, "BAIXO": 15},
            "MUNICIPIO_TOMADOR": {"ESQUERDA": 3, "CIMA": 20, "DIREITA": 97, "BAIXO": 30},
            "NUMERO_NFS": {"ESQUERDA": 65, "CIMA": 2, "DIREITA": 97, "BAIXO": 15},
            "PRESTADOR_CNPJ": {"ESQUERDA": 3, "CIMA": 11, "DIREITA": 97, "BAIXO": 21},
            "VALOR_NFS": {"ESQUERDA": 60, "CIMA": 60, "DIREITA": 97, "BAIXO": 70},
        },
        filter_default=replace(CFG_FALLBACK, preto=1.22, ruido=1.0, contraste=1.18, gama=1.12, borda=1.1),
        filters_by_topic={
            "DATA":            replace(CFG_FALLBACK, preto=1.30, ruido=2.5, contraste=1.08, gama=1.12, borda=1.2),
            "MUNICIPIO_TOMADOR":replace(CFG_FALLBACK, preto=1.22, ruido=1.0, contraste=1.18, gama=1.12, borda=1.1),
            "NUMERO_NFS":      replace(CFG_FALLBACK, preto=1.26, ruido=1.0, contraste=1.25, gama=1.18, borda=1.3),
            "PRESTADOR_CNPJ":  replace(CFG_FALLBACK, preto=1.5, ruido=2.0, contraste=1.5, gama=1.20, borda=1.0, zoom=1.5),
            "VALOR_NFS":       replace(CFG_FALLBACK, preto=1.32, ruido=1.0, contraste=1.30, gama=1.18, borda=1.4, saida_1bit=True),
        },
    ),
    LayoutSpec(
        name="LAGOA DA CONFUSAO",
        keywords=["Lagoa da Confusao", "Lagoa da Confusão", "LAGOA DA CONFUSAO", "LAGOA DA CONFUSÃO",
                  "LAGOA CONFUSAO", "LAGOA DA CONFUSAO - TO", "LAGOA DA CONFUSÃO - TO", "AVENIDA VITORINO PANTA",
                  "19.607.195/0001-60", "77.493-000", "77493-000"],
        anti_keywords=[
            "34.301.768/0001-17", "BARBARA EVELIN", "DANFSE", "DANFS", "Número da DPS", "Incentivador Cultural", "ISS a reter"
        ],
        crops={
            "DATA": {"ESQUERDA": 65, "CIMA": 2, "DIREITA": 97, "BAIXO": 15},
            "MUNICIPIO_TOMADOR": {"ESQUERDA": 3, "CIMA": 24, "DIREITA": 97, "BAIXO": 40},
            "NUMERO_NFS": {"ESQUERDA": 65, "CIMA": 2, "DIREITA": 97, "BAIXO": 15},
            "PRESTADOR_CNPJ": {"ESQUERDA": 3, "CIMA": 15, "DIREITA": 97, "BAIXO": 28},
            "VALOR_NFS": {"ESQUERDA": 60, "CIMA": 45, "DIREITA": 97, "BAIXO": 70},
        },
        filter_default=replace(CFG_FALLBACK, preto=1.22, ruido=1.0, contraste=1.18, gama=1.12, borda=1.1),
        filters_by_topic={
            "DATA":            replace(CFG_FALLBACK, preto=2.5, ruido=4.0, contraste=5.6, gama=1.18, borda=1.0, fechar_furos=3.5),
            "MUNICIPIO_TOMADOR":replace(CFG_FALLBACK, preto=2.0, ruido=3.2, contraste=3.5, gama=1.20, borda=1.0),
            "NUMERO_NFS":      replace(CFG_FALLBACK, preto=2.5, ruido=4.0, contraste=5.6, gama=1.18, borda=1.0, fechar_furos=3.5),
            "PRESTADOR_CNPJ":  replace(CFG_FALLBACK, preto=2.0, ruido=3.2, contraste=3.5, gama=1.20, borda=1.0, zoom=1.5),
            "VALOR_NFS":       replace(CFG_FALLBACK, preto=1.26, ruido=3.5, contraste=3.5, gama=1.25, borda=1.5, saida_1bit=True),
        },
    ),
    LayoutSpec(
        name="LAGOA DA CONFUSAO 2",
        keywords=["Lagoa da Confusao", "Lagoa da Confusão", "LAGOA DA CONFUSAO", "LAGOA DA CONFUSÃO",
                  "LAGOA CONFUSAO", "LAGOA DA CONFUSAO - TO", "LAGOA DA CONFUSÃO - TO", "AVENIDA VITORINO PANTA",
                  "19.607.195/0001-60", "77.493-000", "77493-000"],
        anti_keywords=[
            "danfs", "DANFS", "DANFSE", "DANFS", "Número da DPS", "Cálculo do BM", "INTERMEDIÁRIO DO SERVIÇO",
            "documento auxiliar da nfs-e"
        ],
        crops={
            "DATA": {"ESQUERDA": 65, "CIMA": 2, "DIREITA": 99, "BAIXO": 15},
            "MUNICIPIO_TOMADOR": {"ESQUERDA": 3, "CIMA": 28, "DIREITA": 97, "BAIXO": 43},
            "NUMERO_NFS": {"ESQUERDA": 65, "CIMA": 2, "DIREITA": 99, "BAIXO": 15},
            "PRESTADOR_CNPJ": {"ESQUERDA": 3, "CIMA": 20, "DIREITA": 97, "BAIXO": 33},
            "VALOR_NFS": {"ESQUERDA": 0, "CIMA": 55, "DIREITA": 50, "BAIXO": 100},
        },
        filter_default=replace(CFG_FALLBACK, preto=1.22, ruido=1.0, contraste=1.18, gama=1.12, borda=1.1),
        filters_by_topic={
            "DATA":            replace(CFG_FALLBACK, preto=2.5, ruido=4.0, contraste=5.6, gama=1.18, borda=1.0, fechar_furos=3.5),
            "MUNICIPIO_TOMADOR":replace(CFG_FALLBACK, preto=2.0, ruido=3.2, contraste=3.5, gama=1.20, borda=1.0),
            "NUMERO_NFS":      replace(CFG_FALLBACK, preto=2.5, ruido=4.0, contraste=5.6, gama=1.18, borda=1.0, fechar_furos=3.5),
            "PRESTADOR_CNPJ":  replace(CFG_FALLBACK, preto=2.0, ruido=3.2, contraste=3.5, gama=1.20, borda=1.0, zoom=1.5),
            "VALOR_NFS":       replace(CFG_FALLBACK, preto=1.26, ruido=3.5, contraste=3.5, gama=1.25, borda=1.5, saida_1bit=True),
        },
    ),
    LayoutSpec(
        name="PALMAS",
        keywords=["municipio de palmas", "77.021-900", "77021900", "Palmas/TO",
                  "502 Sul, Paço Municipal"],
        anti_keywords=[
            "danfs", "DANFS", "DANFSE", "DANFS", "Número da DPS", "Cálculo do BM", "INTERMEDIÁRIO DO SERVIÇO",
            "documento auxiliar da nfs-e"
        ],
        crops={
            "DATA": {"ESQUERDA": 3, "CIMA": 7, "DIREITA": 35, "BAIXO": 17},
            "MUNICIPIO_TOMADOR": {"ESQUERDA": 3, "CIMA": 33, "DIREITA": 84, "BAIXO": 45},
            "NUMERO_NFS": {"ESQUERDA": 65, "CIMA": 2, "DIREITA": 97, "BAIXO": 15},
            "PRESTADOR_CNPJ": {"ESQUERDA": 3, "CIMA": 17, "DIREITA": 84, "BAIXO": 32},
            "VALOR_NFS": {"ESQUERDA": 65, "CIMA": 77, "DIREITA": 102, "BAIXO": 89},
        },
        filter_default=replace(CFG_FALLBACK, preto=1.27, ruido=1.0, contraste=1.22, gama=1.15, borda=1.2),
        filters_by_topic={
            "DATA":            replace(CFG_FALLBACK, preto=1.20, ruido=1.0, contraste=1.10, gama=1.10, borda=1.0),
            "MUNICIPIO_TOMADOR":replace(CFG_FALLBACK, preto=1.27, ruido=1.0, contraste=1.22, gama=1.15, borda=1.2),
            "NUMERO_NFS":      replace(CFG_FALLBACK, preto=3.0, ruido=25, contraste=8, gama=1.0, borda=2.0, fechar_furos=1.0, matar_isolados_vizinhos=8, repetir_isolados_passes=2, matar_pretos_ate=500),
            "PRESTADOR_CNPJ":  replace(CFG_FALLBACK, preto=1.30, ruido=1.2, contraste=1.30, gama=1.0, borda=1.0),
            "VALOR_NFS":       replace(CFG_FALLBACK, preto=1.45, ruido=3.0, contraste=2.0, gama=1.22, borda=1.5, saida_1bit=True),
        },
    ),
    LayoutSpec(
        name="CERES",
        keywords=["CERES", "CERES - GO",
                  "76.300-000", "76300-000", "76.300-000", "76.300000", "76300000",
                  "PREFEITURA MUNICIPAL DE CERES"],
        anti_keywords=[
            "danfs", "DANFS", "DANFSE", "DANFS", "Número da DPS", "Cálculo do BM", "INTERMEDIÁRIO DO SERVIÇO",
            "documento auxiliar da nfs-e"
        ],
        crops={
            "DATA": {"ESQUERDA": 65, "CIMA": 20, "DIREITA": 97, "BAIXO": 35},
            "MUNICIPIO_TOMADOR": {"ESQUERDA": 2, "CIMA": 26, "DIREITA": 98, "BAIXO": 36},
            "NUMERO_NFS": {"ESQUERDA": 65, "CIMA": 20, "DIREITA": 97, "BAIXO": 35},
            "PRESTADOR_CNPJ": {"ESQUERDA": 3, "CIMA": 19, "DIREITA": 97, "BAIXO": 29},
            "VALOR_NFS": {"ESQUERDA": 55, "CIMA": 80, "DIREITA": 97, "BAIXO": 115},
        },
        filter_default=replace(CFG_FALLBACK, preto=1.25, ruido=1.0, contraste=1.20, gama=1.15, borda=1.2),
        filters_by_topic={
            # Replicado por tópico (você pode personalizar cada um):
            "DATA":            replace(CFG_FALLBACK, preto=2.8, ruido=1.8, contraste=5.0, gama=1.10, borda=1.0, fechar_furos=1.2),
            "MUNICIPIO_TOMADOR":replace(CFG_FALLBACK, preto=2.8, ruido=1.8, contraste=5.0, gama=1.12, borda=1.1, fechar_furos=1.5),
            "NUMERO_NFS":      replace(CFG_FALLBACK, preto=2.8, ruido=3.0, contraste=6.0, gama=1.2, borda=1.4, saida_1bit=True, fechar_furos=1.0),
            "PRESTADOR_CNPJ":  replace(CFG_FALLBACK, preto=2.8, ruido=2.0, contraste=4.0, gama=1.20, borda=1.0, fechar_furos=1.5),
            "VALOR_NFS":       replace(CFG_FALLBACK, preto=1.45, ruido=3.0, contraste=2.5, gama=1.22, borda=1.5, saida_1bit=True, matar_pretos_ate=15),
        },
    ),
    LayoutSpec(
        name="GOIANIA",
        keywords=["GOIANIA", "GOIANIA-GO", "AV CASTELO BRANCO", "SECRETARIA MUNICIPAL DE FINANÇAS",
                  "SECRETARIA MUNICIPAL DE FINANCAS"],
        anti_keywords=[
            "danfs", "DANFS", "DANFSE", "DANFS", "Número da DPS", "Cálculo do BM", "INTERMEDIÁRIO DO SERVIÇO",
            "documento auxiliar da nfs-e"
        ],
        crops={
            "DATA": {"ESQUERDA": 60, "CIMA": 3, "DIREITA": 124, "BAIXO": 11},
            "MUNICIPIO_TOMADOR": {"ESQUERDA": 2, "CIMA": 17, "DIREITA": 98, "BAIXO": 27},
            "NUMERO_NFS": {"ESQUERDA": 60, "CIMA": 3, "DIREITA": 124, "BAIXO": 11},
            "PRESTADOR_CNPJ": {"ESQUERDA": 3, "CIMA": 9, "DIREITA": 97, "BAIXO": 19},
            "VALOR_NFS": {"ESQUERDA": 55, "CIMA": 65, "DIREITA": 97, "BAIXO": 90},
        },
        filter_default=replace(CFG_FALLBACK, preto=1.25, ruido=1.0, contraste=1.20, gama=1.15, borda=1.2),
        filters_by_topic={
            # Replicado por tópico (você pode personalizar cada um):
            "DATA":            replace(CFG_FALLBACK, preto=2.8, ruido=3., contraste=6.0, gama=1.2, borda=1.4, saida_1bit=True, fechar_furos=1.0),
            "MUNICIPIO_TOMADOR":replace(CFG_FALLBACK, preto=2.8, ruido=1.8, contraste=5.0, gama=1.12, borda=1.1, fechar_furos=1.2),
            "NUMERO_NFS":      replace(CFG_FALLBACK, preto=2.8, ruido=3.0, contraste=6.0, gama=1.2, borda=1.4, saida_1bit=True, fechar_furos=1.0),
            "PRESTADOR_CNPJ":  replace(CFG_FALLBACK, preto=2.8, ruido=2.0, contraste=3.0, gama=1.20, borda=1.0, fechar_furos=1.2),
            "VALOR_NFS":       replace(CFG_FALLBACK, preto=1.45, ruido=3.5, contraste=3.0, gama=1.22, borda=1.5, saida_1bit=True),
        },
    ),
    LayoutSpec(
        name="FORMOSO DO ARAGUAIA",
        keywords=["FORMOSO DO ARAGUAIA","02.075.216/0001-60", "77.470-000", "77470-000",
                  "PREFEITURA MUNICIPAL DE FORMOSO DO ARAGUAIA", "AV. HERMINIO AZEVEDO SAORES",
                  "FORMOSO DO ARAGUAIA - TO"],
        anti_keywords=[
            "danfs", "DANFS", "DANFSE", "DANFS", "Número da DPS", "Cálculo do BM", "INTERMEDIÁRIO DO SERVIÇO",
            "documento auxiliar da nfs-e"
        ],
        crops={
            "DATA": {"ESQUERDA": 40, "CIMA": 7, "DIREITA": 70, "BAIXO": 15},
            "MUNICIPIO_TOMADOR": {"ESQUERDA": 2, "CIMA": 24, "DIREITA": 98, "BAIXO": 33},
            "NUMERO_NFS": {"ESQUERDA": 0, "CIMA": 5, "DIREITA": 24, "BAIXO": 15},
            "PRESTADOR_CNPJ": {"ESQUERDA": 3, "CIMA": 14, "DIREITA": 97, "BAIXO": 23},
            "VALOR_NFS": {"ESQUERDA": 55, "CIMA": 55, "DIREITA": 97, "BAIXO": 80},
        },
        filter_default=replace(CFG_FALLBACK, preto=1.25, ruido=1.0, contraste=1.20, gama=1.15, borda=1.2),
        filters_by_topic={
            # Replicado por tópico (você pode personalizar cada um):
            "DATA":            replace(CFG_FALLBACK, preto=2.8, ruido=1.8, contraste=5.0, gama=1.10, borda=1.0, fechar_furos=1.2),
            "MUNICIPIO_TOMADOR":replace(CFG_FALLBACK, preto=2.8, ruido=1.8, contraste=5.0, gama=1.12, borda=1.1, fechar_furos=1.2),
            "NUMERO_NFS":      replace(CFG_FALLBACK, preto=2.8, ruido=3.0, contraste=6.0, gama=1.2, borda=1.4, saida_1bit=True, fechar_furos=1.0),
            "PRESTADOR_CNPJ":  replace(CFG_FALLBACK, preto=2.8, ruido=2.0, contraste=3.0, gama=1.20, borda=1.0, fechar_furos=1.2),
            "VALOR_NFS":       replace(CFG_FALLBACK, preto=1.45, ruido=3.0, contraste=2.5, gama=1.22, borda=1.5, saida_1bit=True, matar_pretos_ate=15),
        },
    ),
    LayoutSpec(
        name="URUACU 2",
        keywords=["ESPECIAL MODELO", "MODELO UNICO", "ESPECIAL MODELO ÚNICO", "AUT (AIDF)", "ESPECIAL MODELO ÚNICO - AUT (AIDF)",
                   "CÓD CONTROLE INTERNO", "COD CONTROLE INTERNO", "Serviços autorizadas 1 a 7", "Servicos autorizadas 1 a 7"],
        anti_keywords=[
            "danfs", "DANFS", "DANFSE", "DANFS", "Número da DPS", "Cálculo do BM", "INTERMEDIÁRIO DO SERVIÇO",
            "documento auxiliar da nfs-e"
        ],
        crops={
            "DATA": {"ESQUERDA": 55, "CIMA": 75, "DIREITA": 97, "BAIXO": 95},
            "MUNICIPIO_TOMADOR": {"ESQUERDA": 2, "CIMA": 26, "DIREITA": 98, "BAIXO": 34},
            "NUMERO_NFS": {"ESQUERDA": 55, "CIMA": 7, "DIREITA": 95, "BAIXO": 15},
            "PRESTADOR_CNPJ": {"ESQUERDA": 3, "CIMA": 15, "DIREITA": 97, "BAIXO": 25},
            "VALOR_NFS": {"ESQUERDA": 55, "CIMA": 50, "DIREITA": 97, "BAIXO": 75},
        },
        filter_default=replace(CFG_FALLBACK, preto=1.25, ruido=1.0, contraste=1.20, gama=1.15, borda=1.2),
        filters_by_topic={
            # Replicado por tópico (você pode personalizar cada um):
            "DATA":            replace(CFG_FALLBACK, preto=2.8, ruido=1.8, contraste=5.0, gama=1.10, borda=1.0, fechar_furos=1.2),
            "MUNICIPIO_TOMADOR":replace(CFG_FALLBACK, preto=2.8, ruido=1.8, contraste=5.0, gama=1.12, borda=1.1, fechar_furos=1.2),
            "NUMERO_NFS":      replace(CFG_FALLBACK, preto=2.8, ruido=3.0, contraste=6.0, gama=1.2, borda=1.4, saida_1bit=True, fechar_furos=1.0),
            "PRESTADOR_CNPJ":  replace(CFG_FALLBACK, preto=2.8, ruido=2.0, contraste=3.0, gama=1.20, borda=1.0, fechar_furos=1.2),
            "VALOR_NFS":       replace(CFG_FALLBACK, preto=1.45, ruido=3.0, contraste=5.0, gama=1.22, borda=1.5, saida_1bit=True),
        },
    ),
    LayoutSpec(
        name="URUACU",
        keywords=["URUACU", "URUAÇU", "01.219807/0001-82", "AV. HERMINIO AZEVEDO SAORES"],
        anti_keywords=[
            "ESPECIAL MODELO", "MODELO UNICO", "MODELO ÚNICO", "AUT (AIDF)",
            "CÓD CONTROLE INTERNO", "COD CONTROLE INTERNO",
            "SERVIÇOS AUTORIZADAS 1 A 7", "SERVICOS AUTORIZADAS 1 A 7"
        ],
        crops={
            "DATA": {"ESQUERDA": 68, "CIMA": 4, "DIREITA": 98, "BAIXO": 15},
            "MUNICIPIO_TOMADOR": {"ESQUERDA": 2, "CIMA": 25, "DIREITA": 98, "BAIXO": 35},
            "NUMERO_NFS": {"ESQUERDA": 68, "CIMA": 4, "DIREITA": 98, "BAIXO": 15},
            "PRESTADOR_CNPJ": {"ESQUERDA": 3, "CIMA": 14, "DIREITA": 97, "BAIXO": 26},
            "VALOR_NFS": {"ESQUERDA": 70, "CIMA": 42, "DIREITA": 100, "BAIXO": 62},
        },
        filter_default=replace(CFG_FALLBACK, preto=1.25, ruido=1.0, contraste=1.20, gama=1.15, borda=1.2),
        filters_by_topic={
            "DATA":            replace(CFG_FALLBACK, preto=2.8, ruido=1.8, contraste=5.0, gama=1.10, borda=1.0, fechar_furos=1.2, matar_isolados_vizinhos=6, repetir_isolados_passes=3, matar_pretos_ate=600),
            "MUNICIPIO_TOMADOR":replace(CFG_FALLBACK, preto=2.8, ruido=2.5, contraste=6.0, gama=1.12, borda=1.1, fechar_furos=1.2),
            "NUMERO_NFS":      replace(CFG_FALLBACK, preto=2.8, ruido=1.8, contraste=5.0, gama=1.10, borda=1.0, fechar_furos=1.2, matar_isolados_vizinhos=6, repetir_isolados_passes=3, matar_pretos_ate=600),
            "PRESTADOR_CNPJ":  replace(CFG_FALLBACK, preto=2.8, ruido=2.5, contraste=4.0, gama=1.20, borda=1.0, fechar_furos=1.2),
            "VALOR_NFS":       replace(CFG_FALLBACK, preto=1.45, ruido=3.0, contraste=4.0, gama=1.22, borda=1.5, saida_1bit=True, matar_pretos_ate=15),
        },
    ),
    LayoutSpec(
        name="TASSO FRAGOSO",
        keywords=["TASSO FRAGOSO", "TASSO FRAGOSO - MA", "TASSO FRAGOSO-MA", "65820-000", "65820000",
                  "MUNICIPIO DE TASSO FRAGOSO"],
        anti_keywords=[
            "danfs", "DANFS", "DANFSE", "DANFS", "Número da DPS", "Cálculo do BM", "INTERMEDIÁRIO DO SERVIÇO",
            "documento auxiliar da nfs-e"
        ],
        crops={
            "DATA": {"ESQUERDA": 68, "CIMA": 4, "DIREITA": 95, "BAIXO": 15},
            "MUNICIPIO_TOMADOR": {"ESQUERDA": 2, "CIMA": 35, "DIREITA": 98, "BAIXO": 50},
            "NUMERO_NFS": {"ESQUERDA": 68, "CIMA": 4, "DIREITA": 95, "BAIXO": 15},
            "PRESTADOR_CNPJ": {"ESQUERDA": 2, "CIMA": 23, "DIREITA": 98, "BAIXO": 35},
            "VALOR_NFS": {"ESQUERDA": 0, "CIMA": 65, "DIREITA": 30, "BAIXO": 80},
        },
        filter_default=replace(CFG_FALLBACK, preto=1.22, ruido=1.0, contraste=1.18, gama=1.12, borda=1.1),
        filters_by_topic={
            "DATA":            replace(CFG_FALLBACK, preto=2.5, ruido=3.0, contraste=6.5, gama=1.18, borda=1.0, fechar_furos=2.5),
            "MUNICIPIO_TOMADOR":replace(CFG_FALLBACK, preto=2.0, ruido=3.5, contraste=6.5, gama=1.20, borda=1.0),
            "NUMERO_NFS":      replace(CFG_FALLBACK, preto=2.5, ruido=3.0, contraste=6.5, gama=1.18, borda=1.0, fechar_furos=1.5, matar_pretos_ate=200),
            "PRESTADOR_CNPJ":  replace(CFG_FALLBACK, preto=2.0, ruido=3.5, contraste=6.0, gama=1.20, borda=1.0, zoom=1.5),
            "VALOR_NFS":       replace(CFG_FALLBACK, preto=2.0, ruido=3.0, contraste=6.0, gama=1.25, borda=1.5, saida_1bit=True),
        },
    ),
    LayoutSpec(
        name="PARAISO DO TOCANTINS",
        keywords=["PARAISO DO TOCANTINS", "PARAÍSO DO TOCANTINS", "77.600-000", "77600000", "Av. Transbrasiliana",
                  "Paraíso do Tocantins - TO", "Paraíso do"],
        anti_keywords=[
            "danfs", "DANFS", "DANFSE", "DANFS", "Número da DPS", "Cálculo do BM", "INTERMEDIÁRIO DO SERVIÇO",
            "documento auxiliar da nfs-e"
        ],
        crops={
            "DATA": {"ESQUERDA": 0, "CIMA": 10, "DIREITA": 30, "BAIXO": 18},
            "MUNICIPIO_TOMADOR": {"ESQUERDA": 2, "CIMA": 35, "DIREITA": 98, "BAIXO": 45},
            "NUMERO_NFS": {"ESQUERDA": 65, "CIMA": 2, "DIREITA": 97, "BAIXO": 15},
            "PRESTADOR_CNPJ": {"ESQUERDA": 2, "CIMA": 24, "DIREITA": 98, "BAIXO": 33},
            "VALOR_NFS": {"ESQUERDA": 55, "CIMA": 68, "DIREITA": 97, "BAIXO": 103},
        },
        filter_default=replace(CFG_FALLBACK, preto=1.25, ruido=1.0, contraste=1.20, gama=1.15, borda=1.2),
        filters_by_topic={
            # Replicado por tópico (você pode personalizar cada um):
            "DATA":            replace(CFG_FALLBACK, preto=2.8, ruido=1.8, contraste=5.0, gama=1.10, borda=1.0, fechar_furos=1.2),
            "MUNICIPIO_TOMADOR":replace(CFG_FALLBACK, preto=2.8, ruido=1.8, contraste=5.0, gama=1.12, borda=1.1, fechar_furos=1.2),
            "NUMERO_NFS":      replace(CFG_FALLBACK, preto=3.0, ruido=6.0, contraste=8, gama=1.0, borda=2.0, fechar_furos=1.0, matar_isolados_vizinhos=6, repetir_isolados_passes=2, matar_pretos_ate=500),
            "PRESTADOR_CNPJ":  replace(CFG_FALLBACK, preto=2.8, ruido=2.0, contraste=3.0, gama=1.20, borda=1.0, fechar_furos=1.2),
            "VALOR_NFS":       replace(CFG_FALLBACK, preto=1.45, ruido=3.0, contraste=2.5, gama=1.22, borda=1.5, saida_1bit=True, matar_pretos_ate=15),
        },
    ),
    LayoutSpec(
        name="MONTIVIDIU DO NORTE",
        keywords=["MONTIVIDIU DO NORTE", "MUNICIPAL MONTIVIDIU DO NORTE", "Rua Rita Candida", "76.465-000", "76465-000", "3384-6282"],
        anti_keywords=[
            "danfs", "DANFS", "DANFSE", "DANFS", "Número da DPS", "Cálculo do BM", "INTERMEDIÁRIO DO SERVIÇO",
            "documento auxiliar da nfs-e"
        ],
        crops={
            "DATA": {"ESQUERDA": 65, "CIMA": 2, "DIREITA": 97, "BAIXO": 15},
            "MUNICIPIO_TOMADOR": {"ESQUERDA": 3, "CIMA": 24, "DIREITA": 97, "BAIXO": 35},
            "NUMERO_NFS": {"ESQUERDA": 65, "CIMA": 2, "DIREITA": 97, "BAIXO": 15},
            "PRESTADOR_CNPJ": {"ESQUERDA": 3, "CIMA": 13, "DIREITA": 97, "BAIXO": 26},
            "VALOR_NFS": {"ESQUERDA": 60, "CIMA": 46, "DIREITA": 97, "BAIXO": 58},
        },
        filter_default=replace(CFG_FALLBACK, preto=1.22, ruido=1.0, contraste=1.18, gama=1.12, borda=1.1),
        filters_by_topic={
            "DATA":            replace(CFG_FALLBACK, preto=3.0, ruido=2.5, contraste=6.0, gama=1.18, borda=1.0, fechar_furos=3.5),
            "MUNICIPIO_TOMADOR":replace(CFG_FALLBACK, preto=2.0, ruido=3.5, contraste=4.0, gama=1.20, borda=1.0),
            "NUMERO_NFS":      replace(CFG_FALLBACK, preto=3.0, ruido=2.5, contraste=6.0, gama=1.18, borda=1.0, fechar_furos=3.5),
            "PRESTADOR_CNPJ":  replace(CFG_FALLBACK, preto=2.5, ruido=3.5, contraste=4.0, gama=1.20, borda=1.0, zoom=1.5),
            "VALOR_NFS":       replace(CFG_FALLBACK, preto=1.26, ruido=3.5, contraste=3.5, gama=1.25, borda=2.5, saida_1bit=True),
        },
    ),
    LayoutSpec(
        name="RIALMA",
        keywords=["RIALMA", "MUNICIPAL DE RIALMA", "01.135.904/0001-97"],
        anti_keywords=[
            "danfs", "DANFS", "DANFSE", "DANFS", "Número da DPS", "Cálculo do BM", "INTERMEDIÁRIO DO SERVIÇO",
            "documento auxiliar da nfs-e"
        ],
        crops={
            "DATA": {"ESQUERDA": 68, "CIMA": 0, "DIREITA": 97, "BAIXO": 25},
            "MUNICIPIO_TOMADOR": {"ESQUERDA": 2, "CIMA": 23, "DIREITA": 98, "BAIXO": 35},
            "NUMERO_NFS": {"ESQUERDA": 68, "CIMA": 0, "DIREITA": 97, "BAIXO": 25},
            "PRESTADOR_CNPJ": {"ESQUERDA": 3, "CIMA": 12, "DIREITA": 97, "BAIXO": 23},
            "VALOR_NFS": {"ESQUERDA": 45, "CIMA": 55, "DIREITA": 97, "BAIXO": 80},
        },
        filter_default=replace(CFG_FALLBACK, preto=1.25, ruido=1.0, contraste=1.20, gama=1.15, borda=1.2),
        filters_by_topic={
            # Replicado por tópico (você pode personalizar cada um):
            "DATA":            replace(CFG_FALLBACK, preto=2.8, ruido=2.2, contraste=5.3, gama=1.10, borda=1.0, fechar_furos=1.2, matar_isolados_vizinhos=6, repetir_isolados_passes=2, matar_pretos_ate=400),
            "MUNICIPIO_TOMADOR":replace(CFG_FALLBACK, preto=2.8, ruido=1.8, contraste=5.0, gama=1.12, borda=1.1, fechar_furos=1.5),
            "NUMERO_NFS":      replace(CFG_FALLBACK, preto=2.8, ruido=3.0, contraste=6.0, gama=1.2, borda=1.4, saida_1bit=True, fechar_furos=1.0, matar_isolados_vizinhos=6, repetir_isolados_passes=2, matar_pretos_ate=400),
            "PRESTADOR_CNPJ":  replace(CFG_FALLBACK, preto=2.8, ruido=2.5, contraste=5.0, gama=1.20, borda=1.0, fechar_furos=1.5),
            "VALOR_NFS":       replace(CFG_FALLBACK, preto=5.0, ruido=3.0, contraste=25, gama=2.0, borda=1.5, saida_1bit=True, fechar_furos=1.0, matar_isolados_vizinhos=6, repetir_isolados_passes=2, matar_pretos_ate=50),
        }
    ),
    LayoutSpec(
        name="GURUPI",
        keywords=["gurupi", "municipio de gurupi", "77.405-070"],
        anti_keywords=[
            "danfs", "DANFS", "DANFSE", "DANFS", "Número da DPS", "Cálculo do BM", "INTERMEDIÁRIO DO SERVIÇO",
            "documento auxiliar da nfs-e"
        ],
        crops={
            "DATA": {"ESQUERDA": 3, "CIMA": 12, "DIREITA": 35, "BAIXO": 22},
            "MUNICIPIO_TOMADOR": {"ESQUERDA": 3, "CIMA": 34, "DIREITA": 84, "BAIXO": 46},
            "NUMERO_NFS": {"ESQUERDA": 65, "CIMA": 2, "DIREITA": 97, "BAIXO": 15},
            "PRESTADOR_CNPJ": {"ESQUERDA": 3, "CIMA": 21, "DIREITA": 84, "BAIXO": 35},
            "VALOR_NFS": {"ESQUERDA": 65, "CIMA": 77, "DIREITA": 102, "BAIXO": 89},
        },
        filter_default=replace(CFG_FALLBACK, preto=1.27, ruido=1.0, contraste=1.22, gama=1.15, borda=1.2),
        filters_by_topic={
            "DATA":            replace(CFG_FALLBACK, preto=1.20, ruido=1.0, contraste=1.10, gama=1.10, borda=1.0),
            "MUNICIPIO_TOMADOR":replace(CFG_FALLBACK, preto=1.27, ruido=1.0, contraste=1.22, gama=1.15, borda=1.2),
            "NUMERO_NFS":      replace(CFG_FALLBACK, preto=1.8, ruido=6.0, contraste=2.0, gama=1.0, borda=2.7,fechar_furos=2.8, matar_pretos_ate=50),
            "PRESTADOR_CNPJ":  replace(CFG_FALLBACK, preto=1.28, ruido=1.2, contraste=1.20, gama=1.0, borda=1.0),
            "VALOR_NFS":       replace(CFG_FALLBACK, preto=1.45, ruido=3.0, contraste=2.0, gama=1.22, borda=1.5, saida_1bit=True),
        },
    ),
]

# Garante que todo tópico tenha um filtro (se não definido, copia o default do layout)
for _lay in LAYOUTS:
    for _topic in _lay.crops.keys():
        if _topic not in _lay.filters_by_topic:
            _lay.filters_by_topic[_topic] = replace(_lay.filter_default)

# >>> FORÇA ANTI-RISCO nas NFS (mesmo comportamento do Recebimento)
for _lay in LAYOUTS:
    for _topic, _cfg in _lay.filters_by_topic.items():
        if _topic in ("DATA", "NUMERO_NFS", "VALOR_NFS"):
            _cfg.matar_isolados_vizinhos = max(getattr(_cfg, "matar_isolados_vizinhos", 0), 5)
            _cfg.matar_barras_horiz_altura_max = 100
            _cfg.matar_barras_horiz_min_largura_frac = 0.25

# ================== OCR / CROP ==================

def _page_header_text(png_path: str, frac_altura: float = 0.22) -> str:
    """OCR rápido do topo da página (cabeçalho)."""
    try:
        with Image.open(png_path) as im:
            w, h = im.size
            hh = max(1, int(h * frac_altura))
            header = im.crop((0, 0, w, hh))
            return pytesseract.image_to_string(header, lang=OCR_LANG, config="--psm 6") or ""
    except Exception:
        return ""

def _page_footer_text(png_path: str, frac_altura: float = 0.22) -> str:
    """OCR rápido do rodapé da página."""
    try:
        with Image.open(png_path) as im:
            w, h = im.size
            hh = max(1, int(h * frac_altura))
            footer = im.crop((0, h - hh, w, h))
            return pytesseract.image_to_string(footer, lang=OCR_LANG, config="--psm 6") or ""
    except Exception:
        return ""

def _count_hits(text_norm: str, keywords: List[str]) -> int:
    t = (text_norm or "").casefold()
    return sum(1 for kw in keywords if kw and kw.casefold() in t)

def _score_layout_zones(full_text: str, header_text: str, footer_text: str, lay: "LayoutSpec", rival_names: List[str]) -> int:
    """
    Score por zonas:
      + 3x header + 3x footer + 1x full
      + bônus p/ keywords com dígitos em header/footer
      - 2x nomes de layouts rivais no header/footer
      - PENALIZAÇÃO forte se anti_keywords do layout aparecerem
    """
    t_full   = (full_text   or "").casefold()
    t_header = (header_text or "").casefold()
    t_footer = (footer_text or "").casefold()

    base_full    = _count_hits(t_full, lay.keywords)
    base_header  = _count_hits(t_header, lay.keywords)
    base_footer  = _count_hits(t_footer, lay.keywords)

    anti_header_names = sum(1 for rn in rival_names if rn and rn.casefold() in t_header and rn.casefold() != lay.name.casefold())
    anti_footer_names = sum(1 for rn in rival_names if rn and rn.casefold() in t_footer and rn.casefold() != lay.name.casefold())

    ak = getattr(lay, "anti_keywords", []) or []
    anti_full   = _count_hits(t_full,   ak)
    anti_header = _count_hits(t_header, ak)
    anti_footer = _count_hits(t_footer, ak)

    kw_digits_header = sum(1 for kw in lay.keywords if any(ch.isdigit() for ch in kw) and kw.casefold() in t_header)
    kw_digits_footer = sum(1 for kw in lay.keywords if any(ch.isdigit() for ch in kw) and kw.casefold() in t_footer)

    penalty = (4 * anti_header) + (4 * anti_footer) + (1 * anti_full) + (2 * (anti_header_names + anti_footer_names))

    return (3 * base_header) + (3 * base_footer) + base_full + kw_digits_header + kw_digits_footer - penalty

def _probe_layout_on_page(png_path: str, lay: "LayoutSpec") -> int:
    """
    Tie-breaker: OCR 'sondas' em recortes âncora do layout candidato (MUNICIPIO_TOMADOR, DATA, NUMERO_NFS).
    """
    probe_texts: List[str] = []
    if "MUNICIPIO_TOMADOR" in lay.crops:
        try:
            probe_texts.append(_ocr_from_relbox(png_path, lay.crops["MUNICIPIO_TOMADOR"], _cfg_for(lay, "MUNICIPIO_TOMADOR")))
        except Exception:
            pass
    for k in ("DATA", "NUMERO_NFS"):
        if k in lay.crops:
            try:
                probe_texts.append(_ocr_from_relbox(png_path, lay.crops[k], _cfg_for(lay, k)))
            except Exception:
                pass

    joined = " \n ".join(t for t in probe_texts if t)
    t_norm = (joined or "").casefold()
    hits_kw  = _count_hits(t_norm, lay.keywords)
    hits_num = sum(1 for kw in lay.keywords if any(ch.isdigit() for ch in kw) and kw.casefold() in t_norm)
    return (2 * hits_kw) + (2 * hits_num)

def _ocr_text(img: Image.Image) -> str:
    """OCR direto (a imagem já vem tratada)."""
    try:
        txt = pytesseract.image_to_string(img, lang=OCR_LANG, config="--psm 6")
    except Exception:
        txt = ""
    return (txt or "").strip()

def convert_pdf_to_pngs(pdf_path: str, out_dir: str, dpi: int = DPI) -> List[str]:
    os.makedirs(out_dir, exist_ok=True)
    with pdfplumber.open(pdf_path) as pdf:
        pngs = []
        base = os.path.splitext(os.path.basename(pdf_path))[0]
        for i, page in enumerate(pdf.pages, start=1):
            if _cancelled():
                raise KeyboardInterrupt("cancelado pelo usuário")
            img = to_image(page, dpi=dpi)
            out = os.path.join(out_dir, f"{base}_p{i:02d}.png")
            img.save(out, format="PNG", optimize=True)
            pngs.append(out)
        return pngs

# ====== LAYOUT: detecção com fallback PORANGATU + report dos não reconhecidos ======

def infer_layout_from_text(text: str) -> LayoutSpec:
    """Mantida por compatibilidade (retorna o melhor score, sem flag)."""
    t = (text or "").casefold()
    best = LAYOUTS[0]; score_best = -1
    for lay in LAYOUTS:
        score = sum(1 for kw in lay.keywords if kw in t)
        if score > score_best:
            best, score_best = lay, score
    return best


def detect_layout_with_default(text: str) -> Tuple[LayoutSpec, bool]:
    """Retorna (layout, matched); se nenhum keyword bater, usa PORANGATU e matched=False."""
    t = (text or "").casefold()
    scores: List[Tuple[int, LayoutSpec]] = []
    for lay in LAYOUTS:
        score = sum(1 for kw in lay.keywords if kw in t)
        scores.append((score, lay))
    score_best, best = max(scores, key=lambda x: x[0])
    if score_best > 0:
        return best, True
    # fallback para PORANGATU
    porangatu = next((l for l in LAYOUTS if l.name.upper() == "PORANGATU"), LAYOUTS[0])
    return porangatu, False


def _cfg_for(layout: LayoutSpec, topic: str) -> ConfigLimpezaScan:
    """Retorna o filtro específico do tópico ou o padrão do layout."""
    return layout.filters_by_topic.get(topic, layout.filter_default)


def save_crops_and_overlay(png_path: str, layout: LayoutSpec, out_dir: str) -> Tuple[Dict[str, str], Optional[str]]:
    """
    Gera crops (um PNG TRATADO por TÓPICO usando o **filtro do TÓPICO** do **Layout**) + OVERLAY.
    Retorna: (dict campo->path_crop_tratado, path_overlay)
    """
    os.makedirs(out_dir, exist_ok=True)
    base = os.path.splitext(os.path.basename(png_path))[0]
    with Image.open(png_path) as im:
        w, h = im.size
        draw_im = im.copy()
        draw = ImageDraw.Draw(draw_im)
        font = ImageFont.load_default()

        crop_paths: Dict[str, str] = {}
        for field, relbox in layout.crops.items():
            x0, y0, x1, y1 = _rel2abs(w, h, _as_box(relbox), pad=CROP_PAD)
            raw = im.crop((x0, y0, x1, y1))

            topic_cfg = _cfg_for(layout, field)
            _zoom = getattr(topic_cfg, "zoom", 1.0)
            if _zoom != 1.0:
                _Resampling = getattr(Image, "Resampling", Image)
                _resample = getattr(_Resampling, "LANCZOS", getattr(Image, "BICUBIC", 3))
                new_w = max(1, int(raw.width  * _zoom))
                new_h = max(1, int(raw.height * _zoom))
                raw = raw.resize((new_w, new_h), _resample)

            treated = limpar_imagem_escaner(raw, topic_cfg)

            out_path = os.path.join(out_dir, f"{base}_{layout.name}_{field}.png")
            treated.save(out_path, format="PNG", optimize=True)
            crop_paths[field] = out_path

            # desenha overlay
            draw.rectangle((x0, y0, x1, y1), outline="red", width=3)
            label = f"{field}"
            bbox = draw.textbbox((0, 0), label, font=font)
            tw, th = bbox[2]-bbox[0], bbox[3]-bbox[1]
            tx, ty = x0 + 6, max(0, y0 - th - 2)
            draw.rectangle((tx-3, ty-1, tx+tw+3, ty+th+1), fill="white")
            draw.text((tx, ty), label, fill="black", font=font)

        overlay_path = os.path.join(out_dir, f"{base}_{layout.name}_OVERLAY.png")
        draw_im.save(overlay_path, format="PNG", optimize=True)

    return crop_paths, overlay_path

def p_info(msg):  print(f"ℹ️ {msg}", flush=True)
def p_ok(msg):    print(f"✅ {msg}", flush=True)
def p_warn(msg):  print(f"⚠️ {msg}", flush=True)
def p_err(msg):   print(f"❌ {msg}", flush=True)
def p_title(msg): print(f"📌 {msg}", flush=True)

def _line(char="─", n=72): 
    return char * n

def _box_title(t: str) -> str:
    return f"\n{_line('═',72)}\n📄 {t}\n{_line('─',72)}"

# ================== PARSERS POR TÓPICO (ESTRITOS) ==================


def parse_numero_nf_from(topic_text: str) -> str:
    if not topic_text: return ""
    t = topic_text.replace("\n", " ")
    m = re.search(r"(?i)\b(n[uú]mero|n[oº°]\.?)[\s]*[:#\-]?[\s]*([A-Z0-9\.\-\/]{4,25})", t)
    cand = ""
    if m:
        cand = m.group(2).strip().strip(".-:/#")
    if not cand:
        tokens = re.findall(r"[A-Z0-9\-\/\.]{5,25}", t.upper())
        for tok in tokens:
            dig = only_digits(tok)
            if len(dig) == 14:   # CNPJ
                continue
            if RE_DATEPT.search(tok):  # data
                continue
            if len(dig) >= 6:
                cand = tok.strip(".-:/#")
                break
    return cand


def parse_cidade_from(topic_text: str) -> str:
    if not topic_text:

        return ""
    lines = [ln.strip() for ln in topic_text.splitlines() if ln.strip()]
    # ignora linhas de cadastro/identidade que poluem a detecção
    bad = re.compile(
        r"(?i)\b("
        r"endere[cç]o|logradouro|cpf|cnpj|cep|bairro|complemento|inscri[cç][aã]o|"
        r"tomador|prestador|raz[aã]o\s*social|nome\s*\/?\s*raz[aã]o|e-?mail|empresa|fantasia"
        r")\b"
    )

    candidates = [ln for ln in lines if not bad.search(ln)]
    # tenta a ÚLTIMA ocorrência '... Cidade - UF' na linha
    for ln in candidates:
        pass  # placeholder para manter a ordem de leitura mental
    for ln in candidates:
        matches = list(re.finditer(r"(.+?)[\s\-\/,]+([A-Z]{2})\b", ln))
        matches = [m for m in matches if m.group(2) in UF_SIGLAS]
        if matches:

            m = matches[-1]  # última ocorrência com UF
            left, uf = m.group(1), m.group(2)
            # pega o último segmento "tipo cidade" antes do "- UF"
            segs = re.split(r"[,\(\)\-\/|]", left)
            segs = [re.sub(r"[^A-Za-zÀ-ÿ\s]", " ", s).strip() for s in segs if s.strip()]
            if segs:
                cand_city = re.sub(r"\s+", " ", segs[-1]).strip()
                return f"{cand_city} - {uf}"
            
    # fallback: linha mais alfabética "limpa"
    if candidates:
        candidates.sort(key=lambda s: sum(1 for ch in s if ch.isalpha()), reverse=True)
        city = re.sub(r"[^A-Za-zÀ-ÿ\s]", " ", candidates[0])
        city = re.sub(r"\s+", " ", city).strip()
        return city
    return ""

def parse_valor_from(topic_text: str) -> str:
    if not topic_text: return ""
    text = topic_text.replace("\n", " ")
    values = [m.group(1) for m in RE_MONEY_BR.finditer(text)]
    if not values:
        return ""
    def to_num(v: str) -> float:
        return float(v.replace(".", "").replace(",", "."))
    values.sort(key=lambda v: to_num(v), reverse=True)
    return values[0]

# --- AGORA (com validação de CNPJ):
def parse_prestador_from(topic_text: str) -> Tuple[str, str]:
    if not topic_text:
        return "", ""
    cnpj = ""
    m_cnpj = RE_CNPJ.search(topic_text)
    if m_cnpj:
        raw = m_cnpj.group(0)
        cnpj_digits = only_digits(raw)
        if cnpj_is_valid(cnpj_digits):
            cnpj = mask_cnpj(cnpj_digits)
        else:
            cnpj = "INVALIDO"
    else:
        m_cpf = RE_CPF.search(topic_text)
        if m_cpf:
            cnpj = "INVALIDO"
    lines = [ln.strip() for ln in topic_text.splitlines() if ln.strip()]
    no_cnpj = [ln for ln in lines if not RE_CNPJ.search(ln)]
    no_cnpj.sort(key=lambda s: sum(1 for ch in s if ch.isalpha()), reverse=True)
    nome = no_cnpj[0] if no_cnpj else ""
    nome = re.sub(r"(?i)\b(raz[aã]o\s+social|nome\s*\/??\s*raz[aã]o\s*social|prestador(?: de servi[cç]os)?|empresa)\s*[:\-]?\s*", "", nome).strip()
    return cnpj, nome

def parse_data_from(topic_text: str) -> str:
    return normalize_date_ddmmyyyy(topic_text or "")

# ================== ESTRUTURA DA NOTA ==================

@dataclass
class NFSeData:
    numero_nf: str = ""
    municipio: str = ""
    valor_total: str = ""
    cnpj_prestador: str = ""
    nome_prestador: str = ""
    data_emissao: str = ""

    def as_cli_line(self, pdf_file: str, layout_name: str) -> str:
        return (
            f"[OK] {pdf_file} | LAYOUT={layout_name} | Nº={self.numero_nf or '-'} | "
            f"Data={self.data_emissao or '-'} | Cidade={self.municipio or '-'} | "
            f"Prestador={self.nome_prestador or '-'} ({self.cnpj_prestador or '-'}) | "
            f"Valor={'R$ ' + self.valor_total if self.valor_total else '-'}"
        )

# ================== INTEGRAÇÃO COM OPENAI (RESUMO POR IA) ==================

OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")

# --- INÍCIO: Config OpenAI robusta (SDK novo e legado) ---
_api_key = (os.getenv("OPENAI_API_KEY") or "").strip()

_openai_client = None     # SDK novo (>=1.0)
_openai_legacy = None     # SDK legado (<1.0)
_OPENAI_ENV_ERR = None

# Tenta SDK novo
try:
    try:
        from openai import OpenAI  # >= 1.0
        _openai_client = OpenAI(api_key=_api_key) if _api_key else OpenAI()
    except ImportError:
        _openai_client = None
        _OPENAI_ENV_ERR = "OpenAI SDK >=1.0 not found or not installed properly."
except Exception as e:
    _openai_client = None
    _OPENAI_ENV_ERR = e

# Se falhar, tenta SDK legado
if _openai_client is None:
    try:
        import openai as _openai_legacy  # < 1.0
        if _api_key:
            _openai_legacy.api_key = _api_key
    except Exception as e:
        _openai_legacy = None
        _OPENAI_ENV_ERR = e
# --- FIM: Config OpenAI robusta ---


def _prompt_resumo_ia(lote: str, pasta: str, pdf_file: str, layout_name: str, topicos: Dict[str, List[Tuple[int, str]]]) -> str:
    """Monta um prompt **robusto** com regras e mapeamentos, e o bloco bruto igual ao TXT."""
    # Bloco bruto (mesmo formato do TXT)
    s = io.StringIO()
    s.write(f"== LOTE: {lote} ==\n")
    s.write(f"Pasta: {pasta}\n")
    s.write("TOPICOS (OCR completo por recorte/página):\n")
    for topic in topicos.keys():
        if topic == "RECEBIMENTO":
            # NÃO incluir o texto bruto do Recebimento para não contaminar outros campos.
            # A data preferida virá via [DATA_PREFERIDA].
            continue
        s.write(f"\n[{topic}]\n")
        for page_idx, txt in topicos.get(topic, []):
            s.write(f"  - Página {page_idx}:\n")
            if (txt or "").strip():
                for ln in txt.splitlines():
                    s.write(f"    {ln}\n")
            else:
                s.write("    (vazio)\n")

    bloco_txt = s.getvalue()

    regras = rf"""
Você é um extrator de campos de **NFS-e**. Receberá um BLOCO DE TEXTO BRUTO de OCR
(igual ao que vai para o TXT), já dividido por TÓPICOS. Sua tarefa é identificar e
organizar os campos a seguir e devolver **apenas** este **formato final**:

RESUMO FINAL DA NOTA (valores extraídos do seu TÓPICO):
  Número NFS-e........: <numero>
  Data de Emissão.....: <dd/mm/aaaa>
  Fazenda.............: <nome da fazenda mapeada>
  Prestador...........: <nome do prestador>
  CNPJ Prestador......: <cnpj no formato 00.000.000/0000-00 ou 'INVALIDO'>
  Valor Total/Líquido.: <valor no padrão 0.000,00>

**REGRAS IMPORTANTES**
- Trabalhe **somente** com o texto do BLOCO (não invente dados). Corrija OCR óbvio
  (ex.: S/$/& confundindo com 5; letras trocadas por acentos; espaços indevidos em números).
- **[NUMERO_NFS]**: quando existir "Chave de Acesso"/número muito longo e **também** aparecer
  uma linha explícita "Número da NFS-e" com um número pequeno, **o número da nota é o pequeno**.
  Ex.: se aparecer "ONG0005" e "Número da NFS-e 5", retorne **5** ou algi como "00000013", retorne **13**,
  ou se for algo como "00258" entao retorne **258**. Se só houver coisas que parecem
  datas/séries (ex.: 202500/2025000), retorne **-**.
- **[DATA – PRIORIDADE]**: se houver o tópico **[RECEBIMENTO]** (ou texto indicando Recebimento de Mercadoria),
  a **Data de Emissão** do resumo deve ser a **Dt. Vencimento** extraída do RECEBIMENTO. Aceite datas como
  "07/04/20 25" e normalize para **07/04/2025**. **Somente** se NÃO existir Recebimento, use a data da nota.
  Se houver também um tópico **[DATA_PREFERIDA]**, ele confirma a data escolhida.
- **[MUNICIPIO_TOMADOR]**: determine **apenas o nome da FAZENDA mapeada**, nunca retorne cidade/UF.
  Use tolerância a erros de OCR e **similaridade**.
  Regras de mapeamento (tolerantes a variações, acentos e caixa):
  - 1) Se encontrar um código numérico que exista no FARM_MAPPING, use esse mapeamento:
      {FARM_MAPPING}
  - 2) Caso não haja código, mapeie pela cidade/município aproximado usando:
      {CODIGOS_CIDADES}
  - Em ambos os casos, **retorne somente o valor mapeado** (ex.: "Nova Gloria":"Aliança" então retorne Aliança).
  - Se não houver confiança suficiente no mapeamento, retorne "-".
  - Use **EXCLUSIVAMENTE** o texto do tópico **[MUNICIPIO_TOMADOR]** para identificar município/fazenda; 
    **ignore** ocorrências de cidade/UF em quaisquer outros tópicos.
  - Se não houver mapeamento seguro pela cidade ou código numérico do FARM_MAPPING, retorne **"-"**.

- **[PRESTADOR CNPJ]**: se houver um **CPF** no lugar do CNPJ (ex.: "018.732.331/32" ou "018.732.331-32"),
  o campo **CNPJ Prestador** deve ser **'INVALIDO'**. Se houver um CNPJ válido, normalize para
  **00.000.000/0000-00**. Para o **NOME**, pegue a melhor linha que represente o nome empresarial, evitando
  rótulos como "Prestador do Serviço", "Razão Social", etc.
- **[VALOR_NFS]**: avalie todos os valores monetários; **o valor da nota é o que faz
  sentido** (geralmente o maior e/ou aquele sob "Valor Líquido da NFS-e"). Normalize separadores e **retorne só um**.

  - **[VALOR_NFS — ANTI-CONFUSÃO (ALÍQUOTA x VALOR)]**:
  - Trate como **percentual (não dinheiro)** qualquer número com vírgula e **3–4 casas decimais** entre **0 e 100** (ex.: `5,0000`, `0,0500`, `7,5000`) quando:
    - estiver rotulado ou muito próximo (até 2 linhas) de **ALIQ**, **ALÍQ**, **ALÍQUOTA**, **%**, **PERC**, **ISS**, **ISSQN**, **PIS**, **COFINS**, **INSS**, **CSLL**; ou
    - aparecer na mesma coluna/bloco que itens de tributos.
  - Considere **dinheiro** quando:
    - houver rótulos como **VALOR LÍQUIDO**, **VALOR TOTAL**, **VALOR DA NFS-e**, **VALOR DO SERVIÇO**, **VALOR DO DOCUMENTO**, **VALOR A PAGAR** (priorize o **valor monetário mais próximo** ao rótulo); ou
    - o padrão for típico de moeda brasileira: `(\d{1,3}(\.\d{3})*,\d{2})` com ou sem `R$`.
  - **Nunca** use valores associados a **BASE DE CÁLCULO**, **DEDUÇÃO(ÕES)**, **DESCONTO(S)**, **RETENÇÃO(ÕES)**, **ALÍQUOTA**, **%**, **IMPOSTOS** para preencher o **Valor Total/Líquido**.
  - NUNCA, JAMAIS RETORNE "5,0000" OU SEMELHANTE, ANALISE OUTRO SE ACHAR ESSE E SE NAO ACHAR, RETORNE "-".
  - **Exemplo prático**:  
    ```
    ALIQ
    5,0000
    VALOR LÍQUIDO
    41.250,00
    ou
    3.306,00

    ```
    → **Valor Total/Líquido = 41.250,00** (e **não** `5,0000`).

- **[NUMERO_NFS — DESAMBIGUAÇÃO E NORMALIZAÇÃO]**:
  - Ignore **tokens com dígitos separados por espaço** (ex.: `20 50 0`) — **não** são número de NFS-e.
  - Se houver **zeros à esquerda**, remova-os (ex.: `0000013` → **13**).
  - Para **cadeias longas misturando códigos** onde o **número da nota aparece ao final** (ex.: `090900013`), **se houver rótulo claro** na mesma linha ou na linha imediatamente acima/abaixo (**"Número da NFS-e"**, **"Nº"**, **"No"**), **extraia o grupo final de até 5 dígitos**, remova zeros à esquerda e retorne (ex.: `090900013` → **13**).
  - Priorize o número **explicitamente rotulado** como **"Número da NFS-e" / "Nº / No"**. Se coexistir com **Chave de Acesso / Código de Verificação**, **o número da nota é o pequeno** (já normalizado sem zeros à esquerda).
  - Rejeite falsos positivos: **CNPJ** (14 dígitos), **CEP** (8 dígitos contíguos), **datas** (dd/mm/aaaa), **chave de acesso**, **cód. verificação**. Sem candidato confiável, retorne **"-"**.
  - NUNCA, JAMAIS RETORNE "202500", ANALISE OUTRO SE ACHAR ESSE

- **[RECEBIMENTO — ISOLAMENTO]**:
  - O tópico **[RECEBIMENTO]** serve **exclusivamente** para definir a **Data de Emissão** (usando a **Dt. Vencimento**). **Não** utilize nada de RECEBIMENTO para **Número**, **Prestador**, **CNPJ** ou **Valor**.

- **[NORMALIZAÇÕES]**:
  - **Percentuais**: podem ter 1–4 casas decimais (ex.: `5,0`, `5,00`, `5,0000`) e **não** devem preencher **Valor Total/Líquido**.
  
- Se algum campo não puder ser determinado com segurança, use "-".

**Contexto**: LAYOUT DETECTADO = {layout_name}; ARQUIVO = {pdf_file}; LOTE = {lote}.
**Importante**: Responda **somente** com o bloco final acima, sem explicações extras.
"""

    return regras + "\n\n" + bloco_txt

def ia_resumir_nota(lote: str, pasta: str, pdf_file: str, layout_name: str, topicos: Dict[str, List[Tuple[int, str]]]) -> str:
    """Chama a OpenAI (SDK novo OU legado) para gerar o resumo final a partir do texto bruto dos tópicos."""
    prompt = _prompt_resumo_ia(lote, pasta, pdf_file, layout_name, topicos)

    # Se nenhum cliente disponível:
    if _openai_client is None and _openai_legacy is None:
        return (
            "RESUMO FINAL DA NOTA (gerado pela IA):\n"
            f"  [ERRO] OpenAI não configurado. Verifique a instalação da lib 'openai' e a chave. Detalhe: {_OPENAI_ENV_ERR}\n"
        )

    try:
        # SDK NOVO (>=1.0)
        if _openai_client is not None:
            resp = _openai_client.chat.completions.create(
                model=OPENAI_MODEL,
                temperature=0,
                messages=[
                    {"role": "system", "content": "Você é um extrator de campos de NFS-e extremamente preciso."},
                    {"role": "user", "content": prompt},
                ],
            )
            out = (resp.choices[0].message.content or "").strip()

        # SDK LEGADO (<1.0)
        else:
            # fallback de modelo para legados que não suportam 'gpt-4o-mini'
            _model = OPENAI_MODEL or "gpt-3.5-turbo"
            if "4o" in _model and "gpt-3.5-turbo" not in _model:
                _model = "gpt-3.5-turbo"
            resp = _openai_legacy.ChatCompletion.create(
                model=_model,
                temperature=0,
                messages=[
                    {"role": "system", "content": "Você é um extrator de campos de NFS-e extremamente preciso."},
                    {"role": "user", "content": prompt},
                ],
            )
            out = (resp["choices"][0]["message"]["content"] or "").strip()

        if not out:
            out = "[ERRO] Resposta vazia da IA"

        # Garante prefixo esperado
        if not out.lstrip().startswith("RESUMO FINAL DA NOTA"):
            out = "RESUMO FINAL DA NOTA (gerado pela IA):\n" + out

        return out

    except Exception as e:
        return (
            "RESUMO FINAL DA NOTA (gerado pela IA):\n"
            f"  [ERRO] Falha na chamada OpenAI: {e}\n"
        )

# ================== LOTES ==================


def iter_lotes(base_dir: str) -> List[Tuple[str, str, List[str]]]:
    lots: List[Tuple[str, str, List[str]]] = []
    subdirs = [d for d in glob.glob(os.path.join(base_dir, "*")) if os.path.isdir(d)]
    for d in sorted(subdirs):
        pdfs = sorted([f for f in os.listdir(d) if f.lower().endswith(".pdf")])
        if pdfs:
            lots.append((os.path.basename(d), d, pdfs))
    base_pdfs = sorted([f for f in os.listdir(base_dir) if f.lower().endswith(".pdf")])
    if base_pdfs:
        lots.append((os.path.basename(base_dir), base_dir, base_pdfs))
    return lots

# ================== PIPELINE PRINCIPAL ==================
def main():
    lots = iter_lotes(BASE_DIR)
    if _cancelled():
        print("⛔ Cancelado antes de iniciar o processamento.", flush=True)
        return
    if not lots:
        print(f"Nenhum PDF encontrado em '{BASE_DIR}' ou subpastas.", flush=True)
        sys.exit(0)

    for lot_name, lot_dir, pdfs in lots:
        if _cancelled():
            print(f"⛔ Cancelado pelo usuário — encerrando no início do lote '{lot_name}'.", flush=True)
            return
        rows_lanc: List[Dict[str, str]] = []
        unknown_cities = set()
        log_path = _unique_path(str(Path(lot_dir) / f"{lot_name}_leitura_detalhada_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"))
        with open(log_path, "w", encoding="utf-8") as log_f:
            log_f.write(f"== LOTE: {lot_name} ==\n")
            log_f.write(f"Pasta: {lot_dir}\n")
            log_f.write(f"Arquivos PDF: {len(pdfs)}\n\n")
        

            for pdf_file in pdfs:
                _t_ini_nota = time.perf_counter()
                if _cancelled():
                    print("⛔ Cancelado — interrompendo leitura das notas deste lote.", flush=True)
                    return
                pdf_path = os.path.join(lot_dir, pdf_file)
                # LOG no console: qual arquivo está sendo processado
                print(_box_title(f"Lendo nota • Lote: {lot_name} • Arquivo: {pdf_file}"), flush=True)
                p_info(f"Caminho: {pdf_path}")

                # 1) PDF → PNGs @ DPI configurado
                try:
                    png_dir = os.path.join(lot_dir, "_png")
                    pngs = convert_pdf_to_pngs(pdf_path, png_dir, dpi=DPI)
                except Exception as e:
                    log_f.write(f"[ERRO] Falha ao converter '{pdf_file}': {e}\n")
                    continue

                if not pngs:
                    log_f.write(f"[AVISO] '{pdf_file}' sem páginas.\n")
                    continue

                # 2) OCR rápido em TODAS as páginas (texto completo + cabeçalho + rodapé)
                #    e escolha de layout/página pelo score robusto
                quick_pages: List[Tuple[int, str, str, str]] = []  # (idx, full_text, header_text, footer_text)
                for i, png in enumerate(pngs, start=1):
                    if _cancelled():
                        raise KeyboardInterrupt("cancelado pelo usuário")
                    try:
                        if _cancelled():
                            raise KeyboardInterrupt("cancelado pelo usuário")
                        full = pytesseract.image_to_string(
                            Image.open(png), lang=OCR_LANG, config="--psm 6", timeout=3
                        )
                    except Exception:
                        full = ""
                    header = _page_header_text(png)
                    footer = _page_footer_text(png)
                    quick_pages.append((i, full or "", header or "", footer or ""))

                rival_names = [lay.name for lay in LAYOUTS]
                layout = LAYOUTS[0]
                best_score = -10**9
                nota_page_idx = 1
                raw_scores: List[Tuple[int, "LayoutSpec", int]] = []  # (page_idx, layout, score)

                for page_idx, full_text, header_text, footer_text in quick_pages:
                    for lay in LAYOUTS:
                        s = _score_layout_zones(full_text, header_text, footer_text, lay, rival_names)
                        raw_scores.append((page_idx, lay, s))
                        if s > best_score:
                            layout, best_score, nota_page_idx = lay, s, page_idx

                # Tie-breaker na página escolhida, se disputa apertada (≤ 1 ponto)
                page_best = [t for t in raw_scores if t[0] == nota_page_idx]
                if page_best:
                    page_best.sort(key=lambda t: t[2], reverse=True)
                    top_score = page_best[0][2]
                    contenders = [t for t in page_best if t[2] >= top_score - 1]
                    if len(contenders) > 1:
                        probes: List[Tuple["LayoutSpec", int]] = []
                        for _, lay_cand, _ in contenders[:4]:
                            pscore = _probe_layout_on_page(pngs[nota_page_idx - 1], lay_cand)
                            probes.append((lay_cand, pscore))
                        probes.sort(key=lambda x: x[1], reverse=True)
                        if probes and probes[0][0] is not layout and probes[0][1] > 0:
                            layout = probes[0][0]

                # 3) Processar APENAS a página da NOTA: crops TRATADOS + overlay + OCR
                crop_root = os.path.join(lot_dir, "_crops", os.path.splitext(pdf_file)[0])
                os.makedirs(crop_root, exist_ok=True)

                by_topic_texts: Dict[str, List[Tuple[int, str]]] = {k: [] for k in layout.crops.keys()}
                overlays_rel: List[str] = []
                receb_txt: str = ""
                receb_overlays_rel: List[str] = []

                nota_png = pngs[nota_page_idx - 1]
                crop_paths, overlay_path = save_crops_and_overlay(nota_png, layout, crop_root)
                if overlay_path:
                    overlays_rel.append(os.path.relpath(overlay_path, lot_dir))
                for topic, cp in crop_paths.items():
                    try:
                        txt = _ocr_text(Image.open(cp))
                    except Exception as e:
                        txt = f"[OCR ERROR] {e}"
                    by_topic_texts[topic].append((nota_page_idx, txt))

                # Guarda OCR bruto por página para heurística de vencimento/recebimento
                quick_texts_by_page = {idx: f"{header}\n{full}\n{footer}".strip() for idx, full, header, footer in quick_pages}

                # 4) Cabeçalho do TXT detalhado
                sep = "=" * 100
                log_f.write(f"{sep}\nARQUIVO PDF: {pdf_file}\n")
                log_f.write(f"LAYOUT DETECTADO: {layout.name}\n\n")

                # 5) EXTRAÇÃO (estrita por tópico) + prints condicionais
                data = NFSeData()
                def found(topic: str, value: str):
                    if not PRINT_ONLY_IA and value:
                        print(f"[{pdf_file}] {topic}: {value}", flush=True)

                # DATA (nota)
                for _, txt in by_topic_texts.get("DATA", []):
                    v = parse_data_from(txt)
                    if v:
                        data.data_emissao = v
                        found("DATA", v)
                        break

                # MUNICÍPIO
                for _, txt in by_topic_texts.get("MUNICIPIO_TOMADOR", []):
                    v = parse_cidade_from(txt)
                    if v:
                        data.municipio = v
                        found("MUNICIPIO/UF", v)
                        break

                # === Preferência ABSOLUTA: Recebimento da PÁGINA ANTERIOR à NOTA ===
                nota_data = data.data_emissao
                venc_receb = ""

                if len(pngs) > 1:
                    prev_idx = nota_page_idx - 1
                    if prev_idx >= 1:
                        try:
                            # processa SOMENTE a página imediatamente anterior como "Recebimento"
                            rec_png = pngs[prev_idx - 1]
                            receb_layout = _make_recebimento_layout()
                            rec_crop_paths, rec_overlay_path = save_crops_and_overlay(rec_png, receb_layout, crop_root)
                            if rec_overlay_path:
                                receb_overlays_rel.append(os.path.relpath(rec_overlay_path, lot_dir))

                            try:
                                receb_txt = _ocr_text(Image.open(rec_crop_paths["DATA"]))
                            except Exception as e:
                                receb_txt = f"[OCR ERROR] {e}"

                            # guarda o texto lido do recebimento (para o TXT e para a IA)
                            by_topic_texts.setdefault("RECEBIMENTO", []).append((prev_idx, receb_txt or ""))

                            # extrai Dt. Vencimento (aceita "07/04/20 25" etc.)
                            venc_receb = parse_vencimento_from_recebimento(receb_txt) or ""
                        except Exception:
                            pass

                # define data final: Recebimento (página anterior) > data da nota
                data.data_emissao = venc_receb or nota_data
                if data.data_emissao and data.data_emissao != nota_data:
                    found("DATA (vencimento preferida)", data.data_emissao)

                # disponibiliza a data escolhida para a IA
                by_topic_texts.setdefault("DATA_PREFERIDA", []).append((nota_page_idx, data.data_emissao or "-"))                

                # NÚMERO
                for _, txt in by_topic_texts.get("NUMERO_NFS", []):
                    v = parse_numero_nf_from(txt)
                    if v:
                        data.numero_nf = v
                        found("NUMERO_NFS", v)
                        break

                # PRESTADOR
                for _, txt in by_topic_texts.get("PRESTADOR_CNPJ", []):
                    cnpj, nome = parse_prestador_from(txt)
                    ok = False
                    if cnpj:
                        data.cnpj_prestador = cnpj; found("CNPJ_PRESTADOR", cnpj); ok = True
                    if nome:
                        data.nome_prestador = nome; found("NOME_PRESTADOR", nome); ok = True
                    if ok: break

                # VALOR
                for _, txt in by_topic_texts.get("VALOR_NFS", []):
                    v = parse_valor_from(txt)
                    if v:
                        data.valor_total = v
                        found("VALOR_TOTAL", v)
                        break

                # 6) TÓPICOS (OCR completo por recorte/página) — inclui RECEBIMENTO se houver
                log_f.write("TOPICOS (OCR completo por recorte/página):\n")
                ordered_topics = list(layout.crops.keys()) + [k for k in by_topic_texts.keys() if k not in layout.crops]
                for topic in ordered_topics:
                    log_f.write(f"\n[{topic}]\n")
                    items = by_topic_texts.get(topic, [])
                    if not items:
                        log_f.write("  (sem texto extraído)\n")
                        continue
                    for page_idx, txt in items:
                        log_f.write(f"  - Página {page_idx}:\n")
                        if (txt or "").strip():
                            for ln in (txt or "").splitlines():
                                log_f.write(f"    {ln}\n")
                        else:
                            log_f.write("    (vazio)\n")

                # OVERLAYS GERADOS (NFS + RECEBIMENTO)
                all_overlays = []
                all_overlays.extend(overlays_rel)
                all_overlays.extend(receb_overlays_rel)
                if all_overlays:
                    log_f.write("\nOVERLAYS GERADOS (NFS + RECEBIMENTO):\n")
                    for rel in all_overlays:
                        log_f.write(f"  * {rel}\n")

                # 7) RESUMO FINAL DA NOTA — GERADO PELA IA (usa o mesmo texto que foi pro TXT, MAS sem [RECEBIMENTO])
                topicos_ia = {k: v for k, v in by_topic_texts.items() if k != "RECEBIMENTO"}
                resumo_ia = ia_resumir_nota(lot_name, lot_dir, pdf_file, layout.name, topicos_ia)

                # 7.1) Guardar a resposta bruta da IA no TXT (auditoria)
                log_f.write("\n" + (resumo_ia or "").strip() + "\n")

                # 7.2) Extrai campos do resumo e prepara duração
                campos = parse_campos_from_resumo(resumo_ia)
                
                elapsed_sec = time.perf_counter() - _t_ini_nota
                elapsed_str = f"{elapsed_sec:.0f}s" if elapsed_sec >= 1 else f"{elapsed_sec:.1f}s"
                
                # ——— Bloco compacto por nota (título forte + linhas com emojis, organizado)
                print(f"📄 {pdf_file}", flush=True)
                print(f"   📁 Lote: {lot_name}", flush=True)
                print(f"   📍 Caminho: {pdf_path}", flush=True)
                print(f"   🧩 Layout: {layout.name} | 🧱 Páginas: {len(pngs)} | 🎚️ DPI: {DPI}", flush=True)
                print(f"   🔎 Nº: {campos.get('numero') or '-'} | 🗓️ Emissão: {campos.get('data') or '-'}", flush=True)
                print(f"   🏷️ Prestador: {campos.get('prestador') or '-'}", flush=True)
                print(f"   🧾 CNPJ: {campos.get('cnpj') or '-'} | 💰 Valor: {campos.get('valor') or '-'}", flush=True)
                print(f"   🕒 Duração: {elapsed_str}", flush=True)
                
                # 7.3) RESUMO FINAL DA NOTA (em formato “kv” para aparecer em negrito na UI)
                resumo_fmt = _format_resumo_campos(campos)
                print(resumo_fmt, flush=True)
                log_f.write("\n" + resumo_fmt + "\n")


                # 7.3) Se a IA não extraiu nada útil, também mostre o bruto no console para diagnóstico
                def _nz(v): 
                    v = (v or "").strip()
                    return v if v else "-"

                if all((_nz(campos.get(k)) == "-") for k in ("numero","data","fazenda","prestador","cnpj","valor")):
                    print("⚠️ A IA não conseguiu extrair campos desta nota. Resposta bruta:", flush=True)
                    print((resumo_ia or "").strip(), flush=True)

                # ===== CAMPOS PARA A PLANILHA (SEMPRE adiciona, mesmo se faltarem dados) =====
                fazenda_ia = (campos.get("fazenda","") or "").strip()
                cod_faz = codigo_fazenda_from_nome(fazenda_ia) or "-"
                if cod_faz == "-":
                    log_f.write(f"  [AVISO] Fazenda não mapeada no RESUMO: '{fazenda_ia}' -> usando '-'.\n")

                # CNPJ: 'INVALIDO' se não passar na validação; se válido, só dígitos
                cnpj_clean_digits = only_digits(campos.get("cnpj", ""))
                cnpj_out = cnpj_clean_digits if (cnpj_clean_digits and cnpj_is_valid(cnpj_clean_digits)) else "INVALIDO"

                # Valor: tenta IA; se vazio, tenta maior valor do tópico VALOR_NFS; se ainda vazio, deixa em branco
                txt_valor_blocos = " ".join(t for _, t in by_topic_texts.get("VALOR_NFS", []))
                valor_cent = money_br_to_centavos(campos.get("valor", ""))
                if not valor_cent:
                    valor_topico = parse_valor_from(txt_valor_blocos)
                    if valor_topico:
                        valor_cent = money_br_to_centavos(valor_topico)

                # Data: IA > data escolhida no pipeline > vazio (formato dd-mm-aaaa)
                data_escolhida = (campos.get("data","") or data.data_emissao or "")
                data_dash = date_slash_to_dash(data_escolhida)

                # Número: IA > extração estrita > vazio
                numero_nf_digits = re.sub(r"[^\d]+", "", (campos.get("numero","") or data.numero_nf or ""))

                # Descrição/Historico
                desc = f"PAGAMENTO NF {numero_nf_digits or (campos.get('numero','') or '-')} {(campos.get('prestador','') or data.nome_prestador).strip()}".strip()

                # Linha SEMPRE vai para a planilha (mesmo que falte algo)
                rows_lanc.append({
                    # A:Data | B:CodFazenda | C:Conta | D:NumeroNF | E:Historico | F:CNPJ | G:Tipo | H:Padrao | I:Valor1 | J:Valor2 | K:Flag | L:CaminhoNF
                    "data": data_dash or "",
                    "codfaz": cod_faz or "-",
                    "conta": "001",
                    "numero": numero_nf_digits or (campos.get("numero","") or ""),
                    "historico": desc,
                    "cnpj": (only_digits(campos.get("cnpj","")) if cnpj_out != "INVALIDO" else "INVALIDO"),
                    "tipo": "2",
                    "padrao": "000",
                    "valor1": valor_cent or "",
                    "valor2": valor_cent or "",
                    "flag": "N",
                    "caminho": pdf_path,   # NOVO: caminho COMPLETO do PDF
                })

                # 8) Saída no terminal (o resumo formatado já foi exibido acima)
                print("·" * 72, flush=True)  # separador visual entre notas
                print("", flush=True)

            p_ok(f"Lote '{lot_name}' finalizado. TXT detalhado: {log_path}")

            # ===== NOVO: gera planilha 'lancamentos.xlsx' (SEM MACRO) =====
            if rows_lanc:
                try:
                    xlsx_path = _unique_path(os.path.join(lot_dir, "lancamentos.xlsx"))
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "lancamentos"

                    # NOVO LAYOUT (espelho do TXT + coluna final CaminhoNF):
                    # A:Data | B:CodFazenda | C:Conta | D:NumeroNF | E:Historico | F:CNPJ | G:Tipo | H:Padrao | I:Valor1 | J:Valor2 | K:Flag | L:CaminhoNF
                    headers = ["Data","CodFazenda","Conta","NumeroNF","Historico","CNPJ","Tipo","Padrao","Valor1","Valor2","Flag","CaminhoNF"]
                    ws.append(headers)

                    for r in rows_lanc:
                        ws.append([
                            r["data"],
                            r["codfaz"],
                            r["conta"],
                            r["numero"],
                            r["historico"],
                            r["cnpj"],
                            r["tipo"],
                            r["padrao"],
                            r["valor1"],
                            r["valor2"],
                            r["flag"],
                            r["caminho"],   # NOVO: última coluna (L)
                        ])

                    end_row = ws.max_row
                    end_col = ws.max_column
                    ref = f"A1:{get_column_letter(end_col)}{end_row}"
                    tbl = Table(displayName="lancamentos_tbl", ref=ref)
                    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
                    ws.add_table(tbl)

                    # larguras amigáveis (CaminhoNF largo no final)
                    widths = [12,12,8,14,60,22,6,8,12,12,6,60]
                    for i, wth in enumerate(widths, start=1):
                        ws.column_dimensions[get_column_letter(i)].width = wth

                    wb.save(xlsx_path)
                    print(f"[OK] Planilha criada (sem macro): {xlsx_path}", flush=True)

                except Exception as e:
                    print(f"[ERRO] Falha ao criar planilha: {e}", flush=True)

            if unknown_cities:
                print("[ATENÇÃO] Cidades/municípios sem mapeamento em CODIGOS_CIDADES:", flush=True)
                for c in sorted(unknown_cities):
                    print(f"  - {c}", flush=True)

# -*- coding: utf-8 -*-
"""
Automação NFS-e Digitalizadas – UI (estilo Automação Energia)
- Mesma identidade visual e layout
- Botões:
    1) 📄 Separar Nota por Nota  -> chama "Separador PDF Nota por Nota.py"
    2) 📊 Gerar planilha NFS-e   -> chama ESTE arquivo (main) ajustando BASE_DIR
    3) 📥 Importar TXT da NFS-e  -> chama "gerar_txt_lancamentos.py" e (opcional) importa no sistema
- Log estilizado + Cancelar/Limpar/Salvar
- Definir/Salvar API Key (sem hardcode)
- Configurações salvas em JSON único: <MAIN>/json/config_nfs_digitalizadas.json

Dep.: PySide6, openpyxl, pytesseract (indireto para separador)
"""

import os
import sys
import json
import importlib.util
import traceback
from pathlib import Path
from datetime import datetime
from contextlib import contextmanager

from PySide6.QtCore import (Qt, QThread, Signal, QCoreApplication, QTimer)
from PySide6.QtGui import (QIcon, QFont, QColor, QTextCursor, QPixmap, QTextOption, QCloseEvent)
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFrame, QLabel, QToolButton, QPushButton, QTextEdit,
    QFileDialog, QMessageBox, QDialog, QLineEdit, QDialogButtonBox, QFormLayout, QGroupBox,
    QSizePolicy, QTabWidget, QCheckBox
)

# ─────────────────────────────
# Ícone + CSS (MESMO do Automação Energia)
# ─────────────────────────────
ICON_PATH = Path(__file__).parent / "image" / "logo.png"

STYLE_SHEET = """
QMainWindow, QWidget { background-color: #1B1D1E; font-family: 'Segoe UI', Arial, sans-serif; color: #E0E0E0; }
QLineEdit, QDateEdit, QComboBox, QTextEdit { color: #E0E0E0; background-color: #2B2F31; border: 1px solid #1e5a9c; border-radius: 6px; padding: 6px; }
QLineEdit::placeholder { color: #5A5A5A; }
QPushButton { background-color: #1e5a9c; color: #FFFFFF; border: none; border-radius: 6px; padding: 8px 16px; font-weight: bold; }
QPushButton:hover, QPushButton:pressed { background-color: #002a54; }
QPushButton#danger { background-color: #C0392B; }
QPushButton#danger:hover { background-color: #E74C3C; }
QPushButton#success { background-color: #27AE60; }
QPushButton#success:hover { background-color: #2ECC71; }
QGroupBox { border: 1px solid #11398a; border-radius: 6px; margin-top: 10px; font-weight: bold; background-color: #0d1b3d; }
QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px; color: #ffffff; }
QTabWidget::pane { border: 1px solid #1e5a9c; border-radius: 4px; background: #212425; margin-top: 5px; }
QTabBar::tab { background: #2A2C2D; color: #E0E0E0; padding: 8px 16px; border: 1px solid #1e5a9c; border-top-left-radius: 4px; border-top-right-radius: 4px; margin-right: 2px; }
QTabBar::tab:selected { background: #1e5a9c; color: #FFFFFF; border-bottom: 2px solid #002a54; }
QStatusBar { background-color: #212425; color: #7F7F7F; border-top: 1px solid #1e5a9c; }
/* ===== Apenas tela de Configurações (objectName=tab_config) ===== */
QWidget#tab_config QGroupBox {
    background: transparent;
    border: 1px solid #11398a;
    border-radius: 6px;
    margin-top: 14px;
}
QWidget#tab_config QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    padding: 0 6px;
    background-color: #1B1D1E;
    color: #ffffff;
}
QWidget#tab_config QLabel { border: none; background: transparent; }
/* Cards */
QFrame.card { border:1px solid #1e5a9c; border-radius:12px; }
"""

# ─────────────────────────────
# Helpers
# ─────────────────────────────
def _main_root() -> Path:
    # pasta onde está este arquivo: .../Importação NFs Digitalizadas
    return Path(__file__).resolve().parent

def _json_dir() -> Path:
    p = _main_root() / "json"
    p.mkdir(parents=True, exist_ok=True)
    return p

def _logs_dir() -> Path:
    p = _main_root() / "logs"
    p.mkdir(parents=True, exist_ok=True)
    return p

def _cfg_path() -> Path:
    return _json_dir() / "config_nfs_digitalizadas.json"

def _load_cfg() -> dict:
    fp = _cfg_path()
    if fp.exists():
        try:
            return json.loads(fp.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}

def _save_cfg(cfg: dict):
    _cfg_path().write_text(json.dumps(cfg or {}, indent=4, ensure_ascii=False), encoding="utf-8")

@contextmanager
def _temp_argv(argv_list):
    old = sys.argv[:]
    try:
        sys.argv = argv_list[:]
        yield
    finally:
        sys.argv = old

def _load_module_from(path: Path, mod_name: str):
    import importlib.util, sys
    spec = importlib.util.spec_from_file_location(mod_name, str(path))
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Falha ao carregar módulo: {path}")
    mod = importlib.util.module_from_spec(spec)
    # 👇 REGISTRA antes de exec_module — isso vale p/ 'Separador...' e 'gerar_txt_lancamentos.py'
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)  # type: ignore
    return mod

def _unique_path(p):
    """Se o arquivo existir, cria ' (1)', ' (2)', ... até ficar único."""
    p = Path(p)
    if not p.exists():
        return p
    stem, suffix, parent = p.stem, p.suffix, p.parent
    i = 1
    while True:
        cand = parent / f"{stem} ({i}){suffix}"
        if not cand.exists():
            return cand
        i += 1
# ─────────────────────────────
# Dialogs
# ─────────────────────────────
class ApiKeyDialog(QDialog):
    def __init__(self, api_key_default="", parent=None):
        super().__init__(parent)
        self.setWindowTitle("🔑 Definir API da OpenAI")
        self.setModal(True)
        self.setFixedSize(560, 180)
        lay = QVBoxLayout(self)
        form = QFormLayout()
        self.key_edit = QLineEdit(api_key_default)
        self.key_edit.setPlaceholderText("sk-...")
        self.key_edit.setEchoMode(QLineEdit.Password)
        form.addRow("API Key:", self.key_edit)
        lay.addLayout(form)
        btns = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept); btns.rejected.connect(self.reject)
        lay.addWidget(btns)

    def get_key(self) -> str:
        return self.key_edit.text().strip()

class ConfigDialog(QDialog):
    """
    Caixinha “Configurações” com os 3 tópicos na ORDEM pedida:
      1) Separar Nota por Nota
         - Caminho das notas:
         - Caminho das Notas Separadas:
      2) Gerar Planilha NFS-e
         - Caminho das NFS-e:
      3) Importar TXT da NFS-e
         - Caminho da Planilha:
    """
    def __init__(self, cfg: dict, parent=None):
        super().__init__(parent)
        self.setObjectName("tab_config")
        self.setWindowTitle("⚙️ Configurações")
        self.setModal(True)
        self.setFixedSize(720, 420)
        self.cfg = cfg or {}

        root = QVBoxLayout(self)

        # 1) Separar Nota por Nota
        grp1 = QGroupBox("Separar Nota por Nota")
        f1 = QFormLayout(grp1); f1.setContentsMargins(12,18,12,12)
        self.ed_sep_src = QLineEdit(self.cfg.get("separar.source_folder", ""))
        bt_sep_src = QPushButton("Procurar"); bt_sep_src.clicked.connect(self._pick_sep_src)
        row = QHBoxLayout(); row.addWidget(self.ed_sep_src); row.addWidget(bt_sep_src)
        f1.addRow("Caminho das notas:", row)

        self.ed_sep_out = QLineEdit(self.cfg.get("separar.separated_base_folder", ""))
        bt_sep_out = QPushButton("Procurar"); bt_sep_out.clicked.connect(self._pick_sep_out)
        row2 = QHBoxLayout(); row2.addWidget(self.ed_sep_out); row2.addWidget(bt_sep_out)
        f1.addRow("Caminho das Notas Separadas:", row2)

        # 2) Gerar Planilha NFS-e
        grp2 = QGroupBox("Gerar Planilha NFS-e")
        f2 = QFormLayout(grp2); f2.setContentsMargins(12,18,12,12)
        self.ed_plan_base = QLineEdit(self.cfg.get("planilha.base_dir", ""))
        bt_plan = QPushButton("Procurar"); bt_plan.clicked.connect(self._pick_plan_base)
        row3 = QHBoxLayout(); row3.addWidget(self.ed_plan_base); row3.addWidget(bt_plan)
        f2.addRow("Caminho das NFS-e:", row3)

        # 3) Importar TXT da NFS-e
        grp3 = QGroupBox("Importar TXT da NFS-e")
        f3 = QFormLayout(grp3); f3.setContentsMargins(12,18,12,12)
        self.ed_txt_xlsx = QLineEdit(self.cfg.get("txt.xlsx_path", ""))
        bt_txt = QPushButton("Procurar"); bt_txt.clicked.connect(self._pick_txt_xlsx)
        row4 = QHBoxLayout(); row4.addWidget(self.ed_txt_xlsx); row4.addWidget(bt_txt)
        f3.addRow("Caminho da Planilha:", row4)

        root.addWidget(grp1)
        root.addWidget(grp2)
        root.addWidget(grp3)

        btns = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept); btns.rejected.connect(self.reject)
        root.addWidget(btns)

    def _pick_sep_src(self):
        d = QFileDialog.getExistingDirectory(self, "Selecione a pasta com PDFs (notas originais)")
        if d: self.ed_sep_src.setText(d)

    def _pick_sep_out(self):
        d = QFileDialog.getExistingDirectory(self, "Selecione a pasta 'Notas Separadas'")
        if d: self.ed_sep_out.setText(d)

    def _pick_plan_base(self):
        d = QFileDialog.getExistingDirectory(self, "Selecione a pasta base para gerar PLANILHA (lotes)")
        if d: self.ed_plan_base.setText(d)

    def _pick_txt_xlsx(self):
        p,_ = QFileDialog.getOpenFileName(self, "Selecione a planilha (lancamentos.xlsx)", "", "Planilhas (*.xlsx)")
        if p: self.ed_txt_xlsx.setText(p)

    def get_config(self) -> dict:
        return {
            "separar.source_folder": self.ed_sep_src.text().strip(),
            "separar.separated_base_folder": self.ed_sep_out.text().strip(),
            "planilha.base_dir": self.ed_plan_base.text().strip(),
            "txt.xlsx_path": self.ed_txt_xlsx.text().strip(),
        }

# ---------- LOG STREAM (classificação por emojis/palavras) ----------
import io

def _classify_line_kind(s: str) -> str:
    import re
    st = (s or "").strip()
    tl = st.lower()

    if not st:
        return "blank"

    # Linhas de separador (vão virar <divider> visual)
    if re.match(r"^([=\-·\.─—_]|[═─·]){6,}$", st):
        return "divider"

    # Cabeçalhos principais: tratamos como "title" para negrito forte e faixa colorida
    if st.startswith("📘 RESUMO FINAL DA NOTA"):
        return "title"
    if st.startswith("📄 "):  # <— NOVO: cada nota vira um bloco com título forte
        return "title"

    # Linhas key: value do resumo (ex.: "  Número.............: 12")
    if re.match(r"^\s*(número|numero|data\s+de\s+emiss[aã]o|fazenda|prestador|cnpj\s+prestador|valor\s+total/l[ií]quido)\s*\.{0,}\s*:", st, flags=re.I):
        return "kv"  # <— NOVO: tipo especial para renderizar a chave em negrito

    # Se a linha já começa com emoji conhecido, mantemos "raw"
    if st[:1] in {"📌","🔎","🧾","✅","⚠️","❌","ℹ️","📄","📘"}:
        return "raw"

    # Heurísticas gerais
    if any(k in tl for k in ("falha", "erro", "exception", "traceback")): return "error"
    if any(k in tl for k in ("aviso", "atenção", "warning")):            return "warning"
    if any(k in tl for k in ("concluído", "finalizado", "pronto", "sucesso")): return "success"
    if any(k in tl for k in ("iniciando", "gerando", "processando", "lendo nota", "lendo arquivo")): return "title"
    return "info"

class _EmittingTextIO(io.TextIOBase):
    """Redireciona prints do worker para o log da UI em tempo real."""
    def __init__(self, on_line):
        super().__init__()
        self._buf = ""
        self._on_line = on_line

    def write(self, s):
        self._buf += s
        while "\n" in self._buf:
            line, self._buf = self._buf.split("\n", 1)
            if line is None:
                continue
            kind = _classify_line_kind(line)
            self._on_line(line, kind)
        return len(s)

    def flush(self):
        if self._buf:
            line = self._buf
            self._buf = ""
            self._on_line(line, _classify_line_kind(line))

# ─────────────────────────────
# Workers (QThread) – tudo com print redirecionado para o LOG
# ─────────────────────────────
class BaseWorker(QThread):
    log_sig = Signal(str, str)         # (msg, tipo)
    finished_sig = Signal(str)         # status final
    step_sig = Signal(int, int)        # (i, total) — opcional p/ barras futuras

    def __init__(self, cfg: dict, parent=None):
        super().__init__(parent)
        self.cfg = cfg or {}
        self._cancel = False

    def cancel(self):
        self._cancel = True
        try:
            self.requestInterruption()  # sinaliza interrupção no QThread
        except Exception:
            pass


    def _emit(self, msg: str, kind: str="info"):
        self.log_sig.emit(msg, kind)

    @contextmanager
    def _capture_prints(self, prefix: str = ""):
        """
        Redireciona stdout/stderr para o LOG em tempo real, com classificação por tipo.
        """
        stream = _EmittingTextIO(lambda line, kind: self._emit(f"{prefix}{line}", kind))
        old_out, old_err = sys.stdout, sys.stderr
        sys.stdout = sys.stderr = stream
        try:
            yield stream
        finally:
            try:
                stream.flush()
            except Exception:
                pass
            sys.stdout, sys.stderr = old_out, old_err

class WorkerSeparar(BaseWorker):
    """Executa o 'Separar Nota por Nota' chamando o script 'Separador PDF Nota por Nota.py'."""
    def run(self):
        try:
            src = self.cfg.get("separar.source_folder") or ""
            out = self.cfg.get("separar.separated_base_folder") or ""
            if not src or not out or not Path(src).exists():
                self._emit("Defina corretamente os caminhos do separador.", "warning")
                self.finished_sig.emit("Erro")
                return

            mod_path = Path(__file__).parent / "Separador PDF Nota por Nota.py"
            mod = _load_module_from(mod_path, "sep_nfs")
            
            # ✅ injeta callback de cancelamento para o script de separação
            setattr(mod, "is_cancelled", lambda: self._cancel)

            # Palavras chave do próprio script (respeitando defaults do arquivo original)
            general_keywords = getattr(mod, "DEFAULT_GENERAL_KEYWORDS", None) or [
                "NOTA FISCAL DE SERVIÇO", "Recibo Provisório de Serviços", "RPS", "SERVIÇO"
            ]
            ignore_keywords = getattr(mod, "DEFAULT_IGNORE_KEYWORDS", None) or [
                "CANCELADA", "CANCELADO"
            ]
            names_keywords  = getattr(mod, "DEFAULT_NAMES_KEYWORDS", None) or ["Gilson","Lucas","Adriana","Cleuber"]

            self._emit("Iniciando separação (OCR) — isso pode levar alguns minutos…", "title")
            with self._capture_prints():
                mod.process_pdfs(
                    source_folder=src,
                    separated_base_folder=out,
                    general_keywords=general_keywords,
                    ignore_keywords=ignore_keywords,
                    names_keywords=names_keywords,
                    dpi=300,
                    lang="por"
                )
            self._emit("Separação concluída.", "success")
            self.finished_sig.emit("Concluído")
        except Exception as e:
            self._emit(f"Falha geral:\n{traceback.format_exc()}", "error")
            self.finished_sig.emit("Erro")

class WorkerPlanilha(BaseWorker):
    """Executa ESTE script (automacao_NFS Digitalizada.py) para gerar planilha por lote."""
    def run(self):
        try:
            base = self.cfg.get("planilha.base_dir") or ""
            if not base or not Path(base).exists():
                self._emit("Defina 'Caminho das NFS-e' nas Configurações.", "warning")
                self.finished_sig.emit("Erro")
                return

            # Carrega o próprio módulo de OCR/extração (este arquivo)
            nfs_path = Path(__file__)  # este arquivo possui 'main()'
            nfs_mod = _load_module_from(nfs_path, "nfs_core")

            # Injeta BASE_DIR e callback de cancelamento ANTES de rodar main()
            setattr(nfs_mod, "BASE_DIR", str(base))
            setattr(nfs_mod, "is_cancelled", lambda: self._cancel)

            self._emit("Gerando planilhas de lançamentos (um arquivo 'lancamentos.xlsx' por lote)…", "title")
            with self._capture_prints():
                nfs_mod.main()
            self._emit("Geração de planilha finalizada.", "success")
            self.finished_sig.emit("Concluído")

        except Exception:
            self._emit(f"Falha geral ao gerar planilha:\n{traceback.format_exc()}", "error")
            self.finished_sig.emit("Erro")

class WorkerTxtEImport(BaseWorker):
    """Gera TXT a partir da planilha (gerar_txt_lancamentos.py) e oferece importar no sistema."""
    def run(self):
        try:
            xlsx = self.cfg.get("txt.xlsx_path") or ""
            if not xlsx or not Path(xlsx).exists():
                self._emit("Defina corretamente 'Caminho da Planilha' (lancamentos.xlsx).", "warning")
                self.finished_sig.emit("Erro")
                return

            out_txt = str(Path(xlsx).with_name("saida_lancamentos.txt"))
            mod_path = Path(__file__).parent / "gerar_txt_lancamentos.py"
            mod = _load_module_from(mod_path, "txt_gen")

            # ✅ injeta callback de cancelamento para o gerador de TXT
            setattr(mod, "is_cancelled", lambda: self._cancel)

            setattr(mod, "is_cancelled", lambda: self._cancel)  # ✅ permite Cancelar no TXT

            self._emit("Gerando TXT a partir da planilha…", "title")
            with self._capture_prints():
                with _temp_argv(["gerar_txt_lancamentos.py", xlsx, out_txt]):
                    mod.main()
            self._emit(f"TXT gerado em: {out_txt}", "success")
            self._emit("Processo concluído.", "success")
            self.finished_sig.emit(out_txt)
        except SystemExit as e:
            # o script pode chamar sys.exit — tratamos como erro/ok conforme código
            code = int(getattr(e, "code", 1) or 1)
            if code == 0:
                self.finished_sig.emit("Concluído")
            else:
                self._emit("Falha ao gerar TXT (script finalizou com erro).", "error")
                self.finished_sig.emit("Erro")
        except Exception:
            self._emit(f"Erro ao gerar/importar TXT:\n{traceback.format_exc()}", "error")
            self.finished_sig.emit("Erro")

# ─────────────────────────────
# UI Principal (estilo idêntico ao Automação Energia)
# ─────────────────────────────
class AutomacaoNFSDigitalizadasUI(QWidget):
    def __init__(self):
        super().__init__()
        self.setObjectName('tab_automacao_nfs_digitalizadas')
        self.setWindowTitle("Automação NFS-e Digitalizadas")
        if ICON_PATH.exists():
            self.setWindowIcon(QIcon(str(ICON_PATH)))
        self.setStyleSheet(STYLE_SHEET)

        self.cfg = _load_cfg()
        if self.cfg.get("api_key"):
            os.environ["OPENAI_API_KEY"] = self.cfg["api_key"]

        self.worker = None

        root = QVBoxLayout(self)
        root.setContentsMargins(14,14,14,14)
        root.setSpacing(12)

        header = self._build_header()
        root.addWidget(header)

        top = QFrame(); top.setProperty("class", "card")
        top.setStyleSheet("QFrame{border:1px solid #1e5a9c; border-radius:12px;}")
        lay = QVBoxLayout(top); lay.setContentsMargins(14,12,14,12); lay.setSpacing(10)

        actions = QHBoxLayout(); actions.setSpacing(10)

        self.btn_sep = QPushButton("📄 Separar Nota por Nota")
        self.btn_sep.setObjectName("success")
        self.btn_sep.clicked.connect(self._start_separar)
        actions.addWidget(self.btn_sep)

        self.btn_plan = QPushButton("📊 Gerar planilha NFS-e")
        self.btn_plan.clicked.connect(self._start_planilha)
        actions.addWidget(self.btn_plan)

        self.btn_txt = QPushButton("📥 Importar TXT da NFS-e")
        self.btn_txt.clicked.connect(self._start_txt_import)
        actions.addWidget(self.btn_txt)

        self.btn_cancel = QPushButton("⛔ Cancelar")
        self.btn_cancel.setObjectName("danger")
        self.btn_cancel.setEnabled(False)
        self.btn_cancel.clicked.connect(self._cancel_worker)
        actions.addWidget(self.btn_cancel)

        actions.addStretch()

        self.btn_log_clear = QToolButton(); self.btn_log_clear.setText("🧹 Limpar Log")
        self.btn_log_clear.setStyleSheet("QToolButton{background:#0d1b3d; border:1px solid #1e5a9c; border-radius:8px; padding:6px 10px;} QToolButton:hover{background:#123764;}")
        self.btn_log_clear.clicked.connect(self._log_clear)
        actions.addWidget(self.btn_log_clear)

        self.btn_log_save = QToolButton(); self.btn_log_save.setText("💾 Salvar Log")
        self.btn_log_save.setStyleSheet("QToolButton{background:#0d1b3d; border:1px solid #1e5a9c; border-radius:8px; padding:6px 10px;} QToolButton:hover{background:#123764;}")
        self.btn_log_save.clicked.connect(self._log_save)
        actions.addWidget(self.btn_log_save)

        # ⚠️ Pedido: remover a caixa "Usar OCR" da interface (não existe aqui)

        lay.addLayout(actions)
        root.addWidget(top)

        # —— Card do LOG, igual ao Energia —— 
        log_card = QFrame(); log_card.setObjectName("logCard")
        log_card.setStyleSheet("#logCard{background:#212425; border:1px solid #1e5a9c; border-radius:10px;} #logCard QLabel{border:none; background:transparent; color:#E0E0E0;}")
        llay = QVBoxLayout(log_card); llay.setContentsMargins(12,10,12,12); llay.setSpacing(8)

        title = QLabel("📝 Histórico")
        f = QFont(); f.setBold(True); f.setPointSize(12)
        title.setFont(f); title.setStyleSheet("padding:2px 6px;")
        llay.addWidget(title, alignment=Qt.AlignLeft)

        body = QFrame(); body.setObjectName("logBody")
        body.setStyleSheet("#logBody{background:#2B2F31; border:none; border-radius:8px;}")
        body_lay = QVBoxLayout(body); body_lay.setContentsMargins(12,12,12,12); body_lay.setSpacing(0)

        self.log = QTextEdit(readOnly=True)
        self.log.setFrameStyle(QFrame.NoFrame)
        self.log.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.log.setStyleSheet("QTextEdit{background:transparent; border:none; padding:0; margin:0;} QTextEdit::viewport{background:transparent; border:none; padding:0; margin:0;}")
        self.log.document().setDocumentMargin(2)
        self.log.setLineWrapMode(QTextEdit.WidgetWidth)
        self.log.setWordWrapMode(QTextOption.WrapAnywhere)
        self.log.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.log.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        body_lay.addWidget(self.log, 1)

        llay.addWidget(body)
        root.addWidget(log_card, 1)

        footer = QLabel("🧩 Automação NFS-e Digitalizadas — v1.0")
        footer.setAlignment(Qt.AlignCenter)
        footer.setStyleSheet("font-size:11px; color:#7F7F7F; padding-top:4px;")
        root.addWidget(footer)

    # —— Header (igual ao Energia)
    def _build_header(self) -> QFrame:
        header = QFrame()
        header.setStyleSheet("QFrame{border:1px solid #1e5a9c; border-radius:16px;}")
        lay = QHBoxLayout(header); lay.setContentsMargins(18,16,18,16); lay.setSpacing(14)

        icon = QLabel()
        if ICON_PATH.exists():
            pix = QPixmap(str(ICON_PATH)).scaled(44,44, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            icon.setPixmap(pix)
        else:
            icon.setText("🧩"); icon.setStyleSheet("font-size:34px; border:none;")
        lay.addWidget(icon, 0, Qt.AlignVCenter)

        title = QLabel("AUTOMAÇÃO NFS-E DIGITALIZADAS")
        f = QFont(); f.setPointSize(20); f.setBold(True)
        title.setFont(f)
        subtitle = QLabel("Separe os PDFs, gere a planilha e importe o TXT em poucos cliques.")
        title.setStyleSheet("border:none;"); subtitle.setStyleSheet("border:none;")
        title.setWordWrap(True); subtitle.setWordWrap(True)
        title.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        subtitle.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        box = QVBoxLayout()
        box.addWidget(title); box.addWidget(subtitle)
        lay.addLayout(box, 1)

        btn_cfg = QToolButton(); btn_cfg.setText("⚙️ Configurar"); btn_cfg.clicked.connect(self._open_config)
        btn_key = QToolButton(); btn_key.setText("🔑 Definir API da OpenAI"); btn_key.clicked.connect(self._open_key)
        btn_close = QToolButton(); btn_close.setText("✖ Fechar"); btn_close.clicked.connect(self._close_self_tab)

        right = QHBoxLayout(); right.setSpacing(8)
        right.addWidget(btn_cfg); right.addWidget(btn_key); right.addWidget(btn_close)
        lay.addLayout(right, 0)
        return header

    # —— Ações dos botões
    def _start_separar(self):
        cfg = self._require_config(["separar.source_folder", "separar.separated_base_folder"])
        if not cfg: return
        self._start_worker(WorkerSeparar(cfg))

    def _start_planilha(self):
        cfg = self._require_config(["planilha.base_dir"])
        if not cfg: return
        self._start_worker(WorkerPlanilha(cfg))

    def _start_txt_import(self):
        cfg = self._require_config(["txt.xlsx_path"])
        if not cfg: return
        self._start_worker(WorkerTxtEImport(cfg), ask_import=True)

    def _start_worker(self, worker: BaseWorker, ask_import: bool=False):
        if self.worker and self.worker.isRunning():
            self.log_msg("Outro processo ainda está em execução.", "warning")
            return
        self._log_divider()
        self.worker = worker
        self.worker.log_sig.connect(self.log_msg)
        self.worker.finished_sig.connect(lambda status: self._on_finished(status, ask_import))
        self.btn_cancel.setEnabled(True)
        self.worker.start()

    def _cancel_worker(self):
        if self.worker and self.worker.isRunning():
            self.log_msg("Solicitando cancelamento…", "warning")
            self.worker.cancel()  # seta a flag + requestInterruption()
            # Fallback “duro” se ainda estiver rodando após 1,5s
            QTimer.singleShot(1500, lambda: (self.worker.isRunning() and self.worker.terminate()))
        else:
            self.log_msg("Nenhum processo em execução para cancelar.", "info")


    def _on_finished(self, status: str, ask_import: bool):
        self.btn_cancel.setEnabled(False)
        if status and status.lower().endswith(".txt") and ask_import:
            out_txt = status
            # Pergunta para importar no sistema
            resp = QMessageBox.question(
                self, "Importar agora?",
                f"O TXT foi gerado em:\n{out_txt}\n\nDeseja importar os lançamentos agora no sistema?",
                QMessageBox.Yes | QMessageBox.No
            )
            if resp == QMessageBox.Yes:
                self._import_txt_no_sistema(out_txt)
        elif status == "Concluído":
            self.log_msg("Processo concluído.", "success")
        else:
            if status not in ("Concluído",):
                self.log_msg(f"Finalizado: {status}", "warning")

    def _import_txt_no_sistema(self, txt_path: str):
        try:
            mw = self.window()
            if not mw or not hasattr(mw, "_import_lancamentos_txt"):
                raise RuntimeError("Janela principal não encontrada para importar.")
            mw._import_lancamentos_txt(txt_path)
            if hasattr(mw, "carregar_lancamentos"):
                mw.carregar_lancamentos()
            if hasattr(mw, "dashboard"):
                try: mw.dashboard.load_data()
                except Exception: pass
            QMessageBox.information(self, "OK", "Lançamentos importados com sucesso.")
            self.log_msg(f"Importado: {txt_path}", "success")
        except Exception as e:
            QMessageBox.warning(self, "Falha ao importar", str(e))
            self.log_msg(f"Falha ao importar: {e}", "error")

    # —— Config + Key
    def _open_config(self):
        dlg = ConfigDialog(self.cfg, self)
        if dlg.exec():
            self.cfg.update(dlg.get_config())
            _save_cfg(self.cfg)
            self.log_msg("Configurações salvas.", "success")

    def _open_key(self):
        dlg = ApiKeyDialog(self.cfg.get("api_key", ""), self)
        if dlg.exec():
            key = dlg.get_key()
            if not key:
                QMessageBox.warning(self, "API Key", "Informe uma chave válida.")
                return
            self.cfg["api_key"] = key
            os.environ["OPENAI_API_KEY"] = key
            _save_cfg(self.cfg)
            self.log_msg("API Key definida com sucesso.", "success")

    # —— Utilidades
    def _require_config(self, keys: list[str]) -> dict | None:
        missing = [k for k in keys if not (self.cfg.get(k) or "").strip()]
        if missing:
            QMessageBox.warning(self, "Configurações",
                                "Complete as configurações antes de continuar.\n\nItens faltando:\n- " +
                                "\n- ".join(missing))
            return None
        return self.cfg

    def _close_self_tab(self):
        parent = self.parent()
        while parent and not isinstance(parent, QTabWidget):
            parent = parent.parent()
        if parent:
            idx = parent.indexOf(self)
            if idx != -1:
                parent.removeTab(idx)
        else:
            self.close()

    # —— Log estilizado (mesmo visual do Energia, com emojis)
    def log_msg(self, message: str, msg_type: str = "info"):
        msg = (message or "")
        if msg_type == "blank":
            return
        if msg_type == "divider":
            self._log_divider()
            return

        now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        palette = {
            "info":   {"emoji":"ℹ️","text":"#FFFFFF","accent":"#3A3C3D","weight":"600"},  # ← peso um pouco maior
            "success":{"emoji":"✅","text":"#A7F3D0","accent":"#2F7D5D","weight":"700"},
            "warning":{"emoji":"⚠️","text":"#FFFFFF","accent":"#8A6D3B","weight":"700"},
            "error":  {"emoji":"❌","text":"#FF6B6B","accent":"#7A2E2E","weight":"800"},
            "title":  {"emoji":"📌","text":"#FFFFFF","accent":"#1e5a9c","weight":"800"},
            "raw":    {"emoji":"","text":"#E0E0E0","accent":"#3A3C3D","weight":"500"},
            "kv":     {"emoji":"ℹ️","text":"#FFFFFF","accent":"#3A3C3D","weight":"600"},  # ← NOVO
        }
        p = palette.get(msg_type, palette["info"])

        # Se a linha já começa com emoji conhecido, não adicionar outro
        lead = msg.lstrip()[:1]
        add_emoji = (p["emoji"] and lead not in {"📌","🔎","🧾","✅","⚠️","❌","ℹ️","📄","📘"})

        # NOVO: se for "kv", deixa a parte da chave em negrito (antes dos dois-pontos)
        msg_render = msg
        if msg_type == "kv":
            import re
            m = re.match(r"^(\s*)([^:]+:)(\s*)(.*)$", msg)
            if m:
                indent, key, sp, rest = m.groups()
                msg_render = f'{indent}<span style="font-weight:800;">{key}</span>{sp}{rest}'

        emoji_html = f' <span style="margin:0 6px 0 8px;">{p["emoji"]}</span>' if add_emoji else " "
        html = (
            f'<div style="border-left:3px solid {p["accent"]}; padding:6px 10px; margin:2px 0;">'
            f'<span style="opacity:.7; font-family:monospace;">[{now}]</span>'
            f'{emoji_html}'
            f'<span style="color:{p["text"]}; font-weight:{p["weight"]}; white-space:pre-wrap;">{msg_render}</span>'
            f'</div>'
        )
        self.log.append(html)
        sb = self.log.verticalScrollBar()
        if sb:
            sb.setValue(sb.maximum())

    def _log_divider(self):
        self.log.append('<div style="border-top:1px solid #3A3C3D; margin:10px 0;"></div>')

    def _log_clear(self):
        self.log.clear()
        self.log.moveCursor(QTextCursor.Start)
        if self.log.verticalScrollBar():
            self.log.verticalScrollBar().setValue(self.log.verticalScrollBar().minimum())
        self.log_msg("Log limpo.", "info")

    def _log_save(self):
        try:
            out_dir = _logs_dir()
            fname = out_dir / f"nfs_digitalizadas_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            with open(fname, "w", encoding="utf-8") as f:
                f.write(self.log.toPlainText())
            self.log_msg(f"Log salvo em: {fname}", "success")
        except Exception as e:
            self.log_msg(f"Falha ao salvar log: {e}", "error")

# ─────────────────────────────
# ========  PARTE "CORE" ORIGINAL (main)  ========
# OBS: mantida aqui para o WorkerPlanilha poder chamar nfs_core.main()
#      Se já existir no seu arquivo antigo, mantenha toda a sua lógica original abaixo.
#      (Este cabeçalho e UI ficam no topo; a sua lógica de OCR/planilha continua valendo.)
# ─────────────────────────────

# ====== A PARTIR DAQUI: mantenha o SEU código de OCR/Layouts normalmente ======
# Seu arquivo original já contém BASE_DIR, main(), etc.
# Se por acaso remover ou renomear, garanta que exista uma função main() executável
# que use a variável global BASE_DIR como pasta base de leitura e grave lancamentos.xlsx por lote.

# --------- ENTRYPOINT ---------
# NÃO redefina main() aqui — use o pipeline real definido acima.
try:
    BASE_DIR  # garante que exista, caso rode direto o arquivo
except NameError:
    BASE_DIR = r"C:\Users\conta\Downloads"

if __name__ == "__main__":
    # chama o main() REAL (o do pipeline definido acima)
    main()
