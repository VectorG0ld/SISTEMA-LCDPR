#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import os
import re
from openpyxl import load_workbook

# ---------- LOG UI ----------
from datetime import datetime

# ---------- LOG (compatível com a UI) ----------
def log_info(msg): print(msg, flush=True)
def log_ok(msg):   print(msg, flush=True)
def log_warn(msg): print(msg, flush=True)
def log_err(msg):  print(msg, flush=True)

def _cancelled():
    try:
        cb = globals().get("is_cancelled", None)
        return bool(cb and callable(cb) and cb())
    except Exception:
        return False

DEFAULT_XLSX = r"C:\Users\conta\Downloads\PNG_\lancamentos.xlsx"

def only_digits(s: str) -> str:
    return re.sub(r"\D+", "", s or "")

def cnpj_is_valid(cnpj: str) -> bool:
    d = only_digits(cnpj)
    if len(d) != 14 or d == d[0] * 14:
        return False
    pesos1 = [5,4,3,2,9,8,7,6,5,4,3,2]
    soma1 = sum(int(d[i]) * pesos1[i] for i in range(12))
    dv1 = 11 - (soma1 % 11)
    dv1 = 0 if dv1 >= 10 else dv1
    pesos2 = [6,5,4,3,2,9,8,7,6,5,4,3,2]
    soma2 = sum(int(d[i]) * pesos2[i] for i in range(13))
    dv2 = 11 - (soma2 % 11)
    dv2 = 0 if dv2 >= 10 else dv2
    return d[12] == str(dv1) and d[13] == str(dv2)

def norm_date_dash(s: str) -> str:
    """
    Aceita 'dd/mm/aaaa' ou 'dd-mm-aaaa' e retorna sempre 'dd-mm-aaaa'.
    Não formata se não reconhecer.
    """
    s = (s or "").strip()
    m = re.search(r"(\d{2})[\/\-](\d{2})[\/\-](\d{4})", s)
    if not m:
        return s
    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"

def main():
    log_info("Geração de TXT — Lançamentos LCDPR")

    # Caminho da planilha:
    xlsx_path = sys.argv[1] if len(sys.argv) >= 2 else DEFAULT_XLSX
    out_path = sys.argv[2] if len(sys.argv) >= 3 else os.path.join(os.path.dirname(xlsx_path), "saida_lancamentos.txt")

    if _cancelled():
        log_warn("Cancelado pelo usuário.")
        return

    try:
        wb = load_workbook(xlsx_path, data_only=True)
        log_info("Workbook aberto.")
    except Exception as e:
        log_err(f"Falha ao abrir planilha: {e}")
        sys.exit(1)

    sh = wb["lancamentos"] if "lancamentos" in wb.sheetnames else wb[wb.sheetnames[0]]
    log_info(f"Aba utilizada: '{sh.title}'")

    # mapeia cabeçalho
    header_idx = {}
    for c in range(1, sh.max_column + 1):
        name = str(sh.cell(1, c).value or "").strip().lower()
        if name:
            header_idx[name] = c

    def col(name: str, default_if_missing: int) -> int:
        return header_idx.get(name.lower(), default_if_missing)

    i_data    = col("Data",       1)
    i_codfaz  = col("CodFazenda", 2)
    i_conta   = col("Conta",      3)
    i_numero  = col("NumeroNF",   4)
    i_hist    = col("Historico",  5)
    i_cnpj    = col("CNPJ",       6)
    i_tipo    = col("Tipo",       7)
    i_padrao  = col("Padrao",     8)

    tem_caminho = "caminhonf" in header_idx
    i_valor1 = col("Valor1", 10 if tem_caminho else 9)
    i_valor2 = col("Valor2", i_valor1 + 1)
    i_flag   = col("Flag",   i_valor2 + 1)

    rows = []
    total = validas = puladas = 0
    log_info(f"Total de linhas (inclui cabeçalho): {sh.max_row}")

    for r in range(2, sh.max_row + 1):
        if _cancelled():
            log_warn("Cancelado pelo usuário.")
            break
        total += 1

        data      = norm_date_dash(str(sh.cell(r, i_data).value   or "").strip())
        cod_faz   = str(sh.cell(r, i_codfaz).value                or "").strip()
        conta     = str(sh.cell(r, i_conta).value                 or "").strip() or "001"
        numero_nf = str(sh.cell(r, i_numero).value                or "").strip()
        historico = str(sh.cell(r, i_hist).value                  or "").strip()
        cnpj_raw  = str(sh.cell(r, i_cnpj).value                  or "").strip()
        tipo      = str(sh.cell(r, i_tipo).value                  or "").strip() or "2"
        padrao    = str(sh.cell(r, i_padrao).value                or "").strip() or "000"
        valor1    = str(sh.cell(r, i_valor1).value                or "").strip()
        valor2    = str(sh.cell(r, i_valor2).value                or "").strip()
        flag      = str(sh.cell(r, i_flag).value                  or "").strip() or "N"

        faltas = []
        if not data:      faltas.append("Data")
        if not cod_faz:   faltas.append("CodFazenda")
        if not numero_nf: faltas.append("NumeroNF")
        if not valor1:    faltas.append("Valor1")
        if not valor2:    faltas.append("Valor2")

        if faltas:
            puladas += 1
            log_warn(f"Linha {r}: pulada (faltando: {', '.join(faltas)}).")
            continue

        cnpj_out = "INVALIDO"
        if cnpj_raw and cnpj_raw.upper() != "INVALIDO":
            d = only_digits(cnpj_raw)
            if d and cnpj_is_valid(d):
                cnpj_out = d
            else:
                log_warn(f"Linha {r}: CNPJ inválido — gravado como 'INVALIDO'.")

        linha = f"{data}|{cod_faz}|{conta}|{numero_nf}|1|{historico}|{cnpj_out}|{tipo}|{padrao}|{valor1}|{valor2}|{flag}"
        rows.append(linha)
        validas += 1

    if _cancelled():
        log_warn("Processo de TXT cancelado.")
        return

    if not rows:
        log_warn("Nenhuma linha válida encontrada. Nada a escrever.")
        sys.exit(0)

    try:
        with open(out_path, "w", encoding="utf-8", newline="") as f:
            for ln in rows:
                f.write(ln + "\n")
        log_ok(f"TXT gerado: {out_path} ({len(rows)} linha(s)).")
    except Exception as e:
        log_err(f"Falha ao escrever TXT: {e}")
        sys.exit(1)

    log_info(f"Linhas total lidas: {total}")
    log_info(f"Linhas válidas escritas: {validas}")
    log_info(f"Linhas puladas: {puladas}")
    log_ok("Concluído.")

if __name__ == "__main__":
    main()
