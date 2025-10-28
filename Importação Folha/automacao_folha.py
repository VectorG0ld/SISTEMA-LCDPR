# -*- coding: utf-8 -*-
"""
AUTOMAÇÃO FOLHA – ESTILO IMPORTADOR (mesmo do código anexado)
--------------------------------------------------------------
• UI: mesmo tema (linhas azuis só onde necessário), cabeçalho com ❓ Ajuda, ⚙️ Configurar e ✖ Fechar,
  cartões de Controles/Status, Log com layout idêntico e botão ⛔ Cancelar.
• Período compacto: Início (MM/AAAA) → Fim (MM/AAAA).
• Lê planilha (C1 = MM/AAAA; A5.. = info do imóvel/fazenda; B5.. CPF; C5.. Nome; D5.. Líquido; F3/G3/H3 = INSS/IRRF/FGTS).
• Gera 1 TXT único (layout 12 colunas). Conta = "001" fixa. IMÓVEL = nome extraído da coluna A (ex.: "FAZENDA ALIANÇA").
• Funcionários: data do 5º dia útil. Tributos (INSS/IRRF/FGTS): data do 20º dia útil.
• Evita duplicidade por (CPF + Data).
• Cancelável via QThread.

Dependências:
- PySide6
- xlwings (opcional; recalcula fórmulas). Sem xlwings, usa openpyxl (lê valores cacheados).
"""

from __future__ import annotations

import os, re, json, time, calendar, traceback
from pathlib import Path
from datetime import datetime, date, timedelta
from collections import Counter

from PySide6.QtCore import Qt, QThread, Signal, QRegularExpression
from PySide6.QtGui import QFont, QTextCursor, QPixmap, QColor, QTextOption, QRegularExpressionValidator, QIntValidator
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFrame, QLabel, QPushButton, QTextEdit,
    QFileDialog, QMessageBox, QComboBox, QToolButton, QTabWidget, QDialog,
    QDialogButtonBox, QFormLayout, QLineEdit, QSizePolicy, QGroupBox, QListView, QGridLayout
)
import unicodedata
import tempfile
import shutil
import sqlite3

# ============================
# Estilo (igual ao anexo)
# ============================
ICON_PATH = Path(__file__).parent / "image" / "logo.png"

STYLE_SHEET = """
QMainWindow, QWidget { background-color: #1B1D1E; font-family: 'Segoe UI', Arial, sans-serif; color: #E0E0E0; }
QLineEdit, QDateEdit, QComboBox, QTextEdit { color: #E0E0E0; background-color: #2B2F31; border: 1px solid #1e5a9c; border-radius: 6px; padding: 6px; }
QLineEdit::placeholder { color: #5A5A5A; }
QPushButton { background-color: #1e5a9c; color: #FFFFFF; border: none; border-radius: 6px; padding: 8px 16px; font-weight: bold; }
QPushButton:hover, QPushButton:pressed { background-color: #002a54; }
QPushButton#danger { background-color: #C0392B; }
QPushButton#danger:hover { background-color: #E74C3C; }
QPushButton#success { background-color: #27AE60; }
QPushButton#success:hover { background-color: #2ECC71; }
QGroupBox { border: 1px solid #11398a; border-radius: 6px; margin-top: 10px; font-weight: bold; background-color: #0d1b3d; }
QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px; color: #ffffff; }
QTableWidget { background-color: #222426; color: #E0E0E0; border: 1px solid #1e5a9c; border-radius: 4px; gridline-color: #3A3C3D; alternate-background-color: #2A2C2D; }
QHeaderView::section { background-color: #1e5a9c; color: #FFFFFF; padding: 6px; border: none; }
QTabWidget::pane { border: 1px solid #1e5a9c; border-radius: 4px; background: #212425; margin-top: 5px; }
QTabBar::tab { background: #2A2C2D; color: #E0E0E0; padding: 8px 16px; border: 1px solid #1e5a9c; border-top-left-radius: 4px; border-top-right-radius: 4px; margin-right: 2px; }
QTabBar::tab:selected { background: #1e5a9c; color: #FFFFFF; border-bottom: 2px solid #002a54; }
QStatusBar { background-color: #212425; color: #7F7F7F; border-top: 1px solid #1e5a9c; }
/* ===== Diálogo de Configurações ===== */
QWidget#tab_config QGroupBox {
    background: transparent;
    border: 1px solid #11398a;
    border-radius: 6px;
    margin-top: 14px;
}
QWidget#tab_config QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    padding: 0 6px;
    background-color: #1B1D1E;
    color: #ffffff;
}
QWidget#tab_config QLabel { border: none; background: transparent; }
QWidget#tab_config QFrame,
QWidget#tab_config QFrame#card,
QWidget#tab_config QFrame.card {
    background: transparent;
    border: 1px solid #11398a;
    border-radius: 6px;
}
/* ===== Somente no diálogo de Configurar (objectName=tab_config) ===== */
QWidget#tab_config QGroupBox {
    background: transparent;          /* sem fundo azul */
    border: 1px solid #11398a;        /* só a linha azul */
    border-radius: 6px;
    margin-top: 14px;                 /* espaço para o título do groupbox */
}

QWidget#tab_config QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    padding: 0 6px;
    background-color: #1B1D1E;        /* mesma cor do fundo do diálogo */
    color: #ffffff;
}

/* (no dialog Configurar) */
QWidget#tab_config QFrame {            /* todos os frames: sem borda */
    background: transparent;
    border: none;
}
QWidget#tab_config QFrame#card,        /* apenas o card principal com a borda azul */
QWidget#tab_config QFrame.card {
    background: transparent;
    border: 1px solid #11398a;
    border-radius: 6px;
}

/* Labels do Configurar sem borda/fundo */
QWidget#tab_config QLabel {
    border: none;
    background: transparent;
}
/* Evita que textos forcem largura mínima absurda */
QLabel, QTextEdit { min-width: 0; }

/* Log mais dócil: nada de barra horizontal */
#logCard QTextEdit {
    /* sem barra horizontal */
    qproperty-horizontalScrollBarPolicy: 1; /* Qt::ScrollBarAlwaysOff */
}
/* ===== Combos: popup consistente e sem “explodir” tamanho ===== */
QComboBox { combobox-popup: 0; }  /* usa popup rolável */
QComboBox QAbstractItemView {
    background-color: #2B2F31;
    color: #E0E0E0;
    border: 1px solid #1e5a9c;
    outline: none;
}
QComboBox QAbstractItemView::item {
    padding: 6px 10px;
}
QComboBox QAbstractItemView::item:selected {
    background: #1e5a9c;
    color: #FFFFFF;
}

"""

# ===== Mapeamentos de IMÓVEL por TITULAR (coluna A5..) =====
IMOVEL_MAP = {
    "CLEUBER": {
        "FAZENDA FRUTACC": "001",
        "FAZENDA UNIÃO": "002",
        "FAZENDA L3": "003",
        "FAZENDA PRIMAVERA": "004",
        "FAZENDA ALIANÇA": "005",
        "ARMAZEM PRIMAVERA": "006",
        "FAZENDA BARRAGEM GRANDE": "007",
        "FAZENDA ESTRELA": "008",
        "FAZENDA GUARA": "009",
    },
    "ADRIANA": {
        "FAZENDA POUSO DA ANTA": "001",
        "FAZENDA PRIMAVERA III": "002",
    },
    "GILSON": {
        "FAZENDA FORMIGA": "001",
    },
    "LUCAS": {
        "FAZENDA ALIANÇA 2": "001",
    },
}

# ===== Perfis suportados (rótulo -> nome interno da MainWindow) =====
PROFILE_DISPLAY = ["Cleuber", "Gilson", "Adriana", "Lucas"]
PROFILE_MAP = {
    "Cleuber": "Cleuber Marcos",
    "Gilson":  "Gilson Oliveira",
    "Adriana": "Adriana Lucia",
    "Lucas":   "Lucas Laignier",
}

# ===== CNPJs dos TRIBUTOS (usar apenas dígitos) =====
TRIBUTOS_CNPJ = {
    "INSS": re.sub(r"\D", "", "00.394.460/0058-87"),  # 00394460005887
    "IRRF": re.sub(r"\D", "", "29.979.036/0001-40"),  # 29979036000140
    "FGTS": re.sub(r"\D", "", "37.115.367/0001-60"),  # 37115367000160
}

# ============================
# Helpers
# ============================
import re

_ref_re = re.compile(r"REF\.\s*(\d{2}/\d{4})", re.IGNORECASE)

def _extract_ref(h: str) -> str:
    if not h:
        return ""
    m = _ref_re.search(h)
    return (m.group(1) or "").strip() if m else ""

def _digits(s) -> str:
    return re.sub(r"\D", "", str(s or ""))

def _extract_name_from_historico(h: str) -> str:
    """
    Tenta extrair um nome/hint do texto de histórico:
    - primeiro tenta conteúdo entre parênteses;
    - depois tenta o trecho após um traço/dash;
    - fallback: as primeiras palavras do histórico.
    """
    if not h:
        return ""
    try:
        s = str(h).strip()
        # conteúdo entre parênteses
        m = re.search(r"\(([^)]+)\)", s)
        if m:
            return m.group(1).strip()
        # após um traço (–, — ou -)
        m = re.search(r"[-–—]\s*(.+)$", s)
        if m:
            return m.group(1).strip()
        # como fallback, usa até 4 primeiras palavras
        parts = s.split()
        if not parts:
            return ""
        return " ".join(parts[:4]).strip()
    except Exception:
        return ""

def _to_cent(valor) -> str:
    """Converte para centavos sem pontuação. Aceita 123,45 | 123.45 | 'R$ 123,45' | float | int."""
    try:
        s = str(valor).strip().replace("R$", "").replace(" ", "")
        if not s:
            return "0"
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s and "." not in s:
            s = s.replace(",", ".")
        return str(int(round(float(s) * 100)))
    except Exception:
        return "0"

def _fmt_dd_mm_yyyy(d: date) -> str:
    return d.strftime("%d-%m-%Y")

def _next_business_day(d: date) -> date:
    # sábado → segunda, domingo → segunda
    if d.weekday() == 5:  # sat
        return d + timedelta(days=2)
    if d.weekday() == 6:  # sun
        return d + timedelta(days=1)
    return d

def _dia_ajustado(ano: int, mes: int, dia: int) -> date:
    last_day = calendar.monthrange(ano, mes)[1]
    dia = min(dia, last_day)
    return _next_business_day(date(ano, mes, dia))

def _iter_mes_ano(inicio_mm_aaaa: str, fim_mm_aaaa: str):
    im, ia = [int(x) for x in inicio_mm_aaaa.split("/")]
    fm, fa = [int(x) for x in fim_mm_aaaa.split("/")]
    y, m = ia, im
    while (y < fa) or (y == fa and m <= fm):
        yield (m, y)
        m += 1
        if m > 12:
            m = 1
            y += 1

def _make_line(data_br, imovel, conta, num_doc, tipo_doc, historico, cpf, tipo_lanc, cent_ent, cent_sai, cent_saldo, nat) -> str:
    return "|".join([
        data_br, imovel, conta, num_doc, tipo_doc,
        historico, cpf, tipo_lanc, cent_ent, cent_sai, cent_saldo, nat
    ])

# === DEDUP (contra o banco) ===
def _yyyymmdd_from_br(date_br: str) -> int:
    # date_br = "dd-mm-YYYY"
    return int(date_br[6:10] + date_br[3:5] + date_br[0:2])

def _norms_code(code: str) -> list[str]:
    s = (code or '').strip()
    if not s:
        return []
    out = [s]
    if s.isdigit():
        out += [s.zfill(3), (s.lstrip('0') or '0')]
    # remove duplicatas preservando ordem
    seen = set(); res = []
    for x in out:
        if x not in seen:
            seen.add(x); res.append(x)
    return res

def _extract_imovel_name(texto_a: str) -> str:
    """
    A célula A pode vir como "4 - CLEUBER MARCOS DE OLIVEIRA FAZ ALIANÇA" ou conter
    'FAZENDA <NOME>', 'FAZ <NOME>', 'SÍTIO/SITIO <NOME>', 'ARMAZÉM/ARMAZEM <NOME>'.
    Retorna um nome de imóvel padronizado, preferindo o prefixo 'FAZENDA '.
    """
    if not texto_a:
        return ""
    t = str(texto_a).upper()

    # Padrões "FAZENDA XYZ"
    m = re.search(r"(FAZENDA\s+[A-Z0-9ÇÃÕÁÉÍÓÚÂÊÔ ]+)", t)
    if m:
        return m.group(1).strip()

    # Padrões "FAZ XYZ" -> normaliza pra "FAZENDA XYZ"
    m = re.search(r"\bFAZ[\.\s]+([A-Z0-9ÇÃÕÁÉÍÓÚÂÊÔ ]+)", t)
    if m:
        return f"FAZENDA {m.group(1).strip()}"

    # SÍTIO / CHÁCARA / ARMAZÉM
    for prefix in ("SÍTIO", "SITIO", "CHÁCARA", "CHACARA", "ARMAZÉM", "ARMAZEM"):
        m = re.search(rf"({prefix}\s+[A-Z0-9ÇÃÕÁÉÍÓÚÂÊÔ ]+)", t)
        if m:
            return m.group(1).strip()

    # Fallback: tenta pegar o trecho após o hífen
    m = re.search(r"\-\s*([A-Z0-9ÇÃÕÁÉÍÓÚÂÊÔ ]+)", t)
    if m:
        # Se achar "FAZ ALGO", normaliza…
        val = m.group(1).strip()
        if val.startswith("FAZ "):
            return "FAZENDA " + val[4:].strip()
        return val
    return "FAZENDA"

def _mode_or_default(items, default="FAZENDA"):
    try:
        c = Counter([x for x in items if (x and x.strip())])
        if not c:
            return default
        return c.most_common(1)[0][0]
    except Exception:
        return default

def _norm(s: str) -> str:
    """Remove acentos e força UPPER para comparação robusta."""
    s = str(s or "")
    s = unicodedata.normalize("NFKD", s)
    s = s.encode("ASCII", "ignore").decode("ASCII")
    return s.upper().strip()

def _owner_from_text(a_txt: str) -> str | None:
    """Detecta o titular (CLEUBER/ADRIANA/GILSON/LUCAS) no texto da coluna A."""
    TITS = ("CLEUBER", "ADRIANA", "GILSON", "LUCAS")
    t = _norm(a_txt)
    for tt in TITS:
        if tt in t:
            return tt
    return None

def _cod_imovel_from_colA(a_txt: str, owner: str | None = None) -> str:
    """
    Lê a coluna A e tenta identificar a FAZENDA pelo texto livre, mapeando para o CÓDIGO
    cadastrado em IMOVEL_MAP. Se o titular (owner) for detectado, restringe o match ao mapa dele.
    Usa substring e, se necessário, similaridade (fuzzy).
    """
    t = _norm(a_txt or "")
    # normaliza abreviações comuns
    t = t.replace(" FAZ. ", " FAZENDA ").replace(" FAZ ", " FAZENDA ")

    # 0) Se tiver owner, usamos primeiro só o mapa dele
    maps_to_scan = []
    if owner and owner in IMOVEL_MAP:
        maps_to_scan.append(IMOVEL_MAP[owner])
    # 1) Depois, todos os mapas (fallback)
    maps_to_scan.append({faz: code for mp in IMOVEL_MAP.values() for faz, code in mp.items()})

    for mp in maps_to_scan:
        candidates = [(_norm(faz), code) for faz, code in mp.items()]
        # Match direto por substring (mais seguro)
        for n, code in candidates:
            if n in t or t in n:
                return code
        # Fuzzy se não achou por substring
        try:
            from difflib import SequenceMatcher
            best_code, best_ratio = None, 0.0
            for n, code in candidates:
                r = SequenceMatcher(None, n, t).ratio()
                if r > best_ratio:
                    best_ratio, best_code = r, code
            if best_ratio >= 0.60:
                return best_code
        except Exception:
            pass

    # 2) Fallback
    return "001"

# --- Reparador do cache do win32com/pywin32 (para erros CLSIDToPackageMap) ---
def _repair_win32com_genpy_cache():
    """
    Remove caches 'gen_py' conhecidos e força o pywin32 a regenerar
    os wrappers COM. Seguro rodar sempre que precisar.
    """
    try:
        import win32com, importlib, sys
        from pathlib import Path
        import shutil, tempfile, glob

        # 1) gen_py em %LOCALAPPDATA%/Temp
        temp_gen = Path(tempfile.gettempdir()) / "gen_py"
        if temp_gen.exists():
            shutil.rmtree(temp_gen, ignore_errors=True)

        # 2) gen_py dentro do pacote site-packages/win32com
        try:
            gen_pkg = Path(win32com.__gen_path__)  # presente em algumas builds
            if gen_pkg.exists():
                shutil.rmtree(gen_pkg, ignore_errors=True)
        except Exception:
            # fallback: varre por diretórios .../win32com/gen_py*
            try:
                from win32com import __path__ as w32paths
                for p in w32paths:
                    for g in glob.glob(str(Path(p) / "gen_py*")):
                        shutil.rmtree(g, ignore_errors=True)
            except Exception:
                pass

        # 3) recarrega gencache
        try:
            import win32com.client.gencache as gencache
            importlib.reload(gencache)
        except Exception:
            pass
    except Exception:
        # não impede a execução se algo falhar
        pass

# ============================
# Leitura da planilha (xlwings → openpyxl)
# ============================
def _read_sheet_with_xlwings(filepath: str, mes: int, ano: int):
    try:
        import xlwings as xw
    except Exception as e:
        raise RuntimeError("xlwings não disponível. Instale xlwings (pip install xlwings).") from e

    def _open_and_read():
        app = None
        wb = None
        try:
            app = xw.App(visible=False, add_book=False)
            app.display_alerts = False
            app.screen_updating = False

            wb = app.books.open(str(Path(filepath).resolve()), update_links=False, read_only=False)
            sh = wb.sheets[0]

            # define período
            sh.range("C1").value = f"{mes:02d}/{ano}"

            # força recálculo
            try:
                wb.api.CalculateFullRebuild()
            except Exception:
                pass
            try:
                app.api.CalculateFullRebuild()
            except Exception:
                pass
            try:
                sh.api.Calculate()
            except Exception:
                pass

            # leitura segura dos tributos (F3/G3/H3)
            def _to_cent_safe(v):
                try:
                    return _to_cent(v or 0)
                except Exception:
                    return "0"

            inss = irrf = fgts = None
            for _ in range(8):
                time.sleep(0.25)
                vF, vG, vH = sh.range("F3").value, sh.range("G3").value, sh.range("H3").value
                if vF is not None or vG is not None or vH is not None:
                    inss = _to_cent_safe(vF)
                    irrf = _to_cent_safe(vG)
                    fgts = _to_cent_safe(vH)
                    break
            if inss is None:
                inss = _to_cent_safe(sh.range("F3").value)
                irrf = _to_cent_safe(sh.range("G3").value)
                fgts = _to_cent_safe(sh.range("H3").value)

            # linhas A5..D?, com detecção de imóvel e titular
            funcionarios, imoveis = [], []
            r = 5
            while True:
                a_txt = str(sh.range(f"A{r}").value or "").strip()
                cpf   = _digits(sh.range(f"B{r}").value or "")
                nome  = (sh.range(f"C{r}").value or "").strip()
                val   = sh.range(f"D{r}").value
                if not (a_txt or cpf or nome or val):
                    break

                owner = _owner_from_text(a_txt)
                imovel_nome = _extract_imovel_name(a_txt)
                cod_imovel = _cod_imovel_from_colA(a_txt, owner)

                if cod_imovel:
                    imoveis.append(cod_imovel)

                funcionarios.append({
                    "cpf": cpf,
                    "nome": nome,
                    "centavos": _to_cent(val),
                    "imovel": cod_imovel,
                    "imovel_nome": imovel_nome,
                    "titular": owner
                })
                r += 1

            imovel_tributos = _mode_or_default(imoveis, default="001")
            return funcionarios, {"INSS": inss, "IRRF": irrf, "FGTS": fgts}, imovel_tributos
        finally:
            try:
                if wb:
                    wb.close()
            except Exception:
                pass
            try:
                if app:
                    app.quit()
            except Exception:
                pass

    # 1ª tentativa
    try:
        return _open_and_read()
    except Exception as e:
        msg = str(e)
        # Erro típico de cache: ... has no attribute 'CLSIDToPackageMap'
        if "CLSIDToPackageMap" in msg or "gen_py" in msg:
            # limpa cache e tenta mais uma vez
            _repair_win32com_genpy_cache()
            try:
                return _open_and_read()
            except Exception as e2:
                raise RuntimeError(
                    f"Falha no Excel/xlwings ao ler '{filepath}' para {mes:02d}/{ano} "
                    f"(após limpar cache pywin32): {e2}"
                ) from e2
        # outros erros: propaga
        raise RuntimeError(
            f"Falha no Excel/xlwings ao ler '{filepath}' para {mes:02d}/{ano}: {e}"
        ) from e

def _read_sheet_with_openpyxl(filepath: str, mes: int, ano: int):
    from openpyxl import load_workbook
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    # Tenta setar C1; sem recálculo
    try:
        ws["C1"].value = f"{mes:02d}/{ano}"
    except Exception:
        pass

    def cell(addr):
        v = ws[addr].value
        return v if v is not None else 0

    inss = _to_cent(cell("F3"))
    irrf = _to_cent(cell("G3"))
    fgts = _to_cent(cell("H3"))

    funcionarios, imoveis = [], []
    r = 5
    while True:
        a_txt = str(ws[f"A{r}"].value or "").strip()
        cpf   = _digits(ws[f"B{r}"].value or "")
        nome  = str(ws[f"C{r}"].value or "").strip()
        val   = ws[f"D{r}"].value
        if not (a_txt or cpf or nome or val):
            break

        cod_imovel = _cod_imovel_from_colA(a_txt)
        if cod_imovel:
            imoveis.append(cod_imovel)

        funcionarios.append({
            "cpf": cpf, "nome": nome,
            "centavos": _to_cent(val),
            "imovel": cod_imovel  # <- usa o CÓDIGO
        })
        r += 1

    imovel_tributos = _mode_or_default(imoveis, default="001")  # códigos

    return funcionarios, {"INSS": inss, "IRRF": irrf, "FGTS": fgts}, imovel_tributos

def _read_planilha(filepath: str, mes: int, ano: int, force_xlwings: bool = False):
    """
    Se force_xlwings=True, usa xlwings (Excel) e propaga o erro real se falhar.
    Caso contrário, tenta xlwings e, se não houver, usa openpyxl (sem recálculo).
    """
    if force_xlwings:
        # usa xlwings e propaga a exceção real (sem mascarar)
        return _read_sheet_with_xlwings(filepath, mes, ano)

    # modo flexível: tenta xlwings, cai para openpyxl se faltar
    try:
        return _read_sheet_with_xlwings(filepath, mes, ano)
    except Exception:
        return _read_sheet_with_openpyxl(filepath, mes, ano)

class FolhaWorker(QThread):
    # sinais da thread (mensagens, estatísticas e término)
    log_sig = Signal(str, str)           # (mensagem_html_ou_texto, nivel: info/success/warning/error/raw/divider)
    stats_sig = Signal(int, int, int)    # (total, ok, err)
    finished_sig = Signal(str, str)      # (status, caminho_txt)

    def __init__(self, planilha: str, inicio: str, fim: str, db_path: str | None = None, parent=None, force_xlwings: bool = False):
        super().__init__(parent)
        self.planilha = planilha
        self.inicio = inicio
        self.fim = fim
        self._cancel = False
        self.total = 0
        self.ok = 0
        self.err = 0

        self.COD_CONTA = "001"
        self._vistos = set()
        self._linhas = []

        self._db_path = db_path
        self._conn = None

        self._force_xlwings = bool(force_xlwings)  # <- NOVO


    def cancel(self):
        self._cancel = True

    def _emit(self, msg: str, kind: str = "info"):
        self.log_sig.emit(msg, kind)

    def _emit_stats(self):
        self.stats_sig.emit(self.total, self.ok, self.err)

    # ---------- DB helpers (thread-safe p/ SQLite) ----------
    def _conn_ro(self):
        if self._conn is None and self._db_path:
            try:
                uri = f"file:{self._db_path}?mode=ro"
                self._conn = sqlite3.connect(uri, uri=True, check_same_thread=False)
            except Exception:
                self._conn = sqlite3.connect(self._db_path, check_same_thread=False)
        return self._conn

    def _fetch_one(self, sql: str, params: tuple = ()):
        conn = self._conn_ro()
        if not conn:
            return None
        cur = conn.cursor()
        cur.execute(sql, params)
        return cur.fetchone()

    def _imovel_id(self, cod: str):
        if not cod:
            return None
        for alt in _norms_code(cod):
            row = self._fetch_one("SELECT id FROM imovel_rural WHERE cod_imovel=?", (alt,))
            if row:
                return row[0]
        return None

    def _part_id(self, digits: str):
        if not digits:
            return None
        row = self._fetch_one("SELECT id FROM participante WHERE cpf_cnpj=?", (digits,))
        return row[0] if row else None

    def _exists_in_db(self, data_br: str, cod_imovel: str, digits: str, cents_str: str, historico: str) -> bool:
        """
        True se já existe lançamento idêntico (data_ord, imóvel, participante, tipo=2, valor_saida, histórico).
        Se não conseguir resolver FKs ou não houver conexão, NÃO bloqueia.
        """
        id_im = self._imovel_id(cod_imovel)
        pid = self._part_id(digits)
        if not id_im or not pid:
            return False

        data_ord = _yyyymmdd_from_br(data_br)
        try:
            valor = (int(''.join(c for c in str(cents_str) if c.isdigit())) / 100.0)
        except Exception:
            valor = 0.0

        row = self._fetch_one(
            """
            SELECT 1 FROM lancamento
             WHERE data_ord=? AND cod_imovel=? AND id_participante=? AND tipo_lanc=2
               AND ABS(valor_saida - ?) < 0.005
               AND TRIM(COALESCE(historico,'')) = TRIM(?)
             LIMIT 1
            """,
            (data_ord, id_im, pid, valor, historico)
        )
        return bool(row)

    # ---------- Execução ----------
    def run(self):
        try:
            # ---------- helpers locais p/ exibição ----------
            def _fmt_hms(dt: datetime) -> str:
                return dt.strftime("%H:%M:%S")
    
            def _fmt_money_cents(cents_str: str) -> str:
                try:
                    v = int(''.join(ch for ch in str(cents_str) if ch.isdigit())) / 100.0
                except Exception:
                    v = 0.0
                s = f"{v:,.2f}"
                return "R$ " + s.replace(",", "X").replace(".", ",").replace("X", ".")
    
            def _br_slash(date_br_hifen: str) -> str:
                # entra "dd-mm-aaaa", sai "dd/mm/aaaa"
                return (date_br_hifen or "").replace("-", "/")
    
            # ---------- cabeçalho da sessão ----------
            session_start = datetime.now()
            self._emit(f"🧾 <b>Geração de TXT — {session_start.strftime('%d/%m/%Y')}</b>", "raw")
            self._emit(f"⏳ <b>Início:</b> {_fmt_hms(session_start)}", "raw")
            self._emit("", "divider")
    
            # estatísticas de sessão
            self.total = 0
            self.ok = 0
            self.err = 0
            valor_total_sessao = 0.0
            meses_processados = []
    
            # var de saída (linhas do TXT)
            linhas = []
    
            # percorre meses do período
            for m, y in _iter_mes_ano(self.inicio, self.fim):
                if self._cancel:
                    self._emit("Processo cancelado.", "warning")
                    self.finished_sig.emit("Cancelado", "")
                    return
                
                if self._force_xlwings:
                    self._emit("⚙️ Engine: <b>Excel (xlwings)</b> — fórmulas recalculadas.", "raw")
                else:
                    self._emit("⚙️ Engine: <b>Automática</b> (tenta xlwings; se ausente, lê cache via openpyxl).", "raw")

                # leitura da planilha para o mês/ano (seta C1 = MM/AAAA e recalcula se xlwings)
                self._emit(f"🔄 Ajustando C1 → <b>{m:02d}/{y}</b> e lendo dados…", "raw")
                try:
                    funcs, trib, imovel_trib = _read_planilha(self.planilha, m, y, force_xlwings=self._force_xlwings)
                except Exception as e:
                    self.err += 1
                    self._emit_stats()
                    self._emit(f"❌ Erro ao ler {m:02d}/{y}: {e}", "error")
                    continue

                # contabiliza total esperado (funcionários + tributos com valor)
                trib_count = sum(1 for k in ("INSS","IRRF","FGTS") if str(trib.get(k,"0")) != "0")
                self.total += len(funcs) + trib_count
                self._emit_stats()
    
                mes_start = datetime.now()
                meses_processados.append(f"{m:02d}/{y}")
    
                # cabeçalho do mês (clean)
                head = (
                    "<div style='border-top:1px solid #3A3C3D; margin:16px 0 10px 0;'></div>"
                    f"<div style='font-weight:800; font-size:14px; margin:2px 0 12px 0;'>"
                    f"📆 {m:02d}/{y} &nbsp;—&nbsp; Funcionários: <b>{len(funcs)}</b> "
                    f"&nbsp;•&nbsp; Tributos: <b>{trib_count}</b>"
                    "</div>"
                )
                self._emit(head, "raw")
    
                # datas
                data_func = _dia_ajustado(y, m, 5)
                data_trib = _dia_ajustado(y, m, 20)
                data_func_br = _fmt_dd_mm_yyyy(data_func)  # dd-mm-aaaa
                data_trib_br = _fmt_dd_mm_yyyy(data_trib)
    
                subtotal_mes = 0.0
    
                # ---------- funcionários (um card por pagamento) ----------
                for f in funcs:
                    if self._cancel:
                        self._emit("Cancelado pelo usuário.", "warning")
                        self.finished_sig.emit("Cancelado", "")
                        return
    
                    cpf   = _digits(f.get("cpf"))
                    nome  = (f.get("nome") or "").strip()
                    cents = str(f.get("centavos") or "0")
                    imov  = (f.get("imovel") or "001").strip()
    
                    if not nome:
                        self._emit("⚠️ Linha ignorada: sem nome.", "warning")
                        continue
                    if not cpf:
                        self._emit(f"⚠️ Linha ignorada (sem CPF): {nome}", "warning")
                        continue
                    if cents in ("", "0"):
                        self._emit(f"⚠️ Linha ignorada (valor zero): {nome}", "warning")
                        continue
                    
                    # dedupe por sessão
                    key = (cpf, data_func_br)      # CPF + DATA
                    if key in self._vistos:
                        self._emit(f"↩️ DUP ignorado (sessão): {nome} {_br_slash(data_func_br)}", "warning")
                        continue
                    self._vistos.add(key)
    
                    historico = f"FOLHA DE PAGAMENTO REF. {m:02d}/{y} ({nome})"
    
                    # dedupe no banco
                    if self._exists_in_db(data_func_br, imov, cpf, cents, historico):
                        self._emit(f"↩️ DUP no banco: {nome} {_br_slash(data_func_br)}", "warning")
                        continue
                    
                    # linha do TXT (12 colunas)
                    linha = _make_line(
                        data_func_br, imov, self.COD_CONTA,
                        "N", "1", historico, cpf, "2", "000", cents, cents, "N"
                    )
                    linhas.append(linha)
                    self.ok += 1
                    self._emit_stats()
    
                    # valores para exibir
                    money = _fmt_money_cents(cents)
                    subtotal_mes += float(money.replace("R$ ","").replace(".","").replace(",","."))
                    valor_total_sessao += float(money.replace("R$ ","").replace(".","").replace(",","."))
                    agora = _fmt_hms(datetime.now())
    
                    # CARD clean do funcionário
                    self._emit(
                        "<div style='border:1px solid #3A3C3D; border-radius:8px; padding:10px 12px; margin:18px 0;'>"
                        f"<div style='font-weight:700; margin-bottom:6px;'>👤 Funcionário: <b>{nome}</b>"
                        f" <span style='opacity:.75; font-weight:500;'>— CPF: {_digits(cpf)}</span></div>"
                        f"<div>🗓️ Data: <b>{_br_slash(data_func_br)}</b> &nbsp;•&nbsp; 💰 Pago: <b>{money}</b></div>"
                        f"<div>🏠 Imóvel: <b>{imov}</b> &nbsp;•&nbsp; 🧾 Histórico: <span style='opacity:.9;'>{historico}</span></div>"
                        f"<div style='margin-top:6px; opacity:.85;'>✅ OK <b>{agora}</b> &nbsp;•&nbsp; 📌 Pagamento processado</div>"
                        "</div>",
                        "raw"
                    )
    
                # ---------- tributos (cards) ----------
                for rotulo, cents in (("INSS", trib.get("INSS","0")), ("IRRF", trib.get("IRRF","0")), ("FGTS", trib.get("FGTS","0"))):
                    if self._cancel:
                        self._emit("Cancelado pelo usuário.", "warning")
                        self.finished_sig.emit("Cancelado", "")
                        return
                    if not cents or str(cents) == "0":
                        continue
                    
                    cnpj = TRIBUTOS_CNPJ.get(rotulo, "")
                    historico = f"FOLHA DE PAGAMENTO REF. {m:02d}/{y} {rotulo}"
                    imov_use = imovel_trib or "001"
    
                    if self._exists_in_db(data_trib_br, imov_use, cnpj, str(cents), historico):
                        self._emit(f"↩️ DUP no banco: {rotulo} {_br_slash(data_trib_br)}", "warning")
                        continue
                    
                    linha = _make_line(
                        data_trib_br, imov_use, self.COD_CONTA,
                        "N", "1", historico, cnpj, "2", "000", str(cents), str(cents), "N"
                    )
                    linhas.append(linha)
                    self.ok += 1
                    self._emit_stats()
    
                    money = _fmt_money_cents(cents)
                    subtotal_mes += float(money.replace("R$ ","").replace(".","").replace(",","."))
                    valor_total_sessao += float(money.replace("R$ ","").replace(".","").replace(",","."))
                    agora = _fmt_hms(datetime.now())
    
                    # CARD clean do tributo
                    self._emit(
                        "<div style='border:1px dashed #3A3C3D; border-radius:8px; padding:10px 12px; margin:14px 0;'>"
                        f"<div style='font-weight:700; margin-bottom:6px;'>🏛️ Tributo: <b>{rotulo}</b>"
                        f" <span style='opacity:.75; font-weight:500;'>— CNPJ: {_digits(cnpj)}</span></div>"
                        f"<div>🗓️ Data: <b>{_br_slash(data_trib_br)}</b> &nbsp;•&nbsp; 💰 Valor: <b>{money}</b></div>"
                        f"<div>🏠 Imóvel: <b>{imov_use}</b> &nbsp;•&nbsp; 🧾 Histórico: <span style='opacity:.9;'>{historico}</span></div>"
                        f"<div style='margin-top:6px; opacity:.85;'>✅ OK <b>{agora}</b> &nbsp;•&nbsp; 📌 Tributo processado</div>"
                        "</div>",
                        "raw"
                    )
    
                # ---------- subtotal do mês ----------
                mes_end = datetime.now()
                tempo_mes = (mes_end - mes_start).total_seconds()
                subtotal_fmt = "R$ " + f"{subtotal_mes:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                self._emit(f"💵 <b>Subtotal {m:02d}/{y}:</b> <b>{subtotal_fmt}</b> &nbsp;—&nbsp; ⏱️ <b>{int(tempo_mes)}s</b> "
                           f"(<span style='opacity:.8;'>{_fmt_hms(mes_start)} → {_fmt_hms(mes_end)}</span>)", "raw")
                self._emit("", "divider")
    
            # ---------- finalizar: salvamento ----------
            if not linhas:
                self._emit("Nenhuma linha para salvar.", "warning")
                self.finished_sig.emit("Vazio", "")
                return
    
            out_dir = Path(self.planilha).parent
            fname = out_dir / f"folha_{self.inicio.replace('/','-')}_a_{self.fim.replace('/','-')}.txt"
            with open(fname, "w", encoding="utf-8") as f:
                f.write("\n".join(linhas))
    
            # resumo final
            session_end = datetime.now()
            total_secs = int((session_end - session_start).total_seconds())
            total_fmt = "R$ " + f"{valor_total_sessao:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            self._emit(
                "<div style='border-top:1px solid #3A3C3D; margin:16px 0 10px 0;'></div>"
                "<div style='font-weight:800; font-size:14px; margin:2px 0 12px 0;'>🏁 Resumo Final</div>",
                "raw"
            )
            self._emit(f"• 🗓️ <b>Período:</b> {self.inicio} → {self.fim}", "raw")
            self._emit(f"• 📦 <b>Registros processados:</b> {self.ok}", "raw")
            self._emit(f"• 🧾 <b>TXT:</b> <code>{fname}</code>", "raw")
            self._emit(f"• ⏱️ <b>Tempo total:</b> <b>{total_secs}s</b> ({_fmt_hms(session_start)} → {_fmt_hms(session_end)})", "raw")
            self._emit(f"• 💰 <b>Valor total:</b> <b>{total_fmt}</b>", "raw")
    
            # status final + caminho
            self._emit(f"TXT gerado: {fname}", "success")
            self.finished_sig.emit("Concluído", str(fname))
    
        except Exception:
            self.err += 1
            self._emit_stats()
            self._emit(f"Erro inesperado:\n{traceback.format_exc()}", "error")
            self.finished_sig.emit("Erro", "")
        finally:
            try:
                if self._conn:
                    self._conn.close()
            except Exception:
                pass
            
# ============================
# Diálogo de Configurar
# ============================
class ConfigDialog(QDialog):
    def __init__(self, cfg: dict, parent=None):
        super().__init__(parent)
        self.setWindowTitle("⚙️ Configurações — Automação Folha")
        self.setModal(True)
        self.setStyleSheet(STYLE_SHEET)
        self.setFixedWidth(640)
        self.setObjectName("tab_config")  # aplica o CSS acima somente neste diálogo


        self._cfg = dict(cfg or {})
        root = QVBoxLayout(self)
        root.setContentsMargins(14, 14, 14, 14)
        root.setSpacing(10)

        # Card com borda azul (linhas só onde necessário), igual Talão de Energia
        card = QFrame(self)
        card.setObjectName("card")
        card.setStyleSheet("#card{border:1px solid #1e5a9c; border-radius:12px;}")
        card_lay = QVBoxLayout(card)
        card_lay.setContentsMargins(12, 10, 12, 12)
        card_lay.setSpacing(8)

        grp = QGroupBox("Caminhos e Opções", card)
        grp_lay = QFormLayout(grp)
        grp_lay.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        grp_lay.setLabelAlignment(Qt.AlignLeft)
        grp_lay.setFormAlignment(Qt.AlignLeft | Qt.AlignTop)
        grp_lay.setContentsMargins(10, 12, 10, 10)
        grp_lay.setHorizontalSpacing(10)
        grp_lay.setVerticalSpacing(10)

        # Opção: Forçar recálculo com Excel (xlwings)
        from PySide6.QtWidgets import QCheckBox
        self.chk_xlwings = QCheckBox("Recalcular com Excel (xlwings)")
        self.chk_xlwings.setChecked(bool(self._cfg.get("force_xlwings", True)))
        grp_lay.addRow("", self.chk_xlwings)

        # Campo: Planilha de Folha
        self.ed_planilha = QLineEdit(self._cfg.get("folha_xlsx", ""))
        btn_browse = QPushButton("Procurar…")
        def _choose():
            fn, _ = QFileDialog.getOpenFileName(self, "Selecionar Planilha de Folha", "", "Planilhas (*.xlsx *.xlsm)")
            if fn:
                self.ed_planilha.setText(fn)
        btn_browse.clicked.connect(_choose)

        row = QHBoxLayout()
        row.setContentsMargins(0, 0, 0, 0)
        row.setSpacing(8)
        row.addWidget(self.ed_planilha, 1)
        row.addWidget(btn_browse, 0)

        row_w = QFrame(grp)
        row_w.setLayout(row)
        grp_lay.addRow("Planilha de Folha:", row_w)

        card_lay.addWidget(grp)
        root.addWidget(card)

        # Botões (Save/Cancel) alinhados à direita — mesmo padrão
        btns = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel, Qt.Horizontal, self)
        btns.button(QDialogButtonBox.Save).setText("Salvar")
        btns.button(QDialogButtonBox.Cancel).setText("Cancelar")

        def _save():
            self._cfg["folha_xlsx"] = self.ed_planilha.text().strip()
            self._cfg["force_xlwings"] = bool(self.chk_xlwings.isChecked())  # <- NOVO
            self.accept()

        btns.accepted.connect(_save)
        btns.rejected.connect(self.reject)
        root.addWidget(btns)

    def get_config(self) -> dict:
        return dict(self._cfg)

class ProfilePickerDialog(QDialog):
    """Dialogo simples para escolher o PERFIL com botões."""
    def __init__(self, perfis: list[str], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Selecionar perfil de importação")
        self.setModal(True)
        self.setStyleSheet(STYLE_SHEET)
        self.setFixedWidth(520)
        self.selected = None

        root = QVBoxLayout(self)
        root.setContentsMargins(14,14,14,14)
        root.setSpacing(10)

        lbl = QLabel("Escolha o perfil para importar:")
        root.addWidget(lbl)

        grid = QGridLayout()
        grid.setHorizontalSpacing(8)
        grid.setVerticalSpacing(8)

        if perfis:
            cols = 3
            for i, p in enumerate(perfis):
                btn = QPushButton(str(p))
                btn.setMinimumWidth(140)
                btn.clicked.connect(lambda _=False, name=str(p): self._choose(name))
                grid.addWidget(btn, i // cols, i % cols)
            root.addLayout(grid)
        else:
            # Fallback: se não descobrirmos perfis, mostra um campo para digitar
            info = QLabel("Nenhuma lista de perfis encontrada. Digite o nome do perfil:")
            edt = QLineEdit()
            edt.setPlaceholderText("Ex.: Perfil Financeiro")
            def _ok():
                self.selected = edt.text().strip()
                if self.selected:
                    self.accept()
            root.addWidget(info)
            root.addWidget(edt)
            ok = QPushButton("Importar")
            ok.clicked.connect(_ok)
            root.addWidget(ok)

        btns = QDialogButtonBox(QDialogButtonBox.Cancel, Qt.Horizontal, self)
        btns.rejected.connect(self.reject)
        root.addWidget(btns)

    def _choose(self, name: str):
        self.selected = name
        self.accept()

# ============================
# UI Principal (estilo do anexo)
# ============================
class AutomacaoFolhaUI(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("tab_automacao_folha")
        self.setStyleSheet(STYLE_SHEET)

        self.cfg = self._load_config()
        self.worker: FolhaWorker | None = None
        self.stat_total = 0
        self.stat_ok = 0
        self.stat_err = 0
        self._last_txt = ""

        root = QVBoxLayout(self)
        root.setContentsMargins(14,14,14,14)
        root.setSpacing(12)

        root.addWidget(self._build_header())
        top = QHBoxLayout(); top.setSpacing(12)
        top.addWidget(self._build_controls_card(), 3)
        top.addWidget(self._build_stats_card(), 2)
        root.addLayout(top)
        # Log ocupa o resto da tela (mesmo comportamento do Importar Dump)
        log_card = self._build_log_card()
        root.addWidget(log_card, 1)  # stretch=1

        footer = QLabel("🧾 Automação Folha — v1.0")
        footer.setAlignment(Qt.AlignCenter)
        footer.setStyleSheet("font-size:11px; color:#7F7F7F; padding-top:4px;")
        root.addWidget(footer)

    # ---------- Header ----------
    def _build_header(self) -> QFrame:
        header = QFrame()
        header.setStyleSheet("QFrame{border:1px solid #1e5a9c; border-radius:16px;}")
        lay = QHBoxLayout(header)
        lay.setContentsMargins(18,16,18,16)
        lay.setSpacing(14)

        icon = QLabel()
        if ICON_PATH.exists():
            pix = QPixmap(str(ICON_PATH)).scaled(44,44, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            icon.setPixmap(pix)
        else:
            icon.setText("🧾"); icon.setStyleSheet("font-size:34px; border:none;")
        lay.addWidget(icon, 0, Qt.AlignVCenter)

        title = QLabel("AUTOMAÇÃO FOLHA – TXT (12 colunas)")
        f = QFont(); f.setPointSize(20); f.setBold(True); title.setFont(f)
        subtitle = QLabel("Gere e importe TXT da folha com período, log e cancelamento.")

        title.setStyleSheet("border:none;"); subtitle.setStyleSheet("border:none;")
        title.setWordWrap(True)
        subtitle.setWordWrap(True)
        title.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        subtitle.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        box = QVBoxLayout(); box.addWidget(title); box.addWidget(subtitle)
        lay.addLayout(box, 1)

        btn_help = QToolButton(); btn_help.setText("❓ Ajuda")
        btn_help.clicked.connect(lambda: QMessageBox.information(
            self, "Ajuda",
            "• Defina a planilha em ⚙️ Configurar.\n"
            "• Escolha o período (MM/AAAA → MM/AAAA).\n"
            "• A: imóvel/fazenda; B: CPF; C: Funcionário; D: Líquido; F3/G3/H3: INSS/IRRF/FGTS.\n"
            "• Funcionários no 5º dia útil; tributos no 20º dia útil; valores sem pontuação.\n"
            "• Você pode CANCELAR durante a geração.\n\n"
            "FAZENDAS DOMÍNIO:\n\n"
            "CLEUBER:\n"
            "  4  - FAZ ALIANÇA\n"
            "  5  - FAZ PRIMAVERA\n"
            "  6  - ARMAZEM PRIMA\n"
            "  7  - FAZ FORMOSO\n"
            "  8  - FAZ L3\n"
            "  9  - FAZ FRUTACC\n"
            " 10  - FAZ GUARA\n\n"
            "GILSON:\n"
            " 14  - FAZ FORMIGA\n\n"
            "LUCAS:\n"
            " 11  - FAZ ALIANÇA 2\n\n"
            "ADRIANA:\n"
            " 12  - FAZ POUSO DA ANTA\n"
            " 13  - FAZ FRUTACC 3\n"
        ))
        btn_cfg = QToolButton(); btn_cfg.setText("⚙️ Configurar")
        btn_cfg.clicked.connect(self._open_config)

        btn_close = QToolButton(); btn_close.setText("✖ Fechar")
        btn_close.clicked.connect(self._close_self_tab)

        right = QHBoxLayout()
        right.setSpacing(8)
        right.addWidget(btn_help); right.addWidget(btn_cfg); right.addWidget(btn_close)
        lay.addLayout(right, 0)
        return header

    def _get_db_path(self) -> str | None:
        try:
            mw = self.window()
            db = getattr(mw, "db", None)
            if db and hasattr(db, "execute_query"):
                row = db.execute_query("PRAGMA database_list").fetchone()
                if row and len(row) >= 3:
                    return row[2]  # caminho do arquivo .sqlite
        except Exception:
            pass
        return None

    def _close_self_tab(self):
        parent = self.parent()
        while parent and not isinstance(parent, QTabWidget):
            parent = parent.parent()
        if parent:
            idx = parent.indexOf(self)
            if idx != -1:
                parent.removeTab(idx)
        else:
            self.close()

    # ---------- Controls Card ----------
    def _build_controls_card(self) -> QFrame:
        card = QFrame()
        card.setStyleSheet("QFrame{border:1px solid #1e5a9c; border-radius:12px;}")
        lay = QVBoxLayout(card); lay.setContentsMargins(14,12,14,12); lay.setSpacing(10)

        # Período compacto (combos editáveis + popup consistente)
        period = QHBoxLayout()
        period.setSpacing(8)
        period.addWidget(QLabel("Período:"))

        self.cmb_mes_ini = QComboBox()
        self.cmb_mes_fim = QComboBox()
        for m in range(1, 13):
            self.cmb_mes_ini.addItem(f"{m:02d}", m)
            self.cmb_mes_fim.addItem(f"{m:02d}", m)

        self.cmb_ano_ini = QComboBox()
        self.cmb_ano_fim = QComboBox()
        ano_atual = datetime.now().year
        # anos MUITO amplos (ex.: ~160 anos de janela)
        for y in range(ano_atual - 80, ano_atual + 81):
            self.cmb_ano_ini.addItem(str(y), y)
            self.cmb_ano_fim.addItem(str(y), y)

        # tornar editáveis para permitir digitação
        for cmb in (self.cmb_mes_ini, self.cmb_mes_fim, self.cmb_ano_ini, self.cmb_ano_fim):
            cmb.setEditable(True)
            cmb.setInsertPolicy(QComboBox.NoInsert)  # não “insere” novos itens
            cmb.lineEdit().setAlignment(Qt.AlignCenter)

        # validadores: MM (1..12) e AAAA (1900..9999)
        re_mes = QRegularExpression(r"^(0?[1-9]|1[0-2])$")
        val_mes = QRegularExpressionValidator(re_mes)
        val_ano = QIntValidator(1900, 9999)

        self.cmb_mes_ini.lineEdit().setPlaceholderText("MM")
        self.cmb_mes_fim.lineEdit().setPlaceholderText("MM")
        self.cmb_ano_ini.lineEdit().setPlaceholderText("AAAA")
        self.cmb_ano_fim.lineEdit().setPlaceholderText("AAAA")

        self.cmb_mes_ini.lineEdit().setValidator(val_mes)
        self.cmb_mes_fim.lineEdit().setValidator(val_mes)
        self.cmb_ano_ini.lineEdit().setValidator(val_ano)
        self.cmb_ano_fim.lineEdit().setValidator(val_ano)

        # largura e popup “normal”
        for cmb in (self.cmb_mes_ini, self.cmb_mes_fim, self.cmb_ano_ini, self.cmb_ano_fim):
            cmb.setSizeAdjustPolicy(QComboBox.AdjustToContentsOnFirstShow)
            cmb.setMinimumContentsLength(4)
            view = QListView()
            view.setUniformItemSizes(True)
            view.setStyleSheet(
                "QListView { background:#2B2F31; color:#E0E0E0; border:1px solid #1e5a9c; }"
                "QListView::item { padding:6px 10px; }"
                "QListView::item:selected { background:#1e5a9c; color:#FFFFFF; }"
            )
            cmb.setView(view)
            cmb.setMaxVisibleItems(12)

        # larguras equilibradas
        self.cmb_mes_ini.setFixedWidth(72)
        self.cmb_mes_fim.setFixedWidth(72)
        self.cmb_ano_ini.setFixedWidth(92)
        self.cmb_ano_fim.setFixedWidth(92)

        # valores padrão = mês/ano atuais
        self.cmb_mes_ini.setCurrentIndex(datetime.now().month - 1)
        self.cmb_mes_fim.setCurrentIndex(datetime.now().month - 1)
        idx_ini = self.cmb_ano_ini.findData(ano_atual)
        idx_fim = self.cmb_ano_fim.findData(ano_atual)
        self.cmb_ano_ini.setCurrentIndex(idx_ini if idx_ini >= 0 else self.cmb_ano_ini.count() - 1)
        self.cmb_ano_fim.setCurrentIndex(idx_fim if idx_fim >= 0 else self.cmb_ano_fim.count() - 1)

        period.addWidget(self.cmb_mes_ini)
        period.addWidget(self.cmb_ano_ini)
        arrow = QLabel("→")
        arrow.setStyleSheet("padding:0 6px;")
        period.addWidget(arrow)
        period.addWidget(self.cmb_mes_fim)
        period.addWidget(self.cmb_ano_fim)
        period.addStretch()
        lay.addLayout(period)

        # Ações (cores ajustadas)
        actions = QHBoxLayout(); actions.setSpacing(10)
        self.btn_gerar = QPushButton("📤 Ler Planilha → Gerar TXT")
        self.btn_gerar.setObjectName("success")
        self.btn_gerar.clicked.connect(self._start_worker)
        actions.addWidget(self.btn_gerar)

        self.btn_import = QPushButton("📥 Importar Lançamentos (TXT)")
        self.btn_import.clicked.connect(self._importar_txt)
        actions.addWidget(self.btn_import)

        self.btn_cancel = QPushButton("⛔ Cancelar")
        self.btn_cancel.setObjectName("danger")
        self.btn_cancel.setEnabled(False)
        self.btn_cancel.clicked.connect(self._cancel_worker)
        actions.addWidget(self.btn_cancel)

        actions.addStretch()

        self.btn_log_clear = QToolButton(); self.btn_log_clear.setText("🧹 Limpar Log")
        self.btn_log_clear.setStyleSheet("QToolButton{background:#0d1b3d; border:1px solid #1e5a9c; border-radius:8px; padding:6px 10px;} QToolButton:hover{background:#123764;}")
        self.btn_log_clear.clicked.connect(self._log_clear)
        actions.addWidget(self.btn_log_clear)

        self.btn_log_save = QToolButton(); self.btn_log_save.setText("💾 Salvar Log")
        self.btn_log_save.setStyleSheet("QToolButton{background:#0d1b3d; border:1px solid #1e5a9c; border-radius:8px; padding:6px 10px;} QToolButton:hover{background:#123764;}")
        self.btn_log_save.clicked.connect(self._log_save)
        actions.addWidget(self.btn_log_save)

        lay.addLayout(actions)
        return card

    # ---------- Stats Card ----------
    def _build_stats_card(self) -> QFrame:
        card = QFrame(); card.setObjectName("statsCard")
        card.setStyleSheet("#statsCard{border:1px solid #1e5a9c; border-radius:14px;} #statsCard *{border:none; background:transparent;}")
        lay = QVBoxLayout(card); lay.setContentsMargins(14,12,14,12); lay.setSpacing(6)

        title = QLabel("📊 Último Status da Sessão")
        f = QFont(); f.setPointSize(12); f.setBold(True); title.setFont(f)
        lay.addWidget(title)

        self.lbl_last_status = QLabel("Pronto")
        self.lbl_last_status_time = QLabel(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        self.lbl_last_status_time.setAlignment(Qt.AlignRight)

        row = QHBoxLayout(); row.addWidget(self.lbl_last_status); row.addStretch(); row.addWidget(self.lbl_last_status_time)
        lay.addLayout(row)

        chips = QHBoxLayout(); chips.setSpacing(10)
        self.lbl_stat_total = self._make_chip("Total", "#2B2F31", "#E0E0E0")
        self.lbl_stat_ok    = self._make_chip("Sucesso", "#183d2a", "#A7F3D0")
        self.lbl_stat_err   = self._make_chip("Erros", "#3b1f1f", "#FF6B6B")
        chips.addWidget(self.lbl_stat_total); chips.addWidget(self.lbl_stat_ok); chips.addWidget(self.lbl_stat_err); chips.addStretch()
        lay.addLayout(chips)
        return card

    def _make_chip(self, label: str, bg: str, fg: str) -> QLabel:
        w = QLabel(f"{label}: 0")
        w.setAlignment(Qt.AlignCenter)
        w.setStyleSheet(f"QLabel {{ background:{bg}; color:{fg}; border-radius:10px; padding:8px 12px; font-weight:600; }}")
        return w

    # ---------- Log Card (layout idêntico) ----------
    def _build_log_card(self) -> QFrame:
        card = QFrame(); card.setObjectName("logCard")
        card.setStyleSheet("#logCard{background:#212425; border:1px solid #1e5a9c; border-radius:10px;} "
                           "#logCard QLabel{border:none; background:transparent; color:#E0E0E0;}")
        lay = QVBoxLayout(card); lay.setContentsMargins(12,10,12,12); lay.setSpacing(8)
    
        title = QLabel("📝 Histórico")
        f = QFont(); f.setBold(True); f.setPointSize(12)
        title.setFont(f); title.setStyleSheet("padding:2px 6px;")
        lay.addWidget(title, alignment=Qt.AlignLeft)
    
        body = QFrame(); body.setObjectName("logBody")
        body.setStyleSheet("#logBody{background:#2B2F31; border:none; border-radius:8px;}")
        body_lay = QVBoxLayout(body); body_lay.setContentsMargins(12,12,12,12); body_lay.setSpacing(0)
    
        self.log = QTextEdit(readOnly=True)
        self.log.setFrameStyle(QFrame.NoFrame)

        # ocupa todo o espaço, igual ao Importar Dump
        self.log.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.log.setMinimumHeight(0)
        self.log.setMaximumHeight(16777215)

        # zera acolchoamentos para a 1ª linha não “nascer” no meio
        self.log.setStyleSheet(
            "QTextEdit{background:transparent; border:none; padding:0; margin:0;}"
            "QTextEdit::viewport{background:transparent; border:none; padding:0; margin:0;}"
        )
        self.log.document().setDocumentMargin(2)
        self.log.setViewportMargins(0, 0, 0, 0)
        self.log.setContentsMargins(0, 0, 0, 0)

        # mesmas opções de quebra que você já usa
        self.log.setLineWrapMode(QTextEdit.WidgetWidth)
        self.log.setWordWrapMode(QTextOption.WrapAnywhere)
        self.log.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.log.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        body_lay.addWidget(self.log, 1)

        # garante que a primeira mensagem apareça colada no topo
        self.log.clear()
        self.log.moveCursor(QTextCursor.Start)
        if self.log.verticalScrollBar():
            self.log.verticalScrollBar().setValue(self.log.verticalScrollBar().minimum())


        card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        lay.setStretch(0, 0)
        lay.setStretch(1, 1)
        body_lay.setStretch(0, 1)

        lay.addWidget(body)
        return card
    
    # ---------- Log helpers (mesma paleta/estilo) ----------
    def log_msg(self, message: str, msg_type: str = "info", update_status: bool = True):
        now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        palette = {
            "info":   {"emoji":"ℹ️","text":"#FFFFFF","accent":"#3A3C3D","weight":"500"},
            "success":{"emoji":"✅","text":"#A7F3D0","accent":"#2F7D5D","weight":"700"},
            "warning":{"emoji":"⚠️","text":"#FFFFFF","accent":"#8A6D3B","weight":"600"},
            "error":  {"emoji":"❌","text":"#FF6B6B","accent":"#7A2E2E","weight":"800"},
            "title":  {"emoji":"📌","text":"#FFFFFF","accent":"#1e5a9c","weight":"800"},
            "divider":{"emoji":"","text":"","accent":"#3A3C3D","weight":"400"},
        }
        if msg_type == "divider":
            self.log.append('<div style="border-top:1px solid #3A3C3D; margin:10px 0;"></div>')
            return

        # Mensagem crua (HTML pronto, sem timestamp nem moldura)
        if msg_type == "raw":
            self.log.append(message)
            sb = self.log.verticalScrollBar()
            if sb: sb.setValue(sb.maximum())
            return

        p = palette.get(msg_type, palette["info"])
        html = (
            f'<div style="border-left:3px solid {p["accent"]}; padding:6px 10px; margin:2px 0;'
            f'word-break: break-word; overflow-wrap: anywhere; white-space: normal;">'
            f'<span style="opacity:.7; font-family:monospace;">[{now}]</span>'
            f' <span style="margin:0 6px 0 8px;">{p["emoji"]}</span>'
            f'<span style="color:{p["text"]}; font-weight:{p["weight"]};">{message}</span>'
            f'</div>'
        )
        self.log.append(html)
        # mantém no topo quando há só 1ª/2ª linha; senão, rola para o fim
        sb = self.log.verticalScrollBar()
        if sb:
            if self.log.document().blockCount() <= 2:
                sb.setValue(0)
            else:
                sb.setValue(sb.maximum())

        if update_status:
            # status curto (sem paths enormes): tira caminhos e limita tamanho
            def _shorten_for_status(text: str, maxlen: int = 140) -> str:
                # remove/encurta qualquer coisa que pareça path absoluto
                text = re.sub(r'([A-Za-z]:\\\\[^\\s]+|/[^\\s]+)', lambda m: os.path.basename(m.group(0)), text)
                return (text[:maxlen-3] + '...') if len(text) > maxlen else text
    
            short = _shorten_for_status(message)
            self.lbl_last_status.setText(short)
            self.lbl_last_status_time.setText(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    

        self.lbl_last_status.setWordWrap(True)
        self.lbl_last_status.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        self.lbl_last_status_time.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Preferred)
        self.lbl_last_status_time.setMinimumWidth(120)

    def _log_clear(self):
        self.log.clear()                      # sem HTML “fantasma”
        self.log.moveCursor(QTextCursor.Start)
        if self.log.verticalScrollBar():
            self.log.verticalScrollBar().setValue(self.log.verticalScrollBar().minimum())
        self.log_msg("Log limpo.", "info")


    def _log_save(self):
        try:
            out_dir = Path(__file__).parent / "logs"
            out_dir.mkdir(exist_ok=True, parents=True)
            fname = out_dir / f"folha_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            with open(fname, "w", encoding="utf-8") as f:
                f.write(self.log.toPlainText())
            # caminho só no LOG; status curto
            self.log_msg(f"Log salvo em: {fname}", "success", update_status=False)
            self.lbl_last_status.setText("Log salvo.")
            self.lbl_last_status_time.setText(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        except Exception as e:
            self.log_msg(f"Falha ao salvar log: {e}", "error")

    # ---------- Config persistente ----------
    def _cfg_path(self) -> Path:
        p = Path(__file__).parent / "json"
        p.mkdir(parents=True, exist_ok=True)
        return p / "config_folha.json"

    def _load_config(self) -> dict:
        cfg_file = self._cfg_path()
        if cfg_file.exists():
            try:
                return json.loads(cfg_file.read_text(encoding="utf-8"))
            except Exception:
                return {}
        return {}

    def _save_config(self):
        try:
            self._cfg_path().write_text(json.dumps(self.cfg, indent=4, ensure_ascii=False), encoding="utf-8")
            self.log_msg("Configurações salvas.", "success")
        except Exception as e:
            self.log_msg(f"Erro ao salvar config: {e}", "error")

    def _open_config(self):
        dlg = ConfigDialog(self.cfg, self)
        if dlg.exec() == QDialog.Accepted:
            self.cfg.update(dlg.get_config())
            try:
                p = Path(__file__).parent / "json"
                p.mkdir(parents=True, exist_ok=True)
                cfg_file = p / "config_folha.json"
                cfg_file.write_text(json.dumps(self.cfg, indent=4, ensure_ascii=False), encoding="utf-8")
                # caminho só no LOG; status curto
                self.log_msg(f"Configurações salvas em: {cfg_file}", "success", update_status=False)
                self.lbl_last_status.setText("Configurações salvas.")
                self.lbl_last_status_time.setText(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
            except Exception as e:
                self.log_msg(f"Erro ao salvar configurações: {e}", "error")
        
        

    # ---------- Fluxo Geração/Importação ----------
    def _start_worker(self):
        planilha = (self.cfg or {}).get("folha_xlsx", "").strip()
        if not planilha or not Path(planilha).exists():
            QMessageBox.warning(self, "Planilha", "Defina a planilha em ⚙️ Configurar.")
            return

        # lê mês/ano permitindo texto digitado
        def _parse_mes(cmb: QComboBox) -> int:
            t = (cmb.currentText() or "").strip()
            m = re.match(r"^(\d{1,2})$", t)
            if m:
                mi = int(m.group(1))
                if 1 <= mi <= 12:
                    return mi
            d = cmb.currentData()
            if isinstance(d, int):
                return d
            return datetime.now().month
        
        def _parse_ano(cmb: QComboBox) -> int:
            t = (cmb.currentText() or "").strip()
            m = re.match(r"^(\d{4})$", t)
            if m:
                yi = int(m.group(1))
                if 1900 <= yi <= 9999:
                    return yi
            d = cmb.currentData()
            if isinstance(d, int):
                return d
            return datetime.now().year
        
        mes_ini = _parse_mes(self.cmb_mes_ini)
        ano_ini = _parse_ano(self.cmb_ano_ini)
        mes_fim = _parse_mes(self.cmb_mes_fim)
        ano_fim = _parse_ano(self.cmb_ano_fim)
        
        ini = f"{mes_ini:02d}/{ano_ini}"
        fim = f"{mes_fim:02d}/{ano_fim}"

        self._update_stats(0,0,0)
        self.btn_cancel.setEnabled(True)
        self.log_msg(f"Geração do TXT iniciada ({ini} → {fim})…", "title")

        # depois:
        db_path = self._get_db_path()
        force_xlwings = bool((self.cfg or {}).get("force_xlwings", True))
        self.worker = FolhaWorker(planilha, ini, fim, db_path=db_path, parent=self, force_xlwings=force_xlwings)
        self.worker.log_sig.connect(self._on_worker_log)
        self.worker.stats_sig.connect(self._update_stats)
        self.worker.finished_sig.connect(self._on_worker_finished)
        self.worker.start()

    def _on_worker_log(self, msg: str, level: str):
        # Se a mensagem envolve salvar/gerar/importar com caminho, joga só no LOG
        if ("TXT gerado:" in msg) or ("Log salvo em:" in msg) or ("Importado no sistema:" in msg):
            self.log_msg(msg, level, update_status=False)
            # status curto e sem path
            if "TXT gerado:" in msg:
                self.lbl_last_status.setText("TXT gerado.")
            elif "Log salvo em:" in msg:
                self.lbl_last_status.setText("Log salvo.")
            elif "Importado no sistema:" in msg:
                self.lbl_last_status.setText("Importado no sistema.")
            self.lbl_last_status_time.setText(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
            return

        # Demais mensagens: seguem padrão (com proteção interna para não “esticar”)
        self.log_msg(msg, level, update_status=True)

    def _cancel_worker(self):
        if self.worker and self.worker.isRunning():
            self.log_msg("Solicitando cancelamento…", "warning")
            self.worker.cancel()
        else:
            self.log_msg("Nenhum processo em execução para cancelar.", "info")

    def _on_worker_finished(self, status: str, path_txt: str):
        self.btn_cancel.setEnabled(False)
        if status == "Concluído":
            self._last_txt = path_txt
            # caminho completo só no LOG; status curto
            if path_txt:
                self.log_msg(f"TXT gerado: {path_txt}", "success", update_status=False)
            else:
                self.log_msg("TXT gerado.", "success", update_status=False)

            self.lbl_last_status.setText("TXT gerado.")
            self.lbl_last_status_time.setText(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))

            # perguntar se deseja importar agora (via arquivo temporário + seleção de perfil)
            resp = QMessageBox.question(
                self, "Importar agora?",
                "Deseja importar o TXT gerado agora (via arquivo temporário)?",
                QMessageBox.Yes | QMessageBox.No
            )
            if resp == QMessageBox.Yes:
                self._pick_and_import_temp(path_txt)

            
        elif status == "Vazio":
            QMessageBox.information(self, "TXT", "Nenhuma linha foi gerada para o período.")
        elif status == "Cancelado":
            QMessageBox.information(self, "TXT", "Processo cancelado.")
        else:
            QMessageBox.warning(self, "TXT", f"Finalizado com status: {status}")


    def _importar_txt(self):
        path = self._last_txt
        if not path or not os.path.exists(path):
            # permitir seleção manual
            path, _ = QFileDialog.getOpenFileName(self, "Selecione o TXT", "", "TXT (*.txt)")
            if not path:
                return
        try:
            mw = self.window()
            if not hasattr(mw, "_import_lancamentos_txt"):
                raise RuntimeError("A janela principal não expõe _import_lancamentos_txt.")
            mw._import_lancamentos_txt(path)
            if hasattr(mw, "carregar_lancamentos"): mw.carregar_lancamentos()
            if hasattr(mw, "dashboard"):
                try: mw.dashboard.load_data()
                except Exception: pass
            self.log_msg(f"Importado no sistema: {os.path.basename(path)}", "success", update_status=False)
            self.lbl_last_status.setText("Importado no sistema.")
            self.lbl_last_status_time.setText(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))

        except Exception as e:
            self.log_msg(f"Erro ao importar: {e}", "error")

    def _importar_txt_temp(self, src_path: str):
        """Copia o TXT gerado para um arquivo temporário e importa sem pedir caminho."""
        try:
            if not src_path or not os.path.exists(src_path):
                QMessageBox.warning(self, "TXT", "Arquivo gerado não encontrado para importação.")
                return

            base = os.path.basename(src_path)

            # cria um arquivo temporário .txt e copia o conteúdo
            fd, tmp_path = tempfile.mkstemp(prefix="folha_", suffix=".txt")
            os.close(fd)
            shutil.copy2(src_path, tmp_path)

            # importa no sistema pelo caminho temporário (sem diálogo)
            mw = self.window()
            if not hasattr(mw, "_import_lancamentos_txt"):
                raise RuntimeError("A janela principal não expõe _import_lancamentos_txt.")

            self.log_msg(f"Importando (via temporário): {base}", "info", update_status=False)
            mw._import_lancamentos_txt(tmp_path)

            if hasattr(mw, "carregar_lancamentos"):
                mw.carregar_lancamentos()
            if hasattr(mw, "dashboard"):
                try:
                    mw.dashboard.load_data()
                except Exception:
                    pass

            self.log_msg(f"Importado no sistema (via temporário): {base}", "success", update_status=False)
            self.lbl_last_status.setText("Importado no sistema.")
            self.lbl_last_status_time.setText(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))

        except Exception as e:
            self.log_msg(f"Erro ao importar (temporário): {e}", "error")
        finally:
            # tenta remover o temporário (opcional: comente se quiser manter)
            try:
                if 'tmp_path' in locals() and os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception:
                pass

    def _discover_perfis(self) -> list[str]:
        """Descobre a lista de perfis a partir da janela principal/config."""
        mw = self.window()

        # 0) Tenta extrair do combo 'profile_selector' da MainWindow (ordem da UI)
        try:
            from PySide6.QtWidgets import QComboBox
        except Exception:
            QComboBox = None

        if hasattr(mw, "profile_selector") and (
            QComboBox is None or isinstance(mw.profile_selector, QComboBox)
        ):
            try:
                items = [mw.profile_selector.itemText(i) for i in range(mw.profile_selector.count())]
                # Converte nomes internos -> rótulos amigáveis quando possível
                inv = {v: k for k, v in PROFILE_MAP.items()}
                display = [inv.get(txt, txt) for txt in items]
                # Garante que venham exatamente os 4 pedidos, na ordem desejada, se existirem na UI
                ordered = [name for name in PROFILE_DISPLAY if PROFILE_MAP.get(name) in items]
                return ordered or display
            except Exception:
                pass

        # 1) Atributos comuns
        for attr in ("perfis", "profiles", "lista_perfis", "profiles_list"):
            lst = getattr(mw, attr, None)
            if isinstance(lst, (list, tuple)) and lst:
                return [str(x) for x in lst]

        # 2) Métodos comuns
        for meth in ("get_perfis", "listar_perfis", "get_profiles", "list_profiles"):
            if hasattr(mw, meth):
                try:
                    lst = getattr(mw, meth)()
                    if isinstance(lst, (list, tuple)) and lst:
                        return [str(x) for x in lst]
                except Exception:
                    pass

        # 3) Config opcional (json/config_folha.json)
        lst = (self.cfg or {}).get("perfis", [])
        if isinstance(lst, list) and lst:
            return [str(x) for x in lst]

        # 4) Fallback garantido: 4 perfis fixos (como você pediu)
        return list(PROFILE_DISPLAY)

    def _pick_and_import_temp(self, src_path: str):
        """Abre o seletor de PERFIL e importa via temporário para o perfil escolhido."""
        perfis = self._discover_perfis()
        dlg = ProfilePickerDialog(perfis, self)
        if dlg.exec() == QDialog.Accepted and dlg.selected:
            self._importar_txt_temp_to_profile(src_path, dlg.selected)
        else:
            self.log_msg("Importação cancelada pelo usuário (perfil não selecionado).", "warning")
    
    def _importar_txt_temp_to_profile(self, src_path: str, perfil: str):
        """Importa usando um arquivo temporário aplicando o PERFIL selecionado,
        mas ANTES filtra DUPLICADOS (Folha) consultando o Supabase."""
        import os, re, shutil, tempfile
        from pathlib import Path
        from PySide6.QtWidgets import QMessageBox
    
        def _digits(s: str) -> str:
            return re.sub(r"\D", "", s or "")
    
        def _parse_cent(v: str) -> float:
            s = re.sub(r"\D", "", v or "")
            return (int(s) / 100.0) if s else 0.0
    
        try:
            if not src_path or not os.path.exists(src_path):
                QMessageBox.warning(self, "TXT", "Arquivo gerado não encontrado para importação.")
                return
    
            base = os.path.basename(src_path)
            internal = PROFILE_MAP.get(perfil, perfil)
    
            mw = self.window()
            if not hasattr(mw, "_import_lancamentos_txt"):
                raise RuntimeError("A janela principal não expõe _import_lancamentos_txt.")
    
            # Lê todas as linhas do TXT (12 colunas: data|imóvel|conta|num_doc|tipo_doc|hist|cpf|tipo|ent|sai|saldo|nat)
            with open(src_path, "r", encoding="utf-8") as f:
                linhas = [ln.strip() for ln in f if ln.strip()]
    
            # -------- DUP CHECK (Folha) via Supabase ----------
            dups_log = []
            sem_dup = []
    
            # Garantir que estamos no perfil certo ANTES de consultar o BD
            switched = False
            for setter in ("selecionar_perfil", "set_perfil", "set_profile", "setPerfil", "seleciona_perfil", "switch_profile"):
                if hasattr(mw, setter):
                    try:
                        getattr(mw, setter)(internal)
                        switched = True
                        break
                    except Exception:
                        pass
            if not switched:
                for attr in ("perfil_atual", "perfil", "profile", "current_profile"):
                    if hasattr(mw, attr):
                        try:
                            setattr(mw, attr, internal)
                            break
                        except Exception:
                            pass
                        
            # Checagem de duplicidade: (mesmo CPF + mesma Data) e confere valor (saída/entrada)
            for ln in linhas:
                parts = ln.split("|")
                if len(parts) < 12:
                    # linha imprevista: deixa passar (não bloqueia)
                    sem_dup.append(ln)
                    continue
                
                data_br   = (parts[0] or "").strip()  # DD-MM-AAAA
                data_str  = data_br.replace("-", "/")  # DD/MM/AAAA, que é como gravamos em 'lancamento.data'
                historico = parts[5] if len(parts) > 5 else ""
                cpf_cnpj  = _digits(parts[6] if len(parts) > 6 else "")
                # valor (Folha costuma vir em SAÍDA)
                v_ent = _parse_cent(parts[8] if len(parts) > 8 else "")
                v_sai = _parse_cent(parts[9] if len(parts) > 9 else "")
                valor_alvo = v_sai if v_sai > 0 else v_ent
    
                # Sem CPF/Data não tem como checar
                if not cpf_cnpj or not data_str:
                    sem_dup.append(ln)
                    continue
                
                exists = False
                try:
                    # 1) encontra participante
                    pid_rows = (mw.db.sb.table("participante")
                                .select("id,nome")
                                .eq("cpf_cnpj", cpf_cnpj)
                                .limit(1)
                                .execute().data) or []
                    if pid_rows:
                        pid = int(pid_rows[0]["id"])
                        nome = pid_rows[0].get("nome") or ""
    
                        # 2) pega candidatos deste participante nesta data
                        cand = (mw.db.sb.table("lancamento")
                                .select("id,data,tipo_doc,historico,valor_entrada,valor_saida")
                                .eq("id_participante", pid)
                                .eq("data", data_str)          # data exata
                                .order("id", desc=True)
                                .limit(200)
                                .execute().data) or []
    
                        # 3) regra: se for folha (tipo_doc=5) OU histórico começa com "FOLHA"
                        alvo = round(float(valor_alvo or 0), 2)
                        for c in cand:
                            tdoc = int(c.get("tipo_doc") or 0)
                            ve   = round(float(c.get("valor_entrada") or 0), 2)
                            vs   = round(float(c.get("valor_saida") or 0), 2)
                            vcand = vs if vs > 0 else ve

                            hist_lin = (historico or "").strip().upper()
                            hist_db  = (c.get("historico") or "").strip().upper()

                            ref_lin = _extract_ref(hist_lin)  # "REF. 02/2025"
                            ref_db  = _extract_ref(hist_db)
                            same_ref = (ref_lin and ref_db and ref_lin == ref_db) or (not ref_lin and not ref_db and hist_lin == hist_db)

                            if (tdoc == 5 or hist_db.startswith("FOLHA")) and abs(vcand - alvo) < 0.01 and same_ref:
                                exists = True
                                # >>> DIAGNÓSTICO: logar QUAL linha do banco bateu
                                self._emit(
                                    f"↩️ DUP no banco • CPF {cpf_cnpj} • {data_str} • "
                                    f"R$ {alvo:,.2f} • REF {(_extract_ref(historico) or '?')} • "
                                    f"match(id={c.get('id')}, tipo_doc={tdoc}, hist='{c.get('historico')}', "
                                    f"v_ent={c.get('valor_entrada')}, v_sai={c.get('valor_saida')})",
                                    "warning"
                                )
                                dups_log.append({
                                    "cpf": cpf_cnpj,
                                    "data": data_str,
                                    "valor": alvo,
                                    "nome": nome or _extract_name_from_historico(historico) or "",
                                    "hist": historico
                                })
                                break

                except Exception:
                    exists = False
    
                if exists:
                    # não vai para o arquivo que será importado
                    continue
                else:
                    sem_dup.append(ln)
    
            # -------- Importação (apenas não duplicados) ----------
            if not sem_dup:
                self._log_section("DUPLICADOS (Folha)", "🔁")
                self.log.append("<div style='font-family:monospace; color:#ffd166; text-align:center; margin:2px 0 6px 0;'>TODOS OS REGISTROS JÁ EXISTEM (CPF+DATA+VALOR)</div>")
                # Tabela de duplicados
                if dups_log:
                    hdr = ("CPF/CNPJ".ljust(14) + " │ " +
                           "DATA".ljust(10) + " │ " +
                           "VALOR".rjust(12) + " │ " +
                           "NOME/HISTÓRICO")
                    self.log.append("<div style='font-family:monospace;'><b style='color:#ffd166;'>"+hdr+"</b></div>")
                    self.log.append("<div style='font-family:monospace; color:#554a08;'>"
                                    "──────────────┼──────────┼────────────┼────────────────────────────────────────</div>")
                    for d in dups_log:
                        cpf   = f"{(d['cpf'] or '')[:14]:<14}"
                        data  = f"{(d['data'] or ''):<10}"
                        valor = f"{d['valor']:>12.2f}"
                        nomh  = (d["nome"] or d["hist"] or "")[:44]
                        line  = f"{cpf} │ {data} │ {valor} │ {nomh}"
                        self.log.append(f"<span style='font-family:monospace; color:#ffd166;'>{line}</span>")
                QMessageBox.information(self, "Concluído", "Nenhum lançamento novo — todos eram duplicados (Folha).")
                return
    
            # grava um TEMP TXT só com não-duplicados e importa
            fd, tmp_path = tempfile.mkstemp(prefix="folha_", suffix=".txt")
            os.close(fd)
            try:
                with open(tmp_path, "w", encoding="utf-8") as tf:
                    tf.write("\n".join(sem_dup))
    
                self.log_msg(f"Importando (via temporário) no perfil '{perfil}': {base}", "info", update_status=False)
    
                imported = False
                try:
                    mw._import_lancamentos_txt(tmp_path, internal)   # posicional
                    imported = True
                except TypeError:
                    try:
                        mw._import_lancamentos_txt(tmp_path, perfil=internal)  # nomeado
                        imported = True
                    except TypeError:
                        pass
                    
                if not imported:
                    # fallback: já trocamos perfil lá em cima; importa sem arg
                    mw._import_lancamentos_txt(tmp_path)
                    imported = True
    
                # pós-import
                if hasattr(mw, "carregar_lancamentos"):
                    mw.carregar_lancamentos()
                if hasattr(mw, "dashboard"):
                    try:
                        mw.dashboard.load_data()
                    except Exception:
                        pass
                    
                self.log_msg(f"Importado no sistema (perfil '{perfil}', via temporário): {base}", "success", update_status=False)
                self.lbl_last_status.setText("Importado no sistema.")
                self.lbl_last_status_time.setText(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    
                # ------- BLOCO FINAL: Tabela de DUPLICADOS (se houve) -------
                if dups_log:
                    self._log_section("DUPLICADOS (Folha)", "🔁")
                    self.log.append("<div style='font-family:monospace; color:#ffd166; text-align:center; margin:2px 0 6px 0;'>MESMO CPF + MESMA DATA (+ VALOR)</div>")
                    hdr = ("CPF/CNPJ".ljust(14) + " │ " +
                           "DATA".ljust(10) + " │ " +
                           "VALOR".rjust(12) + " │ " +
                           "NOME/HISTÓRICO")
                    self.log.append("<div style='font-family:monospace;'><b style='color:#ffd166;'>"+hdr+"</b></div>")
                    self.log.append("<div style='font-family:monospace; color:#554a08;'>"
                                    "──────────────┼──────────┼────────────┼────────────────────────────────────────</div>")
                    for d in dups_log:
                        cpf   = f"{(d['cpf'] or '')[:14]:<14}"
                        data  = f"{(d['data'] or ''):<10}"
                        valor = f"{d['valor']:>12.2f}"
                        nomh  = (d["nome"] or d["hist"] or "")[:44]
                        line  = f"{cpf} │ {data} │ {valor} │ {nomh}"
                        self.log.append(f"<span style='font-family:monospace; color:#ffd166;'>{line}</span>")
                    self.log.append("<div style='text-align:center;color:#2e3d56;font-family:monospace;'>======================</div>")
    
            finally:
                try:
                    if 'tmp_path' in locals() and os.path.exists(tmp_path):
                        os.remove(tmp_path)
                except Exception:
                    pass
                
        except Exception as e:
            self.log_msg(f"Erro ao importar (perfil '{perfil}'): {e}", "error")
    
    # ---------- Stats ----------
    def _update_stats(self, total: int, ok: int, err: int):
        self.stat_total, self.stat_ok, self.stat_err = total, ok, err
        self.lbl_stat_total.setText(f"Total: {total}")
        self.lbl_stat_ok.setText(f"Sucesso: {ok}")
        self.lbl_stat_err.setText(f"Erros: {err}")

# ============================
# Instalação em MainWindow (opcional)
# ============================
def install_in_mainwindow(main_win):
    # evita duplicidade...
    ui = AutomacaoFolhaUI(main_win)   # <- passe o main_win como parent
    ui.setObjectName('tab_automacao_folha')
    main_win.tabs.addTab(ui, "Automação Folha")
    main_win.tabs.setCurrentWidget(ui)
