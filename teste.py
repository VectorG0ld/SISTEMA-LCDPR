
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Automação NFS-e Digitalizadas — IA IMG64 v2 (sem OCR/Layouts, prompt avançado)
-------------------------------------------------------------------------------
Melhorias desta versão v2:
- Prompt reforçado (multi-instruções) para a IA:
  • Priorizar "data de vencimento/recebimento" (inclusive manuscrita/caneta).
  • NUNCA retornar CNPJ da prefeitura/município/SEFAZ — sempre do PRESTADOR/EMISSOR.
  • Ignorar datas de 2024; extrair somente datas em 2025 (se só houver 2024, deixar vazio).
  • Compreender "sinônimos" e indícios de vencimento: 'vencimento', 'vcto', 'pagar até',
    'recebimento', 'data de pagamento', 'até', carimbo à mão, observação em caneta, etc.
- Lógica local para:
  • Escolha final da data: prioriza vencimento 2025; se ausente, emissão 2025; senão vazio.
  • Higienizar CNPJ e invalidar CNPJ de prefeitura caso a IA falhe (regras anti-prefeitura).
  • Melhor logging (fonte das datas e observações da IA).

Mantido do fluxo original:
- Varredura por lotes (pasta base + subpastas), 1 planilha 'lancamentos.xlsx' por lote.
- UI com botões: Separar, Gerar Planilha, Importar TXT.
- Planilha com as mesmas colunas e estilos (A–L).
- Log detalhado por lote.

Requisitos exemplo:
    pip install pdfplumber pillow openpyxl openai PySide6
"""

from __future__ import annotations

import os
import re
import sys
import io
import glob
import base64
import json
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional
from datetime import datetime
import time

import pdfplumber
from PIL import Image

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from PySide6.QtCore import Qt, QThread, Signal, QObject
from PySide6.QtGui import QIcon
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFileDialog, QLineEdit,
    QTextEdit, QSpinBox, QGridLayout, QFrame, QMessageBox, QApplication
)

from pathlib import Path

# ================== HELPERS BÁSICOS ==================

def _find_in_this_or_parent(filename: str) -> Path:
    here = Path(__file__).resolve().parent
    cand1 = here / filename
    cand2 = here.parent / filename
    if cand1.exists():
        return cand1
    if cand2.exists():
        return cand2
    return here / filename  # retorna caminho padrão (pode não existir)

def _strip_accents(s: str) -> str:
    try:
        import unicodedata
        return "".join(ch for ch in unicodedata.normalize("NFD", s or "") if unicodedata.category(ch) != "Mn")
    except Exception:
        return s or ""

def _norm_simple(s: str) -> str:
    s = _strip_accents(s).lower()
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def _unique_path(base_path: str) -> str:
    if not os.path.exists(base_path):
        return base_path
    root, ext = os.path.splitext(base_path)
    i = 2
    while True:
        cand = f"{root} ({i}){ext}"
        if not os.path.exists(cand):
            return cand
        i += 1

def _cancelled() -> bool:
    try:
        cb = globals().get("is_cancelled", None)
        return bool(cb and callable(cb) and cb())
    except Exception:
        return False

# ================== CONFIG ==================

BASE_DIR = r"C:\Users\conta\Downloads"
DPI = 300
PRINT_ONLY_IA = True
WRITE_TXT_OUTPUT = False

OPENAI_MODEL = os.environ.get("OPENAI_MODEL", "gpt-4o-mini")

# ================== MAPEAMENTOS & REGEX (mantidos/essenciais) ==================

# (exemplo reduzido — substitua/complete com o seu mapeamento completo)
FAZENDA2COD = {
    "ALIANCA": "09",
    "ARM PRIMAVERA": "02",
    "BOA GRANDE": "11",
    "ESTRELA": "04",
    "FRUTACC": "13",
    "FRUTACC III": "12",
    "L3": "16",
    "PRIMAVERA": "01",
    "PRIMAVERA RETIRO": "06",
    "RETIRO PRIMAVERA": "06",
    "RIO BONITO I": "07",
    "RIO BONITO III": "15",
    "RIO BONITO IV": "05",
    "RIO NEGRO": "14",
    "SEDE": "10",
    "SITIO": "03",
    "SITIO ": "03",
    "TOMBO": "08",
    "L2": "17",
    "L4": "18",
}

CODIGOS_CIDADES = {
    "GOIANIA": "5208707",
    "ANAPOLIS": "5201108",
    "RIALMA": "5218608",
    "NOVA GLORIA": "5214904",
    "LAGOA DA CONFUSAO": "1711902",
    "TROMBAS": "5222204",
    "FORMOSO DO ARAGUAIA": "1708205",
    "DUERE": "1707306",
    "BALSAS": "2101400",
    "TASSO FRAGOSO": "2112005",
}

RE_VALOR = re.compile(r"(\d{1,3}(\.\d{3})*,\d{2}|\d+,\d{2})")
RE_DATEPT  = re.compile(r"\b(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\b")
RE_CNPJ    = re.compile(r"\b\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}\b")

def only_digits(s: str) -> str:
    return re.sub(r"\D+", "", s or "")

def valor_to_centavos(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return ""
    s = s.replace(".", "").replace(",", ".")
    try:
        v = float(s)
        return str(int(round(v * 100)))
    except Exception:
        return ""

def date_slash_to_dash_2025_only(s: str) -> str:
    """Converte dd/mm/aaaa ou dd-mm-aaaa para yyyy-mm-dd, retornando vazio se ano != 2025."""
    s = (s or "").strip()
    if not s:
        return ""
    m = RE_DATEPT.search(s)
    if not m:
        return ""
    d, mth, y = m.groups()
    if len(y) == 2:
        y = ("20" + y) if int(y) <= 30 else ("19" + y)
    if y != "2025":
        return ""
    return f"{y}-{int(mth):02d}-{int(d):02d}"

def choose_final_date_2025(data_venc: str, data_emissao: str) -> str:
    d1 = date_slash_to_dash_2025_only(data_venc)
    if d1:
        return d1
    d2 = date_slash_to_dash_2025_only(data_emissao)
    return d2

# ================== PREFEITURA / NFS-E EMITIDA POR MUNICÍPIO ==================

PREFEITURA_TOKENS = [
    "prefeitura", "município de", "municipio de", "pref. municipal", "prefeitura municipal",
    "sefaz", "secretaria da fazenda", "nota fiscal de serviço eletrônica", "nfs-e"
]

def looks_like_prefeitura_razao(razao: str) -> bool:
    t = _norm_simple(razao or "").lower()
    return any(tok in t for tok in PREFEITURA_TOKENS)

def sanitize_cnpj_prestador(razao: str, cnpj: str) -> str:
    if looks_like_prefeitura_razao(razao):
        return "INVALIDO"
    if not cnpj:
        return ""
    m = RE_CNPJ.search(cnpj)
    return only_digits(m.group(0)) if m else "INVALIDO"

# ================== OPENAI — IA com IMG base64 e prompt avançado ==================

PROMPT_IA = (
    "ATUAÇÃO: Você extrai dados de Notas Fiscais de Serviço (NFS-e) brasileiras a partir de IMAGENS (inclusive rabiscos/manuscritos).\n"
    "RESPONDA **APENAS** EM JSON VÁLIDO (um único objeto). Campos obrigatórios:\n"
    "{\n"
    '  "tipo": "nota fiscal" | "recebimento" | "outro",\n'
    '  "numero": "apenas dígitos se possível",\n'
    '  "data_emissao": "dd/mm/aaaa",\n'
    '  "data_vencimento": "dd/mm/aaaa",\n'
    '  "municipio": "município do tomador ou local do serviço (sem UF)",\n'
    '  "prestador": "razão social do PRESTADOR/EMISSOR",\n'
    '  "cnpj_prestador": "CNPJ do PRESTADOR/EMISSOR (nunca o da prefeitura)",\n'
    '  "valor_total": "formato brasileiro, ex.: 12.345,67",\n'
    '  "fontes": { "data_vencimento": "impresso|manuscrito|ambos|indefinido", "observacoes": "texto breve" }\n'
    "}\n\n"
    "REGRAS CRÍTICAS:\n"
    "1) PRIORIZE DATA DE VENCIMENTO/RECEBIMENTO (também escrita à mão/caneta). "
    "   Palavras/indícios: 'vencimento', 'vcto', 'pagar até', 'pagamento', 'recebimento', 'até', carimbo manuscrito.\n"
    "2) **NÃO** retorne CNPJ da prefeitura/município/SEFAZ. Retorne sempre o **CNPJ do PRESTADOR/EMISSOR**.\n"
    "3) **Datas de 2024 DEVEM SER IGNORADAS**. Extraia **somente datas de 2025**. "
    "   Se apenas houver datas de 2024, deixe o campo correspondente **vazio**.\n"
    "4) Se não encontrar um campo com segurança, deixe-o vazio.\n\n"
    "DICAS:\n"
    "- O número da nota normalmente vem como 'NFS-e nº', 'Número da NFS-e', 'RPS', etc. "
    "  Capturar apenas os dígitos quando possível.\n"
    "- O município não deve conter UF (ex.: 'GO', 'TO'). Apenas o nome do município.\n"
    "- Se aparecer logotipo ou timbre da prefeitura, **NÃO** use o CNPJ que apareça perto disso; "
    "  procure o CNPJ do prestador na área do cabeçalho/rodapé/identificação do emissor.\n"
)

def _b64_of_image(path: str) -> str:
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")

def ia_extract_from_images(image_paths: List[str], api_key: str, model: str = OPENAI_MODEL) -> Dict[str, str]:
    from openai import OpenAI
    client = OpenAI(api_key=api_key)

    content = [{"type": "text", "text": PROMPT_IA}]
    for p in image_paths:
        content.append({"type": "image_url", "image_url": {"url": f"data:image/png;base64,{_b64_of_image(p)}", "detail": "high"}})

    try:
        resp = client.chat.completions.create(
            model=model or "gpt-4o-mini",
            response_format={"type": "json_object"},
            temperature=0,
            messages=[{"role": "user", "content": content}],
            max_tokens=1200,
        )
        raw = (resp.choices[0].message.content or "").strip()
        data = json.loads(raw)
        out = {
            "tipo": data.get("tipo", "") or "",
            "numero": data.get("numero", "") or "",
            "data_emissao": data.get("data_emissao", "") or "",
            "data_vencimento": data.get("data_vencimento", "") or "",
            "municipio": data.get("municipio", "") or "",
            "prestador": data.get("prestador", "") or "",
            "cnpj_prestador": data.get("cnpj_prestador", "") or "",
            "valor_total": data.get("valor_total", "") or "",
            "fontes": data.get("fontes", {}) or {},
        }
        return out
    except Exception as e:
        return {"_ia_erro": str(e)}

# ================== PDF → PNG ==================

def convert_pdf_to_pngs(pdf_path: str, out_dir: str, dpi: int = DPI) -> List[str]:
    os.makedirs(out_dir, exist_ok=True)
    with pdfplumber.open(pdf_path) as pdf:
        pngs = []
        base = os.path.splitext(os.path.basename(pdf_path))[0]
        for i, page in enumerate(pdf.pages, start=1):
            if _cancelled():
                raise KeyboardInterrupt("cancelado pelo usuário")
            im = page.to_image(resolution=dpi).original
            out = os.path.join(out_dir, f"{base}_p{i:02d}.png")
            im.save(out, format="PNG", optimize=True)
            pngs.append(out)
    return pngs

# ================== LOG helpers ==================

def _box_title(text: str) -> str:
    line = "═" * max(10, len(text) + 2)
    return f"\n{line}\n {text}\n{line}\n"

def p_info(msg: str): print(f"[INFO] {msg}", flush=True)
def p_ok(msg: str):   print(f"[OK]   {msg}", flush=True)
def p_warn(msg: str): print(f"[AVISO]{msg}", flush=True)
def p_err(msg: str):  print(f"[ERRO] {msg}", flush=True)

# ================== LOTES ==================

def iter_lotes(base_dir: str) -> List[Tuple[str, str, List[str]]]:
    lots: List[Tuple[str, str, List[str]]] = []
    subdirs = [d for d in glob.glob(os.path.join(base_dir, "*")) if os.path.isdir(d)]
    for d in sorted(subdirs):
        pdfs = sorted([f for f in os.listdir(d) if f.lower().endswith(".pdf")])
        if pdfs:
            lots.append((os.path.basename(d), d, pdfs))
    base_pdfs = sorted([f for f in os.listdir(base_dir) if f.lower().endswith(".pdf")])
    if base_pdfs:
        lots.append((os.path.basename(base_dir), base_dir, base_pdfs))
    return lots

# ================== PIPELINE PRINCIPAL ==================

def main():
    lots = iter_lotes(BASE_DIR)
    if _cancelled():
        print("⛔ Cancelado antes de iniciar.", flush=True); return
    if not lots:
        print(f"Nenhum PDF encontrado em '{BASE_DIR}' ou subpastas.", flush=True); sys.exit(0)

    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        print("[ERRO] Defina sua OPENAI_API_KEY (na UI, campo 'OpenAI API Key').", flush=True); return

    for lot_name, lot_dir, pdfs in lots:
        if _cancelled():
            print(f"⛔ Cancelado no início do lote '{lot_name}'.", flush=True); return

        rows_lanc: List[Dict[str, str]] = []
        unknown_cities = set()
        log_path = _unique_path(str(Path(lot_dir) / f"{lot_name}_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"))
        with open(log_path, "w", encoding="utf-8") as log_f:
            log_f.write(f"== LOTE: {lot_name} ==\nPasta: {lot_dir}\nPDFs: {len(pdfs)}\n\n")

            for pdf_file in pdfs:
                if _cancelled():
                    print("⛔ Cancelado — interrompendo notas deste lote.", flush=True); return

                pdf_path = os.path.join(lot_dir, pdf_file)
                print(_box_title(f"Lendo • Lote: {lot_name} • Arquivo: {pdf_file}"), flush=True)
                p_info(f"Caminho: {pdf_path}")

                # 1) PDF → PNG(s)
                png_dir = os.path.join(lot_dir, "_png")
                try:
                    pngs = convert_pdf_to_pngs(pdf_path, png_dir, dpi=DPI)
                except Exception as e:
                    p_err(f"Falha ao converter '{pdf_file}': {e}")
                    log_f.write(f"[ERRO] Falha ao converter '{pdf_file}': {e}\n"); continue

                if not pngs:
                    p_warn(f"'{pdf_file}' sem páginas.")
                    log_f.write(f"[AVISO] '{pdf_file}' sem páginas.\n"); continue

                # 2) IA → JSON bruto
                data = ia_extract_from_images(pngs, api_key=api_key, model=OPENAI_MODEL)

                # 3) Formata resumo + validações
                if "_ia_erro" in data:
                    p_err(f"IA falhou: {data['_ia_erro']}"); log_f.write(f"[ERRO] IA: {data['_ia_erro']}\n"); continue

                resumo_lines = [
                    "RESUMO (IA IMG64 v2):",
                    f"  Tipo:            {data.get('tipo','')}",
                    f"  Nº Nota:         {data.get('numero','')}",
                    f"  Emissão:         {data.get('data_emissao','')}",
                    f"  Vencimento:      {data.get('data_vencimento','')}",
                    f"  Município:       {data.get('municipio','')}",
                    f"  Prestador:       {data.get('prestador','')}",
                    f"  CNPJ Prestador:  {data.get('cnpj_prestador','')}",
                    f"  Valor Total:     {data.get('valor_total','')}",
                ]
                if isinstance(data.get("fontes"), dict):
                    fontes = data["fontes"]
                    resumo_lines.append(f"  Fonte Vencimento: {fontes.get('data_vencimento','')}  Obs: {fontes.get('observacoes','')}")
                resumo_txt = "\n".join(resumo_lines)
                (print if not PRINT_ONLY_IA else print)(resumo_txt, flush=True)
                log_f.write(resumo_txt + "\n")

                # 4) Pós-processamento:
                final_date = choose_final_date_2025(data.get("data_vencimento",""), data.get("data_emissao",""))
                cnpj_out = sanitize_cnpj_prestador(data.get("prestador",""), data.get("cnpj_prestador",""))

                valor_cent = ""
                val = data.get("valor_total","")
                if val and (m := RE_VALOR.search(val)):
                    valor_cent = valor_to_centavos(m.group(1))

                numero_nf_digits = re.sub(r"[^\d]+", "", data.get("numero","") or "")

                cod_faz = "-"
                prest_norm = _norm_simple(data.get("prestador","")).upper()
                for k, v in FAZENDA2COD.items():
                    if k in prest_norm:
                        cod_faz = v; break

                municipio = (data.get("municipio","") or "").strip().upper()
                if municipio and municipio not in CODIGOS_CIDADES:
                    unknown_cities.add(municipio)

                historico = f"PAGAMENTO NF {numero_nf_digits or (data.get('numero','') or '-')} {(data.get('prestador','') or '').strip()}"

                rows_lanc.append({
                    "data": final_date or "",
                    "codfaz": cod_faz,
                    "conta": "001",
                    "numero": numero_nf_digits or (data.get("numero","") or ""),
                    "historico": historico,
                    "cnpj": cnpj_out if cnpj_out != "INVALIDO" else "INVALIDO",
                    "tipo": "2",
                    "padrao": "000",
                    "valor1": valor_cent or "",
                    "valor2": valor_cent or "",
                    "flag": "N",
                    "caminho": pdf_path,
                })

                print("·" * 72, flush=True); print("", flush=True)

            # 5) planilha por lote
            if rows_lanc:
                try:
                    xlsx_path = _unique_path(os.path.join(lot_dir, "lancamentos.xlsx"))
                    wb = Workbook(); ws = wb.active; ws.title = "lancamentos"
                    headers = ["Data","CodFazenda","Conta","NumeroNF","Historico","CNPJ","Tipo","Padrao","Valor1","Valor2","Flag","CaminhoNF"]
                    ws.append(headers)
                    for r in rows_lanc:
                        ws.append([
                            r["data"], r["codfaz"], r["conta"], r["numero"], r["historico"],
                            r["cnpj"], r["tipo"], r["padrao"], r["valor1"], r["valor2"], r["flag"], r["caminho"]
                        ])
                    end_row = ws.max_row; end_col = ws.max_column
                    ref = f"A1:{get_column_letter(end_col)}{end_row}"
                    tbl = Table(displayName="lancamentos_tbl", ref=ref)
                    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
                    ws.add_table(tbl)
                    widths = [12,12,8,14,60,22,6,8,12,12,6,60]
                    for i, wth in enumerate(widths, start=1):
                        ws.column_dimensions[get_column_letter(i)].width = wth
                    wb.save(xlsx_path)
                    print(f"[OK] Planilha criada (sem macro): {xlsx_path}", flush=True)
                except Exception as e:
                    print(f"[ERRO] Falha ao criar planilha: {e}", flush=True)

            if unknown_cities:
                print("[ATENÇÃO] Municípios sem mapeamento em CODIGOS_CIDADES:", flush=True)
                for c in sorted(unknown_cities):
                    print(f"  - {c}", flush=True)

# ================== UI / WORKERS (mantidos) ==================

ICON_PATH = _find_in_this_or_parent("icon-nfs.png")

STYLE_SHEET = """
QWidget#tab_automacao_nfs_digitalizadas {
  background-color: #0F1115; color: #E0E0E0;
  font-family: Segoe UI, Inter, Arial;
  font-size: 13px;
}
QFrame[class="card"] {
  background-color: #12151C;
  border: 1px solid #333845;
  border-radius: 12px;
}
QPushButton {
  background-color: #1e5a9c; border: none; color: white;
  padding: 8px 12px; border-radius: 10px;
}
QPushButton:hover { background-color: #2166b3; }
QPushButton:disabled { background-color: #3A3C3D; color: #888; }
QLineEdit, QSpinBox {
  background-color: #0f141b; color: #E0E0E0;
  border: 1px solid #333845; border-radius: 8px; padding: 6px 8px;
}
QTextEdit {
  background-color: #0f141b; color: #E0E0E0;
  border: 1px solid #333845; border-radius: 8px; padding: 8px;
}
QLabel[role="hint"] {
  color: #9aa4b2; font-size: 12px;
}
"""

class _EmittingTextIO(io.StringIO):
    def __init__(self, emit_fn):
        super().__init__()
        self.emit_fn = emit_fn
        self._buf = ""

    def write(self, s):
        self._buf += s
        while "\n" in self._buf:
            line, self._buf = self._buf.split("\n", 1)
            kind = "raw"
            t = line.strip()
            if t.startswith("[OK]"): kind = "success"
            elif t.startswith("[INFO]"): kind = "info"
            elif t.startswith("[ERRO]"): kind = "error"
            elif t.startswith("[AVISO]"): kind = "warning"
            self.emit_fn(line, kind)

    def flush(self):
        if self._buf:
            self.emit_fn(self._buf, "raw")
            self._buf = ""

class BaseWorker(QThread):
    log_sig = Signal(str, str)
    finished_sig = Signal(str)

    def __init__(self, cfg: dict):
        super().__init__()
        self.cfg = cfg
        self._cancel = False

    def cancel(self): self._cancel = True
    def _emit(self, msg: str, kind: str="info"): self.log_sig.emit(msg, kind)

    def _capture_prints(self, prefix: str = ""):
        class _Ctx:
            def __enter__(_self):
                _self.stream = _EmittingTextIO(lambda line, kind: self._emit(f"{prefix}{line}", kind))
                _self.old_out, _self.old_err = sys.stdout, sys.stderr
                sys.stdout = sys.stderr = _self.stream
                return _self.stream
            def __exit__(_self, exc_type, exc, tb):
                try: _self.stream.flush()
                except Exception: pass
                sys.stdout, sys.stderr = _self.old_out, _self.old_err
        return _Ctx()

class WorkerSeparar(BaseWorker):
    def run(self):
        try:
            src = self.cfg.get("separar.source_folder") or ""
            out = self.cfg.get("separar.separated_base_folder") or ""
            if not src or not out or not Path(src).exists():
                self._emit("Defina corretamente os caminhos do separador.", "warning")
                self.finished_sig.emit("Erro"); return

            mod_path = Path(__file__).parent / "Separador PDF Nota por Nota.py"
            if not mod_path.exists():
                self._emit("Arquivo 'Separador PDF Nota por Nota.py' não encontrado.", "error")
                self.finished_sig.emit("Erro"); return

            import runpy
            self._emit("Iniciando separação…", "title")
            with self._capture_prints():
                runpy.run_path(str(mod_path), run_name="__main__")
            self._emit("Separação concluída.", "success")
            self.finished_sig.emit("Concluído")
        except Exception as e:
            self._emit(f"Falha geral:\n{e}", "error")
            self.finished_sig.emit("Erro")

class WorkerPlanilha(BaseWorker):
    def run(self):
        try:
            base = self.cfg.get("planilha.base_dir") or ""
            if not base or not Path(base).exists():
                self._emit("Defina 'Caminho das NFS-e' nas Configurações.", "warning")
                self.finished_sig.emit("Erro"); return
            import importlib.util
            spec = importlib.util.spec_from_file_location("nfs_core", __file__)
            mod  = importlib.util.module_from_spec(spec); spec.loader.exec_module(mod)
            setattr(mod, "BASE_DIR", str(base))
            setattr(mod, "DPI", int(self.cfg.get("dpi") or 300))
            setattr(mod, "is_cancelled", lambda: self._cancel)
            os.environ["OPENAI_API_KEY"] = self.cfg.get("api_key","")
            self._emit("Gerando planilhas de lançamentos…", "title")
            with self._capture_prints():
                mod.main()
            self._emit("Geração de planilha finalizada.", "success")
            self.finished_sig.emit("Concluído")
        except Exception as e:
            self._emit(f"Falha geral ao gerar planilha:\n{e}", "error")
            self.finished_sig.emit("Erro")

class WorkerTxtEImport(BaseWorker):
    def run(self):
        try:
            xlsx = self.cfg.get("txt.xlsx_path") or ""
            if not xlsx or not Path(xlsx).exists():
                self._emit("Defina 'Caminho da Planilha' (lancamentos.xlsx).", "warning")
                self.finished_sig.emit("Erro"); return
            out_txt = str(Path(xlsx).with_name("saida_lancamentos.txt"))
            mod_path = Path(__file__).parent / "gerar_txt_lancamentos.py"
            if not mod_path.exists():
                self._emit("Arquivo 'gerar_txt_lancamentos.py' não encontrado.", "error")
                self.finished_sig.emit("Erro"); return
            import runpy
            self._emit("Gerando TXT…", "title")
            with self._capture_prints():
                runpy.run_path(str(mod_path), run_name="__main__")
            self._emit(f"TXT gerado em: {out_txt}", "success")
            self.finished_sig.emit(out_txt)
        except SystemExit as e:
            code = int(getattr(e, "code", 1) or 1)
            if code == 0: self.finished_sig.emit("Concluído")
            else:
                self._emit("Falha ao gerar TXT (script finalizou com erro).", "error")
                self.finished_sig.emit("Erro")
        except Exception as e:
            self._emit(f"Erro ao gerar/importar TXT:\n{e}", "error")
            self.finished_sig.emit("Erro")

class AutomacaoNFSDigitalizadasUI(QWidget):
    def __init__(self):
        super().__init__()
        self.setObjectName('tab_automacao_nfs_digitalizadas')
        self.setWindowTitle("Automação NFS-e Digitalizadas (IA IMG64 v2)")
        icon = _find_in_this_or_parent("icon-nfs.png")
        if icon.exists():
            self.setWindowIcon(QIcon(str(icon)))
        self.setStyleSheet(STYLE_SHEET)

        self.cfg = {
            "api_key": os.environ.get("OPENAI_API_KEY", ""),
            "planilha.base_dir": "",
            "txt.xlsx_path": "",
            "separar.source_folder": "",
            "separar.separated_base_folder": "",
            "dpi": DPI,
        }

        root = QVBoxLayout(self); root.setContentsMargins(14,14,14,14); root.setSpacing(12)
        hdr = QHBoxLayout()
        title = QLabel("Automação NFS-e Digitalizadas — IA (IMG base64) v2")
        title.setStyleSheet("font-size:18px; font-weight:700; color:#E0E0E0;")
        hdr.addWidget(title, 1)
        root.addLayout(hdr)

        top = QFrame(); top.setProperty("class", "card")
        lay = QVBoxLayout(top); lay.setContentsMargins(14,12,14,12); lay.setSpacing(10)

        grid = QGridLayout(); grid.setHorizontalSpacing(8); grid.setVerticalSpacing(8)
        grid.addWidget(QLabel("OpenAI API Key:"), 0, 0)
        self.ed_api = QLineEdit(self.cfg["api_key"]); self.ed_api.setEchoMode(QLineEdit.Password)
        grid.addWidget(self.ed_api, 0, 1, 1, 3)

        grid.addWidget(QLabel("Caminho das NFS-e:"), 1, 0)
        self.ed_base = QLineEdit(self.cfg["planilha.base_dir"])
        btn_sel_base = QPushButton("Selecionar…"); btn_sel_base.clicked.connect(self._pick_base_dir)
        grid.addWidget(self.ed_base, 1, 1, 1, 2); grid.addWidget(btn_sel_base, 1, 3)

        grid.addWidget(QLabel("Caminho da Planilha (lancamentos.xlsx):"), 2, 0)
        self.ed_xlsx = QLineEdit(self.cfg["txt.xlsx_path"])
        btn_sel_xlsx = QPushButton("Selecionar…"); btn_sel_xlsx.clicked.connect(self._pick_xlsx)
        grid.addWidget(self.ed_xlsx, 2, 1, 1, 2); grid.addWidget(btn_sel_xlsx, 2, 3)

        grid.addWidget(QLabel("DPI (PDF→PNG):"), 3, 0)
        self.sp_dpi = QSpinBox(); self.sp_dpi.setRange(72, 800); self.sp_dpi.setValue(self.cfg["dpi"])
        grid.addWidget(self.sp_dpi, 3, 1)
        lay.addLayout(grid)

        actions = QHBoxLayout()
        self.btn_separar = QPushButton("📄 Separar Nota por Nota")
        self.btn_planilha = QPushButton("📊 Gerar planilha NFS-e")
        self.btn_txt      = QPushButton("📥 Importar TXT da NFS-e")
        actions.addWidget(self.btn_separar); actions.addWidget(self.btn_planilha); actions.addWidget(self.btn_txt)
        lay.addLayout(actions)

        self.log = QTextEdit(); self.log.setReadOnly(True); self.log.setMinimumHeight(260)
        lay.addWidget(self.log)
        root.addWidget(top)

        self.btn_separar.clicked.connect(self._do_separar)
        self.btn_planilha.clicked.connect(self._do_planilha)
        self.btn_txt.clicked.connect(self._do_txt)

        self._worker: Optional[BaseWorker] = None

    def _pick_base_dir(self):
        d = QFileDialog.getExistingDirectory(self, "Escolha a pasta com os PDFs/Lotes")
        if d: self.ed_base.setText(d)

    def _pick_xlsx(self):
        file,_ = QFileDialog.getOpenFileName(self, "Selecione o 'lancamentos.xlsx'", filter="Excel (*.xlsx)")
        if file: self.ed_xlsx.setText(file)

    def _do_separar(self):
        if self._worker and self._worker.isRunning():
            QMessageBox.warning(self, "Aguarde", "Já existe um processo em execução."); return
        self._sync_cfg()
        self._worker = WorkerSeparar(self.cfg)
        self._worker.log_sig.connect(self._log_msg); self._worker.finished_sig.connect(self._finish)
        self._worker.start()

    def _do_planilha(self):
        if self._worker and self._worker.isRunning():
            QMessageBox.warning(self, "Aguarde", "Já existe um processo em execução."); return
        self._sync_cfg()
        self._worker = WorkerPlanilha(self.cfg)
        self._worker.log_sig.connect(self._log_msg); self._worker.finished_sig.connect(self._finish)
        self._worker.start()

    def _do_txt(self):
        if self._worker and self._worker.isRunning():
            QMessageBox.warning(self, "Aguarde", "Já existe um processo em execução."); return
        self._sync_cfg()
        self._worker = WorkerTxtEImport(self.cfg)
        self._worker.log_sig.connect(self._log_msg); self._worker.finished_sig.connect(self._finish)
        self._worker.start()

    def _finish(self, msg: str):
        self._log_msg(f"Concluído: {msg}", "success")
        self._worker = None

    def _log_msg(self, msg: str, kind: str="info"):
        now = datetime.now().strftime("%H:%M:%S")
        palette = {
            "info":    {"emoji":"ℹ️","color":"#E0E0E0"},
            "success": {"emoji":"✅","color":"#C8FACC"},
            "warning": {"emoji":"⚠️","color":"#FFD580"},
            "error":   {"emoji":"❌","color":"#FF9CA3"},
            "raw":     {"emoji":"","color":"#B0B0B0"},
            "title":   {"emoji":"📌","color":"#FFFFFF"},
        }
        p = palette.get(kind, palette["info"])
        lead = (p["emoji"] + " ") if p["emoji"] else ""
        self.log.append(f'<div style="border-left:3px solid #3A3C3D; padding:6px 10px; margin:2px 0;">'
                        f'<span style="opacity:.7; font-family:monospace;">[{now}]</span> '
                        f'{lead}<span style="color:{p["color"]}; white-space:pre-wrap;">{msg}</span></div>')
        sb = self.log.verticalScrollBar()
        if sb: sb.setValue(sb.maximum())

    def _sync_cfg(self):
        self.cfg["api_key"] = self.ed_api.text().strip()
        self.cfg["planilha.base_dir"] = self.ed_base.text().strip()
        self.cfg["txt.xlsx_path"] = self.ed_xlsx.text().strip()
        self.cfg["dpi"] = int(self.sp_dpi.value())

# --------- ENTRYPOINT ---------
if __name__ == "__main__":
    if len(sys.argv) == 1:
        app = QApplication(sys.argv)
        ui = AutomacaoNFSDigitalizadasUI()
        ui.resize(980, 680)
        ui.show()
        sys.exit(app.exec())
    else:
        if len(sys.argv) >= 2 and os.path.isdir(sys.argv[1]):
            BASE_DIR = sys.argv[1]
        main()
